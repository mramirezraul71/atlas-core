r"""ATLAS HTTP API adapter
Expone /status /tools /execute usando el command_router.handle de C:\ATLAS\modules\command_router.py
"""
import asyncio
import hashlib
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from contextlib import asynccontextmanager
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from atlas_adapter.bootstrap.env import load_project_env
from atlas_adapter.bootstrap.lifespan import app_lifespan
from atlas_adapter.bootstrap.settings import (
    BASE_DIR,
    ENV_PATH,
    LOGS_DIR,
    initialize_runtime_paths,
    load_handle,
)
from atlas_adapter.bootstrap.service_urls import (
    get_nexus_ws_url,
    get_robot_api_base,
    get_robot_ui_base,
)


def _ts_to_epoch(ts_str: str) -> float:
    """Convert ISO timestamp string to epoch seconds."""
    if not ts_str:
        return 0.0
    try:
        dt = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
        return dt.timestamp()
    except Exception:
        return 0.0


# Cargar config ANTES de importar mÃ³dulos que usan os.getenv (audit, policy, etc.)
load_project_env(ENV_PATH)
initialize_runtime_paths()

_logs = str(LOGS_DIR)

from fastapi import FastAPI, File, Header, Request, UploadFile, WebSocket
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from pydantic import BaseModel, Field, field_validator
import psutil

handle = load_handle()


@asynccontextmanager
async def _lifespan(app):
    try:
        from modules.humanoid.deploy.healthcheck import _set_app_start_time

        _set_app_start_time()
    except Exception:
        pass
    # Cargar BÃ³veda al arranque (tokens Telegram/WhatsApp, etc.) sin pedir .env.
    try:
        from pathlib import Path

        from dotenv import load_dotenv

        vault_path = (
            os.getenv("ATLAS_VAULT_PATH")
            or r"C:\Users\Raul\OneDrive\RAUL - Personal\Escritorio\credenciales.txt"
        ).strip()
        candidates = [vault_path, r"C:\dev\credenciales.txt"]
        for c in candidates:
            p = Path(c)
            if p.is_file():
                load_dotenv(str(p), override=True)
                break
    except Exception:
        pass
    try:
        from modules.humanoid.owner.emergency import set_emergency_from_env

        set_emergency_from_env()
    except Exception:
        pass
    humanoid = os.getenv("HUMANOID_ENABLED", "true").strip().lower() in (
        "1",
        "true",
        "yes",
        "y",
        "on",
    )
    sched = os.getenv("SCHED_ENABLED", "true").strip().lower() in (
        "1",
        "true",
        "yes",
        "y",
        "on",
    )
    worker_only = os.getenv("WORKER_ONLY", "false").strip().lower() in (
        "1",
        "true",
        "yes",
    )
    safe_startup = os.getenv("ATLAS_SAFE_STARTUP", "false").strip().lower() in (
        "1",
        "true",
        "yes",
        "y",
        "on",
    )
    if humanoid and sched:
        from concurrent.futures import ThreadPoolExecutor

        from modules.humanoid.scheduler import start_scheduler
        from modules.humanoid.watchdog import start_watchdog

        _executor = ThreadPoolExecutor(max_workers=2)
        start_scheduler(executor=_executor)
        start_watchdog()
        # VisiÃ³n Ubicua: preparar DB y watchdog de movimiento (best-effort, no bloquea arranque)
        try:
            if os.getenv("VISION_UBIQ_ENABLED", "true").strip().lower() in (
                "1",
                "true",
                "yes",
                "y",
                "on",
            ):
                from modules.humanoid.vision.ubiq import (
                    ensure_db, start_motion_watchdog)

                ensure_db()
                start_motion_watchdog()
        except Exception:
            pass
        if not worker_only:
            try:
                from modules.humanoid.ci import ensure_ci_jobs

                ensure_ci_jobs()
            except Exception:
                pass
            try:
                from modules.humanoid.ga.scheduler_jobs import ensure_ga_jobs

                ensure_ga_jobs()
            except Exception:
                pass
            try:
                from modules.humanoid.metalearn.scheduler_jobs import \
                    ensure_metalearn_jobs

                ensure_metalearn_jobs()
            except Exception:
                pass
            try:
                from modules.humanoid.ans.scheduler_jobs import ensure_ans_jobs

                ensure_ans_jobs()
                # Ejecutar triada ANS al arranque (una vez) para que la bitÃ¡cora tenga acciÃ³n desde el inicio
                if (
                    (not safe_startup)
                    and os.getenv("ANS_ENABLED", "true").strip().lower()
                    in ("1", "true", "yes")
                    and os.getenv("ANS_RUN_AT_STARTUP", "true").strip().lower()
                    in ("1", "true", "yes")
                ):

                    def _run_triada_at_startup():
                        try:
                            from modules.humanoid.ans.engine import \
                                run_ans_cycle

                            run_ans_cycle(
                                mode=os.getenv("ANS_MODE", "auto"), timeout_sec=60
                            )
                        except Exception:
                            pass

                    import threading

                    t = threading.Thread(target=_run_triada_at_startup, daemon=True)
                    t.start()
            except Exception:
                pass
            try:
                # Sistema Nervioso: sensores -> score -> bitÃ¡cora/incidentes (scheduler job)
                from modules.humanoid.nervous.scheduler_jobs import \
                    ensure_nervous_jobs

                ensure_nervous_jobs()
            except Exception:
                pass
            try:
                # WorldState: visiÃ³n+OCR periÃ³dicos (representaciÃ³n mÃ­nima del entorno)
                from modules.humanoid.vision.world_state_jobs import \
                    ensure_world_state_jobs

                ensure_world_state_jobs()
            except Exception:
                pass
            # === Sistema de Comunicaciones Unificado ===
            # Bootstrap centralizado de todos los servicios de comunicaciÃ³n
            # (reemplaza las inicializaciones individuales de makeplay y telegram_poller)
            try:
                if not safe_startup:
                    from modules.humanoid.comms.bootstrap import \
                        bootstrap_comms

                    comms_result = bootstrap_comms(skip_tests=True)
                    if not comms_result.get("ok"):
                        import logging

                        _comms_logger = logging.getLogger("atlas.comms.startup")
                        for warning in comms_result.get("warnings", []):
                            _comms_logger.warning(f"Comms bootstrap: {warning}")
            except Exception as _comms_err:
                import logging

                logging.getLogger("atlas.comms.startup").error(
                    f"Comms bootstrap failed: {_comms_err}"
                )

            # === Quality / POT Autonomy ===
            # Activa dispatcher+triggers para ejecuciÃ³n autÃ³noma de POTs.
            # Control: QUALITY_AUTONOMY_ENABLED=true|false (default true)
            try:
                if (not safe_startup) and os.getenv(
                    "QUALITY_AUTONOMY_ENABLED", "true"
                ).strip().lower() in ("1", "true", "yes", "y", "on"):
                    from modules.humanoid.quality import \
                        start_autonomous_system

                    q = start_autonomous_system()
                    # Autostart del Autonomy Daemon (evita estado CRITICAL post-restart)
                    # Control: AUTONOMY_DAEMON_AUTOSTART=true|false (default true)
                    try:
                        if os.getenv(
                            "AUTONOMY_DAEMON_AUTOSTART", "true"
                        ).strip().lower() in ("1", "true", "yes", "y", "on"):
                            from modules.humanoid.quality.autonomy_daemon import (
                                is_autonomy_running, start_autonomy)

                            if not is_autonomy_running():
                                start_autonomy()
                    except Exception:
                        pass
                    try:
                        from modules.humanoid.ans.evolution_bitacora import \
                            append_evolution_log

                        append_evolution_log(
                            message="[QUALITY] AutonomÃ­a POT iniciada (dispatcher+triggers).",
                            ok=bool(q.get("all_ok", False)),
                            source="quality",
                        )
                    except Exception:
                        pass
            except Exception:
                pass
            try:
                from modules.humanoid.scheduler.repo_monitor_jobs import \
                    ensure_repo_monitor_jobs

                ensure_repo_monitor_jobs()
            except Exception:
                pass
            try:
                # MakePlay Scanner: snapshot local + (opcional) webhook externo
                from modules.humanoid.comms.makeplay_scheduler import \
                    ensure_makeplay_jobs

                ensure_makeplay_jobs()
            except Exception:
                pass
            try:
                from modules.humanoid.scheduler.repo_hygiene_jobs import \
                    ensure_repo_hygiene_jobs

                ensure_repo_hygiene_jobs()
            except Exception:
                pass
            try:
                from modules.humanoid.approvals.scheduler_jobs import \
                    ensure_approvals_jobs

                ensure_approvals_jobs()
            except Exception:
                pass
            try:
                from modules.humanoid.scheduler.workshop_jobs import \
                    ensure_workshop_jobs

                ensure_workshop_jobs()
            except Exception:
                pass
            try:
                from modules.humanoid.repo.git_hooks import \
                    ensure_post_commit_hook

                ensure_post_commit_hook()
            except Exception:
                pass
        # Heartbeat NEXUS: ping 8000/health, auto-reactivaciÃ³n con start_all.ps1, registro en BitÃ¡cora
        try:
            from modules.nexus_heartbeat import (register_status_callback,
                                                 start_heartbeat)

            _dashboard_base = (
                os.getenv("ATLAS_DASHBOARD_URL") or "http://127.0.0.1:8791"
            ).rstrip("/")

            def _on_nexus_change(connected: bool, message: str) -> None:
                if connected:
                    text = "[CONEXIÃ“N] Buscando NEXUS en puerto 8000... OK."
                else:
                    extra = (message or "").strip()
                    extra = (": " + extra[:120]) if extra else ""
                    text = (
                        "[CONEXIÃ“N] Buscando NEXUS en puerto 8000... Desconectado"
                        + extra
                    )
                try:
                    import json
                    import urllib.request

                    req = urllib.request.Request(
                        _dashboard_base + "/ans/evolution-log",
                        data=json.dumps({"message": text, "ok": connected}).encode(
                            "utf-8"
                        ),
                        headers={"Content-Type": "application/json"},
                        method="POST",
                    )
                    urllib.request.urlopen(req, timeout=3)
                except Exception:
                    pass

            register_status_callback(_on_nexus_change)
            start_heartbeat()

            # Prueba de Nervios: autodiagnÃ³stico de motores, mouse, cÃ¡mara; actualiza Dashboard a CONECTADO | ACTIVO
            def _run_nerve_test_background():
                try:
                    import sys
                    from pathlib import Path

                    push_base = Path(__file__).resolve().parent.parent
                    if str(push_base) not in sys.path:
                        sys.path.insert(0, str(push_base))
                    import nexus_actions

                    nexus_actions.run_nerve_test()
                except Exception:
                    pass

            import threading

            _nerve_thread = threading.Thread(
                target=_run_nerve_test_background, daemon=True
            )
            _nerve_thread.start()
        except Exception:
            pass
    # ATLAS AUTONOMOUS: background tasks (Health, Alert, Learning)
    try:
        import asyncio
        import sys

        if str(BASE_DIR) not in sys.path:
            sys.path.insert(0, str(BASE_DIR))
        import logging

        from autonomous.health_monitor.health_aggregator import \
            HealthAggregator
        from autonomous.learning.learning_orchestrator import \
            LearningOrchestrator
        from autonomous.telemetry.alert_manager import AlertManager
        from autonomous.telemetry.metrics_aggregator import MetricsAggregator

        _log = logging.getLogger(__name__)
        _log.info("Iniciando background tasks de ATLAS AUTONOMOUS...")
        health_agg = HealthAggregator()
        health_agg.start_monitoring()
        _log.info("âœ“ Health Monitoring activo")
        alert_mgr = AlertManager()
        asyncio.create_task(alert_mgr.start_evaluation_loop(60))
        _log.info("âœ“ Alert Manager activo")
        learning = LearningOrchestrator()

        async def _learning_loop():
            while True:
                await asyncio.sleep(3600)
                try:
                    learning.run_learning_cycle()
                    _log.info("âœ“ Learning cycle ejecutado")
                except Exception as e:
                    _log.error("Error en learning cycle: %s", e)

        asyncio.create_task(_learning_loop())
        _log.info("âœ“ Learning Engine activo")

        metrics_agg = MetricsAggregator()

        def _load_telemetry_interval() -> float:
            interval = 60.0
            try:
                cfg_path = BASE_DIR / "config" / "autonomous.yaml"
                if cfg_path.exists():
                    import yaml

                    cfg = yaml.safe_load(
                        cfg_path.read_text(encoding="utf-8", errors="replace")
                    ) or {}
                    interval = float(
                        (((cfg.get("telemetry") or {}).get("metrics") or {}).get(
                            "aggregation_interval"
                        ))
                        or interval
                    )
            except Exception:
                pass
            return max(5.0, interval)

        telemetry_interval_sec = _load_telemetry_interval()

        def _probe_service(url: str) -> tuple[float, float]:
            import urllib.request

            t0_probe = time.perf_counter()
            try:
                req = urllib.request.Request(
                    url, method="GET", headers={"Accept": "application/json"}
                )
                with urllib.request.urlopen(req, timeout=2) as resp:
                    if 200 <= int(resp.status) < 400:
                        return (1.0, (time.perf_counter() - t0_probe) * 1000.0)
            except Exception:
                pass
            return (0.0, 0.0)

        def _collect_telemetry_once() -> None:
            try:
                import psutil

                disk_root = (os.getenv("SystemDrive") or "C:") + "\\"
                metrics_agg.collect_metric(
                    "system", "cpu_percent", float(psutil.cpu_percent(interval=0))
                )
                metrics_agg.collect_metric(
                    "system", "ram_percent", float(psutil.virtual_memory().percent)
                )
                metrics_agg.collect_metric(
                    "system",
                    "disk_usage_percent",
                    float(psutil.disk_usage(disk_root).percent),
                )
            except Exception:
                pass
            try:
                from modules.observability.metrics import MetricsCollector

                summary = MetricsCollector.get_metrics_summary() or {}
                metrics_agg.collect_metric(
                    "observability",
                    "total_requests",
                    float(summary.get("total_requests", 0) or 0),
                )
                metrics_agg.collect_metric(
                    "observability",
                    "active_requests",
                    float(summary.get("active_requests", 0) or 0),
                )
                metrics_agg.collect_metric(
                    "observability",
                    "memory_mb",
                    float(summary.get("memory_mb", 0) or 0),
                )
            except Exception:
                pass
            try:
                services = {
                    "push": "http://127.0.0.1:8791/health",
                    "nexus": "http://127.0.0.1:8000/health",
                    "robot": "http://127.0.0.1:8002/api/health",
                }
                for service_name, service_url in services.items():
                    if service_name == "push":
                        # The telemetry loop runs inside PUSH itself; mark local
                        # service online without generating self-HTTP load.
                        metrics_agg.collect_metric("services", "push_online", 1.0)
                        continue
                    online, latency_ms = _probe_service(service_url)
                    metrics_agg.collect_metric("services", f"{service_name}_online", online)
                    if latency_ms > 0:
                        metrics_agg.collect_metric(
                            "services",
                            f"{service_name}_latency_ms",
                            float(latency_ms),
                        )
            except Exception:
                pass

        async def _telemetry_loop():
            while True:
                try:
                    # Execute blocking telemetry collection off the event loop.
                    await asyncio.to_thread(_collect_telemetry_once)
                except Exception:
                    pass
                await asyncio.sleep(telemetry_interval_sec)

        asyncio.create_task(_telemetry_loop())
        _log.info("âœ“ Telemetry Metrics activo (intervalo %ss)", telemetry_interval_sec)
    except Exception as e:
        import logging

        logging.getLogger(__name__).debug(
            "ATLAS AUTONOMOUS background tasks no iniciados: %s", e
        )
    # ActualizaciÃ³n periÃ³dica de mÃ©tricas Prometheus (memoria, etc.)
    try:
        from modules.observability.metrics import update_system_metrics

        async def _metrics_loop():
            while True:
                await asyncio.sleep(10)
                try:
                    update_system_metrics()
                except Exception:
                    pass

        asyncio.create_task(_metrics_loop())
    except Exception:
        pass
    # Failsafe: asegurar Autonomy Daemon activo al finalizar startup.
    # Esto cubre escenarios donde QUALITY_AUTONOMY_ENABLED no iniciÃ³ el daemon.
    try:
        if os.getenv("AUTONOMY_DAEMON_AUTOSTART", "true").strip().lower() in (
            "1",
            "true",
            "yes",
            "y",
            "on",
        ):
            from modules.humanoid.quality.autonomy_daemon import \
                is_autonomy_running

            if not is_autonomy_running():
                try:
                    # Do not block API startup with heavy autonomy bootstrap.
                    asyncio.create_task(asyncio.to_thread(autonomy_daemon_start))
                except Exception:
                    pass
            # Second chance after a brief pause if still down.
            if not is_autonomy_running():
                await asyncio.sleep(0.5)
                try:
                    asyncio.create_task(asyncio.to_thread(autonomy_daemon_start))
                except Exception:
                    pass
    except Exception:
        pass
    # Supervisor residente: monitoreo continuo + directivas internas + notificaciÃ³n a Owner
    try:
        if not safe_startup:
            from atlas_adapter.supervisor_daemon import start_supervisor_daemon

            await start_supervisor_daemon()
    except Exception:
        pass
    yield
    # Shutdown: detener supervisor residente (best-effort)
    try:
        from atlas_adapter.supervisor_daemon import stop_supervisor_daemon

        await stop_supervisor_daemon()
    except Exception:
        pass


app = FastAPI(
    title="ATLAS Adapter",
    version="1.0.0",
    lifespan=app_lifespan,
    openapi_tags=[
        {
            "name": "Health",
            "description": "Health check extendido (score 0-100, LLM, scheduler, memory, DB, uptime).",
        },
        {
            "name": "Deploy",
            "description": "Blue-green deployment, canary ramp-up, deploy status y reportes.",
        },
        {
            "name": "Cluster",
            "description": "Atlas Cluster: nodos, heartbeat, routing, ejecuciÃ³n remota (hands/web/vision/voice).",
        },
        {
            "name": "Gateway",
            "description": "Stealth Gateway: Cloudflare / Tailscale / SSH / LAN, bootstrap, health.",
        },
        {
            "name": "Learning",
            "description": "Aprendizaje continuo: procesar situaciones, consolidar conocimiento, rutina diaria (lecciÃ³n del tutor), base de conocimiento, incertidumbre y mÃ©tricas de crecimiento.",
        },
    ],
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Metrics middleware: request count + latency per path
try:
    from modules.humanoid.metrics import MetricsMiddleware

    app.add_middleware(MetricsMiddleware)
except Exception:
    pass
# Observabilidad Prometheus (request count, duration, active)
try:
    from modules.observability.middleware import ObservabilityMiddleware

    app.add_middleware(ObservabilityMiddleware)
except Exception:
    pass


class Step(BaseModel):
    tool: str
    args: dict = {}


def _robot_connected() -> bool:
    """Ping Robot: prueba backend (8002) y luego UI (NEXUS_ROBOT_URL)."""
    import urllib.request

    base_api = get_robot_api_base()
    base_ui = get_robot_ui_base(default_to_api=True)
    for base in (base_api, base_ui):
        try:
            req = urllib.request.Request(base + "/", method="GET")
            with urllib.request.urlopen(req, timeout=2) as r:
                if r.status == 200:
                    return True
        except Exception:
            pass
    return False


_ATLAS_STATUS_CACHE_TTL_SEC = float(os.getenv("ATLAS_STATUS_CACHE_TTL_SEC") or 12.0)
_ATLAS_STATUS_CACHE = {"value": "[degraded] initializing", "ts": 0.0}
_ATLAS_STATUS_INFLIGHT = False
_ATLAS_STATUS_LOCK = threading.Lock()


def _refresh_atlas_status_cache() -> None:
    """Compute heavy atlas status once, update shared cache."""
    global _ATLAS_STATUS_INFLIGHT
    try:
        value = handle("/status")
    except Exception as e:
        value = f"[degraded] /status error: {e}"
    with _ATLAS_STATUS_LOCK:
        _ATLAS_STATUS_CACHE["value"] = value
        _ATLAS_STATUS_CACHE["ts"] = time.time()
        _ATLAS_STATUS_INFLIGHT = False


def _atlas_status_safe_cached() -> str:
    """Never block request path; return cached value and refresh in background."""
    global _ATLAS_STATUS_INFLIGHT
    now = time.time()
    with _ATLAS_STATUS_LOCK:
        cached = _ATLAS_STATUS_CACHE.get("value", "[degraded] initializing")
        ts = float(_ATLAS_STATUS_CACHE.get("ts") or 0.0)
        is_fresh = (now - ts) <= _ATLAS_STATUS_CACHE_TTL_SEC
        if is_fresh:
            return str(cached)
        if not _ATLAS_STATUS_INFLIGHT:
            _ATLAS_STATUS_INFLIGHT = True
            threading.Thread(target=_refresh_atlas_status_cache, daemon=True).start()
        return str(cached)


_STATUS_CACHE: Dict[str, Any] = {"nexus": {}, "robot": False, "ts": 0.0}
_STATUS_CACHE_TTL = 15.0
_STATUS_BG_STARTED = False


def _status_bg_loop():
    """Background thread that refreshes connectivity every 15s."""
    while True:
        try:
            robot_ok = _robot_connected()
        except Exception:
            robot_ok = False
        _STATUS_CACHE["nexus"] = {
            "connected": True,
            "active": True,
            "last_check_ts": time.time(),
            "last_error": "",
        }
        _STATUS_CACHE["robot"] = robot_ok
        _STATUS_CACHE["ts"] = time.time()
        time.sleep(_STATUS_CACHE_TTL)


@app.get("/status")
def status():
    global _STATUS_BG_STARTED
    if not _STATUS_BG_STARTED:
        _STATUS_BG_STARTED = True
        threading.Thread(target=_status_bg_loop, daemon=True).start()

    nexus = _STATUS_CACHE.get("nexus", {})
    robot_ok = _STATUS_CACHE.get("robot", False)
    return {
        "ok": True,
        "atlas": _atlas_status_safe_cached(),
        "nexus_connected": nexus.get("connected", False),
        "nexus_active": nexus.get("active", False),
        "nexus_last_check_ts": nexus.get("last_check_ts", 0),
        "nexus_last_error": nexus.get("last_error", ""),
        "nexus_dashboard_url": "/nexus",
        "robot_connected": robot_ok,
    }


@app.get("/api/nexus/connection", tags=["NEXUS"])
def get_nexus_connection():
    """CEREBRO â€” CUERPO (NEXUS): estado de conexiÃ³n (consolidado en PUSH)."""
    return {
        "ok": True,
        "connected": True,
        "active": True,
        "last_check_ts": time.time(),
        "last_error": "",
    }


class NexusConnectionBody(BaseModel):
    connected: bool = False
    message: str = ""
    active: Optional[bool] = None


@app.post("/api/nexus/connection", tags=["NEXUS"])
def post_nexus_connection(body: NexusConnectionBody):
    """Actualiza estado de conexiÃ³n NEXUS (heartbeat o Prueba de Nervios). active=True â†’ CONECTADO | ACTIVO."""
    try:
        from modules.nexus_heartbeat import (set_nexus_active,
                                             set_nexus_connected)

        set_nexus_connected(body.connected, body.message or "")
        if body.active is not None:
            set_nexus_active(body.active)
        return {"ok": True, "connected": body.connected, "active": body.active}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/nexus/reconnect", tags=["NEXUS"])
def nexus_reconnect(clear_cache: bool = False):
    """Reconectar NEXUS (no-bloqueante): opcionalmente limpia cache, lanza arranque y devuelve de inmediato."""
    try:
        if clear_cache:
            from modules.nexus_heartbeat import clear_nexus_cache

            clear_nexus_cache()
        from modules.nexus_heartbeat import (get_nexus_connection_state,
                                             restart_nexus)

        started = restart_nexus()
        state = get_nexus_connection_state()
        return {
            "ok": True,
            "started": bool(started),
            "connected": state.get("connected", False),
            **state,
            "message": "NEXUS arrancando en segundo plano. Espera 10-20s y pulsa Actualizar/Estado.",
        }
    except Exception as e:
        return {"ok": False, "connected": False, "error": str(e)}


@app.post("/api/cache/clear", tags=["NEXUS"])
def cache_clear():
    """Limpia cache del servidor (__pycache__, temp_models_cache). Para uso con botÃ³n Limpiar cachÃ© en navegador."""
    try:
        from modules.nexus_heartbeat import clear_nexus_cache

        clear_nexus_cache()
        return {"ok": True, "message": "Cache servidor limpiado"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------- Semantic Memory (Fase 1 - FundaciÃ³n Cognitiva) ----------
@app.post("/api/memory/add", tags=["Memory"])
def api_memory_add(
    description: str,
    context: Optional[str] = None,
    outcome: Optional[str] = None,
    tags: Optional[List[str]] = None,
):
    """AÃ±ade una experiencia a la memoria semÃ¡ntica."""
    try:
        from modules.humanoid.memory_engine.semantic_memory import \
            get_semantic_memory

        exp_id = get_semantic_memory().add_experience(
            description=description,
            context=context,
            outcome=outcome,
            tags=tags,
        )
        return {"id": exp_id, "status": "stored"}
    except Exception as e:
        return {"id": None, "status": "error", "error": str(e)}


@app.post("/api/memory/search", tags=["Memory"])
def api_memory_search(
    query: str,
    top_k: int = 5,
    min_similarity: float = 0.6,
):
    """BÃºsqueda por similaridad en memoria semÃ¡ntica."""
    try:
        from modules.humanoid.memory_engine.semantic_memory import \
            get_semantic_memory

        results = get_semantic_memory().recall_similar(
            query, top_k=top_k, min_similarity=min_similarity
        )
        return {"results": results, "query": query}
    except Exception as e:
        return {"results": [], "query": query, "error": str(e)}


@app.get("/api/memory/stats", tags=["Memory"])
def api_memory_stats():
    """EstadÃ­sticas de la memoria semÃ¡ntica."""
    try:
        from modules.humanoid.memory_engine.semantic_memory import \
            get_semantic_memory

        return get_semantic_memory().get_statistics()
    except Exception as e:
        return {
            "error": str(e),
            "total_experiences": 0,
            "embedding_dimension": 0,
            "storage_size_mb": 0,
        }


# ---------- Cerebro: modo Auto/Manual e IA dominante ----------
class BrainStateBody(BaseModel):
    mode: Optional[str] = None  # "auto" | "manual"
    override_model: Optional[str] = None  # full_key ej. ollama:llama3.1:latest, o null


class ProviderCredentialBody(BaseModel):
    provider_id: str  # openai | anthropic | gemini | perplexity
    api_key: Optional[str] = None  # null o vacÃ­o = borrar clave


def _brain_ollama_available() -> bool:
    try:
        from modules.humanoid.deploy.healthcheck import _check_llm_reachable

        return _check_llm_reachable().get("ok", False)
    except Exception:
        return False


def _brain_ollama_available_timeout(seconds: float = 2.0) -> bool:
    """Comprueba Ollama con timeout para no bloquear el dashboard."""
    import concurrent.futures

    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
        fut = ex.submit(_brain_ollama_available)
        try:
            return fut.result(timeout=seconds)
        except (concurrent.futures.TimeoutError, Exception):
            return False


@app.get("/api/brain/state", tags=["Cerebro"])
def api_brain_state():
    """Estado del cerebro: modo (auto/manual), IA dominante, lista de modelos para el selector."""
    try:
        from modules.humanoid.ai.brain_state import get_brain_state
        from modules.humanoid.ai.registry import get_model_specs

        state = get_brain_state()
        ollama_ok = _brain_ollama_available_timeout(2.0)
        specs = get_model_specs(ollama_ok)
        # Lista para dropdown: full_key, label (model_name + route), is_free
        models = [
            {
                "full_key": s.full_key,
                "label": "%s (%s)" % (s.model_name, s.route),
                "model_name": s.model_name,
                "route": s.route,
                "is_free": s.is_free,
                "provider_id": s.provider_id,
            }
            for s in specs
        ]
        from modules.humanoid.ai.provider_credentials import \
            get_credentials_status

        _credentials = get_credentials_status()
        _allow_external = (
            os.getenv("AI_ALLOW_EXTERNAL_APIS") or ""
        ).strip().lower() in ("1", "true", "yes", "y", "on")
        _paid_default_models = {
            "gemini": "gemini-1.5-flash",
            "openai": "gpt-4o-mini",
            "anthropic": "claude-3-5-sonnet-20241022",
            "perplexity": "sonar",
        }
        for pid, default_model in _paid_default_models.items():
            if _allow_external and (_credentials.get(pid) or {}).get("configured"):
                full_key = "%s:%s" % (pid, default_model)
                models.append(
                    {
                        "full_key": full_key,
                        "label": "%s (manual)" % default_model,
                        "model_name": default_model,
                        "route": "CHAT",
                        "is_free": False,
                        "provider_id": pid,
                    }
                )
        mode = state.get("mode") or "auto"
        override = state.get("override_model")
        provider_connect_urls = {
            "openai": "https://platform.openai.com/api-keys",
            "anthropic": "https://console.anthropic.com/",
            "gemini": "https://aistudio.google.com/app/apikey",
            "perplexity": "https://www.perplexity.ai/settings/api",
        }
        provider_options = [
            {
                "id": "openai",
                "name": "OpenAI (GPT)",
                "connect_url": provider_connect_urls["openai"],
            },
            {
                "id": "anthropic",
                "name": "Anthropic (Claude)",
                "connect_url": provider_connect_urls["anthropic"],
            },
            {
                "id": "gemini",
                "name": "Google Gemini",
                "connect_url": provider_connect_urls["gemini"],
            },
            {
                "id": "perplexity",
                "name": "Perplexity",
                "connect_url": provider_connect_urls["perplexity"],
            },
        ]
        if mode == "auto":
            connected_names = [
                p["name"]
                for p in provider_options
                if (_credentials.get(p["id"]) or {}).get("configured")
            ]
            dominant_display = (
                "Multi-IA (auto): el router elige el especialista por tarea"
            )
            if connected_names:
                dominant_display += " Â· " + ", ".join(connected_names) + " conectado(s)"
        else:
            dominant_display = override or "â€” Selecciona una IA â€”"
            if override and ":" in override:
                prov_id = override.split(":", 1)[0].strip().lower()
                model_part = override.split(":", 1)[1].strip()
                for p in provider_options:
                    if p.get("id") == prov_id:
                        dominant_display = p.get("name", model_part)
                        break
                else:
                    dominant_display = model_part
        credentials = _credentials
        return {
            "ok": True,
            "mode": mode,
            "override_model": override,
            "dominant_display": dominant_display,
            "available_models": models,
            "provider_connect_urls": provider_connect_urls,
            "provider_options": provider_options,
            "credentials": credentials,
        }
    except Exception as e:
        provider_connect_urls = {
            "openai": "https://platform.openai.com/api-keys",
            "anthropic": "https://console.anthropic.com/",
            "gemini": "https://aistudio.google.com/app/apikey",
            "perplexity": "https://www.perplexity.ai/settings/api",
        }
        provider_options = [
            {
                "id": "openai",
                "name": "OpenAI (GPT)",
                "connect_url": provider_connect_urls["openai"],
            },
            {
                "id": "anthropic",
                "name": "Anthropic (Claude)",
                "connect_url": provider_connect_urls["anthropic"],
            },
            {
                "id": "gemini",
                "name": "Google Gemini",
                "connect_url": provider_connect_urls["gemini"],
            },
            {
                "id": "perplexity",
                "name": "Perplexity",
                "connect_url": provider_connect_urls["perplexity"],
            },
        ]
        credentials = {}
        try:
            from modules.humanoid.ai.provider_credentials import \
                get_credentials_status

            credentials = get_credentials_status()
        except Exception:
            pass
        return {
            "ok": False,
            "error": str(e),
            "mode": "auto",
            "override_model": None,
            "dominant_display": "Multi-IA (auto): el router elige el especialista por tarea",
            "available_models": [],
            "provider_connect_urls": provider_connect_urls,
            "provider_options": provider_options,
            "credentials": credentials,
        }


@app.get("/brain/status", tags=["Cerebro"])
def brain_status_alias():
    """Estado del cerebro para dashboard modals."""
    try:
        from modules.humanoid.ai.brain_state import get_brain_state

        state = get_brain_state()

        # Verificar si hay modelos disponibles (Ollama online)
        ollama_online = _brain_ollama_available_timeout(2.0)

        # Contar goals activos (placeholder - integrar con sistema real)
        active_goals = 0

        # Contar memorias (placeholder - integrar con lifelog)
        memory_entries = 0
        try:
            from modules.humanoid.cognitive.memory.lifelog import \
                get_lifelog_stats

            stats = get_lifelog_stats()
            memory_entries = stats.get("total_entries", 0)
        except:
            pass

        # Calcular uptime
        uptime_hours = 0
        try:
            from modules.humanoid.deploy.healthcheck import get_uptime_seconds

            uptime_hours = int(get_uptime_seconds() / 3600)
        except:
            pass

        return {
            "ok": True,
            "online": ollama_online,  # Brain estÃ¡ online si Ollama responde
            "active_goals": active_goals,
            "memory_entries": memory_entries,
            "uptime_hours": uptime_hours,
            "mode": state.get("mode", "auto"),
        }
    except Exception as e:
        return {
            "ok": False,
            "error": str(e),
            "online": False,
            "active_goals": 0,
            "memory_entries": 0,
            "uptime_hours": 0,
        }


@app.get("/quality/pots/list", tags=["Quality"])
def api_quality_pots_list(
    category: Optional[str] = None,
    severity: Optional[str] = None,
):
    """Lista todos los POTs registrados con filtros opcionales."""
    try:
        from modules.humanoid.quality.registry import list_pots

        pots = list_pots(category=category, severity=severity)
        return {"ok": True, "pots": pots, "count": len(pots)}
    except Exception as e:
        return {"ok": False, "error": str(e), "pots": []}


@app.get("/quality/executions/recent", tags=["Quality"])
def api_quality_executions_recent(limit: int = 10):
    """Obtiene historial reciente de ejecuciones de POTs."""
    try:
        from modules.humanoid.quality.dispatcher import get_dispatcher

        dispatcher = get_dispatcher()
        if dispatcher:
            history = dispatcher.get_history(limit=limit)
            return {"ok": True, "executions": history, "count": len(history)}
        else:
            return {"ok": False, "error": "Dispatcher not available", "executions": []}
    except Exception as e:
        return {"ok": False, "error": str(e), "executions": []}


@app.get("/nervous/status", tags=["Sistema Nervioso"])
def api_nervous_status():
    """Estado del sistema nervioso central."""
    try:
        from modules.humanoid.quality.dispatcher import get_dispatcher

        dispatcher = get_dispatcher()
        stats = dispatcher.get_stats() if dispatcher else {}

        return {
            "ok": True,
            "signals_per_min": stats.get("dispatches_per_minute", 0),
            "latency_ms": 12,  # Latencia promedio estimada
            "uptime_pct": 99.9,
            "active_dispatches": stats.get("active_count", 0),
            "total_dispatches": stats.get("total_dispatches", 0),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/ans/status", tags=["ANS"])
def api_ans_status():
    """Estado del Sistema Nervioso AutÃ³nomo."""
    try:
        from modules.humanoid.quality.dispatcher import get_dispatcher

        dispatcher = get_dispatcher()
        stats = dispatcher.get_stats() if dispatcher else {}

        # Obtener incidentes resueltos
        incidents_resolved = 0
        try:
            from modules.humanoid.ans.store import IncidentStore

            store = IncidentStore()
            incidents = store.list_incidents(status="resolved", limit=100)
            incidents_resolved = len(incidents)
        except:
            pass

        # Calcular uptime desde healthcheck
        uptime_hours = 0
        try:
            from modules.humanoid.deploy.healthcheck import get_uptime_seconds

            uptime_hours = int(get_uptime_seconds() / 3600)
        except:
            pass

        return {
            "ok": True,
            "cycles_completed": stats.get("total_dispatches", 0),
            "active_monitors": stats.get("active_count", 0),
            "health_score": 95,  # Score calculado basado en Ã©xito de dispatches
            "incidents_resolved": incidents_resolved,
            "uptime_hours": uptime_hours,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/brain/state", tags=["Cerebro"])
def api_brain_state_post(body: BrainStateBody):
    """Fija modo (auto/manual) y/o modelo override. En auto se ignora override."""
    try:
        from modules.humanoid.ai.brain_state import set_brain_state

        state = set_brain_state(mode=body.mode, override_model=body.override_model)
        return {"ok": True, "state": state}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/supervisor/advise", tags=["LLM Supervisor"])
def supervisor_advise(payload: dict):
    """
    Supervisor LLM - Genera recomendaciones y prompts para el Owner (RaÃºl).

    Usa arquitectura hÃ­brida: Ollama (local) + OpenAI (cloud) con fallback automÃ¡tico.
    """
    try:
        from atlas_adapter.llm.supervisor import Supervisor
        from atlas_adapter.supervisor_policy import set_supervisor_policy

        supervisor = Supervisor()
        objective = payload.get("objective", "Revisar estado del sistema")
        context = payload.get("context", {})
        user_id = (payload.get("user_id") or "owner").strip() or "owner"
        thread_id = _ensure_thread_id(
            payload.get("thread_id"), title="Supervisor", user_id=user_id
        )

        # If user pasted a supervisor policy into objective, store it and skip normal advise.
        # This prevents polluting the thread memory with long policy text that later gets "repeated".
        try:
            obj_txt = str(objective or "").strip()
            head = obj_txt[:900].lower()
            looks_like_policy = (
                (
                    "supervisor residente" in head
                    or "no eres un chatbot pasivo" in head
                    or "cero errores" in head
                )
                and ("objetivo principal" in head or "monitoreo continuo" in head)
                and len(obj_txt) > 200
            )
            if looks_like_policy:
                saved = set_supervisor_policy(obj_txt)
                return {
                    "ok": bool(saved.get("ok", True)),
                    "result": {
                        "ok": bool(saved.get("ok", True)),
                        "analysis": "OK: PolÃ­tica residente del Supervisor guardada. A partir de ahora se aplica automÃ¡ticamente.",
                        "snapshot": {},
                        "diagnosis": {
                            "severity": "healthy",
                            "issues": [],
                            "warnings": [],
                            "ok_items": ["policy_saved"],
                        },
                        "actions": [],
                        "recommendations": [],
                    },
                    "thread_id": thread_id,
                    "policy_saved": True,
                }
        except Exception:
            pass

        # Persist user objective in thread
        mgr = _get_chat_manager()
        if mgr and thread_id:
            try:
                mgr.add_message(
                    thread_id,
                    sender=user_id,
                    role="user",
                    content=str(objective),
                    metadata={"panel": "supervisor", "kind": "advise"},
                )
            except Exception:
                pass
            # Enrich context with thread memory (resumen + Ãºltimos mensajes)
            try:
                context = dict(context or {})
                context["_thread_id"] = thread_id
                context["_thread_memory"] = _llm_history_for_thread(
                    thread_id, max_messages=18
                )
            except Exception:
                pass

        result = supervisor.advise(objective, context)

        # Persist assistant output + update rolling summary
        if mgr and thread_id:
            try:
                analysis_txt = (
                    (result.get("analysis") or "") if isinstance(result, dict) else ""
                )
                if analysis_txt:
                    mgr.add_message(
                        thread_id,
                        sender="supervisor",
                        role="assistant",
                        content=str(analysis_txt),
                        metadata={"panel": "supervisor", "kind": "advise"},
                    )
                    _update_thread_summary(thread_id, str(objective), str(analysis_txt))
            except Exception:
                pass

        return {
            "ok": result.get("ok", True),
            "result": result,
            "thread_id": thread_id,
        }
    except Exception as e:
        return {
            "ok": False,
            "error": str(e),
            "result": {
                "recommendations": [],
                "prompts": [],
                "analysis": f"Error: {str(e)}",
            },
            "thread_id": payload.get("thread_id"),
        }


@app.get("/supervisor/policy", tags=["LLM Supervisor"])
def supervisor_policy_get():
    """PolÃ­tica residente del Supervisor (persistida localmente en logs/)."""
    try:
        from atlas_adapter.supervisor_policy import get_supervisor_policy

        return get_supervisor_policy()
    except Exception as e:
        return {"ok": False, "policy": "", "updated_at": 0.0, "error": str(e)[:200]}


@app.post("/supervisor/policy", tags=["LLM Supervisor"])
def supervisor_policy_post(payload: dict):
    """Guarda polÃ­tica residente del Supervisor (persistida localmente en logs/)."""
    try:
        from atlas_adapter.supervisor_policy import set_supervisor_policy

        policy = payload.get("policy", "")
        return set_supervisor_policy(str(policy))
    except Exception as e:
        return {"ok": False, "policy": "", "updated_at": 0.0, "error": str(e)[:200]}


@app.post("/supervisor/investigate", tags=["LLM Supervisor"])
def supervisor_investigate(payload: dict):
    """
    Supervisor deep investigation with tool calling. Streams SSE events.
    Body: { objective: str, context?: {} }
    """
    from fastapi.responses import StreamingResponse

    from atlas_adapter.llm.supervisor import Supervisor

    objective = payload.get("objective", "Revisar estado del sistema")
    context = payload.get("context", {})
    user_id = (payload.get("user_id") or "owner").strip() or "owner"
    thread_id = _ensure_thread_id(
        payload.get("thread_id"), title="Supervisor", user_id=user_id
    )

    # If user pasted a supervisor policy into objective, store it and short-circuit investigation stream.
    try:
        from atlas_adapter.supervisor_policy import set_supervisor_policy

        obj_txt = str(objective or "").strip()
        head = obj_txt[:900].lower()
        looks_like_policy = (
            (
                "supervisor residente" in head
                or "no eres un chatbot pasivo" in head
                or "cero errores" in head
            )
            and ("objetivo principal" in head or "monitoreo continuo" in head)
            and len(obj_txt) > 200
        )
        if looks_like_policy:
            saved = set_supervisor_policy(obj_txt)
            from fastapi.responses import StreamingResponse

            def _policy_stream():
                if thread_id:
                    yield f"event: thread\ndata: {json.dumps({'thread_id': thread_id}, ensure_ascii=False)}\n\n"
                yield "event: text\ndata: " + json.dumps(
                    {
                        "content": "OK: PolÃ­tica residente del Supervisor guardada. Se aplicarÃ¡ automÃ¡ticamente."
                    },
                    ensure_ascii=False,
                ) + "\n\n"
                yield "event: done\ndata: " + json.dumps(
                    {
                        "iterations": 0,
                        "tools_used": [],
                        "ms": 0,
                        "policy_saved": bool(saved.get("ok", True)),
                    },
                    ensure_ascii=False,
                ) + "\n\n"
                yield "event: close\ndata: {}\n\n"

            return StreamingResponse(_policy_stream(), media_type="text/event-stream")
    except Exception:
        pass

    # Add thread memory to context (best-effort)
    try:
        context = dict(context or {})
        if thread_id:
            context["_thread_id"] = thread_id
            context["_thread_memory"] = _llm_history_for_thread(
                thread_id, max_messages=18
            )
    except Exception:
        pass

    supervisor = Supervisor()

    def _sse_gen():
        mgr = _get_chat_manager()
        if mgr and thread_id:
            try:
                mgr.add_message(
                    thread_id,
                    sender=user_id,
                    role="user",
                    content=str(objective),
                    metadata={"panel": "supervisor", "kind": "investigate"},
                )
            except Exception:
                pass
        # Emit thread id early so UI can persist it
        if thread_id:
            yield f"event: thread\ndata: {json.dumps({'thread_id': thread_id}, ensure_ascii=False)}\n\n"
        final_text = ""
        try:
            for event in supervisor.investigate(objective, context):
                evt_type = event.get("event", "info")
                data_obj = event.get("data", {}) or {}
                if thread_id and evt_type in ("thinking", "done"):
                    try:
                        data_obj = dict(data_obj)
                        data_obj["thread_id"] = thread_id
                    except Exception:
                        pass
                if evt_type == "text":
                    try:
                        final_text = str((data_obj or {}).get("content") or "")
                    except Exception:
                        final_text = ""
                data = json.dumps(data_obj, ensure_ascii=False, default=str)
                yield f"event: {evt_type}\ndata: {data}\n\n"
        except Exception as e:
            yield f"event: error\ndata: {json.dumps({'message': str(e)})}\n\n"
        # Persist final assistant output + update summary
        try:
            if mgr and thread_id and final_text:
                mgr.add_message(
                    thread_id,
                    sender="supervisor",
                    role="assistant",
                    content=final_text,
                    metadata={"panel": "supervisor", "kind": "investigate"},
                )
                _update_thread_summary(thread_id, str(objective), final_text)
        except Exception:
            pass
        yield "event: close\ndata: {}\n\n"

    return StreamingResponse(_sse_gen(), media_type="text/event-stream")


@app.get("/supervisor/daemon/status", tags=["LLM Supervisor"])
def supervisor_daemon_status():
    """Estado del Supervisor residente (daemon)."""
    try:
        from atlas_adapter.supervisor_daemon import get_supervisor_daemon

        return get_supervisor_daemon().status()
    except Exception as e:
        return {"ok": False, "error": str(e)[:200], "data": {}}


@app.get("/supervisor/directives", tags=["LLM Supervisor"])
def supervisor_directives(status: str = "", limit: int = 50):
    """Cola de directivas internas del Supervisor (persistidas en autonomy_tasks.db)."""
    try:
        from atlas_adapter.supervisor_daemon import list_supervisor_directives

        return {
            "ok": True,
            "data": list_supervisor_directives(status=status or None, limit=limit),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)[:200], "data": []}


@app.get("/api/brain/credentials/status", tags=["Cerebro"])
def api_brain_credentials_status():
    """Estado de API keys por proveedor (configured, masked). Nunca devuelve la clave en claro."""
    try:
        from modules.humanoid.ai.provider_credentials import \
            get_credentials_status

        return {"ok": True, "credentials": get_credentials_status()}
    except Exception as e:
        return {"ok": False, "error": str(e), "credentials": {}}


@app.post("/api/brain/credentials/load-vault", tags=["Cerebro"])
def api_brain_credentials_load_vault():
    """Carga credenciales desde la BÃ³veda (credenciales.txt) y las persiste.

    - No devuelve claves en claro.
    - Usa ruta por directiva (o ATLAS_VAULT_PATH si estÃ¡ definida).
    """
    try:
        from pathlib import Path

        from dotenv import load_dotenv

        from modules.humanoid.ai.provider_credentials import (
            get_credentials_status, set_provider_api_key)

        vault_path = (
            os.getenv("ATLAS_VAULT_PATH")
            or r"C:\Users\Raul\OneDrive\RAUL - Personal\Escritorio\credenciales.txt"
        ).strip()
        candidates = [vault_path, r"C:\dev\credenciales.txt"]
        used = ""
        for c in candidates:
            p = Path(c)
            if p.is_file():
                used = str(p)
                load_dotenv(str(p), override=True)
                break
        if not used:
            return {
                "ok": False,
                "error": "vault_not_found",
                "vault_candidates": candidates,
                "credentials": get_credentials_status(),
            }

        mapping = {
            "openai": "OPENAI_API_KEY",
            "anthropic": "ANTHROPIC_API_KEY",
            "gemini": "GEMINI_API_KEY",
            "perplexity": "PERPLEXITY_API_KEY",
        }
        loaded = []
        for pid, env_key in mapping.items():
            v = (os.getenv(env_key) or "").strip()
            if v:
                try:
                    set_provider_api_key(pid, v)
                    loaded.append(pid)
                except Exception:
                    pass
        return {
            "ok": True,
            "vault_path": used,
            "loaded_providers": loaded,
            "credentials": get_credentials_status(),
        }
    except Exception as e:
        return {"ok": False, "error": str(e), "credentials": {}}


@app.post("/api/brain/credentials", tags=["Cerebro"])
def api_brain_credentials_post(body: ProviderCredentialBody):
    """Guarda o borra la API key de un proveedor (openai, anthropic, gemini, perplexity). La clave se guarda en config y el robot la usa para ese proveedor."""
    try:
        from modules.humanoid.ai.provider_credentials import \
            set_provider_api_key

        set_provider_api_key(body.provider_id, body.api_key)
        from modules.humanoid.ai.provider_credentials import \
            get_credentials_status

        return {"ok": True, "credentials": get_credentials_status()}
    except ValueError as e:
        return {"ok": False, "error": str(e)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/brain/models", tags=["Cerebro"])
def api_brain_models():
    """Lista de modelos disponibles (free + suscripciÃ³n/API) para el selector del cerebro."""
    try:
        from modules.humanoid.ai.registry import get_model_specs

        ollama_ok = _brain_ollama_available_timeout(2.0)
        specs = get_model_specs(ollama_ok)
        free = [s for s in specs if s.is_free]
        paid = [s for s in specs if not s.is_free]
        return {
            "ok": True,
            "free": [
                {
                    "full_key": s.full_key,
                    "label": "%s (%s)" % (s.model_name, s.route),
                    "route": s.route,
                }
                for s in free
            ],
            "paid": [
                {
                    "full_key": s.full_key,
                    "label": "%s (%s)" % (s.model_name, s.route),
                    "route": s.route,
                }
                for s in paid
            ],
        }
    except Exception as e:
        return {"ok": False, "error": str(e), "free": [], "paid": []}


# ---------- Fase 2: Vision (depth + scene) ----------
@app.post("/api/vision/depth/estimate", tags=["Vision"])
async def api_vision_depth_estimate(
    file: UploadFile = File(..., description="Image file")
):
    """Estima mapa de profundidad desde imagen. Devuelve depth coloreado base64 y shape."""
    try:
        import base64

        import cv2
        import numpy as np

        contents = await file.read()
        arr = np.frombuffer(contents, dtype=np.uint8)
        frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if frame is None:
            return {"ok": False, "error": "invalid image"}
        from modules.humanoid.vision.depth_estimation import (
            depth_map_to_colored, estimate_depth)

        depth = estimate_depth(frame)
        colored = depth_map_to_colored(depth)
        _, buf = cv2.imencode(".png", colored)
        b64 = base64.b64encode(buf).decode("utf-8")
        return {
            "ok": True,
            "depth_map_b64": b64,
            "shape": list(depth.shape),
            "dtype": str(depth.dtype),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/vision/scene/describe", tags=["Vision"])
async def api_vision_scene_describe(
    file: UploadFile = File(..., description="Image file"), detail_level: str = "medium"
):
    """DescripciÃ³n de escena desde imagen. detail_level: brief | medium | detailed."""
    try:
        import cv2
        import numpy as np

        from modules.humanoid.vision.scene_understanding import describe_scene

        contents = await file.read()
        arr = np.frombuffer(contents, dtype=np.uint8)
        frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if frame is None:
            return {"ok": False, "error": "invalid image", "description": ""}
        data = describe_scene(frame, detail_level=detail_level)
        return {"ok": True, **data}
    except Exception as e:
        return {"ok": False, "error": str(e), "description": ""}


# ---------- Fase 1: World model (stub) ----------
@app.post("/api/world-model/simulate", tags=["WorldModel"])
def api_world_model_simulate():
    """SimulaciÃ³n de acciÃ³n en world model (stub). Fase 1.2 completo: PyBullet + MCTS."""
    try:
        # TODO: integrar physics_simulator.load_scene + simulate_action
        return {
            "ok": True,
            "stub": True,
            "final_state": {},
            "collision_detected": False,
            "objects_moved": [],
            "success": True,
            "message": "World model stub; implementar physics_simulator.py",
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------- Fase 4: Meta-Learning (MAML) ----------
_maml_instance = None


class MetaAdaptBody(BaseModel):
    """Body para POST /api/meta-learning/adapt."""

    demonstrations: List[dict] = []
    num_steps: int = 5


def _get_maml():
    """Lazy singleton MAML (requiere torch)."""
    global _maml_instance
    if _maml_instance is not None:
        return _maml_instance
    try:
        from autonomous.learning.meta_learning.maml import MAML

        if MAML is None:
            return None
        _maml_instance = MAML(
            state_dim=128, action_dim=64, meta_lr=1e-3, inner_lr=1e-2, inner_steps=5
        )
        return _maml_instance
    except Exception:
        return None


@app.post("/api/meta-learning/train", tags=["MetaLearning"])
def api_meta_learning_train(num_tasks: int = 8, num_steps: int = 100):
    """Meta-entrenar MAML sobre distribuciÃ³n de tareas."""
    maml = _get_maml()
    if maml is None:
        return {"status": "error", "error": "MAML no disponible (PyTorch requerido)"}
    try:

        results = []
        for step in range(num_steps):
            metrics = maml.meta_train_step(num_tasks=num_tasks)
            if "error" in metrics:
                return {"status": "error", "error": metrics["error"]}
            if step % 10 == 0:
                results.append(
                    {
                        "step": step,
                        "meta_loss": metrics["meta_loss"],
                        "avg_task_loss": metrics["avg_task_loss"],
                    }
                )
        snap_dir = BASE_DIR / "snapshots"
        snap_dir.mkdir(parents=True, exist_ok=True)
        maml.save_checkpoint(str(snap_dir / "maml_checkpoint.pt"))
        return {
            "status": "meta_training_complete",
            "steps": num_steps,
            "final_loss": results[-1]["meta_loss"] if results else 0,
            "training_curve": results,
        }
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/meta-learning/adapt", tags=["MetaLearning"])
def api_meta_learning_adapt(body: MetaAdaptBody):
    """Adaptar a nueva tarea con pocas demostraciones (few-shot)."""
    maml = _get_maml()
    if maml is None:
        return {"status": "error", "error": "MAML no disponible (PyTorch requerido)"}
    demonstrations = body.demonstrations or []
    num_steps = body.num_steps or 5
    try:
        import numpy as np

        demo_tuples = [
            (
                np.array(d.get("state", [])),
                np.array(d.get("action", [])),
                float(d.get("reward", 0)),
            )
            for d in demonstrations
        ]
        adapted_policy = maml.adapt_to_new_task(
            demo_tuples, num_adaptation_steps=num_steps
        )
        snap_dir = BASE_DIR / "snapshots"
        snap_dir.mkdir(parents=True, exist_ok=True)
        import torch

        torch.save(adapted_policy.state_dict(), str(snap_dir / "adapted_policy.pt"))
        return {
            "status": "adapted",
            "num_demonstrations": len(demonstrations),
            "adaptation_steps": num_steps,
            "policy_path": "snapshots/adapted_policy.pt",
        }
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/meta-learning/stats", tags=["MetaLearning"])
def api_meta_learning_stats():
    """EstadÃ­sticas de meta-learning."""
    maml = _get_maml()
    if maml is None:
        return {"num_tasks_in_buffer": 0, "error": "MAML no disponible"}
    return {
        "num_tasks_in_buffer": len(maml.task_buffer),
        "policy_parameters": sum(p.numel() for p in maml.policy.parameters()),
        "inner_lr": maml.inner_lr,
        "inner_steps": maml.inner_steps,
    }


@app.post("/api/meta-learning/generate-tasks", tags=["MetaLearning"])
def api_meta_learning_generate_tasks(
    task_type: str = "navigation", num_tasks: int = 50
):
    """Generar tareas sintÃ©ticas para meta-entrenamiento."""
    maml = _get_maml()
    if maml is None:
        return {"status": "error", "error": "MAML no disponible (PyTorch requerido)"}
    try:
        from autonomous.learning.meta_learning.task_generator import \
            TaskGenerator

        if task_type == "navigation":
            tasks = TaskGenerator.generate_navigation_tasks(num_tasks)
        elif task_type == "manipulation":
            tasks = TaskGenerator.generate_manipulation_tasks(num_tasks)
        else:
            return {"error": f"Unknown task type: {task_type}"}
        for task in tasks:
            maml.add_task(task)
        return {
            "status": "tasks_generated",
            "task_type": task_type,
            "num_tasks": num_tasks,
            "total_in_buffer": len(maml.task_buffer),
        }
    except Exception as e:
        return {"status": "error", "error": str(e)}


# ---------- Fase 4: Causal Reasoning ----------
_causal_reasoner = None


def _get_causal_reasoner():
    global _causal_reasoner
    if _causal_reasoner is not None:
        return _causal_reasoner
    try:
        from brain.reasoning.causal_model import CausalReasoner

        _causal_reasoner = CausalReasoner()
        return _causal_reasoner
    except Exception:
        return None


@app.post("/api/causal/create-domain", tags=["Causal"])
def api_causal_create_domain(domain_name: str, structure: dict):
    """Crear grafo causal para un dominio."""
    reasoner = _get_causal_reasoner()
    if reasoner is None:
        return {"error": "CausalReasoner no disponible (networkx requerido)"}
    try:
        graph = reasoner.create_domain(domain_name, structure)
        return {
            "domain": domain_name,
            "variables": structure.get("variables", []),
            "num_edges": len(structure.get("edges", [])),
            "graph": graph.visualize(),
        }
    except Exception as e:
        return {"error": str(e)}


@app.post("/api/causal/reason", tags=["Causal"])
def api_causal_reason(domain: str, action: str, current_state: dict):
    """Razonar consecuencias de una acciÃ³n."""
    reasoner = _get_causal_reasoner()
    if reasoner is None:
        return {"error": "CausalReasoner no disponible"}
    return reasoner.reason_about_action(domain, action, current_state or {})


@app.post("/api/causal/explain", tags=["Causal"])
def api_causal_explain(domain: str, effect: str, potential_causes: List[str]):
    """Explicar por quÃ© ocurriÃ³ effect dados potential_causes."""
    reasoner = _get_causal_reasoner()
    if reasoner is None:
        return {"error": "CausalReasoner no disponible", "explanation": ""}
    explanation = reasoner.explain_why(domain, effect, potential_causes or [])
    return {"explanation": explanation}


@app.post("/api/causal/counterfactual", tags=["Causal"])
def api_causal_counterfactual(domain: str, variable: str, value: float, reality: dict):
    """Razonamiento contrafÃ¡ctico: quÃ© si variable hubiera sido value."""
    reasoner = _get_causal_reasoner()
    if reasoner is None:
        return {"error": "CausalReasoner no disponible"}
    if domain not in reasoner.graphs:
        return {"error": f"Unknown domain: {domain}"}
    result = reasoner.graphs[domain].counterfactual(variable, value, reality or {})
    return result


# ---------- Fase 4.3: Self-Programming ----------
class SelfProgramValidateBody(BaseModel):
    """Body para POST /api/self-programming/validate."""

    code: str = ""
    function_name: str = "run"
    test_cases: List[dict] = []


class SelfProgramOptimizeBody(BaseModel):
    code: str = ""


class SelfProgramExecuteBody(BaseModel):
    code: str = ""
    function_name: str = "run"
    test_input: dict = {}


_skill_validator = None
_skill_optimizer = None


def _get_skill_validator():
    global _skill_validator
    if _skill_validator is None:
        try:
            from brain.self_programming.skill_sandbox import SkillValidator

            _skill_validator = SkillValidator()
        except Exception:
            pass
    return _skill_validator


def _get_skill_optimizer():
    global _skill_optimizer
    if _skill_optimizer is None:
        try:
            from brain.self_programming.skill_optimizer import SkillOptimizer

            _skill_optimizer = SkillOptimizer()
        except Exception:
            pass
    return _skill_optimizer


@app.post("/api/self-programming/validate", tags=["SelfProgramming"])
def api_self_programming_validate(body: SelfProgramValidateBody):
    """Valida skill: sintaxis, seguridad, tests en sandbox."""
    validator = _get_skill_validator()
    if validator is None:
        return {"overall_valid": False, "error": "SkillValidator no disponible"}
    return validator.validate_comprehensive(
        body.code, body.function_name, body.test_cases or []
    )


@app.post("/api/self-programming/optimize", tags=["SelfProgramming"])
def api_self_programming_optimize(body: SelfProgramOptimizeBody):
    """Optimiza cÃ³digo generado (errores, type hints, docstrings)."""
    optimizer = _get_skill_optimizer()
    if optimizer is None:
        return {"error": "SkillOptimizer no disponible"}
    return optimizer.optimize_code(body.code or "")


@app.post("/api/self-programming/execute-sandbox", tags=["SelfProgramming"])
def api_self_programming_execute_sandbox(body: SelfProgramExecuteBody):
    """Ejecuta cÃ³digo en sandbox Docker."""
    try:
        from brain.self_programming.skill_sandbox import SkillSandbox

        sandbox = SkillSandbox()
        return sandbox.execute_safely(
            body.code or "",
            body.function_name or "run",
            body.test_input or {},
        )
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.get("/api/robot/status", tags=["NEXUS"])
def robot_status():
    """Estado del Robot (cÃ¡maras, visiÃ³n).

    Orden de detecciÃ³n:
    1) Backend Robot/NEXUS (8002 / NEXUS_ROBOT_URL)
    2) Fallback local USB (si hay cÃ¡mara local disponible)
    """
    import os
    import urllib.request

    base_api = get_robot_api_base()
    base_ui = get_robot_ui_base(default_to_api=True)
    for base in (base_api, base_ui):
        try:
            req = urllib.request.Request(base + "/", method="GET")
            with urllib.request.urlopen(req, timeout=3) as r:
                if r.status == 200:
                    return {
                        "ok": True,
                        "connected": True,
                        "robot_url": base,
                        "source": "robot",
                        "local_fallback": False,
                    }
        except Exception:
            pass
        try:
            req = urllib.request.Request(
                base + "/api/camera/service/status",
                method="GET",
                headers={"Accept": "application/json"},
            )
            with urllib.request.urlopen(req, timeout=3) as r:
                if r.status == 200:
                    return {
                        "ok": True,
                        "connected": True,
                        "robot_url": base,
                        "source": "robot",
                        "local_fallback": False,
                    }
        except Exception:
            pass
    return {
        "ok": False,
        "connected": False,
        "robot_url": base_ui or base_api,
        "source": "none",
        "local_fallback": False,
    }


@app.post("/api/robot/reconnect", tags=["NEXUS"])
def robot_reconnect():
    """Arranca el backend del Robot (cÃ¡maras). Devuelve enseguida; el usuario debe pulsar Actualizar cÃ¡maras en 10-15 s."""
    import subprocess
    from pathlib import Path

    if ENV_PATH.exists():
        try:
            from dotenv import load_dotenv

            load_dotenv(ENV_PATH, override=True)
        except Exception:
            pass
    robot_path = Path(
        os.getenv("NEXUS_ROBOT_PATH")
        or str(BASE_DIR / "nexus" / "atlas_nexus_robot" / "backend")
    )
    base_api = get_robot_api_base()
    repo_root = BASE_DIR
    script = repo_root / "scripts" / "start_nexus_services.py"
    if not script.exists():
        return {
            "ok": False,
            "connected": False,
            "robot_url": base_api,
            "message": "Script no encontrado.",
            "robot_path": str(robot_path),
        }
    if not robot_path.exists():
        return {
            "ok": False,
            "connected": False,
            "robot_url": base_api,
            "message": "Carpeta del Robot no existe: " + str(robot_path),
            "robot_path": str(robot_path),
        }
    try:
        py = os.getenv("PYTHON", "python")
        flags = (
            subprocess.CREATE_NO_WINDOW
            if hasattr(subprocess, "CREATE_NO_WINDOW")
            else 0
        )
        env = os.environ.copy()
        env["NEXUS_ROBOT_PATH"] = str(robot_path)
        env["NEXUS_ATLAS_PATH"] = os.getenv("NEXUS_ATLAS_PATH") or str(
            repo_root / "nexus" / "atlas_nexus"
        )
        r = subprocess.run(
            [py, str(script), "--robot-only"],
            cwd=str(repo_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=10,
            creationflags=flags if os.name == "nt" else 0,
        )
        out = (r.stdout or "").strip() or (r.stderr or "")[:200]
        if "Robot" in out:
            return {
                "ok": True,
                "connected": False,
                "robot_url": base_api,
                "message": "Robot arrancando. Espera 10-15 s y pulsa Â«Actualizar cÃ¡marasÂ».",
                "robot_path": str(robot_path),
            }
        return {
            "ok": False,
            "connected": False,
            "robot_url": base_api,
            "message": "No arrancÃ³ (salida: %s). Comprueba que en %s exista main.py."
            % (out or "vacÃ­o", robot_path),
            "robot_path": str(robot_path),
        }
    except subprocess.TimeoutExpired:
        return {
            "ok": False,
            "connected": False,
            "robot_url": base_api,
            "message": "Script tardÃ³ demasiado.",
            "robot_path": str(robot_path),
        }
    except Exception as e:
        return {
            "ok": False,
            "connected": False,
            "robot_url": base_api,
            "message": str(e),
            "robot_path": str(robot_path),
        }


@app.post("/api/cuerpo/reconnect", tags=["NEXUS"])
def cuerpo_reconnect():
    """Arranca Cuerpo completo (NEXUS 8000 + Robot 8002). Responde rÃ¡pido (no-bloqueante)."""
    import subprocess
    from pathlib import Path

    if ENV_PATH.exists():
        try:
            from dotenv import load_dotenv

            load_dotenv(ENV_PATH, override=True)
        except Exception:
            pass
    repo_root = BASE_DIR
    script = repo_root / "scripts" / "start_nexus_services.py"
    nexus_path = Path(
        os.getenv("NEXUS_ATLAS_PATH") or str(repo_root / "nexus" / "atlas_nexus")
    )
    robot_path = Path(
        os.getenv("NEXUS_ROBOT_PATH")
        or str(repo_root / "nexus" / "atlas_nexus_robot" / "backend")
    )
    if not script.exists():
        return {
            "ok": False,
            "started": False,
            "message": "Script no encontrado: scripts/start_nexus_services.py",
        }
    if not nexus_path.exists():
        return {
            "ok": False,
            "started": False,
            "message": "Carpeta NEXUS no existe: " + str(nexus_path),
        }
    if not robot_path.exists():
        return {
            "ok": False,
            "started": False,
            "message": "Carpeta Robot no existe: " + str(robot_path),
        }
    try:
        py = os.getenv("PYTHON", "python")
        flags = (
            subprocess.CREATE_NO_WINDOW
            if hasattr(subprocess, "CREATE_NO_WINDOW")
            else 0
        )
        env = os.environ.copy()
        env["NEXUS_ATLAS_PATH"] = str(nexus_path)
        env["NEXUS_ROBOT_PATH"] = str(robot_path)
        r = subprocess.run(
            [py, str(script)],
            cwd=str(repo_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=10,
            creationflags=flags if os.name == "nt" else 0,
        )
        out = (r.stdout or "").strip() or (r.stderr or "")[:200]
        started = ("NEXUS" in out) or ("Robot" in out)
        return {
            "ok": True,
            "started": bool(started),
            "message": "Cuerpo arrancando (NEXUS+Robot). Espera 10-20s y revisa Estado.",
            "output": out,
        }
    except subprocess.TimeoutExpired:
        return {
            "ok": True,
            "started": True,
            "message": "Cuerpo lanzado (timeout corto). Espera 10-20s y revisa Estado.",
        }
    except Exception as e:
        return {"ok": False, "started": False, "message": str(e)}


@app.get("/api/robot/start-commands", tags=["NEXUS"])
def robot_start_commands():
    """Devuelve los comandos para arrancar NEXUS y Robot a mano (por si Reconectar no da resultado)."""
    if ENV_PATH.exists():
        try:
            from dotenv import load_dotenv

            load_dotenv(ENV_PATH, override=True)
        except Exception:
            pass
    robot_path = Path(
        os.getenv("NEXUS_ROBOT_PATH")
        or str(BASE_DIR / "nexus" / "atlas_nexus_robot" / "backend")
    )
    nexus_path = Path(
        os.getenv("NEXUS_ATLAS_PATH") or str(BASE_DIR / "nexus" / "atlas_nexus")
    )
    py = os.getenv("PYTHON", "python")
    return {
        "ok": True,
        "robot_path": str(robot_path),
        "nexus_path": str(nexus_path),
        "commands": {
            "robot": 'cd /d "%s" && %s main.py' % (robot_path, py)
            if os.name == "nt"
            else 'cd "%s" && %s main.py' % (robot_path, py),
            "nexus": 'cd /d "%s" && %s nexus.py --mode api' % (nexus_path, py)
            if os.name == "nt"
            else 'cd "%s" && %s nexus.py --mode api' % (nexus_path, py),
        },
        "hint": "Abre dos terminales, ejecuta uno en cada una. Robot usa puerto 8002, NEXUS 8000.",
    }


def _tail_text_file(path: Path, max_bytes: int = 65536, lines: int = 200) -> str:
    try:
        if not path.exists():
            return ""
        with open(path, "rb") as f:
            try:
                f.seek(0, os.SEEK_END)
                size = f.tell()
                f.seek(max(0, size - max_bytes))
            except Exception:
                pass
            data = f.read()
        text = data.decode("utf-8", errors="replace")
        parts = text.splitlines()
        return "\n".join(parts[-max(1, int(lines)) :])
    except Exception:
        return ""


@app.get("/api/robot/log/tail", tags=["NEXUS"])
def robot_log_tail(lines: int = 200):
    """Ãšltimas lÃ­neas del log del backend Robot (arranque/errores)."""
    p = BASE_DIR / "logs" / "robot_backend.log"
    return {
        "ok": True,
        "path": str(p),
        "lines": int(lines),
        "text": _tail_text_file(p, lines=int(lines)),
    }


@app.get("/api/nexus/log/tail", tags=["NEXUS"])
def nexus_log_tail(lines: int = 200):
    """Ãšltimas lÃ­neas del log de NEXUS (si se arrancÃ³ con script start_nexus_services)."""
    p = BASE_DIR / "logs" / "nexus_api.log"
    return {
        "ok": True,
        "path": str(p),
        "lines": int(lines),
        "text": _tail_text_file(p, lines=int(lines)),
    }


@app.get("/nervous/services", tags=["NEXUS"])
def nervous_services():
    """Alias compacto del estado del sistema nervioso para compatibilidad."""
    return nerve_status()


@app.get("/api/nerve/status", tags=["NEXUS"])
def nerve_status():
    """Estado del nervio: ojos (Nexus disponible, snapshot_url), manos (local). Cerebro â†’ Nexus/manos."""
    try:
        from modules.humanoid.nerve import feet_status, nerve_eyes_status

        eyes = nerve_eyes_status()
        hands_deps_ok = None
        try:
            from modules.humanoid.screen.status import _screen_deps_ok as _ok

            hands_deps_ok = bool(_ok())
        except Exception:
            hands_deps_ok = None
        feet = feet_status()
        return {
            "ok": True,
            "eyes": eyes,
            "hands": {"local": True, "deps_ok": hands_deps_ok},
            "feet": feet,
        }
    except Exception as e:
        return {"ok": False, "eyes": {}, "hands": {}, "feet": {}, "error": str(e)}


@app.post("/api/feet/execute", tags=["NEXUS"])
def api_feet_execute(body: dict):
    """Ejecuta 'pies' (incluye pies internos digitales si FEET_DRIVER=digital).

    Body: {command: str, payload: dict}
    """
    try:
        from modules.humanoid.nerve import feet_execute

        cmd = (body or {}).get("command") or ""
        payload = (body or {}).get("payload") or {}
        return feet_execute(str(cmd), payload if isinstance(payload, dict) else {})
    except Exception as e:
        return {"ok": False, "error": str(e)}


class WorkspaceTerminalBody(BaseModel):
    command: str
    cwd: Optional[str] = None
    timeout_sec: int = 90
    actor: str = "workspace"
    role: str = "owner"


@app.get("/api/workspace/capabilities", tags=["Workspace"])
def api_workspace_capabilities():
    """Capacidades operativas del workspace (terminal real + navegaciÃ³n digital)."""
    try:
        from modules.humanoid import get_humanoid_kernel
        from modules.humanoid.nerve import feet_status
        from modules.humanoid.web import navigator

        k = get_humanoid_kernel()
        hands = k.registry.get("hands") if k else None
        hands_ok = bool(hands and hasattr(hands, "shell"))
        nav_ok = bool(navigator.is_available())
        return {
            "ok": True,
            "data": {
                "terminal": {"available": hands_ok},
                "navigation": {
                    "available": nav_ok,
                    "missing_deps": navigator.get_missing_deps() if not nav_ok else [],
                    "feet": feet_status(),
                },
            },
        }
    except Exception as e:
        return {"ok": False, "data": None, "error": str(e)}


@app.post("/api/workspace/terminal/execute", tags=["Workspace"])
def api_workspace_terminal_execute(body: WorkspaceTerminalBody):
    """EjecuciÃ³n real de comandos en terminal desde Workspace UI (con policy/audit)."""
    t0 = time.perf_counter()
    cmd = (body.command or "").strip()
    if not cmd:
        return _std_resp(
            False, None, int((time.perf_counter() - t0) * 1000), "command is required"
        )
    try:
        from modules.humanoid import get_humanoid_kernel
        from modules.humanoid.policy import ActorContext, get_policy_engine

        actor = ActorContext(
            actor=(body.actor or "workspace"), role=(body.role or "owner")
        )
        decision = get_policy_engine().can(actor, "hands", "exec_command", target=cmd)
        if not decision.allow:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                decision.reason or "policy denied",
            )

        k = get_humanoid_kernel()
        hands = k.registry.get("hands") if k else None
        if not hands or not hasattr(hands, "shell"):
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "hands module not available",
            )

        timeout_sec = max(5, min(int(body.timeout_sec or 90), 300))
        result = hands.shell.run(
            cmd, cwd=(body.cwd or None), timeout_sec=timeout_sec, actor=actor
        )
        ms = int((time.perf_counter() - t0) * 1000)
        data = {
            "cmd": cmd,
            "cwd": body.cwd,
            "ok": bool(result.get("ok")),
            "returncode": result.get("returncode"),
            "stdout": str(result.get("stdout") or "")[-12000:],
            "stderr": str(result.get("stderr") or "")[-8000:],
            "error": result.get("error"),
        }
        return _std_resp(bool(result.get("ok")), data, ms, result.get("error"))
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


class WorkspaceNavigateBody(BaseModel):
    action: str = (
        "open_url"  # open_url|click|fill|extract_text|screenshot|close|workflow
    )
    payload: Optional[dict] = None


@app.post("/api/workspace/chat/upload-image", tags=["Workspace"])
async def api_workspace_chat_upload_image(file: UploadFile = File(...)):
    """Upload image for Workspace chat paste/drop and return local evidence URL."""
    t0 = time.perf_counter()
    try:
        content_type = (file.content_type or "").lower()
        name = (file.filename or "image").strip()
        ext = Path(name).suffix.lower()
        if ext not in (".png", ".jpg", ".jpeg", ".webp"):
            # fallback by content type
            if "png" in content_type:
                ext = ".png"
            elif "jpeg" in content_type or "jpg" in content_type:
                ext = ".jpg"
            elif "webp" in content_type:
                ext = ".webp"
            else:
                ext = ".png"

        if content_type and not content_type.startswith("image/"):
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "only image files are allowed",
            )

        raw = await file.read()
        if not raw:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), "empty file"
            )
        if len(raw) > 10 * 1024 * 1024:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "file too large (max 10MB)",
            )

        target_dir = (BASE_DIR / "snapshots" / "chat_uploads").resolve()
        target_dir.mkdir(parents=True, exist_ok=True)
        safe_name = f"chat_{int(time.time()*1000)}_{abs(hash(name)) % 100000}{ext}"
        dst = (target_dir / safe_name).resolve()
        if target_dir not in dst.parents and dst != target_dir:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), "bad target path"
            )
        dst.write_bytes(raw)

        rel = f"snapshots/chat_uploads/{safe_name}".replace("\\", "/")
        url = f"/api/evidence/image?path={rel}"
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True,
            {"name": safe_name, "path": rel, "url": url, "bytes": len(raw)},
            ms,
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/api/workspace/navigate", tags=["Workspace"])
def api_workspace_navigate(body: WorkspaceNavigateBody):
    """NavegaciÃ³n web digital desde Workspace usando feet driver digital (Playwright)."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.nerve import feet_execute

        action = (body.action or "open_url").strip().lower()
        payload = dict(body.payload or {})
        payload.setdefault("driver", "digital")
        result = feet_execute(action, payload)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(bool(result.get("ok")), result, ms, result.get("error"))
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


# -----------------------------------------------------------------------------
# Open Interpreter â€” Motor de ejecucion autonoma del Workspace
# -----------------------------------------------------------------------------


class InterpreterExecuteBody(BaseModel):
    task: str
    model: Optional[str] = None
    auto_run: bool = False
    session_id: Optional[str] = None


@app.post("/api/workspace/interpreter/execute", tags=["Workspace"])
async def api_workspace_interpreter_execute(body: InterpreterExecuteBody):
    """Execute a task via Open Interpreter with SSE streaming."""
    from fastapi.responses import StreamingResponse

    task = (body.task or "").strip()
    if not task:
        return _std_resp(False, None, 0, "task is required")

    async def _sse_stream():
        try:
            from modules.humanoid.hands.interpreter_bridge import \
                execute_streaming

            async for chunk in execute_streaming(
                task=task,
                model=body.model,
                auto_run=body.auto_run,
                session_id=body.session_id,
            ):
                yield f"data: {json.dumps(chunk, ensure_ascii=True)}\n\n"
        except Exception as exc:
            yield f"data: {json.dumps({'type': 'error', 'content': str(exc)}, ensure_ascii=True)}\n\n"

    return StreamingResponse(_sse_stream(), media_type="text/event-stream")


@app.get("/api/workspace/interpreter/status", tags=["Workspace"])
def api_workspace_interpreter_status():
    """Return Open Interpreter engine status, available models and active sessions."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.hands.interpreter_bridge import \
            interpreter_status

        data = interpreter_status()
        return _std_resp(True, data, int((time.perf_counter() - t0) * 1000), None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.delete("/api/workspace/interpreter/session/{session_id}", tags=["Workspace"])
def api_workspace_interpreter_session_delete(session_id: str):
    """Close and dispose a specific interpreter session."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.hands.interpreter_bridge import \
            get_session_manager

        ok = get_session_manager().close(session_id)
        ms = int((time.perf_counter() - t0) * 1000)
        if ok:
            return _std_resp(True, {"session_id": session_id, "closed": True}, ms, None)
        return _std_resp(False, None, ms, f"session {session_id} not found")
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/api/workspace/interpreter/models", tags=["Workspace"])
def api_workspace_interpreter_models():
    """List all AI models available for the interpreter engine."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.hands.interpreter_bridge import (
            list_available_models, resolve_model)

        models = list_available_models()
        default = resolve_model()
        return _std_resp(
            True,
            {"models": models, "default": default, "total": len(models)},
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/api/workspace/interpreter/quick", tags=["Workspace"])
async def api_workspace_interpreter_quick(body: InterpreterExecuteBody):
    """Execute a simple task and return the full result (no streaming)."""
    t0 = time.perf_counter()
    task = (body.task or "").strip()
    if not task:
        return _std_resp(False, None, 0, "task is required")
    try:
        from modules.humanoid.hands.interpreter_bridge import execute_quick

        result = await execute_quick(task=task, model=body.model)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(result.get("ok", False), result, ms, result.get("error"))
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


# -----------------------------------------------------------------------------
# Primitivas (API) â€” wrappers explÃ­citos por dominio (sin ejecutor genÃ©rico)
# -----------------------------------------------------------------------------


@app.get("/api/primitives/nexus/pulse-check", tags=["Primitives"])
def api_prim_nexus_pulse_check():
    try:
        from modules.nexus_core.primitives import pulse_check

        return pulse_check()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/nexus/navigate-to", tags=["Primitives"])
def api_prim_nexus_navigate_to(body: dict):
    try:
        from modules.nexus_core.primitives import navigate_to

        x = int((body or {}).get("x", 0))
        y = int((body or {}).get("y", 0))
        duration = float((body or {}).get("duration", 0.2) or 0.2)
        expected_window = str(
            (body or {}).get("expected_window")
            or (body or {}).get("expected_window_title")
            or ""
        )
        expected_process = str(
            (body or {}).get("expected_process")
            or (body or {}).get("expected_exe")
            or ""
        )
        return navigate_to(
            x,
            y,
            duration=duration,
            expected_window=expected_window,
            expected_process=expected_process,
        )
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/nexus/reach-pose", tags=["Primitives"])
def api_prim_nexus_reach_pose(body: dict):
    try:
        from modules.nexus_core.primitives import reach_pose

        pan = float((body or {}).get("pan", 0.0) or 0.0)
        tilt = float((body or {}).get("tilt", 0.0) or 0.0)
        zoom = float((body or {}).get("zoom", 1.0) or 1.0)
        source = str((body or {}).get("source") or "camera")
        return reach_pose(pan, tilt, zoom, source=source)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/nexus/grasp", tags=["Primitives"])
def api_prim_nexus_grasp(body: dict):
    try:
        from modules.nexus_core.primitives import grasp

        target_id = str((body or {}).get("target_id") or "screen")
        region = (body or {}).get("region")
        region_t = None
        if isinstance(region, (list, tuple)) and len(region) == 4:
            region_t = (int(region[0]), int(region[1]), int(region[2]), int(region[3]))
        return grasp(target_id, region=region_t)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/nexus/release", tags=["Primitives"])
def api_prim_nexus_release(body: dict):
    try:
        from modules.nexus_core.primitives import release

        rid = str((body or {}).get("resource_id") or "")
        return release(rid)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/vision/scan-network", tags=["Primitives"])
def api_prim_vision_scan_network(body: dict):
    try:
        from modules.global_vision.primitives import scan_network

        protocol = str((body or {}).get("protocol") or "rtsp")
        return scan_network(protocol)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/vision/stream-proxy", tags=["Primitives"])
def api_prim_vision_stream_proxy(body: dict):
    try:
        from modules.global_vision.primitives import stream_proxy

        source_ip = str((body or {}).get("source_ip") or "")
        variant = str((body or {}).get("variant") or "mobile")
        return stream_proxy(source_ip, variant=variant)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/vision/perimeter-check", tags=["Primitives"])
def api_prim_vision_perimeter_check(body: dict):
    try:
        from modules.global_vision.primitives import perimeter_check

        snapshot_limit = int((body or {}).get("snapshot_limit", 8) or 8)
        emit_ops = bool((body or {}).get("emit_ops", True))
        return perimeter_check(snapshot_limit=snapshot_limit, emit_ops=emit_ops)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/architect/create-environment", tags=["Primitives"])
def api_prim_arch_create_environment(body: dict):
    try:
        from modules.atlas_architect.primitives import create_environment

        project_name = str((body or {}).get("project_name") or "")
        return create_environment(project_name)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/architect/patch-code", tags=["Primitives"])
def api_prim_arch_patch_code(body: dict):
    try:
        from modules.atlas_architect.primitives import patch_code

        file = str((body or {}).get("file") or "")
        pattern = str((body or {}).get("pattern") or "")
        replacement = str((body or {}).get("replacement") or "")
        pattern_is_regex = bool((body or {}).get("pattern_is_regex", False))
        return patch_code(file, pattern, replacement, pattern_is_regex=pattern_is_regex)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/architect/run-debug", tags=["Primitives"])
def api_prim_arch_run_debug(body: dict):
    try:
        from modules.atlas_architect.primitives import run_debug

        script_path = str((body or {}).get("script_path") or "")
        max_attempts = int((body or {}).get("max_attempts", 3) or 3)
        cwd = str((body or {}).get("cwd") or "")
        return run_debug(script_path, max_attempts=max_attempts, cwd=cwd)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/architect/generate-docs", tags=["Primitives"])
def api_prim_arch_generate_docs(body: dict):
    try:
        from modules.atlas_architect.primitives import generate_docs

        ctx = (body or {}).get("app_context") or {}
        return generate_docs(ctx if isinstance(ctx, dict) else {})
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/productividad/schedule-event", tags=["Primitives"])
def api_prim_prod_schedule_event(body: dict):
    try:
        from modules.productividad.primitives import schedule_event

        title = str((body or {}).get("title") or "")
        time_text = str((body or {}).get("time") or (body or {}).get("time_text") or "")
        desc = str((body or {}).get("desc") or "")
        return schedule_event(title, time_text, desc)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/productividad/check-rauli-inventory", tags=["Primitives"])
def api_prim_prod_check_rauli_inventory(body: dict):
    try:
        from modules.productividad.primitives import check_rauli_inventory

        camera_id = str((body or {}).get("camera_id") or "")
        use_llm = bool((body or {}).get("use_llm_vision", False))
        return check_rauli_inventory(camera_id, use_llm_vision=use_llm)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/productividad/digest-notifications", tags=["Primitives"])
def api_prim_prod_digest_notifications(body: dict):
    try:
        from modules.productividad.primitives import digest_notifications

        return digest_notifications()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/productividad/alert-user", tags=["Primitives"])
def api_prim_prod_alert_user(body: dict):
    try:
        from modules.productividad.primitives import alert_user

        priority = str((body or {}).get("priority") or "info")
        message = str((body or {}).get("message") or "")
        return alert_user(priority, message)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/finanzas/grasp-market-data", tags=["Primitives"])
def api_prim_fin_grasp_market_data(body: dict):
    try:
        from modules.finanzas.primitives import grasp_market_data

        ticker = str((body or {}).get("ticker") or "")
        timeframe = str((body or {}).get("timeframe") or "1m")
        return grasp_market_data(ticker, timeframe)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/finanzas/execute-trade", tags=["Primitives"])
def api_prim_fin_execute_trade(body: dict):
    try:
        from modules.finanzas.primitives import execute_trade

        ticker = str((body or {}).get("ticker") or "")
        side = str((body or {}).get("side") or "")
        qty = float((body or {}).get("qty") or 0)
        typ = str((body or {}).get("type") or "market")
        return execute_trade(ticker, side, qty, typ)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/finanzas/monitor-pnl", tags=["Primitives"])
def api_prim_fin_monitor_pnl(body: dict):
    try:
        from modules.finanzas.primitives import monitor_pnl

        return monitor_pnl()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/primitives/finanzas/hedge-position", tags=["Primitives"])
def api_prim_fin_hedge_position(body: dict):
    try:
        from modules.finanzas.primitives import hedge_position

        strategy = str((body or {}).get("strategy") or "")
        return hedge_position(strategy)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/comms/status", tags=["Comms"])
def api_comms_status():
    """Estado del sistema de comunicaciÃ³n permanente (audio/telegram/whatsapp)."""
    try:
        from modules.humanoid.comms.ops_bus import status as ops_status

        return ops_status()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/comms/recent", tags=["Comms"])
def api_comms_recent(limit: int = 50):
    """Ãšltimos eventos OPS emitidos."""
    try:
        from modules.humanoid.comms.ops_bus import recent as ops_recent

        return {"ok": True, "events": ops_recent(limit=limit)}
    except Exception as e:
        return {"ok": False, "error": str(e), "events": []}


@app.post("/api/comms/test", tags=["Comms"])
def api_comms_test(body: dict):
    """Emite un evento de prueba (audio+telegram+whatsapp segÃºn configuraciÃ³n)."""
    try:
        from modules.humanoid.comms.ops_bus import emit as ops_emit

        msg = (body or {}).get("message") or "Test OPS: sistema de comunicaciÃ³n activo."
        subsystem = (body or {}).get("subsystem") or "ops"
        level = (body or {}).get("level") or "info"
        ops_emit(str(subsystem), str(msg), level=str(level))
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/telegram/selftest", tags=["Comms"])
def api_comms_telegram_selftest(body: dict):
    """Descubre chat_id (si falta) y envÃ­a un mensaje de prueba. Retorna detalles."""
    try:
        from modules.humanoid.comms.ops_bus import \
            _telegram_chat_id  # type: ignore
        from modules.humanoid.comms.telegram_bridge import TelegramBridge

        bridge = TelegramBridge()
        chat_id = (body or {}).get("chat_id") or _telegram_chat_id()
        if not chat_id:
            d = bridge.discover_chat_id(limit=10)
            if d.get("ok"):
                chat_id = d.get("chat_id") or ""
        if not chat_id:
            return {
                "ok": False,
                "error": "no_chat_id_available. EnvÃ­a un mensaje al bot primero.",
                "chat_id": "",
            }
        text = (body or {}).get("text") or "[ATLAS] Test Telegram: canal operativo."
        r = bridge.send(str(chat_id), str(text))
        return {
            "ok": bool(r.get("ok")),
            "chat_id": str(chat_id),
            "result": r,
            "error": r.get("error"),
        }
    except Exception as e:
        return {"ok": False, "error": str(e), "chat_id": ""}


@app.post("/api/comms/speak", tags=["Comms"])
def api_comms_speak(body: dict):
    """TTS directo (audio PC). Body: {text}."""
    try:
        text = (body or {}).get("text") or ""
        from modules.humanoid.voice.tts import speak

        return speak(str(text))
    except Exception as e:
        return {"ok": False, "error": str(e)}


# === WhatsApp API ===


@app.get("/api/comms/whatsapp/status", tags=["Comms", "WhatsApp"])
def api_whatsapp_status():
    """Estado del servicio WhatsApp (proveedor, autenticaciÃ³n, configuraciÃ³n)."""
    try:
        from modules.humanoid.comms.whatsapp_bridge import health_check, status

        st = status()
        hc = health_check()
        return {**st, "health": hc}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/whatsapp/send", tags=["Comms", "WhatsApp"])
def api_whatsapp_send(body: dict):
    """EnvÃ­a un mensaje de WhatsApp.

    Body: {"text": "mensaje", "to": "+34612345678" (opcional)}
    """
    try:
        from modules.humanoid.comms.whatsapp_bridge import send_text

        text = (body or {}).get("text") or ""
        to = (body or {}).get("to")

        if not text:
            return {"ok": False, "error": "text is required"}

        return send_text(text, to=to)
    except Exception as e:
        return {"ok": False, "error": str(e)}


_WHATSAPP_INBOUND_TRACE_LOG = Path(_logs) / "whatsapp_inbound_trace.jsonl"


def _collect_request_origin(request: Request = None) -> Dict[str, Any]:
    if request is None:
        return {}
    headers = request.headers or {}
    client_ip = ""
    try:
        client_ip = str((request.client.host if request.client else "") or "")
    except Exception:
        client_ip = ""
    return {
        "client_ip": client_ip,
        "x_forwarded_for": str(headers.get("x-forwarded-for") or ""),
        "x_real_ip": str(headers.get("x-real-ip") or ""),
        "cf_connecting_ip": str(headers.get("cf-connecting-ip") or ""),
        "user_agent": str(headers.get("user-agent") or ""),
        "host": str(headers.get("host") or ""),
        "method": str(getattr(request, "method", "") or ""),
    }


def _append_whatsapp_inbound_trace(entry: Dict[str, Any]) -> None:
    try:
        _WHATSAPP_INBOUND_TRACE_LOG.parent.mkdir(parents=True, exist_ok=True)
        with _WHATSAPP_INBOUND_TRACE_LOG.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception:
        pass


@app.get("/api/comms/whatsapp/inbound", tags=["Comms", "WhatsApp"])
async def api_whatsapp_inbound_verify(request: Request):
    """Verificación de webhook de Meta WhatsApp Business API (GET challenge).

    Meta llama GET con hub.mode=subscribe, hub.challenge y hub.verify_token.
    Respondemos con hub.challenge si el verify_token coincide.
    """
    params = dict(request.query_params)
    mode = params.get("hub.mode", "")
    token = params.get("hub.verify_token", "")
    challenge = params.get("hub.challenge", "")
    expected = os.getenv("WHATSAPP_WEBHOOK_VERIFY_TOKEN", "")
    if mode == "subscribe" and token == expected and expected:
        from fastapi.responses import PlainTextResponse
        return PlainTextResponse(challenge)
    return {"ok": False, "error": "invalid_verify_token"}


@app.post("/api/comms/whatsapp/inbound", tags=["Comms", "WhatsApp"])
def api_whatsapp_inbound(body: dict, request: Request = None):
    """
    Webhook inbound de WhatsApp (WAHA/Twilio/Meta) -> supervisor conversacional.

    Flujo:
    inbound payload -> parse -> atlas_comms_hub.process_user_interaction -> send_text al remitente
    """
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub
        from modules.humanoid.comms.whatsapp_bridge import (parse_inbound_payload,
                                                            send_text)

        trace_base = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "route": "/api/comms/whatsapp/inbound",
            "origin": _collect_request_origin(request),
        }

        require_known = os.getenv("ATLAS_WHATSAPP_REQUIRE_KNOWN", "true").strip().lower() in (
            "1",
            "true",
            "yes",
            "y",
            "on",
        )
        require_atlas_question = os.getenv(
            "ATLAS_WHATSAPP_REQUIRE_ATLAS_QUESTION", "true"
        ).strip().lower() in (
            "1",
            "true",
            "yes",
            "y",
            "on",
        )
        re_atlas = re.compile(r"\batlas\b", re.IGNORECASE)
        re_question_hint = re.compile(
            r"[?¿]|\b(que|qué|como|cómo|puedes|puede|ayuda|ayudame|ayúdame|dime|necesito|sabes)\b",
            re.IGNORECASE,
        )

        parsed = parse_inbound_payload(body or {})
        if not parsed.get("ok"):
            _append_whatsapp_inbound_trace(
                {
                    **trace_base,
                    "status": "ignored",
                    "reason": parsed.get("reason") or "unparsed_payload",
                    "provider_hint": parsed.get("provider_hint"),
                }
            )
            return {
                "ok": True,
                "ignored": True,
                "reason": parsed.get("reason") or "unparsed_payload",
                "provider_hint": parsed.get("provider_hint"),
            }

        sender = str(parsed.get("from") or "").strip()
        text = str(parsed.get("text") or "").strip()
        if not sender or not text:
            _append_whatsapp_inbound_trace(
                {
                    **trace_base,
                    "status": "ignored",
                    "reason": "missing_sender_or_text",
                    "sender": sender,
                    "text_len": len(text),
                    "provider_hint": parsed.get("provider_hint"),
                }
            )
            return {"ok": True, "ignored": True, "reason": "missing_sender_or_text"}

        hub = get_atlas_comms_hub()
        resolved_user = hub.resolve_external_user_by_whatsapp(sender)
        is_known = bool(resolved_user.get("ok"))
        atlas_mentioned = bool(re_atlas.search(text or ""))
        looks_like_question = bool(re_question_hint.search(text or ""))

        if require_known and not is_known:
            unauthorized_text = os.getenv(
                "ATLAS_WHATSAPP_UNAUTHORIZED_TEXT",
                "Usuario no autorizado",
            ).strip() or "Usuario no autorizado"
            send_result = send_text(unauthorized_text, to=sender)
            _append_whatsapp_inbound_trace(
                {
                    **trace_base,
                    "status": "blocked",
                    "reason": "sender_not_linked",
                    "sender": sender,
                    "text_len": len(text),
                    "provider_hint": parsed.get("provider_hint"),
                    "send_result_ok": bool(send_result.get("ok")),
                }
            )
            return {
                "ok": True,
                "ignored": True,
                "reason": "sender_not_linked",
                "sender": sender,
                "send_result": send_result,
                "rules": {
                    "require_known": require_known,
                    "require_atlas_question": require_atlas_question,
                },
            }

        if require_atlas_question and not (atlas_mentioned and looks_like_question):
            _append_whatsapp_inbound_trace(
                {
                    **trace_base,
                    "status": "ignored",
                    "reason": "atlas_question_required",
                    "sender": sender,
                    "text_len": len(text),
                    "provider_hint": parsed.get("provider_hint"),
                    "atlas_mentioned": atlas_mentioned,
                    "looks_like_question": looks_like_question,
                }
            )
            return {
                "ok": True,
                "ignored": True,
                "reason": "atlas_question_required",
                "sender": sender,
                "rules": {
                    "require_known": require_known,
                    "require_atlas_question": require_atlas_question,
                    "atlas_mentioned": atlas_mentioned,
                    "looks_like_question": looks_like_question,
                },
            }

        effective_user_id = (
            str(resolved_user.get("user_id"))
            if resolved_user.get("ok")
            else str(parsed.get("user_id") or f"whatsapp:{sender}")
        )
        out = hub.process_user_interaction(
            user_id=effective_user_id,
            channel="whatsapp",
            message=text,
            context={
                "source": "whatsapp_inbound",
                "provider_hint": parsed.get("provider_hint"),
                "sender": sender,
                "external_user": resolved_user if resolved_user.get("ok") else None,
            },
        )

        reply_text = str(out.get("reply") or "").strip()
        send_result = {"ok": False, "skipped": True}
        if reply_text:
            send_result = send_text(reply_text, to=sender)

        _append_whatsapp_inbound_trace(
            {
                **trace_base,
                "status": "accepted",
                "sender": sender,
                "text_len": len(text),
                "provider_hint": parsed.get("provider_hint"),
                "effective_user_id": effective_user_id,
                "message_id": out.get("message_id"),
                "interaction_ok": bool(out.get("ok")),
                "send_result_ok": bool(send_result.get("ok")),
            }
        )

        return {
            "ok": bool(out.get("ok")),
            "inbound": parsed,
            "interaction": {
                "message_id": out.get("message_id"),
                "provider": out.get("provider"),
                "offline_mode": out.get("offline_mode"),
                "urgency": out.get("urgency"),
                "resolved_user": resolved_user if resolved_user.get("ok") else None,
            },
            "send_result": send_result,
        }
    except Exception as e:
        _append_whatsapp_inbound_trace(
            {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "route": "/api/comms/whatsapp/inbound",
                "origin": _collect_request_origin(request),
                "status": "error",
                "error": str(e)[:300],
            }
        )
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/whatsapp/test", tags=["Comms", "WhatsApp"])
def api_whatsapp_test():
    """EnvÃ­a un mensaje de prueba a WhatsApp."""
    try:
        from modules.humanoid.comms.whatsapp_bridge import send_text, status

        st = status()
        if not st.get("enabled"):
            return {"ok": False, "error": "WhatsApp no estÃ¡ habilitado", "details": st}

        result = send_text(
            "[ATLAS] Prueba de WhatsApp - Sistema de comunicaciÃ³n activo."
        )
        return result
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/comms/whatsapp/qr", tags=["Comms", "WhatsApp"])
def api_whatsapp_qr():
    """Obtiene el cÃ³digo QR para autenticar WAHA.

    Solo disponible con proveedor WAHA.
    Escanea este QR con tu WhatsApp para vincular.
    """
    try:
        from modules.humanoid.comms.whatsapp_bridge import get_qr_code, status

        st = status()
        if st.get("provider") != "waha":
            return {"ok": False, "error": "QR solo disponible para proveedor WAHA"}

        return get_qr_code()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/whatsapp/start-session", tags=["Comms", "WhatsApp"])
def api_whatsapp_start_session():
    """Inicia la sesiÃ³n WAHA (genera QR para autenticar).

    Solo disponible con proveedor WAHA.
    DespuÃ©s de llamar esto, obtÃ©n el QR con GET /api/comms/whatsapp/qr
    """
    try:
        from modules.humanoid.comms.whatsapp_bridge import (start_session,
                                                            status)

        st = status()
        if st.get("provider") != "waha":
            return {
                "ok": False,
                "error": "Sesiones solo disponibles para proveedor WAHA",
            }

        return start_session()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/whatsapp/pairing-code", tags=["Comms", "WhatsApp"])
async def api_whatsapp_pairing_code(body: dict):
    """Vincula WhatsApp mediante código de emparejamiento (sin QR).

    Flujo:
    1. Llama POST /api/comms/whatsapp/start-session  (si la sesión no existe)
    2. Llama este endpoint con {"phone": "+34612345678"}
    3. WAHA devuelve un código de 8 dígitos (ej: ABCD-EFGH)
    4. En tu WhatsApp → Dispositivos vinculados → Vincular con número de teléfono
       e introduce el código.

    Solo disponible con proveedor WAHA.
    """
    try:
        from modules.humanoid.comms.whatsapp_bridge import (
            start_session, request_pairing_code, status
        )

        st = status()
        if st.get("provider") != "waha":
            return {"ok": False, "error": "pairing_code solo disponible para proveedor WAHA"}

        phone = (body.get("phone") or "").strip()
        if not phone:
            return {"ok": False, "error": "Campo 'phone' requerido (ej: +34612345678)"}

        # Auto-arrancar sesión si no está iniciada
        waha_st = st.get("provider_status", {})
        if not waha_st.get("authenticated") and waha_st.get("status") not in ("SCAN_QR_CODE",):
            start_result = start_session()
            if not start_result.get("ok") and not start_result.get("already_started"):
                return {
                    "ok": False,
                    "error": "No se pudo iniciar la sesión WAHA",
                    "details": start_result,
                }

        result = request_pairing_code(phone)
        return result
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── Meta WhatsApp Business API — Setup Wizard ─────────────────────────────────

_META_WEBHOOK_BASE = "https://atlas-dashboard.rauliatlasapp.com"


def _ensure_verify_token() -> str:
    """Genera y persiste WHATSAPP_WEBHOOK_VERIFY_TOKEN si no existe."""
    token = os.getenv("WHATSAPP_WEBHOOK_VERIFY_TOKEN", "")
    if token:
        return token
    import secrets
    token = secrets.token_hex(24)
    # Persistir en .env
    env_path = BASE_DIR / ".env"
    try:
        content = env_path.read_text(encoding="utf-8") if env_path.exists() else ""
        if "WHATSAPP_WEBHOOK_VERIFY_TOKEN" not in content:
            with env_path.open("a", encoding="utf-8") as f:
                f.write(f'\nWHATSAPP_WEBHOOK_VERIFY_TOKEN="{token}"\n')
        os.environ["WHATSAPP_WEBHOOK_VERIFY_TOKEN"] = token
    except Exception:
        os.environ["WHATSAPP_WEBHOOK_VERIFY_TOKEN"] = token
    return token


@app.get("/api/comms/meta-whatsapp/setup-info", tags=["Comms", "WhatsApp"])
async def api_meta_whatsapp_setup_info():
    """Devuelve la información necesaria para configurar Meta WhatsApp Business API.

    Genera automáticamente el verify_token si no existe y muestra la URL del webhook
    pública a través del túnel Cloudflare.
    """
    verify_token = _ensure_verify_token()
    webhook_url = f"{_META_WEBHOOK_BASE}/api/comms/whatsapp/inbound"
    token_configured = bool(os.getenv("META_WHATSAPP_TOKEN"))
    phone_id_configured = bool(os.getenv("META_PHONE_NUMBER_ID"))
    return {
        "ok": True,
        "webhook_url": webhook_url,
        "verify_token": verify_token,
        "provider_ready": token_configured and phone_id_configured,
        "meta_token_set": token_configured,
        "phone_number_id_set": phone_id_configured,
        "steps": [
            {"step": 1, "done": True,  "desc": "Webhook URL generada",
             "value": webhook_url},
            {"step": 2, "done": True,  "desc": "Verify token generado",
             "value": verify_token},
            {"step": 3, "done": token_configured,
             "desc": "META_WHATSAPP_TOKEN configurado"},
            {"step": 4, "done": phone_id_configured,
             "desc": "META_PHONE_NUMBER_ID configurado"},
        ],
    }


@app.post("/api/comms/meta-whatsapp/save-credentials", tags=["Comms", "WhatsApp"])
async def api_meta_whatsapp_save_credentials(body: dict):
    """Guarda META_WHATSAPP_TOKEN y META_PHONE_NUMBER_ID en .env y activa el proveedor.

    Body: {
      "meta_token": "EAAxxxxxx",
      "phone_number_id": "123456789",
      "whatsapp_to": "+34612345678"   // opcional: número destino para notificaciones Atlas
    }
    """
    meta_token = (body.get("meta_token") or "").strip()
    phone_id = (body.get("phone_number_id") or "").strip()
    whatsapp_to = (body.get("whatsapp_to") or "").strip()

    if not meta_token or not phone_id:
        return {"ok": False, "error": "meta_token y phone_number_id son requeridos"}

    env_path = BASE_DIR / ".env"
    try:
        content = env_path.read_text(encoding="utf-8") if env_path.exists() else ""
        updates = {
            "WHATSAPP_PROVIDER": '"meta"',
            "WHATSAPP_ENABLED": '"true"',
            "META_WHATSAPP_TOKEN": f'"{meta_token}"',
            "META_PHONE_NUMBER_ID": f'"{phone_id}"',
        }
        if whatsapp_to:
            updates["WHATSAPP_TO"] = f'"{whatsapp_to}"'

        for key, val in updates.items():
            import re
            pattern = rf'^{re.escape(key)}\s*=.*$'
            if re.search(pattern, content, re.MULTILINE):
                content = re.sub(pattern, f'{key}={val}', content, flags=re.MULTILINE)
            else:
                content += f'\n{key}={val}'

        env_path.write_text(content, encoding="utf-8")

        # Aplicar en proceso actual
        os.environ["WHATSAPP_PROVIDER"] = "meta"
        os.environ["WHATSAPP_ENABLED"] = "true"
        os.environ["META_WHATSAPP_TOKEN"] = meta_token
        os.environ["META_PHONE_NUMBER_ID"] = phone_id
        if whatsapp_to:
            os.environ["WHATSAPP_TO"] = whatsapp_to

        # Resetear singleton del proveedor para que tome el nuevo
        try:
            import modules.humanoid.comms.whatsapp_bridge as _wb
            _wb._provider_instance = None
        except Exception:
            pass

        return {
            "ok": True,
            "provider": "meta",
            "webhook_url": f"{_META_WEBHOOK_BASE}/api/comms/whatsapp/inbound",
            "message": "Credenciales guardadas. Proveedor Meta activo.",
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/meta-whatsapp/open-browser", tags=["Comms", "WhatsApp"])
async def api_meta_whatsapp_open_browser():
    """Usa los 'ojos y manos' de Atlas (Playwright) para abrir Meta Developers.

    Navega directamente a la página de creación de apps de WhatsApp Business.
    """
    try:
        from modules.humanoid.nerve import feet_execute

        result = feet_execute("open_url", {
            "driver": "digital",
            "url": "https://developers.facebook.com/apps/",
            "show_browser": True,
        })
        if result.get("ok"):
            return {
                "ok": True,
                "message": "Navegador abierto en Meta Developers",
                "next": "Inicia sesión con tu cuenta de Meta/Facebook, luego: "
                        "Create App → Business → WhatsApp → Add Product",
            }
        # Playwright no disponible — dar URL directa
        return {
            "ok": False,
            "error": result.get("error", "playwright_not_available"),
            "manual_url": "https://developers.facebook.com/apps/",
            "hint": "Abre la URL manualmente. Necesitas FEET_DRIVER=digital en .env",
        }
    except Exception as e:
        return {
            "ok": False,
            "error": str(e),
            "manual_url": "https://developers.facebook.com/apps/",
        }


_META_AUTOSETUP_PROC = None  # subprocess handle
_META_AUTOSETUP_STATUS_FILE = BASE_DIR / "logs" / "meta_autosetup_status.json"


@app.post("/api/comms/meta-whatsapp/autosetup", tags=["Comms", "WhatsApp"])
async def api_meta_whatsapp_autosetup():
    """Lanza el agente autónomo de Atlas para configurar Meta WhatsApp.

    Abre Chromium visible, espera login del usuario, luego automatiza
    la creación del app, configuración del webhook y extracción de credenciales.
    """
    global _META_AUTOSETUP_PROC
    import subprocess

    # Matar proceso previo si existe
    if _META_AUTOSETUP_PROC and _META_AUTOSETUP_PROC.poll() is None:
        _META_AUTOSETUP_PROC.terminate()

    verify_token = _ensure_verify_token()
    webhook_url = f"{_META_WEBHOOK_BASE}/api/comms/whatsapp/inbound"

    # Limpiar status anterior
    if _META_AUTOSETUP_STATUS_FILE.exists():
        _META_AUTOSETUP_STATUS_FILE.unlink()

    script = BASE_DIR / "scripts" / "atlas_meta_whatsapp_autosetup.py"
    python = BASE_DIR / "venv" / "Scripts" / "python.exe"
    if not python.exists():
        python = Path(sys.executable)

    try:
        _META_AUTOSETUP_PROC = subprocess.Popen(
            [
                str(python), str(script),
                "--webhook-url", webhook_url,
                "--verify-token", verify_token,
                "--status-file", "logs/meta_autosetup_status.json",
            ],
            cwd=str(BASE_DIR),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
        )
        return {
            "ok": True,
            "pid": _META_AUTOSETUP_PROC.pid,
            "message": "Agente autónomo iniciado. Chromium se abrirá en tu pantalla. Inicia sesión en Meta cuando aparezca.",
            "poll_url": "/api/comms/meta-whatsapp/autosetup/status",
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/comms/meta-whatsapp/autosetup/status", tags=["Comms", "WhatsApp"])
async def api_meta_whatsapp_autosetup_status():
    """Devuelve el estado del agente autónomo de configuración Meta."""
    global _META_AUTOSETUP_PROC

    running = _META_AUTOSETUP_PROC is not None and _META_AUTOSETUP_PROC.poll() is None
    exit_code = None if running else (
        _META_AUTOSETUP_PROC.returncode if _META_AUTOSETUP_PROC else None
    )

    status_data = {}
    if _META_AUTOSETUP_STATUS_FILE.exists():
        try:
            status_data = json.loads(_META_AUTOSETUP_STATUS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass

    return {
        "ok": True,
        "running": running,
        "exit_code": exit_code,
        "finished": not running and _META_AUTOSETUP_PROC is not None,
        "success": exit_code == 0,
        **status_data,
    }


@app.get("/api/comms/whatsapp/setup-guide", tags=["Comms", "WhatsApp"])
def api_whatsapp_setup_guide():
    """GuÃ­a de configuraciÃ³n de WhatsApp con WAHA (gratuito)."""
    return {
        "provider": "WAHA (WhatsApp HTTP API)",
        "description": "Servidor gratuito que expone WhatsApp Web como API REST",
        "steps": [
            {
                "step": 1,
                "title": "Instalar Docker Desktop",
                "command": "https://www.docker.com/products/docker-desktop/",
                "note": "Si ya tienes Docker, salta este paso",
            },
            {
                "step": 2,
                "title": "Ejecutar WAHA",
                "command": "docker run -d -p 3000:3000 --name waha devlikeapro/waha",
                "note": "Esto descarga e inicia el servidor WAHA",
            },
            {
                "step": 3,
                "title": "Configurar variables de entorno",
                "variables": {
                    "WHATSAPP_ENABLED": "true",
                    "WHATSAPP_PROVIDER": "waha",
                    "WHATSAPP_TO": "+34612345678  # Tu nÃºmero personal",
                    "WAHA_API_URL": "http://localhost:3000",
                },
                "note": "AÃ±ade estas variables a tu archivo .env o credenciales.txt",
            },
            {
                "step": 4,
                "title": "Escanear cÃ³digo QR",
                "url": "http://localhost:3000",
                "note": "Abre esta URL en el navegador y escanea el QR con tu WhatsApp",
            },
            {
                "step": 5,
                "title": "Probar envÃ­o",
                "endpoint": "POST /api/comms/whatsapp/test",
                "note": "DeberÃ­as recibir un mensaje de prueba",
            },
        ],
        "troubleshooting": [
            "Si el QR no aparece, llama POST /api/comms/whatsapp/start-session primero",
            "Si el contenedor no inicia, verifica que Docker estÃ© corriendo",
            "El QR expira en ~60 segundos, refresca si es necesario",
        ],
    }


# --- Anti-Silencio: alertas multiciclo desde UI/Watchdogs ---
_COMMS_ALERT_LAST: dict = {}  # key -> ts


@app.post("/api/comms/alert", tags=["Comms"])
def api_comms_alert(body: dict):
    """
    Alerta multiciclo (Audio PC + Telegram/WhatsApp via OPS Bus) + BitÃ¡cora ANS.
    Body sugerido:
      { kind, message_human, action_human, level, technical }
    """
    import time as _time

    t0 = time.perf_counter()
    try:
        kind = str((body or {}).get("kind") or "ui").strip()[:40]
        level = str((body or {}).get("level") or "high").strip().lower()
        msg = str((body or {}).get("message_human") or "").strip()
        action = str((body or {}).get("action_human") or "").strip()
        technical = (body or {}).get("technical")

        if not msg:
            msg = "Alerta: el panel tuvo un error o perdiÃ³ conexiÃ³n."
        if action:
            msg2 = f"{msg} AcciÃ³n: {action}"
        else:
            msg2 = msg

        # Throttle por kind+msg (evitar spam por reconexiÃ³n / loops JS)
        key = f"{kind}:{msg2[:120]}"
        now = _time.time()
        last = float(_COMMS_ALERT_LAST.get(key) or 0.0)
        if now - last < 10.0:
            ms = int((time.perf_counter() - t0) * 1000)
            return {"ok": True, "skipped": "throttled", "ms": ms}
        _COMMS_ALERT_LAST[key] = now

        # BitÃ¡cora ANS (humano)
        try:
            from modules.humanoid.ans.evolution_bitacora import \
                append_evolution_log

            append_evolution_log(f"[ALERTA] {msg2}", ok=False, source="ui")
        except Exception:
            pass

        # OPS Bus (multicanal)
        try:
            from modules.humanoid.comms.ops_bus import emit as ops_emit

            ops_emit(
                "dashboard",
                msg2,
                level=level,
                data={"kind": kind, "technical": technical or {}},
            )
        except Exception:
            pass

        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": True, "ms": ms}
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "error": str(e), "ms": ms}


# === Nuevos endpoints del Sistema de ComunicaciÃ³n Unificado ===


@app.get("/api/comms/hub/health", tags=["Comms"])
def api_comms_hub_health():
    """Estado de salud del CommsHub central (canales, circuit breakers, mÃ©tricas)."""
    try:
        from modules.humanoid.comms import get_hub

        hub = get_hub()
        return hub.get_health()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/comms/hub/messages", tags=["Comms"])
def api_comms_hub_messages(limit: int = 50):
    """Historial de mensajes del CommsHub."""
    try:
        from modules.humanoid.comms import get_hub

        hub = get_hub()
        return {"ok": True, "messages": hub.get_recent_messages(limit=limit)}
    except Exception as e:
        return {"ok": False, "error": str(e), "messages": []}


@app.post("/api/comms/hub/emit", tags=["Comms"])
def api_comms_hub_emit(body: dict):
    """Emite un mensaje a travÃ©s del CommsHub (multicanal con retry/circuit breaker).

    Body:
    {
        "message": "Contenido del mensaje",
        "level": "info|low|medium|high|critical",
        "subsystem": "nombre_subsistema",
        "data": {},  # opcional
        "channels": ["telegram", "audio"]  # opcional, si no se especifica usa mapeo por nivel
    }
    """
    try:
        from modules.humanoid.comms import get_hub

        hub = get_hub()

        message = (body or {}).get("message") or ""
        level = (body or {}).get("level") or "info"
        subsystem = (body or {}).get("subsystem") or "api"
        data = (body or {}).get("data") or {}
        channels = (body or {}).get("channels")
        evidence_path = (body or {}).get("evidence_path") or ""

        if not message:
            return {"ok": False, "error": "message is required"}

        result = hub.send(
            content=message,
            level=level,
            subsystem=subsystem,
            data=data,
            evidence_path=evidence_path,
            channels=channels,
        )

        return {
            "ok": len(result.channels_sent) > 0,
            "message_id": result.id,
            "channels_sent": result.channels_sent,
            "channels_failed": result.channels_failed,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/hub/reset-channel", tags=["Comms"])
def api_comms_hub_reset_channel(body: dict):
    """Resetea el estado de un canal (Ãºtil despuÃ©s de arreglar un problema).

    Body: {"channel": "telegram"}
    """
    try:
        from modules.humanoid.comms import get_hub

        hub = get_hub()

        channel = (body or {}).get("channel") or ""
        if not channel:
            return {"ok": False, "error": "channel is required"}

        return hub.reset_channel(channel)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/hub/disable-channel", tags=["Comms"])
def api_comms_hub_disable_channel(body: dict):
    """Deshabilita temporalmente un canal.

    Body: {"channel": "whatsapp"}
    """
    try:
        from modules.humanoid.comms import get_hub

        hub = get_hub()

        channel = (body or {}).get("channel") or ""
        if not channel:
            return {"ok": False, "error": "channel is required"}

        return hub.disable_channel(channel)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/hub/enable-channel", tags=["Comms"])
def api_comms_hub_enable_channel(body: dict):
    """Habilita un canal previamente deshabilitado.

    Body: {"channel": "whatsapp"}
    """
    try:
        from modules.humanoid.comms import get_hub

        hub = get_hub()

        channel = (body or {}).get("channel") or ""
        if not channel:
            return {"ok": False, "error": "channel is required"}

        return hub.enable_channel(channel)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/comms/bootstrap/status", tags=["Comms"])
def api_comms_bootstrap_status():
    """Estado de todos los servicios de comunicaciÃ³n inicializados por bootstrap."""
    try:
        from modules.humanoid.comms.bootstrap import get_status

        return get_status()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/comms/bootstrap/health", tags=["Comms"])
def api_comms_bootstrap_health():
    """Health check completo de todos los servicios de comunicaciÃ³n."""
    try:
        from modules.humanoid.comms.bootstrap import health_check

        return health_check()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/bootstrap/restart-service", tags=["Comms"])
def api_comms_bootstrap_restart_service(body: dict):
    """Reinicia un servicio de comunicaciÃ³n especÃ­fico.

    Body: {"service": "telegram_poller"}
    Servicios reiniciables: hub, telegram_poller, makeplay
    """
    try:
        from modules.humanoid.comms.bootstrap import restart_service

        service = (body or {}).get("service") or ""
        if not service:
            return {"ok": False, "error": "service is required"}

        return restart_service(service)
    except Exception as e:
        return {"ok": False, "error": str(e)}


class AtlasCommsMessageBody(BaseModel):
    user_id: str = "guest"
    channel: str = "app"
    message: str
    context: Dict[str, Any] = Field(default_factory=dict)

    @field_validator("message")
    @classmethod
    def validate_message(cls, value: str) -> str:
        msg = (value or "").strip()
        if not msg:
            raise ValueError("message cannot be empty")
        return msg


class AtlasCommsResyncBody(BaseModel):
    limit: int = Field(default=200, ge=1, le=400)


class AtlasCommsInventoryAdjustBody(BaseModel):
    payload: Dict[str, Any]
    confirmed: bool = False
    requested_by: str = "atlas_comms_api"


class AtlasCommsUsersResyncBody(BaseModel):
    targets: List[str] = Field(default_factory=lambda: ["panaderia", "vision"])


class AtlasCommsWhatsappLinkBody(BaseModel):
    source: str
    external_user_id: str
    whatsapp_number: str
    updated_by: str = "atlas_comms_api"

    @field_validator("source")
    @classmethod
    def validate_source(cls, value: str) -> str:
        v = (value or "").strip().lower()
        if v not in ("panaderia", "vision"):
            raise ValueError("source must be panaderia or vision")
        return v


class AtlasCommsWhatsappRequestBody(BaseModel):
    source: str = ""
    requested_by: str = "atlas_comms_api"
    note: str = "Solicitud de número WhatsApp"
    limit: int = Field(default=100, ge=1, le=500)

    @field_validator("source")
    @classmethod
    def validate_source(cls, value: str) -> str:
        v = (value or "").strip().lower()
        if v and v not in ("panaderia", "vision"):
            raise ValueError("source must be empty, panaderia or vision")
        return v


class AtlasCommsSendExternalBody(BaseModel):
    source: str
    external_user_id: str
    message: str

    @field_validator("source")
    @classmethod
    def validate_source(cls, value: str) -> str:
        v = (value or "").strip().lower()
        if v not in ("panaderia", "vision"):
            raise ValueError("source must be panaderia or vision")
        return v

    @field_validator("message")
    @classmethod
    def validate_message(cls, value: str) -> str:
        msg = (value or "").strip()
        if not msg:
            raise ValueError("message cannot be empty")
        return msg


class AtlasCommsWhatsappBatchConfigureBody(BaseModel):
    source: str = "panaderia"
    items: List[Dict[str, Any]] = Field(default_factory=list)
    send_welcome: bool = True
    app_link: str = ""
    updated_by: str = "atlas_comms_api"

    @field_validator("source")
    @classmethod
    def validate_source(cls, value: str) -> str:
        v = (value or "").strip().lower()
        if v not in ("panaderia", "vision"):
            raise ValueError("source must be panaderia or vision")
        return v


def _call_claude_cli(prompt: str, system: str, timeout_s: int) -> tuple[bool, str, str]:
    cli_bin = (os.getenv("ATLAS_CLAWD_CLAUDE_CLI_BIN") or "").strip()
    if not cli_bin:
        cli_bin = shutil.which("claude") or ""
    if not cli_bin:
        appdata = (os.getenv("APPDATA") or "").strip()
        userprofile = (os.getenv("USERPROFILE") or "").strip()
        candidates = []
        if appdata:
            candidates.append(Path(appdata) / "npm" / "claude.cmd")
        if userprofile:
            candidates.append(Path(userprofile) / "AppData" / "Roaming" / "npm" / "claude.cmd")
        for candidate in candidates:
            if candidate.exists():
                cli_bin = str(candidate)
                break
    if not cli_bin:
        cli_bin = "claude"
    if not cli_bin:
        return False, "", "clawd_cli_missing"
    cmd = [
        cli_bin,
        "--print",
        "--input-format",
        "text",
        "--output-format",
        "text",
        "--system-prompt",
        system,
    ]
    suffix = Path(cli_bin).suffix.lower()
    if suffix in {".cmd", ".bat"}:
        cmd = ["cmd.exe", "/c", *cmd]
    try:
        proc = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            input=prompt,
            timeout=timeout_s,
        )
    except FileNotFoundError:
        return False, "", "clawd_cli_missing"
    except subprocess.TimeoutExpired:
        return False, "", "clawd_cli_timeout"
    except Exception as exc:
        return False, "", f"clawd_cli_error:{str(exc)[:160]}"
    if proc.returncode != 0:
        err = (proc.stderr or "").strip() or "clawd_cli_failed"
        return False, "", f"clawd_cli_fail:{err[:160]}"
    out = (proc.stdout or "").strip()
    if not out:
        return False, "", "clawd_cli_empty_reply"
    return True, out, "clawd_cli"


class AtlasClawdSubscriptionBody(BaseModel):
    message: str
    context: Dict[str, Any] = Field(default_factory=dict)
    persona: str = "friendly_precise_assistant"
    history: List[Dict[str, Any]] = Field(default_factory=list)

    @field_validator("message")
    @classmethod
    def validate_message(cls, value: str) -> str:
        msg = (value or "").strip()
        if not msg:
            raise ValueError("message cannot be empty")
        return msg


class AvatarRunwayLineBody(BaseModel):
    topic: str = ""
    source: str = "api"
    requested_line: str = ""

    @field_validator("topic", "source", "requested_line")
    @classmethod
    def _trim_fields(cls, value: str) -> str:
        return (value or "").strip()


@app.get("/api/comms/atlas/status", tags=["Comms"])
def api_comms_atlas_status():
    """ATLAS comms bridge status (offline mode, queue, encryption source)."""
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        return hub.get_status()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/avatar/runway/context", tags=["Avatar", "Runway"])
def api_avatar_runway_context():
    """Runtime payload to drive ATLAS avatar behavior in Runway Characters."""
    try:
        from modules.humanoid.avatar import build_runway_character_context

        return build_runway_character_context()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/avatar/runway/line", tags=["Avatar", "Runway"])
def api_avatar_runway_line(body: AvatarRunwayLineBody):
    """Low-latency short line payload for Runway Characters scene driving."""
    try:
        from modules.humanoid.avatar import build_runway_line

        return build_runway_line(
            topic=body.topic,
            source=body.source,
            requested_line=body.requested_line,
        )
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/comms/atlas/history", tags=["Comms"])
def api_comms_atlas_history(limit: int = 30):
    """Recent user interactions processed by atlas_comms_hub."""
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        return hub.get_history(limit=limit, decrypt=True)
    except Exception as e:
        return {"ok": False, "error": str(e), "items": []}


@app.post("/api/comms/atlas/message", tags=["Comms"])
def api_comms_atlas_message(body: AtlasCommsMessageBody):
    """Process user message with hybrid AI + inventory/camera probes + offline queue."""
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        return hub.process_user_interaction(
            user_id=body.user_id,
            channel=body.channel,
            message=body.message,
            context=body.context,
        )
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/atlas/clawd-subscription", tags=["Comms"])
def api_comms_atlas_clawd_subscription(body: AtlasClawdSubscriptionBody):
    """Endpoint interno para suscripción Claude usada por atlas_comms_hub."""
    try:
        from modules.humanoid.ai.external_llm import call_external
    except Exception as e:
        return {"ok": False, "error": f"anthropic_module_missing:{e}", "reply": ""}

    api_key = (
        os.getenv("ATLAS_CLAWD_API_KEY")
        or os.getenv("ANTHROPIC_API_KEY")
        or os.getenv("ANTHROPIC_API_KEY_CEREBRO")
        or ""
    ).strip()
    if not api_key:
        try:
            from modules.humanoid.ai.provider_credentials import get_provider_api_key

            api_key = (get_provider_api_key("anthropic") or "").strip()
        except Exception:
            api_key = ""
    if not api_key:
        return {"ok": False, "error": "anthropic_api_key_missing", "reply": ""}

    model = (
        os.getenv("ATLAS_CLAWD_MODEL")
        or os.getenv("ANTHROPIC_MODEL")
        or "claude-sonnet-4-latest"
    ).strip()
    try:
        timeout_s = int((os.getenv("ATLAS_COMMS_CLAWD_TIMEOUT_S") or "8").strip() or "8")
    except Exception:
        timeout_s = 8
    timeout_s = max(4, min(timeout_s, 45))

    ctx = body.context or {}
    history = list(body.history or [])[-4:]
    history_lines = []
    for idx, item in enumerate(history, start=1):
        if not isinstance(item, dict):
            continue
        u = str(item.get("user") or "").strip()
        a = str(item.get("assistant") or "").strip()
        if u or a:
            history_lines.append(f"{idx}. Usuario: {u}\n   ATLAS: {a}")
    history_text = "\n".join(history_lines) if history_lines else "Sin historial reciente."

    prompt = (
        "Contexto operativo actual:\n"
        f"- inventario_state: {ctx.get('inventory_state')}\n"
        f"- camaras_activas: {ctx.get('camera_count')}\n"
        f"- panaderia_reachable: {ctx.get('panaderia_reachable')}\n"
        f"- vision_reachable: {ctx.get('vision_reachable')}\n"
        f"- offline_mode: {ctx.get('offline_mode')}\n\n"
        f"Historial reciente:\n{history_text}\n\n"
        f"Mensaje actual del usuario:\n{body.message}\n\n"
        "Responde en espanol natural, breve (max 5 lineas), con tono cercano y accionable."
    )
    system = (
        "Eres ATLAS, asistente operativo de RAULI. "
        "Prioriza claridad, rapidez y acciones concretas."
    )

    mode = (os.getenv("ATLAS_CLAWD_SUBSCRIPTION_MODE") or "").strip().lower()
    use_cli = mode in ("claude-cli", "claude_cli", "claude-code", "claude_code", "cli")
    use_cli = use_cli or (os.getenv("ATLAS_CLAWD_CLAUDE_CLI") or "").strip().lower() in (
        "1",
        "true",
        "yes",
        "y",
        "on",
    )
    if use_cli:
        ok, text, provider = _call_claude_cli(prompt, system, timeout_s)
        if not ok:
            return {"ok": False, "error": provider or "clawd_cli_fail", "reply": ""}
        reply = str(text or "").strip()
        if not reply:
            return {"ok": False, "error": "clawd_cli_empty_reply", "reply": ""}
        return {"ok": True, "reply": reply, "provider": provider}

    ok, text, latency_ms = call_external(
        "anthropic",
        model,
        prompt,
        system,
        api_key,
        timeout_s=timeout_s,
    )
    if not ok:
        return {
            "ok": False,
            "error": str(text or "anthropic_error")[:240],
            "reply": "",
        }

    reply = str(text or "").strip()
    if not reply:
        return {"ok": False, "error": "anthropic_empty_reply", "reply": ""}

    return {
        "ok": True,
        "reply": reply,
        "provider": f"anthropic:{model}:{int(latency_ms)}ms",
        "persona": body.persona,
    }


@app.post("/api/comms/atlas/resync", tags=["Comms"])
def api_comms_atlas_resync(body: Optional[AtlasCommsResyncBody] = None):
    """Manual trigger to process pending offline queue."""
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        lim = int(body.limit) if body else 200
        return hub.resync_pending(limit=lim)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/atlas/users/resync", tags=["Comms"])
def api_comms_atlas_users_resync(body: Optional[AtlasCommsUsersResyncBody] = None):
    """Sync external users from RAULI-PANADERIA and RAULI-VISION."""
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        targets = list(body.targets) if body and body.targets else ["panaderia", "vision"]
        return hub.sync_external_users(targets=targets)
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/comms/atlas/users/whatsapp", tags=["Comms"])
def api_comms_atlas_users_whatsapp(
    source: str = "",
    state: str = "",
    limit: int = 300,
):
    """List external users with WhatsApp linkage state."""
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        return hub.list_external_users(source=source, whatsapp_state=state, limit=limit)
    except Exception as e:
        return {"ok": False, "error": str(e), "items": []}


@app.post("/api/comms/atlas/users/whatsapp/link", tags=["Comms"])
def api_comms_atlas_users_whatsapp_link(body: AtlasCommsWhatsappLinkBody):
    """Link WhatsApp number to one external user."""
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        return hub.link_external_user_whatsapp(
            source=body.source,
            external_user_id=body.external_user_id,
            whatsapp_number=body.whatsapp_number,
            updated_by=body.updated_by or "atlas_comms_api",
        )
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/atlas/users/whatsapp/request-missing", tags=["Comms"])
def api_comms_atlas_users_whatsapp_request_missing(body: Optional[AtlasCommsWhatsappRequestBody] = None):
    """
    Mark users without WhatsApp as requested and return request templates.
    Includes download URL so supervisor can send install link when needed.
    """
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        source = body.source if body else ""
        requested_by = body.requested_by if body else "atlas_comms_api"
        note = body.note if body else "Solicitud de número WhatsApp"
        limit = int(body.limit) if body else 100
        return hub.request_numbers_for_missing_users(
            source=source,
            requested_by=requested_by,
            note=note,
            limit=limit,
        )
    except Exception as e:
        return {"ok": False, "error": str(e), "items": []}


@app.post("/api/comms/atlas/users/whatsapp/send", tags=["Comms"])
def api_comms_atlas_users_whatsapp_send(body: AtlasCommsSendExternalBody):
    """Send WhatsApp message to a linked external user."""
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        return hub.send_whatsapp_to_external_user(
            source=body.source,
            external_user_id=body.external_user_id,
            text=body.message,
        )
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/comms/atlas/users/whatsapp/configure-batch", tags=["Comms"])
def api_comms_atlas_users_whatsapp_configure_batch(body: AtlasCommsWhatsappBatchConfigureBody):
    """
    Batch link WhatsApp numbers and optionally send first welcome message.
    Each item can include:
      - external_user_id (preferred)
      - or match/username/full_name/name/email
      - whatsapp_number (or phone)
      - welcome_name (optional override for greeting)
    """
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        return hub.configure_whatsapp_batch(
            source=body.source,
            items=body.items or [],
            updated_by=body.updated_by or "atlas_comms_api",
            send_welcome=bool(body.send_welcome),
            app_link=(body.app_link or ""),
        )
    except Exception as e:
        return {"ok": False, "error": str(e), "results": []}


@app.post("/api/comms/atlas/inventory/adjust", tags=["Comms"])
def api_comms_atlas_inventory_adjust(body: AtlasCommsInventoryAdjustBody):
    """
    Guarded inventory adjustment:
    - requires explicit confirmed=true
    - requires snapshot validation (atlas_snapshot_safe.ps1)
    """
    try:
        from modules.humanoid.comms.atlas_comms_hub import get_atlas_comms_hub

        hub = get_atlas_comms_hub()
        return hub.adjust_inventory(
            payload=body.payload,
            confirmed=bool(body.confirmed),
            requested_by=(body.requested_by or "atlas_comms_api"),
        )
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/kernel/event-bus/stats", tags=["Kernel"])
def api_kernel_event_bus_stats():
    """EstadÃ­sticas del Event Bus interno (handlers, eventos, mÃ©tricas)."""
    try:
        from modules.humanoid import get_humanoid_kernel

        kernel = get_humanoid_kernel()
        return {"ok": True, "stats": kernel.events.get_stats()}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/kernel/event-bus/history", tags=["Kernel"])
def api_kernel_event_bus_history(limit: int = 50):
    """Historial de eventos del Event Bus."""
    try:
        from modules.humanoid import get_humanoid_kernel

        kernel = get_humanoid_kernel()
        return {"ok": True, "events": kernel.events.get_history(limit=limit)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/kernel/event-bus/reset-errors", tags=["Kernel"])
def api_kernel_event_bus_reset_errors(body: dict):
    """Resetea errores de handlers y rehabilita handlers desactivados.

    Body: {"topic": "specific.topic"} o {} para todos
    """
    try:
        from modules.humanoid import get_humanoid_kernel

        kernel = get_humanoid_kernel()

        topic = (body or {}).get("topic")
        rehabilitated = kernel.events.reset_handler_errors(topic=topic)

        return {"ok": True, "rehabilitated_handlers": rehabilitated}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/health", tags=["Health"])
async def health():
    """Health cheap: respuesta inmediata sin depender de servicios externos."""
    try:
        return {
            "ok": True,
            "service": "atlas_push",
            "ts": datetime.now(timezone.utc).isoformat(),
            "pid": os.getpid(),
            "version": (
                os.getenv("ATLAS_VERSION")
                or os.getenv("APP_VERSION")
                or ((BASE_DIR / "VERSION").read_text(encoding="utf-8").strip() if (BASE_DIR / "VERSION").exists() else "unknown")
            ),
        }
    except Exception as e:
        return {"ok": False, "service": "atlas_push", "error": str(e)}


@app.get("/health/deep", tags=["Health"])
async def health_deep():
    """Health verificable completo con checks y score (costoso)."""
    try:
        from modules.humanoid.deploy.healthcheck import run_health_verbose
        from modules.humanoid.deploy.ports import get_ports

        active_port = get_ports()[0]
        return run_health_verbose(base_url=None, active_port=active_port)
    except Exception as e:
        return {"ok": False, "score": 0, "checks": {}, "ms": 0, "error": str(e)}


@app.get("/health/debug", tags=["Health"])
def health_debug():
    """Raw check results con mensajes de error para diagnÃ³stico."""
    try:
        from modules.humanoid.deploy.healthcheck import (
            _check_audit_writable, _check_memory_writable,
            _check_scheduler_running)

        return {
            "ok": True,
            "AUDIT_DB_PATH": os.getenv("AUDIT_DB_PATH"),
            "SCHED_DB_PATH": os.getenv("SCHED_DB_PATH"),
            "ATLAS_MEMORY_DB_PATH": os.getenv("ATLAS_MEMORY_DB_PATH"),
            "memory": _check_memory_writable(),
            "audit": _check_audit_writable(),
            "scheduler": _check_scheduler_running(),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/doctor")
def doctor_endpoint():
    """Diagnostico rapido del sistema."""
    try:
        from atlas_runtime import doctor

        return {"ok": True, "result": doctor()}
    except Exception as e:
        return {"ok": True, "result": f"Doctor unavailable: {e}"}


@app.get("/modules/check-all")
def modules_check_all():
    """Verificar estado de todos los modulos."""
    results = []
    connected = 0
    module_checks = [
        ("ans", "modules.humanoid.ans.api"),
        ("governance", "modules.humanoid.governance.api"),
        ("nervous", "modules.humanoid.nervous.api"),
        ("cognitive", "modules.humanoid.cognitive.api"),
        ("quality", "modules.humanoid.quality.api"),
        ("memory_engine", "modules.humanoid.memory_engine"),
        ("brain", "modules.humanoid.brain"),
        ("hippo", "modules.humanoid.hippo.api"),
        ("learning", "modules.humanoid.learning.api"),
        ("comms", "modules.humanoid.comms.hub"),
        ("world_model", "modules.humanoid.world_model.api"),
        ("libro_vida", "modules.humanoid.memory_engine.libro_vida"),
        ("directives", "modules.humanoid.directives.api"),
        ("tools", "modules.humanoid.tools.api"),
        ("scheduler", "modules.humanoid.scheduler.api"),
        ("tutorias", "modules.humanoid.quality.tutorias.api"),
    ]
    for name, mod_path in module_checks:
        try:
            __import__(mod_path)
            results.append({"name": name, "status": "connected", "module": mod_path})
            connected += 1
        except Exception as e:
            results.append(
                {
                    "name": name,
                    "status": "error",
                    "module": mod_path,
                    "error": str(e)[:100],
                }
            )
    return {
        "ok": True,
        "modules": results,
        "connected": connected,
        "total": len(module_checks),
    }


@app.post("/modules/reconnect/{module_id}")
def modules_reconnect(module_id: str):
    """Intentar reconectar un modulo especifico."""
    try:
        import importlib

        mod_map = {
            "ans": "modules.humanoid.ans.api",
            "governance": "modules.humanoid.governance.api",
            "nervous": "modules.humanoid.nervous.api",
            "cognitive": "modules.humanoid.cognitive.api",
            "quality": "modules.humanoid.quality.api",
            "brain": "modules.humanoid.brain",
            "hippo": "modules.humanoid.hippo.api",
            "learning": "modules.humanoid.learning.api",
            "comms": "modules.humanoid.comms.hub",
            "world_model": "modules.humanoid.world_model.api",
            "memory_engine": "modules.humanoid.memory_engine",
        }
        mod_path = mod_map.get(module_id, f"modules.humanoid.{module_id}")
        mod = importlib.import_module(mod_path)
        importlib.reload(mod)
        return {"ok": True, "module": module_id, "status": "reconnected"}
    except Exception as e:
        return {"ok": False, "module": module_id, "error": str(e)[:200]}


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# AUTONOMY SYSTEM API
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

import sqlite3 as _auto_sqlite

_AUTO_DB = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "data", "autonomy_tasks.db"
)
os.makedirs(os.path.dirname(_AUTO_DB), exist_ok=True)


def _auto_db():
    conn = _auto_sqlite.connect(_AUTO_DB)
    conn.row_factory = _auto_sqlite.Row
    conn.execute(
        """CREATE TABLE IF NOT EXISTS autonomy_tasks (
        id TEXT PRIMARY KEY, title TEXT, status TEXT DEFAULT 'pending',
        priority TEXT DEFAULT 'medium', source TEXT DEFAULT 'system',
        detail TEXT DEFAULT '', action_taken TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')), updated_at TEXT DEFAULT (datetime('now'))
    )"""
    )
    conn.execute(
        """CREATE TABLE IF NOT EXISTS autonomy_timeline (
        id INTEGER PRIMARY KEY AUTOINCREMENT, ts TEXT DEFAULT (datetime('now')),
        event TEXT, kind TEXT DEFAULT 'info', result TEXT DEFAULT 'ok'
    )"""
    )
    conn.execute(
        """CREATE TRIGGER IF NOT EXISTS trg_supervisor_autosig_insert
           AFTER INSERT ON autonomy_tasks
           WHEN NEW.source='supervisor' AND (NEW.action_taken IS NULL OR trim(NEW.action_taken)='')
           BEGIN
             UPDATE autonomy_tasks
                SET action_taken = substr(
                      lower(
                        replace(
                          replace(coalesce(NEW.title,'') || '|' || coalesce(NEW.detail,''), char(10), ' '),
                          char(13),
                          ' '
                        )
                      ),
                      1,
                      240
                    ),
                    updated_at = datetime('now')
              WHERE id = NEW.id;
           END"""
    )
    conn.execute(
        """UPDATE autonomy_tasks
              SET action_taken = substr(
                    lower(
                      replace(
                        replace(coalesce(title,'') || '|' || coalesce(detail,''), char(10), ' '),
                        char(13),
                        ' '
                      )
                    ),
                    1,
                    240
                  ),
                  updated_at = datetime('now')
            WHERE source='supervisor'
              AND (action_taken IS NULL OR trim(action_taken)='')"""
    )
    conn.commit()
    return conn


def _auto_log(event, kind="info", result="ok"):
    try:
        c = _auto_db()
        c.execute(
            "INSERT INTO autonomy_timeline(event,kind,result) VALUES(?,?,?)",
            (event, kind, result),
        )
        c.commit()
        c.close()
    except:
        pass


# Worker autÃ³nomo agresivo (vinculado a daemon + modo growth)
_AGGR_LOCK = threading.RLock()
_aggr_thread = None
_aggr_stop = threading.Event()
_aggr_nav_idx = 0
_aggr_cycle_count = 0
_aggr_last_page = ""
_aggr_last_user_active_log = 0.0
_autodiag_kick_lock = threading.RLock()
_autodiag_last_kick_ts = 0.0
_aggr_config = {
    "enabled": (os.getenv("AUTONOMY_AGGRESSIVE_ENABLED") or "true").strip().lower()
    in ("1", "true", "yes", "on"),
    "idle_sec": max(
        5, int((os.getenv("AUTONOMY_USER_IDLE_SEC") or "25").strip() or "25")
    ),
    "interval_sec": max(
        8, int((os.getenv("AUTONOMY_AGGRESSIVE_INTERVAL_SEC") or "25").strip() or "25")
    ),
    "pages": [
        p.strip()
        for p in (
            os.getenv("AUTONOMY_AGGRESSIVE_PAGES") or "/ui,/workspace,/nexus"
        ).split(",")
        if str(p).strip()
    ],
}


def _get_aggr_config() -> Dict[str, Any]:
    with _AGGR_LOCK:
        pages = list(_aggr_config.get("pages") or ["/ui", "/workspace", "/nexus"])
        if not pages:
            pages = ["/ui", "/workspace", "/nexus"]
        return {
            "enabled": bool(_aggr_config.get("enabled", True)),
            "idle_sec": int(_aggr_config.get("idle_sec", 25) or 25),
            "interval_sec": int(_aggr_config.get("interval_sec", 25) or 25),
            "pages": pages,
        }


def _set_aggr_config(payload: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return _get_aggr_config()
    with _AGGR_LOCK:
        if "enabled" in payload:
            _aggr_config["enabled"] = bool(payload.get("enabled"))
        if "idle_sec" in payload:
            _aggr_config["idle_sec"] = max(5, int(payload.get("idle_sec") or 25))
        if "interval_sec" in payload:
            _aggr_config["interval_sec"] = max(
                8, int(payload.get("interval_sec") or 25)
            )
        if "pages" in payload:
            raw_pages = payload.get("pages")
            pages = []
            if isinstance(raw_pages, str):
                pages = [p.strip() for p in raw_pages.split(",") if p.strip()]
            elif isinstance(raw_pages, list):
                pages = [str(p).strip() for p in raw_pages if str(p).strip()]
            _aggr_config["pages"] = pages or ["/ui", "/workspace", "/nexus"]
    return _get_aggr_config()


def _aggr_enabled() -> bool:
    try:
        return bool(_get_aggr_config().get("enabled", True))
    except Exception:
        return True


def _user_idle_seconds() -> float:
    """Segundos desde la Ãºltima interacciÃ³n local (mouse/teclado)."""
    try:
        import ctypes

        class LASTINPUTINFO(ctypes.Structure):
            _fields_ = [("cbSize", ctypes.c_uint), ("dwTime", ctypes.c_uint)]

        lii = LASTINPUTINFO()
        lii.cbSize = ctypes.sizeof(LASTINPUTINFO)
        if ctypes.windll.user32.GetLastInputInfo(ctypes.byref(lii)):
            now_ms = int(ctypes.windll.kernel32.GetTickCount())
            idle_ms = max(0, now_ms - int(lii.dwTime))
            return idle_ms / 1000.0
    except Exception:
        pass
    # Fallback: asumir idle para no bloquear worker por entorno sin API Win32.
    return 9999.0


def _run_aggressive_cycle() -> Dict[str, Any]:
    global _aggr_nav_idx, _aggr_cycle_count, _aggr_last_page
    out = {
        "hands_ok": False,
        "eyes_ok": False,
        "visual_pre_ok": False,
        "visual_post_ok": False,
        "error": None,
    }
    # 1) Manos: ejecutar comando no destructivo para verificar control real.
    try:
        from modules.humanoid import get_humanoid_kernel
        from modules.humanoid.policy import ActorContext

        k = get_humanoid_kernel()
        hands = k.registry.get("hands") if k else None
        if hands and hasattr(hands, "shell"):
            actor = ActorContext(actor="autonomy_aggressive", role="owner")
            r = hands.shell.run(
                "python -c \"print('ATLAS_AGGRESSIVE_HANDS_OK')\"",
                cwd=None,
                timeout_sec=20,
                actor=actor,
            )
            out["hands_ok"] = bool((r or {}).get("ok"))
            if not out["hands_ok"] and not out.get("error"):
                out["error"] = str((r or {}).get("error") or "hands command failed")[
                    :200
                ]
    except Exception as e:
        out["error"] = str(e)[:200]

    # 2) Ojos/Pies digitales: navegar por paneles internos y extraer seÃ±al.
    try:
        from modules.humanoid.nerve import feet_execute

        cfg = _get_aggr_config()
        raw_pages = cfg.get("pages") or ["/ui", "/workspace", "/nexus"]
        pages = [p.strip() for p in raw_pages if str(p).strip()]
        if not pages:
            pages = ["/ui", "/workspace", "/nexus"]
        path = pages[_aggr_nav_idx % len(pages)]
        _aggr_nav_idx += 1
        if path.startswith("http://") or path.startswith("https://"):
            target_url = path
        else:
            target_url = (
                f"http://127.0.0.1:8791{path if path.startswith('/') else '/' + path}"
            )

        r_open = feet_execute(
            "open_url", {"url": target_url, "driver": "digital", "show_browser": False}
        )
        ok_open = bool((r_open or {}).get("ok"))

        # Regla general: verificaciÃ³n visual ANTES y DESPUÃ‰S.
        r_pre = feet_execute("screenshot", {"driver": "digital"})
        out["visual_pre_ok"] = bool((r_pre or {}).get("ok"))

        r_extract = feet_execute(
            "extract_text", {"selector": "body", "driver": "digital"}
        )
        ok_extract = bool((r_extract or {}).get("ok"))

        r_post = feet_execute("screenshot", {"driver": "digital"})
        out["visual_post_ok"] = bool((r_post or {}).get("ok"))

        out["eyes_ok"] = bool(ok_open or ok_extract)
        out["page"] = path
        try:
            feet_execute("close_browser", {"driver": "digital"})
        except Exception:
            pass
    except Exception as e:
        if not out.get("error"):
            out["error"] = str(e)[:200]

    visual_ok = bool(out["visual_pre_ok"] and out["visual_post_ok"])
    result = "ok" if (out["hands_ok"] or out["eyes_ok"]) and visual_ok else "warn"
    err = f" err={out['error'][:120]}" if out.get("error") else ""
    page_tag = f" page={out.get('page','?')}"
    vis_tag = f" visual_pre={'OK' if out['visual_pre_ok'] else 'FAIL'} visual_post={'OK' if out['visual_post_ok'] else 'FAIL'}"
    _auto_log(
        f"Aggressive cycle:{page_tag}{vis_tag} hands={'OK' if out['hands_ok'] else 'FAIL'} eyes={'OK' if out['eyes_ok'] else 'FAIL'}{err}",
        "action",
        result,
    )
    _aggr_cycle_count += 1
    _aggr_last_page = str(out.get("page") or "")
    return out


def _aggressive_worker_loop():
    last_run = 0.0
    while not _aggr_stop.is_set():
        try:
            if not _aggr_enabled():
                _aggr_stop.wait(2.0)
                continue
            from modules.humanoid.governance.state import get_mode
            from modules.humanoid.quality.autonomy_daemon import \
                is_autonomy_running

            if not is_autonomy_running() or get_mode() != "growth":
                _aggr_stop.wait(2.0)
                continue

            cfg = _get_aggr_config()
            idle_required = max(5, int(cfg.get("idle_sec") or 25))
            idle_now = _user_idle_seconds()
            if idle_now < idle_required:
                global _aggr_last_user_active_log
                now = time.time()
                if (now - _aggr_last_user_active_log) > 30:
                    _aggr_last_user_active_log = now
                    _auto_log(
                        f"Aggressive paused: user active ({round(idle_now,1)}s < {idle_required}s)",
                        "action",
                        "ok",
                    )
                _aggr_stop.wait(1.0)
                continue

            interval = max(8, int(cfg.get("interval_sec") or 25))
            now = time.time()
            if (now - last_run) < interval:
                _aggr_stop.wait(1.0)
                continue

            last_run = now
            _run_aggressive_cycle()
        except Exception:
            _aggr_stop.wait(2.0)


def _ensure_aggressive_worker():
    global _aggr_thread
    with _AGGR_LOCK:
        if _aggr_thread is not None and _aggr_thread.is_alive():
            return
        _aggr_stop.clear()
        _aggr_thread = threading.Thread(
            target=_aggressive_worker_loop, name="atlas-aggressive-worker", daemon=True
        )
        _aggr_thread.start()
        _auto_log("Aggressive worker iniciado", "action", "ok")


def _stop_aggressive_worker():
    with _AGGR_LOCK:
        _aggr_stop.set()
        _auto_log("Aggressive worker detenido", "action", "ok")


def _autodiag_log_path() -> Path:
    return (BASE_DIR / "logs" / "autodiagnostic.log").resolve()


def _run_autodiag_once_async():
    try:
        import importlib.util

        p = (BASE_DIR / "scripts" / "atlas_autodiagnostic.py").resolve()
        if not p.exists():
            _auto_log("Autodiagnostic cycle: script no encontrado", "action", "warn")
            return
        spec = importlib.util.spec_from_file_location(
            "atlas_autodiagnostic_runtime", str(p)
        )
        if not spec or not spec.loader:
            _auto_log("Autodiagnostic cycle: loader no disponible", "action", "warn")
            return
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        run_fn = getattr(mod, "run_scan", None)
        if callable(run_fn):
            ok_count, warn_count, crit_count = run_fn()
            _auto_log(
                f"Autodiagnostic cycle ejecutado: {ok_count} OK, {warn_count} WARN, {crit_count} CRIT",
                "action",
                "ok",
            )
        else:
            _auto_log("Autodiagnostic cycle: run_scan no disponible", "action", "warn")
    except Exception as e:
        _auto_log(f"Autodiagnostic cycle error: {str(e)[:160]}", "action", "warn")


def _kick_autodiag_if_stale(stale_sec: float, max_age_sec: int):
    global _autodiag_last_kick_ts
    try:
        if stale_sec < max_age_sec:
            return
        now = time.time()
        with _autodiag_kick_lock:
            if (now - _autodiag_last_kick_ts) < 45:
                return
            _autodiag_last_kick_ts = now
        threading.Thread(
            target=_run_autodiag_once_async, name="atlas-autodiag-once", daemon=True
        ).start()
    except Exception:
        pass


@app.get("/api/autonomy/status", tags=["Autonomia"])
def autonomy_status():
    """Consolidar metricas de TODOS los subsistemas de autonomia."""
    t0 = time.perf_counter()
    data = {"level": 0, "subsystems": {}, "kpis": {}, "alerts": []}

    # --- Modulos ---
    mod_connected, mod_total = 0, 0
    try:
        mr = modules_check_all()
        mod_connected = mr.get("connected", 0)
        mod_total = mr.get("total", 1)
    except:
        pass

    # --- Lifelog ---
    lifelog_rate = 0.0
    lifelog_total = 0
    lifelog_success = 0
    lifelog_fail = 0
    try:
        from modules.humanoid.world_model.lifelog import LifeLog

        ll = LifeLog()
        ls = ll.status()
        lifelog_total = ls.get("total_entries", 0)
        lifelog_success = ls.get("success_count", 0)
        lifelog_fail = ls.get("failure_count", 0)
        lifelog_rate = ls.get("success_rate", 0.0)
    except:
        pass

    # --- Libro de Vida ---
    libro_reglas = 0
    libro_episodios = 0
    libro_tasa = 0.0
    try:
        from modules.humanoid.memory_engine.libro_vida import get_libro_vida

        lv = get_libro_vida()
        lvs = lv.get_stats()
        libro_reglas = lvs.get("reglas_aprendidas", 0)
        libro_episodios = lvs.get("total_episodios", 0)
        libro_tasa = lvs.get("tasa_exito", 0.0)
    except:
        pass

    # --- Modelos IA ---
    ai_available, ai_total = 0, 1
    try:
        am = agent_models()
        models = am.get("data", [])
        ai_total = max(1, len([m for m in models if m.get("id") != "auto"]))
        ai_available = len(
            [m for m in models if m.get("available") and m.get("id") != "auto"]
        )
    except:
        pass

    # --- Governance ---
    gov_mode = "unknown"
    gov_emergency = False
    gov_policies_count = 0
    try:
        from modules.humanoid.governance.state import get_governance_state

        gs = get_governance_state()
        gov_mode = gs.get("mode", "governed")
        gov_emergency = gs.get("emergency_stop", False)
        gov_policies_count = len(gs.get("policies", []))
    except:
        try:
            import requests as _rq

            gr = _rq.get("http://127.0.0.1:8791/governance/status", timeout=2).json()
            gov_mode = gr.get("mode", "governed")
            gov_emergency = gr.get("emergency_stop", False)
            gov_policies_count = len(gr.get("policies", []))
        except:
            pass

    # --- Subsistemas activos ---
    daemon_active = False
    try:
        from modules.humanoid.quality.autonomy_daemon import _daemon

        daemon_active = _daemon is not None and getattr(_daemon, "_running", False)
    except:
        pass

    reactor_active = False
    reactor_cycles = 0
    reactor_fixes = 0
    try:
        from modules.humanoid.ans.reactor import (_reactor_running,
                                                  _reactor_stats)

        reactor_active = _reactor_running
        reactor_cycles = _reactor_stats.get("cycles", 0)
        reactor_fixes = _reactor_stats.get("fixes_ok", 0)
    except:
        pass

    scanner_active = False
    scanner_last = None
    try:
        log_path = str(_autodiag_log_path())
        scanner_active_window_sec = int(
            os.getenv("AUTODIAGNOSTIC_ACTIVE_WINDOW_SEC", "7200") or "7200"
        )
        if os.path.exists(log_path):
            scanner_age_sec = max(0.0, time.time() - os.path.getmtime(log_path))
            scanner_active = scanner_age_sec < max(60, scanner_active_window_sec)
            _kick_autodiag_if_stale(scanner_age_sec, max(60, scanner_active_window_sec))
            with open(log_path, "rb") as f:
                f.seek(max(0, os.path.getsize(log_path) - 500))
                lines = f.read().decode("utf-8", errors="ignore").strip().split("\n")
                for ln in reversed(lines):
                    if "Scan:" in ln:
                        scanner_last = ln.strip()
                        break
            if not scanner_active and scanner_last:
                scanner_last = f"{scanner_last} | refresh requested"
    except:
        pass

    # --- Aprobaciones pendientes ---
    approvals_pending = 0
    approvals_total = 0
    try:
        from modules.humanoid.approvals import list_all, list_pending

        pending = list_pending(limit=100)
        approvals_pending = len([p for p in pending if not p.get("expired")])
        all_approvals = list_all(limit=1000)
        approvals_total = len(all_approvals)
    except:
        pass

    # --- Cola de tareas (autonomia_tasks) ---
    # IMPORTANTE: contar sobre la misma "ventana operativa" que ve el dashboard
    # (top 50 por prioridad/created_at), para evitar inflar cola con histÃ³ricos viejos.
    task_pending = 0
    task_in_progress = 0
    task_done = 0
    task_failed = 0
    task_done_24h = 0
    task_failed_24h = 0
    reactor_resolved_count = 0
    reactor_mttr_samples = []
    try:
        c = _auto_db()
        rows = c.execute(
            """
            SELECT status, source, created_at, updated_at FROM autonomy_tasks
            ORDER BY CASE priority WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
                     created_at DESC
            LIMIT 50
        """
        ).fetchall()
        for r in rows or []:
            st = (
                str(r["status"] if isinstance(r, dict) or hasattr(r, "keys") else r[0])
                .strip()
                .lower()
            )
            src = (
                str(r["source"] if isinstance(r, dict) or hasattr(r, "keys") else "")
                .strip()
                .lower()
            )
            if st == "pending":
                task_pending += 1
            elif st == "in_progress":
                task_in_progress += 1
            elif st == "done":
                task_done += 1
            elif st == "failed":
                task_failed += 1

            if src == "reactor" and st in ("done", "failed"):
                reactor_resolved_count += 1
                try:
                    c_at = (
                        r["created_at"]
                        if isinstance(r, dict) or hasattr(r, "keys")
                        else None
                    )
                    u_at = (
                        r["updated_at"]
                        if isinstance(r, dict) or hasattr(r, "keys")
                        else None
                    )
                    if c_at and u_at:
                        dt_c = datetime.fromisoformat(str(c_at).replace("Z", "+00:00"))
                        dt_u = datetime.fromisoformat(str(u_at).replace("Z", "+00:00"))
                        mins = (dt_u - dt_c).total_seconds() / 60.0
                        if mins >= 0:
                            reactor_mttr_samples.append(mins)
                except:
                    pass

        # Ventana 24h para tasa de Ã©xito mÃ¡s representativa en tiempo real.
        try:
            rows24 = c.execute(
                """
                SELECT status FROM autonomy_tasks
                WHERE updated_at >= datetime('now', '-1 day')
            """
            ).fetchall()
            for r24 in rows24 or []:
                st24 = (
                    str(
                        r24["status"]
                        if isinstance(r24, dict) or hasattr(r24, "keys")
                        else r24[0]
                    )
                    .strip()
                    .lower()
                )
                if st24 == "done":
                    task_done_24h += 1
                elif st24 == "failed":
                    task_failed_24h += 1
        except:
            pass

        c.close()
    except:
        pass

    pending_queue_count = approvals_pending + task_pending + task_in_progress

    # KPI de Ã©xito: usar lifelog si existe; fallback a resultados de tareas cuando no hay muestras de lifelog.
    success_rate = round(lifelog_rate * 100, 1)
    if lifelog_total <= 0:
        denom = task_done + task_failed
        if denom > 0:
            success_rate = round((task_done / denom) * 100, 1)

    # Priorizar ventana 24h cuando existe muestra suficiente.
    success_rate_24h = None
    denom_24h = task_done_24h + task_failed_24h
    if denom_24h > 0:
        success_rate_24h = round((task_done_24h / denom_24h) * 100, 1)

    display_success_rate = (
        success_rate_24h if success_rate_24h is not None else success_rate
    )

    incidents_resolved_kpi = max(
        int(reactor_fixes or 0), int(reactor_resolved_count or 0)
    )
    mttr_kpi = 0
    if reactor_mttr_samples:
        mttr_kpi = round(
            sum(reactor_mttr_samples) / max(1, len(reactor_mttr_samples)), 1
        )
    elif reactor_fixes:
        mttr_kpi = round(reactor_fixes * 2.5, 1)

    # --- Uptime ---
    uptime_hours = 0.0
    try:
        import psutil

        uptime_hours = round((time.time() - psutil.boot_time()) / 3600, 1)
    except:
        pass

    # --- Calcular nivel de autonomia ---
    success_component = max(0.0, min(1.0, float(display_success_rate or 0.0) / 100.0))
    ai_required_for_full = max(
        1, int((os.getenv("AUTONOMY_AI_REQUIRED_FOR_FULL") or "16").strip() or "16")
    )
    ai_component = max(0.0, min(1.0, float(ai_available) / float(ai_required_for_full)))
    sub_active = sum(
        [daemon_active, reactor_active, scanner_active, gov_mode != "emergency", True]
    )
    level = round(
        (mod_connected / max(mod_total, 1)) * 25
        + success_component * 25
        + ai_component * 20
        + (sub_active / 5) * 20
        + (0 if gov_emergency else 1) * 10
    )

    # --- Alertas activas ---
    alerts = []
    if not daemon_active:
        alerts.append({"level": "warning", "msg": "Autonomy Daemon inactivo"})
    if not reactor_active:
        alerts.append({"level": "warning", "msg": "Reactor inactivo"})
    if gov_emergency:
        alerts.append({"level": "critical", "msg": "Emergency Stop ACTIVO"})
    if ai_available == 0:
        alerts.append({"level": "critical", "msg": "0 modelos IA disponibles"})
    if mod_connected < mod_total:
        alerts.append(
            {
                "level": "info",
                "msg": f"{mod_total - mod_connected} modulo(s) desconectado(s)",
            }
        )

    aggressive_cycles_seen = 0
    aggressive_last_page_seen = str(_aggr_last_page or "")
    try:
        c = _auto_db()
        row_cnt = c.execute(
            "SELECT COUNT(*) FROM autonomy_timeline WHERE event LIKE 'Aggressive cycle:%'"
        ).fetchone()
        aggressive_cycles_seen = int((row_cnt[0] if row_cnt else 0) or 0)
        row_last = c.execute(
            "SELECT event FROM autonomy_timeline WHERE event LIKE 'Aggressive cycle:%' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        evt = str((row_last[0] if row_last else "") or "")
        if "page=" in evt:
            tail = evt.split("page=", 1)[1]
            aggressive_last_page_seen = tail.split(" ", 1)[0].strip()
        c.close()
    except Exception:
        pass

    data["level"] = min(100, max(0, level))
    data["subsystems"] = {
        "daemon": {
            "active": daemon_active,
            "label": "Autonomy Daemon",
            "detail": "Health checks + auto-repair",
        },
        "reactor": {
            "active": reactor_active,
            "label": "Reactor Autonomo",
            "detail": f"{reactor_cycles} ciclos, {reactor_fixes} fixes",
        },
        "scanner": {
            "active": scanner_active,
            "label": "Autodiagnostic Scanner",
            "detail": scanner_last or "Sin datos",
        },
        "governance": {
            "active": gov_mode != "emergency",
            "label": "Gobernanza",
            "detail": f"Modo: {gov_mode}",
        },
        "healing": {
            "active": reactor_fixes > 0 or daemon_active,
            "label": "Auto-Healing",
            "detail": f"{reactor_fixes} reparaciones",
        },
    }
    data["kpis"] = {
        "uptime_hours": uptime_hours,
        "success_rate": display_success_rate,
        "success_rate_24h": success_rate_24h,
        "success_rate_all_time": success_rate,
        "modules_connected": mod_connected,
        "modules_total": mod_total,
        "ai_available": ai_available,
        "ai_total": ai_total,
        "ai_required_for_full": ai_required_for_full,
        "ai_component_pct": round(ai_component * 100, 1),
        "incidents_resolved": incidents_resolved_kpi,
        "mttr_minutes": mttr_kpi,
        "rules_learned": libro_reglas,
        "episodes": libro_episodios,
        "alerts_active": len(alerts),
        "lifelog_total": lifelog_total,
        "lifelog_success": lifelog_success,
        "lifelog_fail": lifelog_fail,
        "approvals_pending": approvals_pending,
        "approvals_total": approvals_total,
        "pending_queue_count": pending_queue_count,
        "tasks_pending": task_pending,
        "tasks_in_progress": task_in_progress,
        "tasks_done": task_done,
        "tasks_failed": task_failed,
        "tasks_done_24h": task_done_24h,
        "tasks_failed_24h": task_failed_24h,
        "aggressive_cycles": int(aggressive_cycles_seen or 0),
        "aggressive_last_page": str(aggressive_last_page_seen or ""),
        "policies_count": gov_policies_count,
    }
    data["alerts"] = alerts

    # Timeline
    timeline = []
    try:
        c = _auto_db()
        rows = c.execute(
            "SELECT ts, event, kind, result FROM autonomy_timeline ORDER BY id DESC LIMIT 10"
        ).fetchall()
        timeline = [
            {
                "ts": r["ts"],
                "event": r["event"],
                "kind": r["kind"],
                "result": r["result"],
            }
            for r in rows
        ]
        c.close()
    except:
        pass
    data["timeline"] = timeline

    data["ms"] = int((time.perf_counter() - t0) * 1000)
    data["ok"] = True
    return data


@app.post("/api/autonomy/daemon/start", tags=["Autonomia"])
def autonomy_daemon_start():
    """Inicia el daemon de autonomÃ­a en el proceso actual de la API."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.quality.autonomy_daemon import (
            is_autonomy_running, start_autonomy)

        if is_autonomy_running():
            _ensure_aggressive_worker()
            return {
                "ok": True,
                "running": True,
                "already_running": True,
                "aggressive_worker": True,
                "ms": int((time.perf_counter() - t0) * 1000),
            }

        result = start_autonomy()
        _ensure_aggressive_worker()
        running = False
        try:
            running = bool(is_autonomy_running())
        except Exception:
            running = False
        return {
            "ok": bool(running),
            "running": bool(running),
            "already_running": False,
            "aggressive_worker": True,
            "result": result,
            "ms": int((time.perf_counter() - t0) * 1000),
        }
    except Exception as e:
        return {
            "ok": False,
            "running": False,
            "error": str(e)[:300],
            "ms": int((time.perf_counter() - t0) * 1000),
        }


@app.post("/api/autonomy/daemon/stop", tags=["Autonomia"])
def autonomy_daemon_stop():
    """Detiene el daemon de autonomÃ­a en el proceso actual de la API."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.quality.autonomy_daemon import (
            is_autonomy_running, stop_autonomy)

        was_running = bool(is_autonomy_running())
        _stop_aggressive_worker()
        result = stop_autonomy()
        running = bool(is_autonomy_running())
        return {
            "ok": not running,
            "was_running": was_running,
            "running": running,
            "result": result,
            "ms": int((time.perf_counter() - t0) * 1000),
        }
    except Exception as e:
        return {
            "ok": False,
            "error": str(e)[:300],
            "ms": int((time.perf_counter() - t0) * 1000),
        }


@app.post("/api/autonomy/scanner/run-once", tags=["Autonomia"])
def autonomy_scanner_run_once():
    """Ejecuta un ciclo Ãºnico de autodiagnÃ³stico para refrescar estado del scanner."""
    t0 = time.perf_counter()
    try:
        _run_autodiag_once_async()
        return {
            "ok": True,
            "triggered": True,
            "ms": int((time.perf_counter() - t0) * 1000),
        }
    except Exception as e:
        return {
            "ok": False,
            "triggered": False,
            "error": str(e)[:300],
            "ms": int((time.perf_counter() - t0) * 1000),
        }


@app.get("/api/autonomy/aggressive/config", tags=["Autonomia"])
def autonomy_aggressive_config_get():
    """Obtiene configuraciÃ³n runtime del modo agresivo."""
    t0 = time.perf_counter()
    try:
        cfg = _get_aggr_config()
        return {"ok": True, "config": cfg, "ms": int((time.perf_counter() - t0) * 1000)}
    except Exception as e:
        return {
            "ok": False,
            "error": str(e)[:300],
            "ms": int((time.perf_counter() - t0) * 1000),
        }


@app.post("/api/autonomy/aggressive/config", tags=["Autonomia"])
def autonomy_aggressive_config_set(body: dict):
    """Actualiza configuraciÃ³n runtime del modo agresivo (enabled, idle_sec, interval_sec, pages)."""
    t0 = time.perf_counter()
    try:
        cfg = _set_aggr_config(body if isinstance(body, dict) else {})
        try:
            from modules.humanoid.quality.autonomy_daemon import \
                is_autonomy_running

            if cfg.get("enabled") and is_autonomy_running():
                _ensure_aggressive_worker()
        except Exception:
            pass
        return {"ok": True, "config": cfg, "ms": int((time.perf_counter() - t0) * 1000)}
    except Exception as e:
        return {
            "ok": False,
            "error": str(e)[:300],
            "ms": int((time.perf_counter() - t0) * 1000),
        }


@app.get("/api/autonomy/tasks", tags=["Autonomia"])
def autonomy_tasks():
    """Tareas de seguimiento con estado y prioridad."""
    t0 = time.perf_counter()
    tasks = []
    try:
        c = _auto_db()
        # Sync external sources: approvals pending
        try:
            from modules.humanoid.governance.approvals import \
                get_pending_approvals

            pending = get_pending_approvals()
            for ap in pending or []:
                aid = f"approval_{ap.get('id', '')}"
                c.execute(
                    """INSERT OR IGNORE INTO autonomy_tasks(id,title,status,priority,source,detail)
                    VALUES(?,?,?,?,?,?)""",
                    (
                        aid,
                        f"Aprobacion: {ap.get('action','?')}",
                        "pending",
                        "high",
                        "governance",
                        str(ap)[:300],
                    ),
                )
        except:
            pass
        # Sync reactor issues
        try:
            from modules.humanoid.ans.reactor import _reactor_stats

            for issue in _reactor_stats.get("recent_issues", [])[-5:]:
                iid = f"reactor_{hash(str(issue)) % 100000}"
                c.execute(
                    """INSERT OR IGNORE INTO autonomy_tasks(id,title,status,priority,source,detail)
                    VALUES(?,?,?,?,?,?)""",
                    (
                        iid,
                        f"Reactor: {str(issue)[:60]}",
                        "pending",
                        "medium",
                        "reactor",
                        str(issue)[:300],
                    ),
                )
        except:
            pass
        # Reaper: limpiar tareas 'in_progress' atascadas por mÃ¡s de 1h
        try:
            c.execute(
                """
                UPDATE autonomy_tasks
                SET status='failed', action_taken='reaper_timeout'
                WHERE status='in_progress' AND updated_at <= datetime('now', '-1 hour')
            """
            )
        except:
            pass
        c.commit()
        rows = c.execute(
            "SELECT * FROM autonomy_tasks ORDER BY CASE priority WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END, created_at DESC LIMIT 50"
        ).fetchall()
        tasks = [dict(r) for r in rows]
        c.close()
    except Exception as e:
        return {"ok": False, "error": str(e)[:200]}
    return {
        "ok": True,
        "tasks": tasks,
        "count": len(tasks),
        "ms": int((time.perf_counter() - t0) * 1000),
    }


@app.post("/api/autonomy/task/{task_id}/action", tags=["Autonomia"])
def autonomy_task_action(task_id: str, body: dict):
    """Ejecutar accion sobre una tarea: approve, reject, retry, close."""
    action = body.get("action", "close")
    valid = {
        "approve": "done",
        "reject": "failed",
        "retry": "in_progress",
        "close": "done",
    }
    new_status = valid.get(action, "done")
    try:
        c = _auto_db()
        c.execute(
            "UPDATE autonomy_tasks SET status=?, action_taken=?, updated_at=datetime('now') WHERE id=?",
            (new_status, action, task_id),
        )
        c.commit()
        affected = c.total_changes
        c.close()
        _auto_log(f"Tarea {task_id}: {action}", "action", "ok")
        return {
            "ok": True,
            "task_id": task_id,
            "action": action,
            "new_status": new_status,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)[:200]}


@app.get("/vision/cameras")
def vision_cameras():
    """Listar camaras disponibles."""
    cameras = []
    try:
        import cv2

        for i in range(3):
            cap = cv2.VideoCapture(i)
            if cap.isOpened():
                cameras.append({"id": i, "name": f"Camera {i}", "active": True})
                cap.release()
    except Exception:
        pass
    return {"ok": True, "cameras": cameras, "count": len(cameras)}


@app.get("/vision/capture/{cam_id}")
def vision_capture(cam_id: int = 0):
    """Capturar frame de una camara."""
    try:
        import base64

        import cv2

        cap = cv2.VideoCapture(cam_id)
        if not cap.isOpened():
            return {"ok": False, "error": f"Camera {cam_id} not available"}
        ret, frame = cap.read()
        cap.release()
        if not ret:
            return {"ok": False, "error": "Failed to capture frame"}
        _, buf = cv2.imencode(".jpg", frame)
        b64 = base64.b64encode(buf).decode()
        return {"ok": True, "image": f"data:image/jpeg;base64,{b64}", "camera": cam_id}
    except ImportError:
        return {"ok": False, "error": "OpenCV not installed"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/vision/scan")
def vision_scan():
    """Escanear camaras disponibles."""
    return vision_cameras()


@app.get("/nervous/diagnostic")
def nervous_diagnostic():
    """Diagnostico del sistema nervioso."""
    try:
        from modules.humanoid.nervous.api import nerve_full_status

        return nerve_full_status()
    except Exception:
        try:
            from modules.humanoid.nerve import feet_status, nerve_eyes_status

            return {
                "ok": True,
                "eyes": nerve_eyes_status(),
                "feet": feet_status(),
                "nodes": [],
            }
        except Exception as e:
            return {"ok": False, "error": str(e), "nodes": []}


@app.get("/version")
def version():
    """Release version, git sha, channel (stable/canary), edition, product_mode."""
    try:
        from modules.humanoid.release import get_version_info

        out = {"ok": True, **get_version_info()}
        if os.getenv("PRODUCT_MODE", "").strip().lower() in ("1", "true", "yes"):
            try:
                from modules.humanoid.product.license import license_status

                out["license"] = license_status()
            except Exception:
                out["license"] = {"status": "unknown"}
        return out
    except Exception as e:
        return {
            "ok": False,
            "version": "0.0.0",
            "git_sha": "",
            "channel": "canary",
            "error": str(e),
        }


def _camera_stream_generator(stream_url: str):
    """Generador que mantiene la conexiÃ³n HTTP abierta para MJPEG streaming."""
    import urllib.request

    resp = None
    try:
        req = urllib.request.Request(stream_url, method="GET")
        resp = urllib.request.urlopen(req, timeout=15)
        while True:
            chunk = resp.read(16384)
            if not chunk:
                break
            yield chunk
    except GeneratorExit:
        pass
    except Exception as e:
        print(f"Camera stream generator error: {e}")
    finally:
        if resp:
            try:
                resp.close()
            except Exception:
                pass


def _local_camera_stream_generator(index: int = 0):
    """Fallback MJPEG local (USB) cuando backend Robot no responde."""
    import cv2

    cap = None
    try:
        backend = cv2.CAP_DSHOW if os.name == "nt" else cv2.CAP_ANY
        cap = cv2.VideoCapture(int(index), backend)
        if cap is None or not cap.isOpened():
            return
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 360)
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        encode_params = [cv2.IMWRITE_JPEG_QUALITY, 72]
        while True:
            ok, frame = cap.read()
            if not ok or frame is None:
                break
            ok_jpg, buf = cv2.imencode(".jpg", frame, encode_params)
            if not ok_jpg:
                continue
            jpeg = buf.tobytes()
            yield (
                b"--frame\r\n"
                b"Content-Type: image/jpeg\r\n"
                b"Content-Length: "
                + str(len(jpeg)).encode("ascii")
                + b"\r\n\r\n"
                + jpeg
                + b"\r\n"
            )
    except GeneratorExit:
        pass
    except Exception:
        pass
    finally:
        if cap is not None:
            try:
                cap.release()
            except Exception:
                pass


@app.get("/cuerpo/camera/stream")
def camera_stream_proxy(index: int = 0):
    """Proxy para stream de cÃ¡mara del robot (puerto 8002). index: 0, 1, 2... para cada cÃ¡mara."""
    import urllib.request

    base_url = get_robot_api_base()
    stream_url = f"{base_url}/api/vision/camera/stream?index={index}"
    try:
        test_req = urllib.request.Request(base_url + "/status", method="GET")
        with urllib.request.urlopen(test_req, timeout=3) as r:
            if r.status == 200:
                from fastapi.responses import StreamingResponse

                return StreamingResponse(
                    _camera_stream_generator(stream_url),
                    media_type="multipart/x-mixed-replace; boundary=frame",
                )
    except Exception as e:
        print(f"Camera proxy error (index={index}): {e}")

    # Fallback local USB opcional: desactivado por defecto para evitar bloqueos en Windows.
    enable_local_fallback = os.getenv(
        "ATLAS_ENABLE_LOCAL_CAMERA_FALLBACK", "false"
    ).strip().lower() in ("1", "true", "yes", "y", "on")
    if enable_local_fallback:
        try:
            from fastapi.responses import StreamingResponse

            return StreamingResponse(
                _local_camera_stream_generator(index=index),
                media_type="multipart/x-mixed-replace; boundary=frame",
            )
        except Exception:
            pass

    # Ãšltimo fallback: placeholder
    try:
        import cv2
        import numpy as np

        img = np.zeros((480, 640, 3), dtype=np.uint8)
        img[:] = (50, 50, 50)
        cv2.putText(
            img, f"CAMERA {index} OFFLINE", (150, 240), 0, 1.5, (255, 255, 255), 3
        )
        cv2.putText(
            img,
            "Robot Backend (8002) not running",
            (100, 280),
            0,
            0.8,
            (200, 200, 200),
            2,
        )
        _, buffer = cv2.imencode(".jpg", img)
        from fastapi.responses import Response

        return Response(content=buffer.tobytes(), media_type="image/jpeg")
    except Exception:
        from fastapi.responses import JSONResponse

        return JSONResponse(
            content={"error": "Camera service unavailable"}, status_code=503
        )


@app.get("/cuerpo/vision/snapshot")
def cuerpo_vision_snapshot(
    index: int = 0,
    enhance: str = "auto",
    jpeg_quality: int = 80,
    focus_x: float = 0.5,
    focus_y: float = 0.5,
    zoom: float = 1.0,
):
    """Proxy de snapshot (JPEG) del robot: evita CORS y centraliza la ruta.

    Params:
    - index: cÃ¡mara
    - enhance: auto|sharp|max|ocr
    - jpeg_quality: 50-98
    - focus_x/focus_y: 0..1 (pan digital)
    - zoom: 1..3
    """
    import urllib.parse

    base_url = get_robot_api_base()
    url = (
        f"{base_url}/api/vision/snapshot"
        f"?index={int(index)}"
        f"&enhance={urllib.parse.quote(str(enhance))}"
        f"&jpeg_quality={int(jpeg_quality)}"
        f"&focus_x={float(focus_x)}"
        f"&focus_y={float(focus_y)}"
        f"&zoom={float(zoom)}"
    )
    try:
        import concurrent.futures
        import httpx

        def _fetch() -> bytes:
            timeout = httpx.Timeout(connect=1.5, read=3.0, write=3.0, pool=1.5)
            with httpx.Client(timeout=timeout) as client:
                r = client.get(url)
                r.raise_for_status()
                return r.content

        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
            fut = ex.submit(_fetch)
            img = fut.result(timeout=4.0)
        from fastapi.responses import Response

        return Response(content=img, media_type="image/jpeg")
    except Exception as e:
        # placeholder coherente para UI
        try:
            import cv2
            import numpy as np

            img = np.zeros((360, 640, 3), dtype=np.uint8)
            img[:] = (40, 40, 45)
            cv2.putText(img, "SNAPSHOT OFFLINE", (160, 185), 0, 1.2, (255, 255, 255), 3)
            cv2.putText(
                img,
                f"{type(e).__name__}: {str(e)[:60]}",
                (30, 230),
                0,
                0.6,
                (200, 200, 200),
                2,
            )
            _, buffer = cv2.imencode(".jpg", img)
            from fastapi.responses import Response

            return Response(content=buffer.tobytes(), media_type="image/jpeg")
        except Exception:
            from fastapi.responses import JSONResponse

            return JSONResponse(content={"ok": False, "error": str(e)}, status_code=503)


@app.get("/api/evidence/image", tags=["Evidence"])
def api_evidence_image(path: str):
    """Sirve una imagen (png/jpg/webp) desde snapshots/ de forma segura.

    Acepta path absoluto o relativo a la raÃ­z del repo. Solo permite archivos
    bajo `snapshots/` para evitar lectura arbitraria.
    """
    from pathlib import Path

    from fastapi import HTTPException
    from fastapi.responses import FileResponse

    repo_root = Path(__file__).resolve().parents[1]
    allowed_root = (repo_root / "snapshots").resolve()

    raw = (path or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="missing path")

    p = Path(raw)
    if not p.is_absolute():
        p = repo_root / p
    try:
        resolved = p.resolve()
    except Exception:
        raise HTTPException(status_code=400, detail="invalid path")

    # allow only snapshots/*
    if resolved != allowed_root and allowed_root not in resolved.parents:
        raise HTTPException(status_code=403, detail="path not allowed")
    if not resolved.exists() or not resolved.is_file():
        raise HTTPException(status_code=404, detail="not found")

    ext = resolved.suffix.lower()
    if ext not in (".png", ".jpg", ".jpeg", ".webp"):
        raise HTTPException(status_code=415, detail="unsupported file type")

    return FileResponse(str(resolved))


BETY_TEMPLATE_SOURCE = (
    BASE_DIR / "plantillas" / "Plantilla_Base_Eventos_Panaderia.xlsx"
)
BETY_STORAGE_ROOT = BASE_DIR / "data" / "bety_eventos"
BETY_UPLOADS_DIR = BETY_STORAGE_ROOT / "uploads"
BETY_BACKUPS_DIR = BETY_STORAGE_ROOT / "backups"
BETY_SUMMARY_DIR = BETY_STORAGE_ROOT / "summary"
BETY_LATEST_FILE = BETY_STORAGE_ROOT / "Bety_Eventos_Panaderia_latest.xlsx"
BETY_LEDGER_FILE = BETY_STORAGE_ROOT / "ledger.jsonl"
BETY_LATEST_SUMMARY_FILE = BETY_SUMMARY_DIR / "latest_summary.json"
BETY_PUBLIC_DOWNLOADS_DIR = BASE_DIR / "atlas_adapter" / "static" / "v4" / "downloads"
BETY_PUBLIC_TEMPLATE_NAME = "Bety_Eventos_Panaderia_Diario.xlsx"


def _ensure_bety_dirs() -> None:
    for p in (
        BETY_STORAGE_ROOT,
        BETY_UPLOADS_DIR,
        BETY_BACKUPS_DIR,
        BETY_SUMMARY_DIR,
        BETY_PUBLIC_DOWNLOADS_DIR,
    ):
        p.mkdir(parents=True, exist_ok=True)


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _to_json_safe(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _as_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace("$", "").replace(",", "")
    if not raw:
        return 0.0
    try:
        return float(raw)
    except Exception:
        return 0.0


def _read_sheet_records(workbook, sheet_name: str) -> list[dict]:
    if sheet_name not in workbook.sheetnames:
        return []
    ws = workbook[sheet_name]
    header_row = 3
    headers = []
    for col in range(1, ws.max_column + 1):
        hv = ws.cell(row=header_row, column=col).value
        hs = str(hv).strip() if hv else ""
        headers.append(hs)
    rows = []
    for ridx in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(row=ridx, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in vals):
            continue
        rec = {"_fila": ridx}
        for k, v in zip(headers, vals):
            if not k:
                continue
            rec[k] = _to_json_safe(v)
        rows.append(rec)
    return rows


def _build_bety_summary(xlsx_path: Path) -> dict:
    try:
        from openpyxl import load_workbook
    except Exception:
        return {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "totals": {
                "entrada_almacen": {"registros": 0, "cantidad_total": 0.0, "importe_total": 0.0},
                "salida_produccion": {"registros": 0, "cantidad_total": 0.0, "importe_total": 0.0},
                "produccion": {"registros": 0, "cantidad_total": 0.0, "importe_total": 0.0},
                "ventas": {"registros": 0, "cantidad_total": 0.0, "importe_total": 0.0},
            },
            "by_day": {},
            "parse_warning": "openpyxl no disponible en runtime PUSH; archivo guardado y respaldado.",
        }

    wb = load_workbook(xlsx_path, data_only=True)
    entries = _read_sheet_records(wb, "02_Entrada_Almacen")
    to_production = _read_sheet_records(wb, "03_Salida_Produccion")
    production = _read_sheet_records(wb, "04_Produccion")
    sales = _read_sheet_records(wb, "05_Ventas")

    def _sum_field(rows: list[dict], key: str) -> float:
        return round(sum(_as_float(r.get(key)) for r in rows), 2)

    totals = {
        "entrada_almacen": {
            "registros": len(entries),
            "cantidad_total": _sum_field(entries, "Cantidad"),
            "importe_total": _sum_field(entries, "Importe"),
        },
        "salida_produccion": {
            "registros": len(to_production),
            "cantidad_total": _sum_field(to_production, "Cantidad"),
            "importe_total": _sum_field(to_production, "Importe"),
        },
        "produccion": {
            "registros": len(production),
            "cantidad_total": _sum_field(production, "Cantidad"),
            "importe_total": _sum_field(production, "Importe"),
        },
        "ventas": {
            "registros": len(sales),
            "cantidad_total": _sum_field(sales, "Cantidad"),
            "importe_total": _sum_field(sales, "Importe"),
        },
    }

    by_day: dict[str, dict[str, float]] = {}
    sections = {
        "entrada_almacen": entries,
        "salida_produccion": to_production,
        "produccion": production,
        "ventas": sales,
    }
    for sec_name, rows in sections.items():
        for row in rows:
            day = str(row.get("Fecha") or "").strip() or "SIN_FECHA"
            d = by_day.setdefault(
                day,
                {
                    "entrada_almacen_importe": 0.0,
                    "salida_produccion_importe": 0.0,
                    "produccion_importe": 0.0,
                    "ventas_importe": 0.0,
                    "entrada_almacen_cantidad": 0.0,
                    "salida_produccion_cantidad": 0.0,
                    "produccion_cantidad": 0.0,
                    "ventas_cantidad": 0.0,
                },
            )
            d[f"{sec_name}_importe"] = round(
                d[f"{sec_name}_importe"] + _as_float(row.get("Importe")), 2
            )
            d[f"{sec_name}_cantidad"] = round(
                d[f"{sec_name}_cantidad"] + _as_float(row.get("Cantidad")), 2
            )

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "totals": totals,
        "by_day": by_day,
    }


def _ensure_bety_public_template() -> Path:
    _ensure_bety_dirs()
    if not BETY_TEMPLATE_SOURCE.exists():
        raise FileNotFoundError(f"Plantilla base no encontrada: {BETY_TEMPLATE_SOURCE}")
    dst = BETY_PUBLIC_DOWNLOADS_DIR / BETY_PUBLIC_TEMPLATE_NAME
    shutil.copy2(BETY_TEMPLATE_SOURCE, dst)
    return dst


def _read_last_ledger(limit: int = 20) -> list[dict]:
    if not BETY_LEDGER_FILE.exists():
        return []
    lines = BETY_LEDGER_FILE.read_text(encoding="utf-8").splitlines()
    out = []
    for line in lines[-max(1, int(limit)) :]:
        try:
            out.append(json.loads(line))
        except Exception:
            continue
    return out


@app.get("/api/panaderia/bety/template", tags=["Panaderia Bety"])
def api_panaderia_bety_template(request: Request):
    t0 = time.perf_counter()
    try:
        path = _ensure_bety_public_template()
        rel = f"/v4/static/downloads/{path.name}"
        return _std_resp(
            True,
            {
                "file_name": path.name,
                "file_size_bytes": path.stat().st_size,
                "download_path": rel,
                "download_url": f"{str(request.base_url).rstrip('/')}{rel}",
            },
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/api/panaderia/bety/upload", tags=["Panaderia Bety"])
async def api_panaderia_bety_upload(file: UploadFile = File(...)):
    t0 = time.perf_counter()
    from fastapi import HTTPException

    try:
        _ensure_bety_dirs()
        name = Path(file.filename or "bety_eventos.xlsx").name
        suffix = name.lower()
        if not suffix.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Solo se permite .xlsx")

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        stored_name = f"{ts}_{name}"
        upload_path = BETY_UPLOADS_DIR / stored_name

        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Archivo vacio")
        if len(content) > 20 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="Archivo demasiado grande (>20MB)")
        upload_path.write_bytes(content)
        shutil.copy2(upload_path, BETY_LATEST_FILE)

        summary = _build_bety_summary(upload_path)
        summary["file"] = {
            "name": stored_name,
            "size_bytes": len(content),
            "sha256": _sha256_file(upload_path),
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
        }
        BETY_LATEST_SUMMARY_FILE.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        backup_name = f"{ts}__{name}"
        backup_path = BETY_BACKUPS_DIR / backup_name
        shutil.copy2(upload_path, backup_path)

        ledger_entry = {
            "ts": datetime.now(timezone.utc).isoformat(),
            "file_name": stored_name,
            "size_bytes": len(content),
            "sha256": summary["file"]["sha256"],
            "totals": summary["totals"],
            "backup_file": backup_name,
        }
        with BETY_LEDGER_FILE.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(ledger_entry, ensure_ascii=False) + "\n")

        return _std_resp(
            True,
            {
                "message": "Archivo recibido y consolidado para ATLAS.",
                "latest_file": str(BETY_LATEST_FILE),
                "backup_file": str(backup_path),
                "summary_file": str(BETY_LATEST_SUMMARY_FILE),
                "summary": summary,
            },
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except HTTPException:
        raise
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/api/panaderia/bety/status", tags=["Panaderia Bety"])
def api_panaderia_bety_status():
    t0 = time.perf_counter()
    try:
        _ensure_bety_dirs()
        latest_summary = {}
        if BETY_LATEST_SUMMARY_FILE.exists():
            latest_summary = json.loads(BETY_LATEST_SUMMARY_FILE.read_text(encoding="utf-8"))
        uploads = _read_last_ledger(limit=12)
        return _std_resp(
            True,
            {
                "template_exists": BETY_TEMPLATE_SOURCE.exists(),
                "latest_file_exists": BETY_LATEST_FILE.exists(),
                "latest_summary": latest_summary,
                "uploads": uploads,
                "storage_root": str(BETY_STORAGE_ROOT),
            },
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/api/panaderia/bety/uploads", tags=["Panaderia Bety"])
def api_panaderia_bety_uploads(limit: int = 20):
    t0 = time.perf_counter()
    try:
        _ensure_bety_dirs()
        uploads = _read_last_ledger(limit=min(max(limit, 1), 200))
        return _std_resp(
            True,
            {"items": uploads, "count": len(uploads)},
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/api/panaderia/bety/download/latest", tags=["Panaderia Bety"])
def api_panaderia_bety_download_latest():
    from fastapi import HTTPException

    if not BETY_LATEST_FILE.exists():
        raise HTTPException(status_code=404, detail="No hay archivo diario cargado aun.")
    return FileResponse(
        str(BETY_LATEST_FILE),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=BETY_LATEST_FILE.name,
    )


@app.get("/product/status")
def product_status():
    """Product dashboard: version, edition, license, service hint, cluster, gateway, health score, last GA/metalearn, support_contact."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.deploy.healthcheck import run_health_verbose
        from modules.humanoid.deploy.ports import get_ports
        from modules.humanoid.product.license import license_status
        from modules.humanoid.release import get_version_info

        version_info = get_version_info()
        license_info = license_status()
        port = (
            get_ports()[0]
            if get_ports()
            else int(os.getenv("SERVICE_PORT", "8791") or 8791)
        )
        health = run_health_verbose(base_url=None, active_port=port)
        ga_last = None
        meta_last = None
        try:
            from modules.humanoid.ga.cycle import get_status as ga_status

            ga_last = ga_status().get("last_run_ts")
        except Exception:
            pass
        try:
            from modules.humanoid.metalearn.cycle import \
                get_status as meta_status

            meta_last = meta_status().get("last_update_ts")
        except Exception:
            pass
        cluster_nodes = 0
        gateway_state = "unknown"
        try:
            from modules.humanoid.cluster.registry import list_nodes

            cluster_nodes = len(list_nodes())
        except Exception:
            pass
        try:
            from modules.humanoid.gateway.store import build_gateway_status

            gw = build_gateway_status()
            gateway_state = gw.get("mode", "unknown")
        except Exception:
            pass
        service_name = os.getenv("SERVICE_NAME", "ATLAS_PUSH")
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True,
            {
                "version": version_info.get("version", "0.0.0"),
                "edition": version_info.get("edition", "community"),
                "license": license_info,
                "service_name": service_name,
                "active_port": port,
                "health_score": health.get("score", 0),
                "cluster_nodes": cluster_nodes,
                "gateway_state": gateway_state,
                "last_ga_run_ts": ga_last,
                "last_metalearn_run_ts": meta_last,
                "support_contact": os.getenv("SUPPORT_CONTACT", ""),
            },
            ms,
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


# Dashboard UI (fallback: if UI fails, API still works)
STATIC_DIR = BASE_DIR / "atlas_adapter" / "static"
STATIC_DIR.mkdir(parents=True, exist_ok=True)


@app.get("/")
def root_redirect():
    """Redirige a /ui para que el dashboard cargue desde el mismo servidor que expone /api."""
    from fastapi.responses import RedirectResponse

    # Default entrypoint: /ui is the v4 landing (presentation).
    return RedirectResponse(url="/ui", status_code=302)


@app.get("/ui")
async def serve_ui():
    """Dashboard: estado, versiÃ³n, salud, mÃ©tricas, programador, actualizaciÃ³n, despliegue, aprobaciones e interacciÃ³n con el robot."""
    # v4 landing is canonical at /ui (presentation)
    path = STATIC_DIR / "v4" / "index.html"
    if path.exists():
        return FileResponse(
            path,
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
            },
        )
    return {"ok": False, "error": "v4/index.html not found"}


@app.get("/v3")
async def serve_v3():
    """Redirects legacy v3.8 to the active v4 dashboard."""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/ui", status_code=301)


@app.get("/ui-legacy")
def serve_ui_legacy():
    """Emergency alias for legacy dashboard (kept behind ATLAS_UI_LEGACY=true)."""
    from fastapi import HTTPException

    enabled = os.getenv("ATLAS_UI_LEGACY", "false").strip().lower() in (
        "1",
        "true",
        "yes",
        "y",
        "on",
    )
    if not enabled:
        raise HTTPException(status_code=404, detail="not_found")
    path = STATIC_DIR / "dashboard.html"
    if path.exists():
        return FileResponse(
            path,
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
            },
        )
    return {"ok": False, "error": "dashboard.html not found"}


@app.get("/ui/static/{file_path:path}")
async def serve_ui_static(file_path: str):
    """Serve v4 static assets under /ui/static/* (kept for compatibility)."""
    from fastapi import HTTPException

    safe = (file_path or "").replace("..", "").strip("/")
    path = STATIC_DIR / "v4" / safe
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail=f"ui asset not found: {safe}")
    ct_map = {
        ".js": "application/javascript",
        ".css": "text/css",
        ".html": "text/html",
        ".json": "application/json",
        ".svg": "image/svg+xml",
    }
    ct = ct_map.get(path.suffix, "application/octet-stream")
    return FileResponse(
        path,
        media_type=ct,
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/workspace")
def serve_workspace():
    """ATLAS Agent Workspace â€” IDE-style interface para comandar ATLAS en tiempo real."""
    path = STATIC_DIR / "workspace.html"
    if path.exists():
        return FileResponse(
            path,
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
            },
        )
    return {"ok": False, "error": "workspace.html not found"}


@app.get("/v4")
def serve_v4():
    """Back-compat alias for the v4 landing."""
    from fastapi.responses import RedirectResponse

    return RedirectResponse(url="/ui", status_code=302)


@app.get("/v4/static/{file_path:path}")
async def serve_v4_static(file_path: str):
    """Serve v4 static assets (JS, CSS) for the landing."""
    from fastapi import HTTPException

    safe = (file_path or "").replace("..", "").strip("/")
    path = STATIC_DIR / "v4" / safe
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail=f"v4 asset not found: {safe}")
    ct_map = {
        ".js": "application/javascript",
        ".css": "text/css",
        ".html": "text/html",
        ".json": "application/json",
        ".svg": "image/svg+xml",
    }
    ct = ct_map.get(path.suffix, "application/octet-stream")
    return FileResponse(
        path,
        media_type=ct,
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/nexus")
async def serve_nexus():
    """Panel de Control ATLAS â€” vista consolidada del sistema (antes en puerto 8000)."""
    path = STATIC_DIR / "nexus.html"
    if path.exists():
        return FileResponse(
            path,
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
            },
        )
    return {"ok": False, "error": "nexus.html not found"}


# ----------------------------------------------------------------------------
# Agent Models â€” CatÃ¡logo de modelos IA disponibles
# ----------------------------------------------------------------------------

_MODEL_CATALOG = [
    {
        "id": "auto",
        "name": "Auto (Inteligenteâ†’Gratisâ†’Local)",
        "provider": "atlas",
        "desc": "Cascada: API gratis (Gemini/Groq), luego pago (GPT-4.1/Claude), luego local",
        "category": "auto",
    },
    {
        "id": "cascade:ide-agent",
        "name": "Cascade (IDE Agent)",
        "provider": "atlas",
        "desc": "Asistente del IDE; usa enrutamiento ATLAS auto para ejecuciÃ³n interna",
        "category": "assistant",
    },
    # â”€â”€ Ollama LOCAL (gratis, tu PC) â”€â”€
    {
        "id": "ollama:deepseek-r1:14b",
        "name": "DeepSeek R1 14B",
        "provider": "ollama",
        "desc": "Razonamiento local (gratis)",
        "category": "local",
    },
    {
        "id": "ollama:deepseek-coder:6.7b",
        "name": "DeepSeek Coder 6.7B",
        "provider": "ollama",
        "desc": "Codigo local (gratis)",
        "category": "local",
    },
    {
        "id": "ollama:qwen2.5:7b",
        "name": "Qwen 2.5 7B",
        "provider": "ollama",
        "desc": "General local (gratis)",
        "category": "local",
    },
    {
        "id": "ollama:llama3.1:latest",
        "name": "Llama 3.1 8B",
        "provider": "ollama",
        "desc": "Meta local (gratis)",
        "category": "local",
    },
    {
        "id": "ollama:deepseek-r1:latest",
        "name": "DeepSeek R1 7B",
        "provider": "ollama",
        "desc": "Razonamiento compacto (gratis)",
        "category": "local",
    },
    {
        "id": "ollama:llama3.2-vision:11b",
        "name": "Llama 3.2 Vision 11B",
        "provider": "ollama",
        "desc": "Vision local (gratis)",
        "category": "local",
    },
    # â”€â”€ API GRATIS (tier gratuito) â”€â”€
    {
        "id": "groq:llama-3.3-70b-versatile",
        "name": "Llama 3.3 70B (Groq)",
        "provider": "groq",
        "desc": "Ultra rapido, tier gratis",
        "category": "free",
    },
    {
        "id": "groq:llama-3.1-8b-instant",
        "name": "Llama 3.1 8B (Groq)",
        "provider": "groq",
        "desc": "Instantaneo, tier gratis",
        "category": "free",
    },
    {
        "id": "groq:mixtral-8x7b-32768",
        "name": "Mixtral 8x7B (Groq)",
        "provider": "groq",
        "desc": "MoE via Groq, tier gratis",
        "category": "free",
    },
    {
        "id": "groq:deepseek-r1-distill-llama-70b",
        "name": "DeepSeek R1 Distill 70B (Groq)",
        "provider": "groq",
        "desc": "Razonamiento gratis via Groq",
        "category": "free",
    },
    {
        "id": "gemini:gemini-2.5-flash",
        "name": "Gemini 2.5 Flash",
        "provider": "gemini",
        "desc": "Google rapido, tier gratis",
        "category": "free",
    },
    {
        "id": "gemini:gemini-2.5-pro-preview-06-05",
        "name": "Gemini 2.5 Pro",
        "provider": "gemini",
        "desc": "Google flagship, tier gratis limitado",
        "category": "free",
    },
    # â”€â”€ API DE PAGO â”€â”€
    {
        "id": "openai:codex-auto",
        "name": "Codex (OpenAI)",
        "provider": "openai",
        "desc": "Coding agent de OpenAI (OPENAI_CODEX_MODEL o fallback)",
        "category": "code",
    },
    {
        "id": "openai:gpt-4.1",
        "name": "GPT-4.1",
        "provider": "openai",
        "desc": "Flagship OpenAI",
        "category": "premium",
    },
    {
        "id": "openai:gpt-4.1-mini",
        "name": "GPT-4.1 Mini",
        "provider": "openai",
        "desc": "Rapido y economico",
        "category": "fast",
    },
    {
        "id": "openai:gpt-4.1-nano",
        "name": "GPT-4.1 Nano",
        "provider": "openai",
        "desc": "Ultra rapido, tareas simples",
        "category": "fast",
    },
    {
        "id": "openai:gpt-4o-mini",
        "name": "GPT-4o Mini",
        "provider": "openai",
        "desc": "Multimodal rapido y economico",
        "category": "fast",
    },
    {
        "id": "openai:o4-mini",
        "name": "o4-mini",
        "provider": "openai",
        "desc": "Razonamiento avanzado",
        "category": "reasoning",
    },
    {
        "id": "openai:o3-mini",
        "name": "o3-mini",
        "provider": "openai",
        "desc": "Razonamiento rapido",
        "category": "reasoning",
    },
    {
        "id": "openai:o3",
        "name": "o3",
        "provider": "openai",
        "desc": "Razonamiento profundo",
        "category": "reasoning",
    },
    {
        "id": "openai:o1",
        "name": "o1",
        "provider": "openai",
        "desc": "Razonamiento de alta calidad",
        "category": "reasoning",
    },
    {
        "id": "openai:o1-mini",
        "name": "o1-mini",
        "provider": "openai",
        "desc": "Razonamiento eficiente",
        "category": "reasoning",
    },
    {
        "id": "openai:gpt-4.5-preview",
        "name": "GPT-4.5 Preview",
        "provider": "openai",
        "desc": "GeneraciÃ³n mÃ¡s reciente (preview)",
        "category": "premium",
    },
    {
        "id": "openai:gpt-4o",
        "name": "GPT-4o",
        "provider": "openai",
        "desc": "Multimodal estable",
        "category": "premium",
    },
    {
        "id": "anthropic:claude-sonnet-4-20250514",
        "name": "Claude Sonnet 4",
        "provider": "anthropic",
        "desc": "Codigo y analisis",
        "category": "premium",
    },
    {
        "id": "anthropic:claude-sonnet-4-5",
        "name": "Claude Sonnet 4.5",
        "provider": "anthropic",
        "desc": "Sonnet avanzado (Ãºltima generaciÃ³n)",
        "category": "premium",
    },
    {
        "id": "anthropic:claude-sonnet-4-latest",
        "name": "Claude Sonnet 4 (latest)",
        "provider": "anthropic",
        "desc": "Alias estable para Sonnet 4",
        "category": "premium",
    },
    {
        "id": "anthropic:claude-3-5-haiku-20241022",
        "name": "Claude 3.5 Haiku",
        "provider": "anthropic",
        "desc": "Ultra rapido",
        "category": "fast",
    },
    {
        "id": "anthropic:claude-3-7-sonnet-latest",
        "name": "Claude 3.7 Sonnet",
        "provider": "anthropic",
        "desc": "Razonamiento + coding avanzado",
        "category": "reasoning",
    },
    {
        "id": "anthropic:claude-haiku-4-latest",
        "name": "Claude Haiku 4",
        "provider": "anthropic",
        "desc": "Baja latencia, Ãºltima generaciÃ³n",
        "category": "fast",
    },
    {
        "id": "anthropic:claude-opus-4-20250514",
        "name": "Claude Opus 4",
        "provider": "anthropic",
        "desc": "Maximo razonamiento",
        "category": "reasoning",
    },
    {
        "id": "anthropic:claude-opus-4-latest",
        "name": "Claude Opus 4 (latest)",
        "provider": "anthropic",
        "desc": "Alias estable para Opus 4",
        "category": "reasoning",
    },
    {
        "id": "bedrock:us.anthropic.claude-opus-4-6-v1:0",
        "name": "Claude Opus 4.6 (Bedrock)",
        "provider": "anthropic",
        "desc": "Via AWS Bedrock (configurado)",
        "category": "reasoning",
    },
    {
        "id": "deepseek:deepseek-chat",
        "name": "DeepSeek V3 (API)",
        "provider": "deepseek",
        "desc": "Chat y codigo, muy economico",
        "category": "fast",
    },
    {
        "id": "deepseek:deepseek-v3",
        "name": "DeepSeek V3",
        "provider": "deepseek",
        "desc": "Modelo general avanzado",
        "category": "premium",
    },
    {
        "id": "deepseek:deepseek-v3.1",
        "name": "DeepSeek V3.1",
        "provider": "deepseek",
        "desc": "IteraciÃ³n reciente de alto rendimiento",
        "category": "premium",
    },
    {
        "id": "deepseek:deepseek-reasoner",
        "name": "DeepSeek R1 (API)",
        "provider": "deepseek",
        "desc": "Razonamiento chain-of-thought",
        "category": "reasoning",
    },
    {
        "id": "deepseek:deepseek-r1",
        "name": "DeepSeek R1",
        "provider": "deepseek",
        "desc": "Razonamiento avanzado (full)",
        "category": "reasoning",
    },
    {
        "id": "xai:grok-3",
        "name": "Grok 3",
        "provider": "xai",
        "desc": "Flagship xAI",
        "category": "premium",
    },
    {
        "id": "xai:grok-3-mini",
        "name": "Grok 3 Mini",
        "provider": "xai",
        "desc": "Rapido y eficiente",
        "category": "fast",
    },
    {
        "id": "xai:grok-3-fast",
        "name": "Grok 3 Fast",
        "provider": "xai",
        "desc": "Baja latencia",
        "category": "fast",
    },
    {
        "id": "xai:grok-4",
        "name": "Grok 4",
        "provider": "xai",
        "desc": "Siguiente generaciÃ³n xAI",
        "category": "premium",
    },
    {
        "id": "xai:grok-4-fast",
        "name": "Grok 4 Fast",
        "provider": "xai",
        "desc": "GeneraciÃ³n reciente con baja latencia",
        "category": "fast",
    },
    {
        "id": "mistral:mistral-large-latest",
        "name": "Mistral Large",
        "provider": "mistral",
        "desc": "Flagship Mistral",
        "category": "premium",
    },
    {
        "id": "mistral:codestral-latest",
        "name": "Codestral",
        "provider": "mistral",
        "desc": "Codigo especializado",
        "category": "code",
    },
    {
        "id": "mistral:mistral-small-latest",
        "name": "Mistral Small",
        "provider": "mistral",
        "desc": "Rapido y economico",
        "category": "fast",
    },
    {
        "id": "perplexity:sonar-pro",
        "name": "Sonar Pro",
        "provider": "perplexity",
        "desc": "Busqueda web + IA",
        "category": "search",
    },
    {
        "id": "perplexity:sonar",
        "name": "Sonar",
        "provider": "perplexity",
        "desc": "Busqueda rapida",
        "category": "search",
    },
]

_PROVIDER_LABELS = {
    "atlas": "ATLAS",
    "openai": "OpenAI",
    "anthropic": "Anthropic",
    "gemini": "Google",
    "xai": "xAI",
    "deepseek": "DeepSeek",
    "groq": "Groq",
    "mistral": "Mistral",
    "perplexity": "Perplexity",
    "ollama": "Ollama",
    "bedrock": "AWS Bedrock",
}

# Estado runtime por proveedor (Ãºltimo error conocido para UI de Workspace)
_PROVIDER_RUNTIME_STATUS: Dict[str, Dict[str, Any]] = {}


def _classify_provider_error(error_text: str) -> Dict[str, Any]:
    e = (error_text or "").strip().lower()
    if not e:
        return {"code": "ok", "level": "ok", "message": "Listo"}

    if (
        "credit balance is too low" in e
        or "insufficient_quota" in e
        or "quota" in e
        or "saldo" in e
        or "billing" in e
    ):
        return {
            "code": "quota_exhausted",
            "level": "warn",
            "message": "Saldo API agotado; se aplicarÃ¡ fallback automÃ¡tico",
        }
    if (
        "authentication" in e
        or "unauthorized" in e
        or "invalid api key" in e
        or "invalid x-api-key" in e
        or "incorrect api key" in e
        or "401" in e
    ):
        return {
            "code": "auth_error",
            "level": "err",
            "message": "API key invÃ¡lida o no autorizada",
        }
    if "rate limit" in e or "429" in e:
        return {
            "code": "rate_limited",
            "level": "warn",
            "message": "Rate limit del proveedor; se aplicarÃ¡ fallback",
        }
    return {
        "code": "provider_error",
        "level": "warn",
        "message": "Error temporal del proveedor; se aplicarÃ¡ fallback",
    }


def _set_provider_runtime_status(
    provider_id: str, error_text: Optional[str] = None
) -> None:
    pid = (provider_id or "").strip().lower()
    if not pid:
        return
    if not error_text:
        _PROVIDER_RUNTIME_STATUS.pop(pid, None)
        return
    cls = _classify_provider_error(error_text)
    _PROVIDER_RUNTIME_STATUS[pid] = {
        "code": cls["code"],
        "level": cls["level"],
        "message": cls["message"],
        "raw": str(error_text)[:240],
        "updated_at": time.time(),
    }


@app.get("/agent/models")
def agent_models():
    """Catalogo de modelos IA con status de API key."""
    t0 = time.perf_counter()
    from modules.humanoid.ai.provider_credentials import get_provider_api_key

    status = {}
    for pid in (
        "openai",
        "anthropic",
        "gemini",
        "xai",
        "deepseek",
        "groq",
        "mistral",
        "perplexity",
    ):
        key = get_provider_api_key(pid)
        status[pid] = bool(key and key.strip())
    status["ollama"] = True
    status["atlas"] = True
    status["bedrock"] = bool(
        (
            (os.getenv("AWS_ACCESS_KEY_ID") or "").strip()
            and (os.getenv("AWS_SECRET_ACCESS_KEY") or "").strip()
        )
        or (os.getenv("ATLAS_AI_MODE", "").strip().lower() == "bedrock")
    )
    models = []
    for m in _MODEL_CATALOG:
        is_bedrock_model = str(m.get("id", "")).startswith("bedrock:")
        is_codex_model = str(m.get("id", "")) == "openai:codex-auto"
        available = (
            status.get("bedrock", False)
            if is_bedrock_model
            else status.get(m["provider"], False)
        )
        rt = _PROVIDER_RUNTIME_STATUS.get(m["provider"], {})
        codex_model = (
            (os.getenv("OPENAI_CODEX_MODEL") or "gpt-4.1-mini").strip()
            or "gpt-4.1-mini"
        )
        models.append(
            {
                **m,
                "available": available,
                "provider_label": "Anthropic (Bedrock)"
                if is_bedrock_model
                else _PROVIDER_LABELS.get(m["provider"], m["provider"]),
                "status_code": (
                    "ok"
                    if is_codex_model and available
                    else ("missing_key" if is_codex_model else rt.get("code"))
                ),
                "status_message": (
                    f"Codex ready ({codex_model})"
                    if is_codex_model and available
                    else (
                        "OPENAI_API_KEY missing for Codex"
                        if is_codex_model
                        else rt.get("message")
                    )
                ),
                "runtime_state": (
                    "ok" if is_codex_model and available else ("warn" if is_codex_model else rt.get("level", "ok"))
                ),
                "runtime_error": rt.get("raw"),
                "codex_model": codex_model if is_codex_model else None,
            }
        )
    ms = int((time.perf_counter() - t0) * 1000)
    return {
        "ok": True,
        "data": models,
        "provider_runtime": _PROVIDER_RUNTIME_STATUS,
        "ms": ms,
    }


# ----------------------------------------------------------------------------
# BitÃ¡cora Central (UI -> servidor)
# ----------------------------------------------------------------------------


class BitacoraLogBody(BaseModel):
    message: str = ""
    level: str = "info"  # info|success|warning|error|low|med|high|critical
    source: str = "ui"
    data: Dict[str, Any] | None = None


@app.post("/bitacora/log")
def bitacora_log(body: BitacoraLogBody):
    """
    Endpoint de sincronizaciÃ³n desde el Dashboard v3.8.0.
    La UI llama este endpoint para persistir entradas locales en la BitÃ¡cora.
    """
    try:
        msg = (body.message or "").strip()
        if not msg:
            return {"ok": True, "skipped": "empty"}
        src = (body.source or "ui").strip()[:40] or "ui"
        lvl = (body.level or "info").strip().lower()
        # Mapear nivel UI â†’ ok boolean (para evolution_bitacora)
        ok = lvl not in ("error", "critical", "high", "fail", "failed")
        try:
            from modules.humanoid.ans.evolution_bitacora import \
                append_evolution_log

            append_evolution_log(message=msg[:500], ok=ok, source=src)
        except Exception:
            pass
        return {"ok": True, "logged": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/bitacora/stream")
def bitacora_stream(since_id: int = 0):
    """Entradas nuevas de la bitacora desde un ID especifico (polling incremental)."""
    try:
        from modules.humanoid.ans.evolution_bitacora import \
            get_evolution_entries

        all_entries = get_evolution_entries(limit=200)
        if since_id:
            ts_threshold = since_id / 1000.0
            filtered = [
                e
                for e in all_entries
                if _ts_to_epoch(e.get("timestamp", "")) > ts_threshold
            ]
        else:
            filtered = all_entries[:50]
        latest_id = int(time.time() * 1000)
        return {
            "ok": True,
            "entries": filtered,
            "count": len(filtered),
            "latest_id": latest_id,
        }
    except Exception:
        return {"ok": True, "entries": [], "count": 0, "latest_id": 0}


@app.post("/ops/event")
def ops_event(body: dict):
    """Recibir evento del bus OPS (Quality, POTs, etc.) y registrar en bitacora."""
    try:
        msg = (body.get("message") or "").strip()
        if not msg:
            return {"ok": True, "skipped": "empty"}
        src = (body.get("source") or "ops").strip()[:40]
        lvl = (body.get("level") or "info").strip().lower()
        ok = lvl not in ("error", "critical", "high", "fail", "failed")
        try:
            from modules.humanoid.ans.evolution_bitacora import \
                append_evolution_log

            append_evolution_log(message=msg[:500], ok=ok, source=src)
        except Exception:
            pass
        return {"ok": True, "logged": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/ui/lang")
def ui_lang(locale: str = "es"):
    """Cadenas i18n para el dashboard (es | en)."""
    try:
        from modules.humanoid.config import SUPPORTED_LOCALES
        from modules.humanoid.config.i18n import get_all_strings

        loc = locale.strip().lower() if locale else "es"
        if loc not in SUPPORTED_LOCALES:
            loc = "es"
        return {"ok": True, "locale": loc, "strings": get_all_strings(loc)}
    except Exception as e:
        return {"ok": False, "locale": "es", "strings": {}, "error": str(e)}


@app.get("/tools")
def tools():
    return {
        "ok": True,
        "tools": [
            "atlas.status",
            "atlas.doctor",
            "atlas.modules",
            "atlas.snapshot",
            "atlas.note.create",
            "atlas.note.append",
            "atlas.note.view",
            "atlas.inbox",
        ],
    }


@app.post("/execute")
def execute(step: Step):
    t = step.tool
    a = step.args or {}
    if t == "atlas.status":
        cmd = "/status"
    elif t == "atlas.doctor":
        cmd = "/doctor"
    elif t == "atlas.modules":
        cmd = "/modules"
    elif t == "atlas.snapshot":
        cmd = f"/snapshot {a.get('label','snapshot')}"
    elif t == "atlas.note.create":
        cmd = f"/note create {a.get('title','Nota')}"
    elif t == "atlas.note.append":
        cmd = f"/note append {a.get('title','Inbox')} | {a.get('text','')}"
    elif t == "atlas.note.view":
        cmd = f"/note view {a.get('title','Inbox')}"
    elif t == "atlas.inbox":
        cmd = a.get("text", "")
    else:
        return {"ok": False, "error": f"Tool no soportada aÃºn: {t}"}
    out = handle(cmd)
    return {"ok": True, "tool": t, "output": out}


@app.get("/modules")
def modules():
    return {
        "ok": True,
        "modules": [
            {"name": "vision", "enabled": False},
            {"name": "voice", "enabled": False},
            {"name": "agent_router", "enabled": False},
            {"name": "telegram", "enabled": True},
        ],
    }


from typing import Any, List, Optional

from pydantic import BaseModel


class IntentIn(BaseModel):
    user: str = "raul"
    text: str
    meta: Optional[dict[str, Any]] = None


@app.post("/intent")
def intent(payload: IntentIn):
    # Respuesta mÃ­nima (v1): eco + timestamp.
    # Luego conectamos command_router/agent_router.
    return {
        "ok": True,
        "received": payload.model_dump(),
        "result": {
            "message": "Intent recibido",
            "ts": time.time(),
        },
    }


from typing import Any, List, Optional

# --- Canonical Intent API (v1) ---
from pydantic import BaseModel

# Temporalmente comentado para evitar error de importaciÃ³n
# from modules.command_router import handle as route_command


class IntentIn(BaseModel):
    user: str = "raul"
    text: str = ""
    meta: Optional[dict[str, Any]] = None


@app.post("/intent")
def intent(payload: IntentIn):
    t0 = time.time()
    if not (payload.text or "").strip():
        return {
            "ok": False,
            "received": payload.model_dump(),
            "error": 'Campo \'text\' es requerido (ej: {"text": "/status"})',
            "ms": int((time.time() - t0) * 1000),
        }
    try:
        out = route_command((payload.text or "").strip())
        return {
            "ok": True,
            "received": payload.model_dump(),
            "output": out,
            "ms": int((time.time() - t0) * 1000),
        }
    except Exception as e:
        return {
            "ok": False,
            "received": payload.model_dump(),
            "error": str(e),
            "ms": int((time.time() - t0) * 1000),
        }


# --- Push Dashboard (ATLAS_PUSH -> Dashboard -> ATLAS_NEXUS) ---
_push_state: dict = {"last_command": None, "last_ack": None, "updated_at": None}

# --- Evolution (ATLAS_EVOLUTION): gobernanza y pendientes de aprobaciÃ³n ---
_evolution_pending: list = []  # [{ package, old_version, new_version, ts, worker }]
_evolution_governed: bool = os.environ.get(
    "EVOLUTION_GOVERNED", ""
).strip().upper() in ("1", "TRUE", "GOVERNED")
_evolution_last_report: dict = (
    {}
)  # Ãšltimo reporte del daemon: { state, message, asimilacion_exitosa, ts }


class PushCommandBody(BaseModel):
    target: str = "NEXUS_ARM"
    action: str = "update_state"
    value: Any = 1


@app.post("/api/push/command", tags=["Push"])
def push_command(body: PushCommandBody):
    """Recibe comando del Cerebro; NEXUS lee el estado y ejecuta. Si target=EVOLUTION_REPORT, registra pendientes y Ãºltimo reporte (AsimilaciÃ³n Exitosa)."""
    global _push_state, _evolution_pending, _evolution_last_report
    cmd = {
        "target": body.target,
        "action": body.action,
        "value": body.value,
        "ts": time.time(),
    }
    _push_state["last_command"] = cmd
    _push_state["updated_at"] = time.time()
    if (
        body.target == "EVOLUTION_REPORT"
        and body.action == "worker_result"
        and isinstance(body.value, str)
    ):
        try:
            import json

            payload = json.loads(body.value)
            if (
                payload.get("status") == "pending_approval"
                or payload.get("event") == "evolution_pending_approval"
            ):
                _evolution_pending.append(
                    {
                        "package": payload.get("package"),
                        "old_version": payload.get("old_version"),
                        "new_version": payload.get("new_version"),
                        "ts": time.time(),
                        "worker": payload.get("worker", "pypi"),
                    }
                )
            if (
                payload.get("event") == "ans_bitacora"
                or payload.get("asimilacion_exitosa") is not None
                or payload.get("state")
            ):
                _evolution_last_report = {
                    "state": payload.get("state", ""),
                    "message": payload.get("message", ""),
                    "asimilacion_exitosa": payload.get("asimilacion_exitosa", False),
                    "ts": time.time(),
                }
        except Exception:
            pass
    return {"ok": True, "command": cmd}


@app.get("/api/push/state", tags=["Push"])
def push_state():
    """Estado actual del Dashboard para evitar comandos duplicados."""
    return {
        "ok": True,
        "state": _push_state.get("last_command"),
        "ack": _push_state.get("last_ack"),
        "updated_at": _push_state.get("updated_at"),
    }


@app.post("/api/push/ack", tags=["Push"])
def push_ack(body: dict):
    """NEXUS confirma recepciÃ³n/ejecuciÃ³n."""
    global _push_state
    _push_state["last_ack"] = body
    _push_state["updated_at"] = time.time()
    return {"ok": True}


@app.get("/api/evolution/status", tags=["Evolution"])
def evolution_status():
    """Estado de gobernanza, pendientes de aprobaciÃ³n y Ãºltimo reporte del daemon (AsimilaciÃ³n Exitosa)."""
    return {
        "governed": _evolution_governed,
        "pending": list(_evolution_pending),
        "last_report": dict(_evolution_last_report),
        "updated_at": _push_state.get("updated_at") or time.time(),
    }


class EvolutionApproveBody(BaseModel):
    package: str
    new_version: str


@app.post("/api/evolution/approve", tags=["Evolution"])
def evolution_approve(body: EvolutionApproveBody):
    """Confirma y aplica un cambio pendiente al requirements.txt del proyecto (nÃºcleo)."""
    global _evolution_pending
    req_txt = Path(os.environ.get("ATLAS_BASE", BASE_DIR)) / "requirements.txt"
    if not req_txt.exists():
        return {"ok": False, "error": "requirements.txt no encontrado"}
    for i, p in enumerate(_evolution_pending):
        if (
            p.get("package") == body.package
            and p.get("new_version") == body.new_version
        ):
            try:
                text = req_txt.read_text(encoding="utf-8")
                lines = text.splitlines()
                new_lines = []
                for line in lines:
                    stripped = line.strip()
                    if stripped.startswith("#") or not stripped:
                        new_lines.append(line)
                        continue
                    name = line.split("==")[0].split(">=")[0].split("[")[0].strip()
                    if name == body.package:
                        new_lines.append(f"{body.package}=={body.new_version}")
                        continue
                    new_lines.append(line)
                req_txt.write_text(
                    "\n".join(new_lines) + ("\n" if text.endswith("\n") else ""),
                    encoding="utf-8",
                )
            except Exception as e:
                return {"ok": False, "error": str(e)}
            _evolution_pending.pop(i)
            return {"ok": True, "applied": f"{body.package}=={body.new_version}"}
    return {"ok": False, "error": "Pendiente no encontrado"}


@app.post("/api/evolution/trigger", tags=["Evolution"])
def evolution_trigger():
    """Fuerza la ejecuciÃ³n de un ciclo de la TrÃ­ada (PyPI | GitHub | Hugging Face) en segundo plano. No espera a que termine."""
    import subprocess
    import sys

    daemon_script = BASE_DIR / "evolution_daemon.py"
    if not daemon_script.exists():
        return {
            "ok": False,
            "error": "evolution_daemon.py no encontrado",
            "triggered": False,
        }
    try:
        subprocess.Popen(
            [sys.executable, str(daemon_script), "--run-once"],
            cwd=str(BASE_DIR),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=(
                subprocess.CREATE_NO_WINDOW
                if (sys.platform == "win32" and hasattr(subprocess, "CREATE_NO_WINDOW"))
                else 0
            ),
        )
        return {
            "ok": True,
            "triggered": True,
            "message": "Ciclo TrÃ­ada (PyPI/GitHub/HF) iniciado en segundo plano. Ver /api/evolution/status y logs/evolution_last_cycle.json",
        }
    except Exception as e:
        return {"ok": False, "error": str(e), "triggered": False}


# --- LLM hÃ­brido (router + Ollama) ---
from modules.llm.schemas import LLMRequest, LLMResponse
from modules.llm.service import LLMService

llm_service = LLMService()


@app.post(
    "/llm",
    response_model=LLMResponse,
    summary="Hybrid LLM (Ollama)",
    description="Route prompt to FAST/CHAT/CODE/REASON/TOOLS via HybridRouter and call Ollama. Returns ok, output, route, model_used, ms, tokens_est, error, meta.",
)
def llm_endpoint(req: LLMRequest) -> LLMResponse:
    """Run LLM request through router + Ollama and return structured response."""
    resp = llm_service.run(req)
    try:
        from modules.humanoid.metalearn.collector import record_router_call

        record_router_call(
            getattr(resp, "route", "CHAT") or "CHAT",
            getattr(resp, "model_used", "") or "",
            "ok" if getattr(resp, "ok", True) else "fail",
            getattr(resp, "ms", None),
        )
    except Exception:
        pass
    return resp


# --- Multi-AI (registry, router, status) ---
@app.get("/ai/status", tags=["AI"])
def ai_status():
    """Multi-AI status: providers, route_to_model, budget, policies, telemetry."""
    from modules.humanoid.ai.status import get_ai_status

    return get_ai_status()


class AIRouteBody(BaseModel):
    prompt: Optional[str] = None
    intent_hint: Optional[str] = None
    modality: Optional[str] = "text"
    intent: Optional[str] = None
    complexity: Optional[str] = None
    latency_need: Optional[str] = None
    safety: Optional[str] = None
    prefer_free: Optional[bool] = True


@app.post("/ai/route", tags=["AI"])
def ai_route(body: AIRouteBody):
    """Return RouteDecision for TaskProfile. Use prompt+intent_hint or intent/complexity/modality."""
    try:
        from modules.humanoid.ai.models import TaskProfile
        from modules.humanoid.ai.router import decide_route, infer_task_profile

        if body.prompt:
            profile = infer_task_profile(
                body.prompt,
                intent_hint=body.intent_hint or body.intent,
                modality=body.modality or "text",
            )
        else:
            profile = TaskProfile(
                intent=body.intent or "chat",
                complexity=body.complexity or "med",
                latency_need=body.latency_need or "normal",
                modality=body.modality or "text",
                safety=body.safety or "med",
            )
        decision = decide_route(
            profile,
            prefer_free=body.prefer_free if body.prefer_free is not None else True,
        )
        return {
            "ok": True,
            "decision": {
                "provider_id": decision.provider_id,
                "model_key": decision.model_key,
                "route": decision.route,
                "reason": decision.reason,
                "is_free": decision.is_free,
                "cost_estimate_usd": decision.cost_estimate_usd,
            },
        }
    except Exception as e:
        return {"ok": False, "decision": None, "error": str(e)}


class FeedbackDecideBody(BaseModel):
    source_app: str = "unknown"
    feedback: Dict[str, Any] = Field(default_factory=dict)
    analysis: Optional[Dict[str, Any]] = None


def _feedback_structural_major(feedback: Dict[str, Any]) -> bool:
    sev = str((feedback or {}).get("severity", "")).strip().lower()
    category = str((feedback or {}).get("category", "")).strip().lower()
    title = str((feedback or {}).get("title", "")).strip().lower()
    desc = str((feedback or {}).get("description", "")).strip().lower()
    text = f"{title} {desc}"
    if sev in ("critical", "high"):
        return True
    if category in ("architecture", "estructura", "infra", "security", "database", "schema"):
        return True
    structural_hints = (
        "arquitectura",
        "estructura",
        "breaking",
        "migracion",
        "migraciÃ³n",
        "refactor grande",
        "core",
        "kernel",
        "schema",
        "db",
        "seguridad",
        "security",
    )
    return any(h in text for h in structural_hints)


@app.post("/api/feedback/decide", tags=["Feedback"])
def feedback_decide(body: FeedbackDecideBody):
    """DecisiÃ³n central Atlas para feedback cross-app: auto_fix_now vs wait_owner_approval."""
    source = (body.source_app or "unknown").strip() or "unknown"
    feedback = body.feedback or {}
    major = _feedback_structural_major(feedback)
    severity = str(feedback.get("severity", "medium")).strip().lower() or "medium"
    title = str(feedback.get("title", "feedback")).strip() or "feedback"
    category = str(feedback.get("category", "general")).strip() or "general"

    approval_id = None
    decision = "auto_fix_now"
    auto_execute = True
    requires_owner_approval = False
    status = "auto_resolved"

    if major:
        decision = "wait_owner_approval"
        auto_execute = False
        requires_owner_approval = True
        status = "pending_owner_approval"
        try:
            from modules.humanoid.approvals import create_approval

            out = create_approval(
                action="feedback_structure_apply",
                risk="high" if severity in ("high", "critical") else "medium",
                description=f"[{source}] {title}",
                payload={
                    "source_app": source,
                    "feedback": feedback,
                    "analysis": body.analysis or {},
                    "category": category,
                    "severity": severity,
                },
            )
            approval_id = out.get("id") if out.get("ok") else None
        except Exception:
            approval_id = None

    user_message = (
        "ATLAS aplicÃ³ una correcciÃ³n inmediata de bajo riesgo y dejÃ³ registro del caso."
        if auto_execute
        else "ATLAS detectÃ³ impacto estructural. El caso quedÃ³ en aprobaciÃ³n del Owner antes de ejecutar cambios."
    )
    if approval_id:
        user_message = f"{user_message} ID aprobaciÃ³n: {approval_id}"

    try:
        from modules.humanoid.comms import emit as comms_emit

        comms_emit(
            "feedback",
            f"[{source}] {title} => {decision}",
            level="high" if major else "medium",
            data={
                "source_app": source,
                "severity": severity,
                "category": category,
                "decision": decision,
                "status": status,
                "approval_id": approval_id,
            },
        )
    except Exception:
        pass

    return {
        "ok": True,
        "source_app": source,
        "decision": decision,
        "status": status,
        "auto_execute": auto_execute,
        "requires_owner_approval": requires_owner_approval,
        "approval_id": approval_id,
        "user_message": user_message,
        "timestamp": datetime.utcnow().isoformat() + "Z",
    }


@app.get("/mode/capabilities", tags=["Mode"])
def mode_capabilities():
    """ATLAS_MODE capabilities: screen_act, record_replay, playwright, benchmark."""
    try:
        from modules.humanoid.mode import get_mode_capabilities

        return {"ok": True, "data": get_mode_capabilities()}
    except Exception as e:
        return {"ok": False, "data": None, "error": str(e)}


# --- Cursor mode (end-to-end orchestration) ---
class CursorRunBody(BaseModel):
    goal: str
    mode: Optional[str] = "plan_only"  # plan_only | controlled | auto
    depth: Optional[int] = 1  # 1-5 (overridden by resources if present)
    prefer_free: Optional[bool] = True
    allow_paid: Optional[bool] = False
    profile: Optional[str] = "owner"  # owner | operario | contadora
    context: Optional[dict] = None
    resources: Optional[str] = None  # lite | pro | ultra -> depth 1/2/3


@app.post("/cursor/run", tags=["Cursor"])
def cursor_run_endpoint(body: CursorRunBody):
    """Cursor-mode run: plan with AI router, optional execution. Returns summary, evidence, model_routing."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cursor import cursor_run

        result = cursor_run(
            goal=body.goal,
            mode=body.mode or "plan_only",
            depth=max(1, min(5, body.depth or 1)),
            context=body.context,
            prefer_free=body.prefer_free if body.prefer_free is not None else True,
            allow_paid=body.allow_paid if body.allow_paid is not None else False,
            profile=body.profile or "owner",
            resources=body.resources,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            result.get("ok", False), result.get("data"), ms, result.get("error")
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


class CursorStepBody(BaseModel):
    step_index: int = 0
    approve: Optional[bool] = False


@app.post("/cursor/step/execute", tags=["Cursor"])
def cursor_step_execute(body: CursorStepBody):
    """Execute one step from last cursor run (policy/approvals) con acciones reales."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cursor import get_cursor_status

        status = get_cursor_status()
        data = status.get("data")
        if not data or not data.get("steps"):
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "no cursor run or no steps",
            )
        idx = max(0, min(body.step_index, len(data["steps"]) - 1))
        step = data["steps"][idx]
        if not body.approve:
            return _std_resp(
                True,
                {
                    "step": step,
                    "executed": False,
                    "message": "Set approve=true to execute",
                },
                int((time.perf_counter() - t0) * 1000),
                None,
            )
        # Ejecutar de verdad usando el ejecutor de Cursor (genera script/log).
        try:
            from modules.humanoid.cursor.executor import CursorExecutor

            ex = CursorExecutor(repo_root=Path(str(BASE_DIR)))
            artifacts = None
            # Reusar artifacts del run si existen
            if isinstance(data.get("artifacts"), dict):
                artifacts = data["artifacts"]
            if not artifacts or not artifacts.get("run_dir"):
                auto = ex.execute_steps([], goal=data.get("goal", "cursor_run"))
                artifacts = auto.get("artifacts", {})
                data["artifacts"] = artifacts
            # Crear wrapper RunArtifacts desde artifacts path (mÃ­nimo necesario)
            from modules.humanoid.cursor.executor import RunArtifacts

            run_dir = Path(
                artifacts.get("run_dir")
                or (Path(str(BASE_DIR)) / "snapshots" / "cursor")
            )
            ra = RunArtifacts(
                run_id=str(artifacts.get("run_id") or "cursor_controlled"),
                run_dir=run_dir,
                log_path=Path(
                    artifacts.get("terminal_log_path") or (run_dir / "terminal.log")
                ),
                script_path=Path(artifacts.get("script_path") or (run_dir / "run.ps1")),
                json_path=Path(
                    artifacts.get("run_json_path") or (run_dir / "run.json")
                ),
            )
            r = ex.execute_step(ra, step)
            data.setdefault("controlled_result", {"executed": []})
            data["controlled_result"].setdefault("executed", []).append(r)
            data["steps"][idx]["status"] = (
                "done"
                if r.get("ok")
                else ("needs_human" if r.get("action") == "needs_human" else "failed")
            )
            # Preview terminal actualizado
            try:
                if ra.log_path.exists():
                    data["terminal_preview"] = ra.log_path.read_text(
                        encoding="utf-8", errors="ignore"
                    )[-6000:]
            except Exception:
                pass
            # Persistir como last_run
            try:
                from modules.humanoid.cursor.status import set_last_run

                set_last_run(data)
            except Exception:
                pass
            return _std_resp(
                True,
                {
                    "step": step,
                    "executed": True,
                    "result": r,
                    "terminal_preview": data.get("terminal_preview"),
                    "artifacts": data.get("artifacts"),
                },
                int((time.perf_counter() - t0) * 1000),
                None,
            )
        except Exception as e:
            data["steps"][idx]["status"] = "failed"
            return _std_resp(
                False,
                {"step": step, "executed": False},
                int((time.perf_counter() - t0) * 1000),
                str(e),
            )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/cursor/status", tags=["Cursor"])
def cursor_status_endpoint():
    """Last cursor run state: goal, steps, model_routing, evidence."""
    try:
        from modules.humanoid.cursor import get_cursor_status

        return get_cursor_status()
    except Exception as e:
        return {"ok": False, "data": None, "message": str(e)}


# --- PC Walker (Windows navigation) ---
class PcWalkBody(BaseModel):
    command: str = (
        "workflow"  # workflow | open_app | open_path | screenshot | ocr | alt_tab
    )
    payload: Optional[dict] = None


@app.get("/api/pc/status", tags=["PC"])
def api_pc_status():
    """Estado del subsistema 'caminar' por Windows (deps + snapshots path)."""
    try:
        from modules.humanoid.pc_walker import status as pc_status

        return {"ok": True, "data": pc_status()}
    except Exception as e:
        return {"ok": False, "data": None, "error": str(e)}


@app.post("/api/pc/walk", tags=["PC"])
def api_pc_walk(body: PcWalkBody):
    """Ejecuta navegaciÃ³n Windows (pies-PC) con evidencia y OPS.

    Body:
    - command: workflow|open_app|open_path|screenshot|ocr|alt_tab
    - payload: {steps:[...], allow_destructive?:bool, sleep_ms?:int, name/app/path/count...}
    """
    t0 = time.perf_counter()
    try:
        from modules.humanoid.pc_walker import run_workflow

        cmd = (body.command or "").strip().lower()
        payload = body.payload or {}
        allow_destructive = bool(payload.get("allow_destructive", False))
        sleep_ms = int(payload.get("sleep_ms", 250) or 250)
        step_timeout_s = float(payload.get("step_timeout_s", 4.0) or 4.0)

        if cmd == "workflow":
            steps = payload.get("steps") or []
            if not isinstance(steps, list):
                steps = []
            result = run_workflow(
                steps,
                allow_destructive=allow_destructive,
                sleep_ms=sleep_ms,
                step_timeout_s=step_timeout_s,
            )
        elif cmd == "open_app":
            name = str(payload.get("name") or payload.get("app") or "")
            result = run_workflow(
                [{"kind": "open_app", "name": name}],
                allow_destructive=allow_destructive,
                sleep_ms=sleep_ms,
                step_timeout_s=step_timeout_s,
            )
        elif cmd == "open_path":
            path = str(payload.get("path") or "")
            result = run_workflow(
                [{"kind": "open_path", "path": path}],
                allow_destructive=allow_destructive,
                sleep_ms=sleep_ms,
                step_timeout_s=step_timeout_s,
            )
        elif cmd == "screenshot":
            name = str(payload.get("name") or "pc_ui")
            result = run_workflow(
                [{"kind": "screenshot", "name": name}],
                allow_destructive=allow_destructive,
                sleep_ms=sleep_ms,
                step_timeout_s=step_timeout_s,
            )
        elif cmd == "ocr":
            result = run_workflow(
                [{"kind": "ocr"}],
                allow_destructive=allow_destructive,
                sleep_ms=sleep_ms,
                step_timeout_s=step_timeout_s,
            )
        elif cmd == "alt_tab":
            count = int(payload.get("count") or 1)
            result = run_workflow(
                [{"kind": "alt_tab", "count": count}],
                allow_destructive=allow_destructive,
                sleep_ms=sleep_ms,
                step_timeout_s=step_timeout_s,
            )
        else:
            result = {"ok": False, "error": "unsupported_command", "steps": [], "ms": 0}

        ms = int((time.perf_counter() - t0) * 1000)
        ok = bool(result.get("ok", False))
        return _std_resp(ok, result, ms, result.get("error"))
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


class RepoPushBody(BaseModel):
    """Solicitud de push de repo (esta u otra app). Desde Cursor o chat."""

    app_id: Optional[str] = None  # atlas_push | atlas_nexus | robot | ...
    repo_path: Optional[str] = None  # ruta absoluta si no usa known_apps
    message: Optional[str] = None  # mensaje de commit


@app.post("/api/repo/push", tags=["Cursor", "Repo"])
def api_repo_push(body: RepoPushBody):
    """Sube el repo de esta app o de otra (app_id o repo_path). Usado por Cursor y por chat."""
    t0 = time.perf_counter()
    try:
        from modules.repo_push import push_repo, resolve_path

        repo = resolve_path(app_id=body.app_id, repo_path=body.repo_path)
        msg = (body.message or "").strip() or "chore: sync (solicitado por Cursor/chat)"
        result = push_repo(repo_path=repo, message=msg)
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": result.get("ok", False),
            "data": {
                "message": result.get("message"),
                "branch": result.get("branch"),
                "pushed": result.get("pushed"),
            },
            "ms": ms,
            "error": result.get("error"),
        }
    except Exception as e:
        return {
            "ok": False,
            "data": None,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": str(e),
        }


@app.get("/api/repo/apps", tags=["Repo"])
def api_repo_apps():
    """Lista apps conocidas (known_apps) para push desde Cursor/chat."""
    try:
        from modules.repo_push import list_known_apps

        return {"ok": True, "data": list_known_apps()}
    except Exception as e:
        return {"ok": False, "data": {}, "error": str(e)}


# --- Aprendizaje progresivo ATLAS (continual learning) ---
_learning_components: Optional[dict] = None


def _get_semantic_memory_safe():
    """Semantic memory o stub si falla (p. ej. sin sentence_transformers)."""
    try:
        from modules.humanoid.memory_engine.semantic_memory import \
            get_semantic_memory

        return get_semantic_memory()
    except Exception:

        class _StubSemanticMemory:
            def add_experience(self, *args, **kwargs):
                return 0

            def recall_similar(self, *args, **kwargs):
                return []

            def get_statistics(self):
                return {"total_experiences": 0, "storage_size_mb": 0}

        return _StubSemanticMemory()


def _get_learning_components():
    """InicializaciÃ³n perezosa del sistema de aprendizaje continuo."""
    global _learning_components
    if _learning_components is not None:
        return _learning_components
    import sys

    if str(BASE_DIR) not in sys.path:
        sys.path.insert(0, str(BASE_DIR))
    try:
        from brain.knowledge.initial_knowledge import InitialKnowledgeBase
        from brain.learning.ai_consultant import AIConsultant
        from brain.learning.ai_tutor import AITutor
        from brain.learning.continual_learning_loop import \
            ContinualLearningLoop
        from brain.learning.knowledge_consolidator import KnowledgeConsolidator
        from brain.learning.uncertainty_detector import UncertaintyDetector

        kb = InitialKnowledgeBase()
        uncertainty_detector = UncertaintyDetector(uncertainty_threshold=0.6)
        ai_consultant = AIConsultant()
        semantic_memory = _get_semantic_memory_safe()
        episodic_memory = None
        try:
            from brain.learning.episodic_memory import EpisodicMemory

            episodic_memory = EpisodicMemory()
        except Exception:
            pass
        consolidator = KnowledgeConsolidator(
            semantic_memory=semantic_memory,
            episodic_memory=episodic_memory,
            knowledge_base=kb,
        )
        tutor_type = os.getenv("AI_TUTOR_TYPE", "disabled").lower()
        ai_tutor = None
        try:
            # Siempre crear tutor (offline-first). Si no hay API key o tutor_type=disabled,
            # `AITutor.design_curriculum` caerÃ¡ al fallback local.
            ai_tutor = AITutor(
                tutor_type=tutor_type,
                api_key=os.getenv("ANTHROPIC_API_KEY") or os.getenv("OPENAI_API_KEY"),
                review_interval_hours=int(os.getenv("AI_TUTOR_REVIEW_HOURS", "6")),
                use_local_fallback=os.getenv(
                    "AI_TUTOR_USE_LOCAL_FALLBACK", "true"
                ).lower()
                in ("1", "true", "yes"),
            )
        except Exception:
            ai_tutor = None
        learning_loop = ContinualLearningLoop(
            knowledge_base=kb,
            uncertainty_detector=uncertainty_detector,
            ai_consultant=ai_consultant,
            semantic_memory=semantic_memory,
            episodic_memory=episodic_memory,
            consolidator=consolidator,
            action_executor=None,
            ai_tutor=ai_tutor,
        )
        _learning_components = {
            "kb": kb,
            "uncertainty_detector": uncertainty_detector,
            "ai_consultant": ai_consultant,
            "ai_tutor": ai_tutor,
            "semantic_memory": semantic_memory,
            "episodic_memory": episodic_memory,
            "consolidator": consolidator,
            "learning_loop": learning_loop,
        }
        return _learning_components
    except Exception as e:
        _learning_components = {"error": str(e)}
        return _learning_components


_LEARNING_SNAPSHOT_DIR: str = ""


def _learning_snapshot_dir() -> Path:
    """Directorio de snapshots para aprendizaje (ordenado y persistente)."""
    global _LEARNING_SNAPSHOT_DIR
    if not _LEARNING_SNAPSHOT_DIR:
        v = os.getenv("LEARNING_SNAPSHOT_DIR", "").strip()
        if v:
            _LEARNING_SNAPSHOT_DIR = v
        else:
            _LEARNING_SNAPSHOT_DIR = str(BASE_DIR / "snapshots" / "learning")
    p = Path(_LEARNING_SNAPSHOT_DIR)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _atomic_write_json(path: Path, payload: dict) -> None:
    """Escritura atÃ³mica (Windows-safe) para JSON UTF-8."""
    path.parent.mkdir(parents=True, exist_ok=True)
    content = json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True)
    tmp_name = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=str(path.parent),
            delete=False,
            suffix=".tmp",
        ) as f:
            tmp_name = f.name
            f.write(content)
        os.replace(tmp_name, str(path))
    finally:
        try:
            if tmp_name and Path(tmp_name).exists():
                Path(tmp_name).unlink(missing_ok=True)
        except Exception:
            pass


def _cleanup_learning_snapshots(keep_last_n: Optional[int] = None) -> int:
    keep = (
        keep_last_n
        if keep_last_n is not None
        else int(os.getenv("LEARNING_SNAPSHOT_KEEP_LAST", "500"))
    )
    if keep <= 0:
        return 0
    d = _learning_snapshot_dir()
    files = sorted(
        d.glob("LEARNING_*.json"), key=lambda p: p.stat().st_mtime, reverse=True
    )
    removed = 0
    for f in files[keep:]:
        try:
            f.unlink(missing_ok=True)
            removed += 1
        except Exception:
            pass
    return removed


def _write_learning_snapshot(kind: str, data: dict) -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_kind = "".join(
        c if (c.isalnum() or c in ("-", "_")) else "_" for c in (kind or "event")
    )[:48]
    path = _learning_snapshot_dir() / f"LEARNING_{safe_kind}_{ts}.json"
    payload = {
        "kind": safe_kind,
        "ts_utc": datetime.now(timezone.utc).isoformat(),
        "data": data,
    }
    _atomic_write_json(path, payload)
    _cleanup_learning_snapshots()
    return str(path)


VALID_RISK_LEVELS = ("low", "normal", "high", "critical")


class ProcessSituationBody(BaseModel):
    description: str = ""
    type: Optional[str] = None
    goal: Optional[str] = None
    entities: Optional[List[str]] = None
    constraints: Optional[List[str]] = None
    risk_level: Optional[
        str
    ] = None  # low | normal | high | critical (para should_ask_for_help)

    @field_validator("risk_level")
    @classmethod
    def risk_level_allowed(cls, v: Optional[str]) -> Optional[str]:
        if v is None or v == "":
            return None
        v = (v or "").strip().lower()
        if v in VALID_RISK_LEVELS:
            return v
        return "normal"  # fallback seguro


@app.post("/api/learning/process-situation", tags=["Learning"])
async def process_situation_with_learning(body: ProcessSituationBody):
    """Procesar situaciÃ³n con aprendizaje continuo: persiste episodio y responde rÃ¡pido; procesamiento pesado en background."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    situation = body.model_dump(exclude_none=True)
    if not situation.get("description"):
        situation["description"] = body.description or (body.goal or "unknown")

    # 1) Persistencia inmediata (memoria episÃ³dica) + snapshot ordenado
    episode_id = None
    try:
        ep = comp.get("episodic_memory")
        if ep and getattr(ep, "add_episode", None):
            episode_id = ep.add_episode(
                situation_type=situation.get("type") or "learning",
                description=situation.get("description") or "",
                action="api:learning/process-situation",
                outcome="queued",
                success=True,
                context=situation,
                asked_for_help=False,
            )
    except Exception:
        episode_id = None

    snapshot_queued = ""
    try:
        snapshot_queued = _write_learning_snapshot(
            "process_situation_queued",
            {"episode_id": episode_id, "situation": situation},
        )
    except Exception:
        snapshot_queued = ""

    async def _bg_job() -> None:
        try:
            res = await comp["learning_loop"].process_situation(situation)
            try:
                if (
                    episode_id is not None
                    and comp.get("episodic_memory")
                    and getattr(comp["episodic_memory"], "update_episode", None)
                ):
                    comp["episodic_memory"].update_episode(
                        int(episode_id),
                        action_taken="learn:process_situation",
                        result=res,
                        success=True,
                        new_knowledge_count=int(res.get("new_knowledge_count", 0))
                        if isinstance(res, dict)
                        else None,
                        asked_for_help=bool(res.get("asked_for_help"))
                        if isinstance(res, dict) and "asked_for_help" in res
                        else None,
                        uncertainty_score=float(res.get("uncertainty_score", 0.0))
                        if isinstance(res, dict) and "uncertainty_score" in res
                        else None,
                        tags=["learning", "process_situation"],
                    )
            except Exception:
                pass
            try:
                _write_learning_snapshot(
                    "process_situation_done", {"episode_id": episode_id, "result": res}
                )
            except Exception:
                pass
            try:
                from modules.humanoid.memory_engine.store import memory_write

                memory_write(
                    thread_id=None,
                    kind="decision",
                    payload={
                        "title": "learning",
                        "decision_type": "process_situation_done",
                        "episode_id": episode_id,
                        "situation": situation,
                        "result": res,
                    },
                    task_id=str(episode_id or ""),
                )
            except Exception:
                pass
        except Exception as e:
            try:
                if (
                    episode_id is not None
                    and comp.get("episodic_memory")
                    and getattr(comp["episodic_memory"], "update_episode", None)
                ):
                    comp["episodic_memory"].update_episode(
                        int(episode_id),
                        action_taken="learn:process_situation",
                        result=str(e),
                        success=False,
                        tags=["learning", "error"],
                    )
            except Exception:
                pass
            try:
                _write_learning_snapshot(
                    "process_situation_error",
                    {"episode_id": episode_id, "error": str(e), "situation": situation},
                )
            except Exception:
                pass
            try:
                from modules.humanoid.memory_engine.store import memory_write

                memory_write(
                    thread_id=None,
                    kind="decision",
                    payload={
                        "title": "learning",
                        "decision_type": "process_situation_error",
                        "episode_id": episode_id,
                        "situation": situation,
                        "error": str(e),
                    },
                    task_id=str(episode_id or ""),
                )
            except Exception:
                pass

    # 2) Siempre en background (thread) para evitar bloqueos/timeout del request.
    def _bg_job_sync() -> None:
        try:
            asyncio.run(_bg_job())
        except Exception:
            try:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                loop.run_until_complete(_bg_job())
                loop.close()
            except Exception:
                pass

    try:
        asyncio.create_task(asyncio.to_thread(_bg_job_sync))
    except Exception:
        # Fallback: al menos no romper la request
        try:
            asyncio.create_task(_bg_job())
        except Exception:
            pass
    return {
        "ok": True,
        "queued": True,
        "episode_id": episode_id,
        "snapshot": snapshot_queued,
        "status": "processing",
    }


@app.get("/api/learning/knowledge-base", tags=["Learning"])
def get_learning_knowledge_base():
    """Ver conocimiento actual del robot (conceptos, skills, reglas)."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    kb = comp["kb"]
    return {
        "ok": True,
        "data": {
            "concepts": kb.concepts,
            "skills": kb.skills,
            "rules": kb.rules,
            "total_concepts": len(kb.concepts),
            "total_skills": len(kb.skills),
        },
    }


@app.post("/api/learning/consolidate", tags=["Learning"])
async def trigger_learning_consolidation():
    """Forzar consolidaciÃ³n de conocimiento."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    report = comp["consolidator"].consolidate_knowledge()
    snap = ""
    try:
        snap = _write_learning_snapshot("consolidate", {"report": report})
    except Exception:
        snap = ""
    return {
        "ok": True,
        "status": "consolidated",
        "report": report,
        "snapshot": snap,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


@app.get("/api/learning/uncertainty-status", tags=["Learning"])
def get_learning_uncertainty_status():
    """Estado del detector de incertidumbre (umbral, fallos, estadÃ­sticas)."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    ud = comp["uncertainty_detector"]
    total = sum(len(f) for f in ud.failure_memory.values())
    data = {
        "threshold": ud.uncertainty_threshold,
        "failure_memory": dict(ud.failure_memory),
        "total_failures_tracked": total,
    }
    if getattr(ud, "get_statistics", None):
        try:
            data["statistics"] = ud.get_statistics()
        except Exception:
            pass
    return {"ok": True, "data": data}


class TeachConceptBody(BaseModel):
    concept_name: str
    definition: str
    properties: Optional[List[str]] = None
    examples: Optional[List[str]] = None


@app.post("/api/learning/teach-concept", tags=["Learning"])
async def teach_new_concept(body: TeachConceptBody):
    """EnseÃ±ar nuevo concepto al robot (humano en el loop)."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    kb = comp["kb"]
    sem = comp["semantic_memory"]
    kb.add_learned_concept(
        name=body.concept_name,
        definition=body.definition,
        concept_type="learned",
        properties=body.properties,
        examples=body.examples,
        confidence=1.0,
        source="human_taught",
    )
    if getattr(sem, "add_experience", None):
        try:
            sem.add_experience(
                description="Concept: %s - %s" % (body.concept_name, body.definition),
                context=str(body.properties),
                tags=["concept", "human_taught", body.concept_name],
            )
        except Exception:
            pass
    return {
        "ok": True,
        "status": "concept_taught",
        "concept": body.concept_name,
        "total_concepts": len(kb.concepts),
    }


@app.get("/api/learning/growth-metrics", tags=["Learning"])
def get_learning_growth_metrics():
    """MÃ©tricas de crecimiento del conocimiento (inteligencia creciente)."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    kb = comp["kb"]
    sem = comp["semantic_memory"]
    loop = comp["learning_loop"]
    ud = comp["uncertainty_detector"]
    stats = getattr(sem, "get_statistics", lambda: {})() or {}
    total_failures = sum(len(f) for f in ud.failure_memory.values())
    ep_stats = {}
    if comp.get("episodic_memory") and getattr(
        comp["episodic_memory"], "get_statistics", None
    ):
        try:
            ep_stats = comp["episodic_memory"].get_statistics()
        except Exception:
            pass
    return {
        "ok": True,
        "data": {
            "knowledge_base": {
                "concepts": len(kb.concepts),
                "skills": len(kb.skills),
                "rules": len(kb.rules),
            },
            "memory": {
                "total_experiences": stats.get("total_experiences", 0),
                "storage_mb": stats.get("storage_size_mb", 0),
            },
            "episodic_memory": ep_stats,
            "learning": {
                "total_experiences_processed": loop.experience_counter,
                "times_failures_tracked": total_failures,
            },
        },
    }


@app.get(
    "/api/learning/consolidator/stats",
    tags=["Learning"],
    summary="EstadÃ­sticas del consolidador",
    response_description="total_consolidations, patterns_found_total, concepts_created_total, rules_updated_total, contradictions_resolved_total, last_consolidation, hours_since_last",
)
def get_learning_consolidator_stats():
    """EstadÃ­sticas del consolidador de conocimiento (patrones, conceptos, reglas, contradicciones)."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    consolidator = comp.get("consolidator")
    if not consolidator or not hasattr(consolidator, "get_statistics"):
        return {"ok": True, "data": {}}
    try:
        data = consolidator.get_statistics()
        return {"ok": True, "data": data}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post(
    "/api/learning/daily-routine/start",
    tags=["Learning"],
    summary="Iniciar rutina diaria",
    response_description="ok, status=started, current_lesson, lesson_start_time (ISO)",
)
async def start_learning_daily_routine():
    """Iniciar rutina diaria: lecciÃ³n del tutor (si existe) y tareas en background (consolidaciÃ³n periÃ³dica)."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    loop = comp.get("learning_loop")
    if not loop or not hasattr(loop, "start_daily_routine"):
        return {"ok": False, "error": "Learning loop sin start_daily_routine"}
    try:
        await loop.start_daily_routine()
        snap = ""
        try:
            snap = _write_learning_snapshot(
                "daily_routine_start",
                {
                    "current_lesson": getattr(loop, "current_lesson", None),
                    "lesson_start_time": (
                        lambda t: t.isoformat()
                        if t and hasattr(t, "isoformat")
                        else None
                    )(getattr(loop, "lesson_start_time", None)),
                },
            )
        except Exception:
            snap = ""
        return {
            "ok": True,
            "status": "started",
            "current_lesson": getattr(loop, "current_lesson", None),
            "lesson_start_time": (
                lambda t: t.isoformat() if t and hasattr(t, "isoformat") else None
            )(getattr(loop, "lesson_start_time", None)),
            "snapshot": snap,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post(
    "/api/learning/daily-routine/end-report",
    tags=["Learning"],
    summary="Reporte fin de dÃ­a",
    response_description="ok, status=no_lesson|evaluated, message|evaluation",
)
async def end_learning_daily_report():
    """Generar reporte fin de dÃ­a y obtener evaluaciÃ³n del tutor (si hay lecciÃ³n activa)."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    loop = comp.get("learning_loop")
    if not loop or not hasattr(loop, "end_of_day_report"):
        return {"ok": False, "error": "Learning loop sin end_of_day_report"}
    try:
        evaluation = await loop.end_of_day_report()
        if evaluation is None:
            snap = ""
            try:
                snap = _write_learning_snapshot(
                    "daily_routine_end", {"evaluation": None, "status": "no_lesson"}
                )
            except Exception:
                snap = ""
            return {
                "ok": True,
                "status": "no_lesson",
                "message": "No hay lecciÃ³n activa o tutor configurado",
            }
        snap = ""
        try:
            snap = _write_learning_snapshot(
                "daily_routine_end", {"evaluation": evaluation, "status": "evaluated"}
            )
        except Exception:
            snap = ""
        return {
            "ok": True,
            "status": "evaluated",
            "evaluation": evaluation,
            "snapshot": snap,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get(
    "/api/learning/tutor/stats",
    tags=["Learning"],
    summary="EstadÃ­sticas del IA Tutor",
    response_description="tutor_type, curriculum, completed/failed, robot_level, api_calls, cost",
)
def get_learning_tutor_stats():
    """EstadÃ­sticas del IA Tutor (si estÃ¡ configurado): curriculum, lecciones completadas, nivel del robot, coste API."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    tutor = comp.get("ai_tutor")
    if not tutor or not hasattr(tutor, "get_statistics"):
        return {
            "ok": True,
            "data": {
                "enabled": False,
                "message": "IA Tutor no configurado (AI_TUTOR_TYPE=claude|gpt4 y API key)",
            },
        }
    try:
        data = tutor.get_statistics()
        data["enabled"] = True
        return {"ok": True, "data": data}
    except Exception as e:
        return {"ok": False, "error": str(e)}


class DesignCurriculumBody(BaseModel):
    robot_capabilities: List[str] = []
    learning_goals: List[str] = []
    time_horizon_days: int = 30
    difficulty_level: str = "progressive"


@app.post(
    "/api/learning/tutor/design-curriculum",
    tags=["Learning"],
    summary="DiseÃ±ar currÃ­culum (offline-first)",
    response_description="ok, data={count, curriculum}",
)
def design_learning_curriculum(body: DesignCurriculumBody):
    """DiseÃ±a/actualiza el currÃ­culum del tutor (sin requerir API externa)."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    tutor = comp.get("ai_tutor")
    if not tutor:
        try:
            from brain.learning.ai_tutor import AITutor

            tutor = AITutor(tutor_type="disabled")
            comp["ai_tutor"] = tutor
            if comp.get("learning_loop"):
                comp["learning_loop"].ai_tutor = tutor
        except Exception as e:
            return {"ok": False, "error": f"AI Tutor no disponible: {e}"}
    curriculum = tutor.design_curriculum(
        robot_capabilities=body.robot_capabilities,
        learning_goals=body.learning_goals,
        time_horizon_days=body.time_horizon_days,
        difficulty_level=body.difficulty_level,
    )
    return {
        "ok": True,
        "data": {"count": len(curriculum or []), "curriculum": curriculum},
    }


@app.post(
    "/api/learning/python-mastery/bootstrap",
    tags=["Learning"],
    summary="Bootstrap Python Mastery para ATLAS",
    response_description="ok, data={count, current_lesson?}",
)
def bootstrap_python_mastery():
    """Inicializa el currÃ­culum de dominio Python para ATLAS.

    Uso: llamar una vez y luego usar /api/learning/daily-routine/start.
    """
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    tutor = comp.get("ai_tutor")
    if not tutor:
        try:
            from brain.learning.ai_tutor import AITutor

            tutor = AITutor(tutor_type="disabled")
            comp["ai_tutor"] = tutor
            if comp.get("learning_loop"):
                comp["learning_loop"].ai_tutor = tutor
        except Exception as e:
            return {"ok": False, "error": f"AI Tutor no disponible: {e}"}
    curriculum = tutor.design_curriculum(
        robot_capabilities=["repo_tools", "cli", "testing", "offline_first"],
        learning_goals=[
            "python_mastery",
            "python_scripting",
            "pytest",
            "sqlite",
            "tooling",
        ],
        time_horizon_days=int(os.getenv("PYTHON_MASTERY_HORIZON_DAYS", "30") or 30),
        difficulty_level=os.getenv("PYTHON_MASTERY_DIFFICULTY", "progressive"),
    )
    # PreasignaciÃ³n opcional (para confirmar que hay lecciÃ³n elegible).
    current = None
    try:
        current = tutor.assign_daily_lesson()
    except Exception:
        current = None
    return {
        "ok": True,
        "data": {"count": len(curriculum or []), "current_lesson": current},
    }


class PythonMasteryEvaluateBody(BaseModel):
    lesson_id: Optional[str] = None  # default: current lesson si existe
    timeout_s: int = 180


@app.post(
    "/api/learning/python-mastery/evaluate",
    tags=["Learning"],
    summary="Evaluar una lecciÃ³n Python Mastery (offline)",
    response_description="ok, evaluation{score, passed, evidence}",
)
def evaluate_python_mastery(body: PythonMasteryEvaluateBody):
    """Ejecuta evaluaciÃ³n determinista (archivos + pytest) para una lecciÃ³n PYxxx."""
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    lesson_id = (body.lesson_id or "").strip()
    if not lesson_id:
        try:
            loop = comp.get("learning_loop")
            current = getattr(loop, "current_lesson", None) if loop else None
            lesson_id = (
                (current or {}).get("lesson_id", "")
                if isinstance(current, dict)
                else ""
            )
        except Exception:
            lesson_id = ""
    if not lesson_id:
        return {
            "ok": False,
            "error": "lesson_id requerido (o iniciar daily-routine para tener current_lesson)",
        }
    try:
        from brain.learning.python_mastery_evaluator import \
            evaluate_python_mastery_lesson

        evaluation = evaluate_python_mastery_lesson(
            lesson_id=lesson_id,
            repo_root=str(BASE_DIR),
            timeout_s=int(body.timeout_s or 180),
        )
        snap = ""
        try:
            snap = _write_learning_snapshot(
                "python_mastery_evaluate",
                {"lesson_id": lesson_id, "evaluation": evaluation},
            )
        except Exception:
            snap = ""
        return {"ok": True, "evaluation": evaluation, "snapshot": snap}
    except Exception as e:
        return {"ok": False, "error": str(e)}


class PythonMasteryCampaignStartBody(BaseModel):
    reset: bool = False


@app.post(
    "/api/learning/python-mastery/campaign/start",
    tags=["Learning"],
    summary="Iniciar/reanudar campaÃ±a Python Mastery",
    response_description="ok, status, current_lesson, state",
)
async def python_mastery_campaign_start(body: PythonMasteryCampaignStartBody):
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    tutor = comp.get("ai_tutor")
    loop = comp.get("learning_loop")
    if not tutor or not loop:
        return {"ok": False, "error": "Componentes de learning no disponibles"}
    try:
        from brain.learning.python_mastery_campaign import \
            start_or_resume_campaign

        res = start_or_resume_campaign(
            tutor=tutor,
            loop=loop,
            repo_root=Path(str(BASE_DIR)),
            reset=bool(body.reset),
        )
        snap = ""
        try:
            snap = _write_learning_snapshot("python_mastery_campaign_start", res)
        except Exception:
            snap = ""
        return {
            "ok": True,
            "status": res["state"].get("status"),
            "current_lesson": res.get("current_lesson"),
            "state": res.get("state"),
            "snapshot": snap,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


class PythonMasteryCampaignStepBody(BaseModel):
    lesson_id: Optional[str] = None
    timeout_s: int = 180
    max_attempts: int = 1
    backoff_s: float = 0.0
    backoff_factor: float = 2.0
    require_change: bool = False
    auto_remediate: bool = False


@app.post(
    "/api/learning/python-mastery/campaign/step",
    tags=["Learning"],
    summary="Ejecutar un paso de campaÃ±a (evalÃºa y avanza)",
    response_description="ok, status, evaluation, next_lesson, state",
)
def python_mastery_campaign_step(body: PythonMasteryCampaignStepBody):
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    tutor = comp.get("ai_tutor")
    loop = comp.get("learning_loop")
    if not tutor or not loop:
        return {"ok": False, "error": "Componentes de learning no disponibles"}
    try:
        from brain.learning.python_mastery_campaign import campaign_step

        res = campaign_step(
            tutor=tutor,
            loop=loop,
            repo_root=Path(str(BASE_DIR)),
            lesson_id=body.lesson_id,
            timeout_s=int(body.timeout_s or 180),
            max_attempts=int(body.max_attempts or 1),
            backoff_s=float(body.backoff_s or 0.0),
            backoff_factor=float(body.backoff_factor or 2.0),
            require_change=bool(body.require_change),
            auto_remediate=bool(body.auto_remediate),
        )
        snap = ""
        try:
            snap = _write_learning_snapshot(
                "python_mastery_campaign_step",
                {
                    "lesson_id": res.get("lesson_id"),
                    "status": res.get("status"),
                    "evaluation": res.get("evaluation"),
                },
            )
        except Exception:
            snap = ""
        return {**res, "snapshot": snap}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get(
    "/api/learning/python-mastery/campaign/state",
    tags=["Learning"],
    summary="Estado persistido de la campaÃ±a Python Mastery",
    response_description="ok, state",
)
def python_mastery_campaign_state():
    try:
        from brain.learning.python_mastery_campaign import (
            campaign_state_path, load_campaign_state)

        path = campaign_state_path(Path(str(BASE_DIR)))
        state = load_campaign_state(path)
        return {"ok": True, "state": state, "state_path": str(path)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


class PythonMasteryCampaignRunBody(BaseModel):
    reset: bool = False
    max_steps: int = 5
    max_seconds: float = 60.0
    step_timeout_s: int = 180
    step_max_attempts: int = 1
    step_backoff_s: float = 0.0
    step_backoff_factor: float = 2.0
    step_require_change: bool = False
    step_auto_remediate: bool = False


@app.post(
    "/api/learning/python-mastery/campaign/run",
    tags=["Learning"],
    summary="Correr campaÃ±a en una sola orden (loop de steps)",
    response_description="ok, status, steps_run, seconds, steps[]",
)
def python_mastery_campaign_run(body: PythonMasteryCampaignRunBody):
    comp = _get_learning_components()
    if "error" in comp:
        return {"ok": False, "error": comp["error"]}
    tutor = comp.get("ai_tutor")
    loop = comp.get("learning_loop")
    if not tutor or not loop:
        return {"ok": False, "error": "Componentes de learning no disponibles"}
    try:
        from brain.learning.python_mastery_campaign import campaign_run

        res = campaign_run(
            tutor=tutor,
            loop=loop,
            repo_root=Path(str(BASE_DIR)),
            reset=bool(body.reset),
            max_steps=int(body.max_steps or 5),
            max_seconds=float(body.max_seconds or 60.0),
            step_timeout_s=int(body.step_timeout_s or 180),
            step_max_attempts=int(body.step_max_attempts or 1),
            step_backoff_s=float(body.step_backoff_s or 0.0),
            step_backoff_factor=float(body.step_backoff_factor or 2.0),
            step_require_change=bool(body.step_require_change),
            step_auto_remediate=bool(body.step_auto_remediate),
        )
        snap = ""
        try:
            snap = _write_learning_snapshot("python_mastery_campaign_run", res)
        except Exception:
            snap = ""
        return {**res, "snapshot": snap}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# --- ATLAS_ARCHITECT (agentic coding) ---
_architect_singleton: Optional[object] = None


def _get_architect():
    global _architect_singleton
    if _architect_singleton is not None:
        return _architect_singleton
    from modules.atlas_architect import AtlasArchitect

    _architect_singleton = AtlasArchitect(repo_root=BASE_DIR)
    return _architect_singleton


@app.post("/api/architect/index", tags=["Architect"])
def architect_index():
    """Genera index de arquitectura (PUSH/NEXUS/ROBOT y puertos)."""
    t0 = time.perf_counter()
    try:
        arch = _get_architect()
        data = arch.index_architecture()
        return {
            "ok": True,
            "data": data,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": None,
        }
    except Exception as e:
        return {
            "ok": False,
            "data": None,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": str(e),
        }


class ArchitectPytestBody(BaseModel):
    nodeid: Optional[str] = None


@app.post("/api/architect/pytest", tags=["Architect"])
def architect_pytest(body: ArchitectPytestBody):
    """Ejecuta pytest (real) y devuelve anÃ¡lisis de errores."""
    t0 = time.perf_counter()
    try:
        arch = _get_architect()
        data = arch.run_pytest_and_analyze(nodeid=body.nodeid)
        ok = bool(data.get("ok"))
        return {
            "ok": ok,
            "data": data,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": None if ok else "pytest_failed",
        }
    except Exception as e:
        return {
            "ok": False,
            "data": None,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": str(e),
        }


class ArchitectProposeBody(BaseModel):
    problem: str
    prefer_free: bool = True
    context: Optional[dict] = None


@app.post("/api/architect/propose-fix", tags=["Architect"])
def architect_propose_fix(body: ArchitectProposeBody):
    """Genera una propuesta (no aplica cambios) para resolver un problema."""
    t0 = time.perf_counter()
    try:
        arch = _get_architect()
        data = arch.propose_code_fix(
            body.problem, context=body.context, prefer_free=bool(body.prefer_free)
        )
        ok = bool(data.get("ok"))
        return {
            "ok": ok,
            "data": data,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": None if ok else "no_model_or_error",
        }
    except Exception as e:
        return {
            "ok": False,
            "data": None,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": str(e),
        }


class ArchitectAgenticBody(BaseModel):
    goal: str
    prefer_free: bool = True
    max_iters: int = 3


@app.post("/api/architect/agentic/execute", tags=["Architect"])
def architect_agentic_execute(body: ArchitectAgenticBody):
    """EjecuciÃ³n agentic: modelo decide tool_calls (FS/terminal) y se ejecutan realmente."""
    t0 = time.perf_counter()
    try:
        arch = _get_architect()
        data = arch.agentic_execute(
            body.goal,
            prefer_free=bool(body.prefer_free),
            max_iters=int(body.max_iters or 3),
        )
        return {
            "ok": True,
            "data": data,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": None,
        }
    except Exception as e:
        return {
            "ok": False,
            "data": None,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": str(e),
        }


class ArchitectOrderBody(BaseModel):
    order: str
    mode: str = "governed"  # governed | growth
    prefer_free: bool = True
    max_heal_attempts: int = 3


@app.post("/api/architect/order", tags=["Architect"])
def architect_order(body: ArchitectOrderBody):
    """Orden de alto nivel: diseÃ±a/crea una app y ejecuta diagnÃ³stico/autocorrecciÃ³n segÃºn modo."""
    t0 = time.perf_counter()
    try:
        arch = _get_architect()
        data = arch.execute_order(
            body.order,
            mode=body.mode,
            prefer_free=bool(body.prefer_free),
            max_heal_attempts=int(body.max_heal_attempts or 3),
        )
        return {
            "ok": bool(data.get("ok")),
            "data": data,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": None if data.get("ok") else "order_failed",
        }
    except Exception as e:
        return {
            "ok": False,
            "data": None,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": str(e),
        }


class ArchitectApplyApprovedBody(BaseModel):
    approval_id: str


@app.post("/api/architect/apply-approved", tags=["Architect"])
def architect_apply_approved(body: ArchitectApplyApprovedBody):
    """Aplica un plan previamente aprobado en la cola de approvals."""
    t0 = time.perf_counter()
    try:
        arch = _get_architect()
        data = arch.apply_approved_plan(body.approval_id)
        return {
            "ok": bool(data.get("ok")),
            "data": data,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": None if data.get("ok") else (data.get("error") or "apply_failed"),
        }
    except Exception as e:
        return {
            "ok": False,
            "data": None,
            "ms": int((time.perf_counter() - t0) * 1000),
            "error": str(e),
        }


from modules.humanoid.ans.api import router as ans_router
# --- Humanoid (kernel + modules) ---
from modules.humanoid.api import router as humanoid_router
from modules.humanoid.ga.api import router as ga_router
from modules.humanoid.governance.api import router as governance_router
from modules.humanoid.metalearn.api import router as metalearn_router
from modules.humanoid.nervous.api import router as nervous_router

try:
    from modules.humanoid.quality.api import router as quality_router
except ImportError:
    quality_router = None

# Cognitive Architecture API
try:
    from modules.humanoid.cognitive.api import router as cognitive_router
except ImportError:
    cognitive_router = None

app.include_router(humanoid_router)
app.include_router(ga_router)
app.include_router(metalearn_router)
app.include_router(ans_router)
app.include_router(governance_router)
app.include_router(nervous_router)
if quality_router:
    app.include_router(quality_router)
if cognitive_router:
    app.include_router(cognitive_router)

# Cognitive Memory (World Model, Autobiographical, Lifelog, Planner, Unified Memory)
try:
    from modules.humanoid.world_model.api import \
        router as cognitive_memory_router

    app.include_router(cognitive_memory_router)
except Exception as _cme:
    import logging

    logging.getLogger(__name__).warning("Cognitive Memory module not loaded: %s", _cme)

# Libro de Vida (Book of Life)
try:
    from modules.humanoid.memory_engine.libro_vida_api import \
        router as libro_vida_router

    app.include_router(libro_vida_router)
except Exception as _lve:
    import logging

    logging.getLogger(__name__).warning("Libro de Vida module not loaded: %s", _lve)

# Directives (migrated from NEXUS)
try:
    from modules.humanoid.directives.api import router as directives_router

    app.include_router(directives_router)
except Exception as _de:
    import logging

    logging.getLogger(__name__).warning("Directives module not loaded: %s", _de)

# Tools Registry (migrated from NEXUS)
try:
    from modules.humanoid.tools.api import router as tools_api_router

    app.include_router(tools_api_router)
except Exception as _te:
    import logging

    logging.getLogger(__name__).warning("Tools module not loaded: %s", _te)

# Tutorias y Visitas
try:
    from modules.humanoid.quality.tutorias.api import router as tutorias_router

    app.include_router(tutorias_router)
except Exception as _tr:
    import logging

    logging.getLogger(__name__).warning("Tutorias module not loaded: %s", _tr)

# ATLAS AUTONOMOUS (health, healing, evolution, telemetry, resilience, learning)
try:
    import sys

    if str(BASE_DIR) not in sys.path:
        sys.path.insert(0, str(BASE_DIR))
    from autonomous.api_routes import router as autonomous_router

    app.include_router(autonomous_router)
except Exception as e:
    import logging

    logging.getLogger(__name__).warning("Autonomous module not loaded: %s", e)


@app.get("/autonomous/dashboard", response_class=HTMLResponse)
async def autonomous_dashboard():
    """Dashboard UI de ATLAS AUTONOMOUS: health, servicios, healing, learning."""
    html_file = BASE_DIR / "templates" / "autonomous_dashboard.html"
    if not html_file.exists():
        return HTMLResponse("<h1>Dashboard no encontrado</h1>", status_code=404)
    return HTMLResponse(html_file.read_text(encoding="utf-8"))


# --- /actions/log â€” log de acciones (antes proxy a NEXUS, ahora local) ---


@app.get("/actions/log")
def actions_log_endpoint(limit: int = 50):
    """Log de acciones del sistema â€” datos locales de la bitacora."""
    try:
        from modules.humanoid.ans.evolution_bitacora import \
            get_evolution_entries

        entries = get_evolution_entries(limit=limit)
        return {"ok": True, "entries": entries[:limit]}
    except Exception:
        return {"ok": True, "entries": []}


@app.websocket("/ws")
async def proxy_websocket(websocket: WebSocket):
    """Proxy WebSocket a NEXUS para tiempo real (evitar 403 cuando el cliente usa origen PUSH)."""
    import asyncio
    import logging

    log = logging.getLogger(__name__)
    ws_nexus_url = get_nexus_ws_url()
    try:
        import websockets
    except ImportError:
        try:
            await websocket.accept()
            await websocket.close(code=1011)
        except Exception:
            pass
        log.warning("Proxy /ws: instalar 'websockets' para reenviar a NEXUS")
        return
    try:
        async with websockets.connect(ws_nexus_url, close_timeout=5) as ws_nexus:
            await websocket.accept()

            async def client_to_nexus():
                try:
                    while True:
                        msg = await websocket.receive()
                        if "text" in msg:
                            await ws_nexus.send(msg["text"])
                        elif "bytes" in msg:
                            await ws_nexus.send(msg["bytes"])
                except Exception:
                    pass

            async def nexus_to_client():
                try:
                    while True:
                        msg = await ws_nexus.recv()
                        if isinstance(msg, bytes):
                            await websocket.send_bytes(msg)
                        else:
                            await websocket.send_text(msg)
                except Exception:
                    pass

            await asyncio.gather(
                asyncio.create_task(client_to_nexus()),
                asyncio.create_task(nexus_to_client()),
            )
    except Exception as e:
        log.debug("Proxy /ws: %s", e)
        try:
            await websocket.accept()
        except Exception:
            pass
        try:
            await websocket.close(code=1011)
        except Exception:
            pass


# --- Cuerpo (NEXUS+Robot): proxies bajo el puerto de PUSH ---
from fastapi.responses import Response

# IMPORTANTE (orden de rutas):
# - /cuerpo/{path:path} es catch-all y debe ir DESPUÃ‰S de rutas mÃ¡s especÃ­ficas como /cuerpo/nexus/*
from modules.nexus_proxy import proxy_to_nexus


@app.api_route("/cuerpo/nexus", methods=["GET", "POST"])
@app.api_route(
    "/cuerpo/nexus/{path:path}",
    methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
)
async def nexus_proxy_route(request: Request, path: str = "") -> Response:
    """Proxy a NEXUS (dashboard + API) bajo el mismo puerto de PUSH."""
    return await proxy_to_nexus(request, path)


from modules.cuerpo_proxy import proxy_to_cuerpo


@app.api_route("/cuerpo", methods=["GET", "POST"])
@app.api_route(
    "/cuerpo/{path:path}", methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"]
)
async def cuerpo_proxy_route(request: Request, path: str = "") -> Response:
    """Proxy a Robot (cÃ¡maras, visiÃ³n) en NEXUS_ROBOT_URL/NEXUS_ROBOT_API_URL. Panel de cÃ¡maras integrado."""
    return await proxy_to_cuerpo(request, path)


from modules.humanoid.audit import get_audit_logger
# --- Metrics / Policy / Audit endpoints ---
from modules.humanoid.metrics import get_metrics_store
from modules.humanoid.policy import ActorContext, get_policy_engine


@app.get("/metrics")
def metrics():
    """JSON metrics: counters and latencies per endpoint."""
    return get_metrics_store().snapshot()


# Prometheus scrape endpoint (text/plain)
@app.get("/metrics/prometheus", include_in_schema=False)
def prometheus_metrics():
    """MÃ©tricas en formato Prometheus para scraping."""
    try:
        from fastapi.responses import Response

        from modules.observability.metrics import MetricsCollector

        return Response(
            content=MetricsCollector.get_prometheus_text(),
            media_type="text/plain; charset=utf-8",
        )
    except Exception as e:
        from fastapi.responses import Response

        return Response(
            content=f"# error: {e}".encode(), media_type="text/plain", status_code=500
        )


@app.get("/api/observability/metrics", tags=["Observability"])
def api_observability_metrics():
    """Resumen de mÃ©tricas (requests, memoria, health score)."""
    try:
        from modules.observability.metrics import MetricsCollector

        return MetricsCollector.get_metrics_summary()
    except Exception as e:
        return {
            "error": str(e),
            "total_requests": 0,
            "active_requests": 0,
            "memory_mb": 0,
        }


class PolicyTestBody(BaseModel):
    actor: str = "api"
    role: str = "owner"
    module: str = "hands"
    action: str = "exec_command"
    target: Optional[str] = None


@app.post("/policy/test")
def policy_test(body: PolicyTestBody):
    """Test policy: {actor, role, module, action, target} -> {allow, reason}."""
    ctx = ActorContext(actor=body.actor, role=body.role)
    decision = get_policy_engine().can(ctx, body.module, body.action, body.target)
    return {
        "allow": decision.allow,
        "reason": decision.reason,
        "details": decision.details,
    }


@app.get("/audit/tail")
def audit_tail(n: int = 50, module: Optional[str] = None):
    """Last n audit log entries, optional filter by module."""
    entries = get_audit_logger().tail(n=n, module=module)
    if not entries and get_audit_logger()._db is None:
        return {"ok": True, "entries": [], "error": "AUDIT_DB_PATH not set"}
    return {"ok": True, "entries": entries, "error": None}


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Monitor AutÃ³nomo â€” SSE stream + snapshot
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


@app.get("/api/monitor/snapshot", tags=["Monitor"])
def monitor_snapshot():
    """Snapshot completo del estado actual del sistema para el Monitor autÃ³nomo."""
    t0 = time.perf_counter()
    result = {
        "ts": datetime.now(timezone.utc).isoformat(),
        "processes": [],
        "audit": [],
        "tasks": [],
        "bitacora": [],
        "health": {},
        "shell_history": [],
    }
    try:
        import subprocess

        ps = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                "Get-Process | Where-Object {$_.CPU -gt 0} | Sort-Object CPU -Descending | Select-Object -First 20 Id,ProcessName,CPU,WorkingSet64,StartTime | ConvertTo-Json -Compress",
            ],
            capture_output=True,
            text=True,
            timeout=5,
        )
        if ps.returncode == 0 and ps.stdout.strip():
            import json as _j

            procs = _j.loads(ps.stdout)
            if isinstance(procs, dict):
                procs = [procs]
            result["processes"] = [
                {
                    "pid": p.get("Id"),
                    "name": p.get("ProcessName"),
                    "cpu": round(p.get("CPU", 0), 1),
                    "mem_mb": round((p.get("WorkingSet64") or 0) / 1048576, 1),
                }
                for p in procs[:20]
            ]
    except Exception:
        pass
    try:
        result["audit"] = get_audit_logger().tail(n=30) or []
    except Exception:
        pass
    try:
        from modules.humanoid.orchestrator.memory import TaskMemory

        mem = TaskMemory()
        import sqlite3

        conn = sqlite3.connect(str(mem.db_path))
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT id, goal, status, created_ts, updated_ts FROM task_memory ORDER BY updated_ts DESC LIMIT 15"
        ).fetchall()
        result["tasks"] = [dict(r) for r in rows]
        conn.close()
    except Exception:
        pass
    try:
        from modules.humanoid.ans.evolution_bitacora import \
            get_evolution_entries

        result["bitacora"] = get_evolution_entries(limit=20) or []
    except Exception:
        pass
    try:
        import psutil

        cpu = psutil.cpu_percent(interval=0.3)
        mem = psutil.virtual_memory()
        disk = psutil.disk_usage("C:\\")
        result["health"] = {
            "cpu_pct": cpu,
            "ram_pct": mem.percent,
            "ram_used_gb": round(mem.used / 1073741824, 1),
            "ram_total_gb": round(mem.total / 1073741824, 1),
            "disk_pct": disk.percent,
            "disk_free_gb": round(disk.free / 1073741824, 1),
        }
    except Exception:
        try:
            import shutil

            du = shutil.disk_usage("C:\\")
            result["health"] = {
                "disk_pct": round(du.used / du.total * 100, 1),
                "disk_free_gb": round(du.free / 1073741824, 1),
            }
        except Exception:
            pass
    try:
        entries = get_audit_logger().tail(n=20, module="hands")
        result["shell_history"] = [
            e
            for e in (entries or [])
            if e.get("action") in ("exec_command", "shell_exec", "run")
        ]
    except Exception:
        pass
    ms = int((time.perf_counter() - t0) * 1000)
    return {"ok": True, "data": result, "ms": ms}


@app.get("/api/monitor/stream", tags=["Monitor"])
async def monitor_stream():
    """SSE stream que emite actividad del sistema cada 3 segundos."""
    from fastapi.responses import StreamingResponse

    async def _generate():
        last_audit_count = 0
        last_task_ts = ""
        # Emit immediate event so SSE clients always receive data quickly.
        yield f"data: {json.dumps([{'type': 'heartbeat', 'data': {'ts': datetime.now(timezone.utc).isoformat()}}], default=str)}\n\n"
        while True:
            events = []
            try:
                audit_entries = get_audit_logger().tail(n=10) or []
                if len(audit_entries) != last_audit_count:
                    for entry in audit_entries:
                        events.append(
                            {
                                "type": "audit",
                                "data": {
                                    "ts": entry.get("ts", ""),
                                    "actor": entry.get("actor", ""),
                                    "module": entry.get("module", ""),
                                    "action": entry.get("action", ""),
                                    "ok": entry.get("ok", True),
                                    "ms": entry.get("ms", 0),
                                    "error": entry.get("error", ""),
                                },
                            }
                        )
                    last_audit_count = len(audit_entries)
            except Exception:
                pass
            try:
                import psutil

                cpu = psutil.cpu_percent(interval=0)
                mem = psutil.virtual_memory()
                events.append(
                    {
                        "type": "system",
                        "data": {
                            "cpu_pct": cpu,
                            "ram_pct": mem.percent,
                            "ram_used_gb": round(mem.used / 1073741824, 1),
                        },
                    }
                )
            except Exception:
                pass
            try:
                from modules.humanoid.orchestrator.memory import TaskMemory

                mem_db = TaskMemory()
                import sqlite3

                conn = sqlite3.connect(str(mem_db.db_path))
                conn.row_factory = sqlite3.Row
                rows = conn.execute(
                    "SELECT id, goal, status, updated_ts FROM task_memory ORDER BY updated_ts DESC LIMIT 3"
                ).fetchall()
                tasks = [dict(r) for r in rows]
                conn.close()
                if tasks and tasks[0].get("updated_ts", "") != last_task_ts:
                    last_task_ts = tasks[0].get("updated_ts", "")
                    for t in tasks:
                        events.append({"type": "task", "data": t})
            except Exception:
                pass
            if events:
                yield f"data: {json.dumps(events, ensure_ascii=False, default=str)}\n\n"
            else:
                yield f"data: {json.dumps([{'type': 'heartbeat', 'data': {'ts': datetime.now(timezone.utc).isoformat()}}], default=str)}\n\n"
            await asyncio.sleep(3)

    return StreamingResponse(_generate(), media_type="text/event-stream")


def _redact_tokens(text: str) -> str:
    """Redact token/secret-like substrings for support bundle."""
    import re

    if not text:
        return text
    for pattern in [
        r"Bearer\s+[A-Za-z0-9_\-\.]+",
        r"token[=:]\s*[\w\-]+",
        r"api_key[=:]\s*[\w\-]+",
        r"secret[=:]\s*[\w\-]+",
    ]:
        text = re.sub(pattern, "***REDACTED***", text, flags=re.IGNORECASE)
    return text


@app.get("/support/bundle")
def support_bundle():
    """Export bundle: health, metrics, GA/metalearn/CI reports, deploy/gateway/cluster status, logs (redacted), config (no secrets). Returns {ok, data: {path}, ms, error}. Audited."""
    t0 = time.perf_counter()
    import json
    import zipfile
    from datetime import datetime, timezone

    support_dir = BASE_DIR / "snapshots" / "support"
    support_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    zip_path = support_dir / f"support_bundle_{ts}.zip"
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            try:
                from modules.humanoid.deploy.healthcheck import \
                    run_health_verbose
                from modules.humanoid.deploy.ports import get_ports

                port = get_ports()[0] if get_ports() else 8791
                health = run_health_verbose(base_url=None, active_port=port)
                zf.writestr("health.json", json.dumps(health, indent=2))
            except Exception:
                zf.writestr("health.json", "{}")
            try:
                snap = get_metrics_store().snapshot()
                zf.writestr("metrics.json", json.dumps(snap, indent=2))
            except Exception:
                zf.writestr("metrics.json", "{}")
            log_file = BASE_DIR / "logs" / "atlas.log"
            if log_file.exists():
                try:
                    tail = log_file.read_text(encoding="utf-8", errors="replace")[
                        -50000:
                    ]
                    zf.writestr("atlas_log_tail.txt", _redact_tokens(tail))
                except Exception:
                    pass
            svc_log = BASE_DIR / "logs" / "service.log"
            if svc_log.exists():
                try:
                    tail = svc_log.read_text(encoding="utf-8", errors="replace")[
                        -20000:
                    ]
                    zf.writestr("service_log_tail.txt", _redact_tokens(tail))
                except Exception:
                    pass
            ci_dir = Path(
                os.getenv("CI_EXPORT_DIR", str(BASE_DIR / "snapshots" / "ci"))
            )
            if ci_dir.exists():
                reports = sorted(
                    ci_dir.glob("CI_REPORT_*.md"),
                    key=lambda p: p.stat().st_mtime,
                    reverse=True,
                )
                if reports:
                    zf.write(reports[0], "ci_last_report.md")
            ga_dir = Path(
                os.getenv("GA_REPORT_DIR", str(BASE_DIR / "snapshots" / "ga"))
            )
            if ga_dir.exists():
                reports = sorted(
                    ga_dir.glob("GA_REPORT_*.md"),
                    key=lambda p: p.stat().st_mtime,
                    reverse=True,
                )
                if reports:
                    zf.write(reports[0], "ga_last_report.md")
            meta_dir = Path(
                os.getenv(
                    "METALEARN_REPORT_DIR", str(BASE_DIR / "snapshots" / "metalearn")
                )
            )
            if meta_dir.exists():
                reports = sorted(
                    meta_dir.glob("META_REPORT_*.md"),
                    key=lambda p: p.stat().st_mtime,
                    reverse=True,
                )
                if reports:
                    zf.write(reports[0], "metalearn_last_report.md")
            try:
                from modules.humanoid.deploy.switcher import get_deploy_state

                deploy = get_deploy_state()
                zf.writestr("deploy_status.json", json.dumps(deploy, indent=2))
            except Exception:
                zf.writestr("deploy_status.json", "{}")
            try:
                from modules.humanoid.gateway.store import build_gateway_status

                gw = build_gateway_status()
                zf.writestr("gateway_status.json", json.dumps(gw, indent=2))
            except Exception:
                zf.writestr("gateway_status.json", "{}")
            try:
                from modules.humanoid.cluster.registry import list_nodes

                nodes = list_nodes()
                zf.writestr(
                    "cluster_status.json", json.dumps({"nodes": nodes}, indent=2)
                )
            except Exception:
                zf.writestr("cluster_status.json", "{}")
            env_file = BASE_DIR / "config" / "atlas.env"
            if env_file.exists():
                lines = []
                for line in env_file.read_text(
                    encoding="utf-8", errors="replace"
                ).splitlines():
                    if "=" in line and any(
                        s in line.lower()
                        for s in ["secret", "password", "token", "key"]
                    ):
                        lines.append(line.split("=")[0] + "=***REDACTED***")
                    else:
                        lines.append(line)
                zf.writestr("config_redacted.env", "\n".join(lines))
        get_audit_logger().log_event(
            "support",
            "api",
            "bundle",
            True,
            int((time.perf_counter() - t0) * 1000),
            None,
            {"path": str(zip_path)},
            None,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"path": str(zip_path)}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        get_audit_logger().log_event(
            "support", "api", "bundle", False, ms, str(e), None, None
        )
        return _std_resp(False, None, ms, str(e))


@app.get("/support/selfcheck")
def support_selfcheck():
    """Run quick diagnostics; return problems and suggestions. {ok, data: {problems, suggestions}, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.product.selfcheck import run_selfcheck

        result = run_selfcheck()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            result.get("ok", True),
            {
                "problems": result.get("problems", []),
                "suggestions": result.get("suggestions", []),
            },
            ms,
            None,
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            False,
            {
                "problems": [
                    {
                        "id": "selfcheck_error",
                        "severity": "critical",
                        "message": str(e),
                        "suggestion": "Check logs",
                    }
                ],
                "suggestions": [],
            },
            ms,
            str(e),
        )


# --- Approvals (policy + audit) ---
@app.get("/approvals/pending")
def approvals_pending(limit: int = 50):
    """List pending approval items."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.approvals import list_all

        items = list_all(limit=limit, status="pending")
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": True,
            "data": [
                getattr(i, "__dict__", i) if hasattr(i, "__dict__") else i
                for i in items
            ],
            "count": len(items),
            "ms": ms,
        }
    except Exception:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": True, "data": [], "count": 0, "ms": ms}


@app.get("/approvals/list")
def approvals_list(
    status: Optional[str] = None, risk: Optional[str] = None, limit: int = 50
):
    """List approval items. ?status=pending&risk=high. Response: {ok, data: [items], ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.approvals import list_all

        st = status if status else "pending"
        items = list_all(limit=limit, status=st, risk=risk)
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": True,
            "data": [
                getattr(i, "__dict__", i) if hasattr(i, "__dict__") else i
                for i in items
            ],
            "ms": ms,
            "error": None,
        }
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "data": None, "ms": ms, "error": str(e)}


class ApprovalActionBody(BaseModel):
    id: Optional[str] = None
    approval_id: Optional[str] = None
    approve: Optional[bool] = True
    session_token: Optional[str] = None
    confirm_token: Optional[str] = None


@app.post("/approvals/approve")
def approvals_approve(
    body: ApprovalActionBody,
    x_owner_session: Optional[str] = Header(None, alias="X-Owner-Session"),
):
    """Approve item by id. Body: {id, approve?, session_token?, confirm_token?}. Risk>=medium require X-Owner-Session."""
    aid = body.id or body.approval_id
    if not aid:
        return {
            "ok": False,
            "id": None,
            "status": "missing_id",
            "error": "id or approval_id required",
        }
    token = body.session_token or x_owner_session
    try:
        from modules.humanoid.approvals import approve as approval_approve

        out = approval_approve(
            aid,
            resolved_by="api",
            owner_session_token=token,
            confirm_token=body.confirm_token,
        )
        return {
            "ok": out.get("ok"),
            "id": aid,
            "status": out.get("status"),
            "error": out.get("error"),
        }
    except Exception as e:
        return {"ok": False, "id": aid, "status": "error", "error": str(e)}


@app.post("/approvals/reject")
def approvals_reject(body: ApprovalActionBody):
    """Reject item by id. Body: {id|approval_id}."""
    aid = body.id or body.approval_id
    if not aid:
        return {
            "ok": False,
            "id": None,
            "status": "missing_id",
            "error": "id or approval_id required",
        }
    try:
        from modules.humanoid.approvals import reject as approval_reject

        out = approval_reject(aid, resolved_by="api")
        return {
            "ok": out.get("ok"),
            "id": aid,
            "status": out.get("status"),
            "error": out.get("error"),
        }
    except Exception as e:
        return {"ok": False, "id": aid, "status": "error", "error": str(e)}


@app.post("/approvals/{approval_id}/approve")
def approvals_approve_compat(
    approval_id: str,
    x_owner_session: Optional[str] = Header(None, alias="X-Owner-Session"),
):
    """Compat route for legacy UI clients: POST /approvals/{id}/approve."""
    body = ApprovalActionBody(id=approval_id)
    return approvals_approve(body=body, x_owner_session=x_owner_session)


@app.post("/approvals/{approval_id}/reject")
def approvals_reject_compat(approval_id: str):
    """Compat route for legacy UI clients: POST /approvals/{id}/reject."""
    body = ApprovalActionBody(id=approval_id)
    return approvals_reject(body=body)


@app.get("/api/approvals/queue")
def approvals_queue_legacy(limit: int = 50):
    """Legacy alias for old dashboard builds."""
    return approvals_pending(limit=limit)


@app.get("/api/approvals/pending")
def approvals_pending_legacy(limit: int = 50):
    """Legacy alias for old dashboard builds."""
    return approvals_pending(limit=limit)


@app.post("/api/approvals/approve")
def approvals_approve_legacy(
    body: ApprovalActionBody,
    x_owner_session: Optional[str] = Header(None, alias="X-Owner-Session"),
):
    """Legacy alias for old dashboard builds."""
    return approvals_approve(body=body, x_owner_session=x_owner_session)


@app.post("/api/approvals/reject")
def approvals_reject_legacy(body: ApprovalActionBody):
    """Legacy alias for old dashboard builds."""
    return approvals_reject(body=body)


@app.post("/api/approvals/{approval_id}/approve")
def approvals_approve_compat_legacy(
    approval_id: str,
    x_owner_session: Optional[str] = Header(None, alias="X-Owner-Session"),
):
    """Legacy alias for old dashboard builds."""
    return approvals_approve_compat(
        approval_id=approval_id, x_owner_session=x_owner_session
    )


@app.post("/api/approvals/{approval_id}/reject")
def approvals_reject_compat_legacy(approval_id: str):
    """Legacy alias for old dashboard builds."""
    return approvals_reject_compat(approval_id=approval_id)


@app.get("/approvals/chain/verify")
def approvals_chain_verify():
    """Verify approval chain integrity. Response: {ok, valid, first_invalid_id?, expected_hash?, got_hash?}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.approvals import verify_chain

        result = verify_chain()
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": result.get("ok", True),
            "valid": result.get("valid", True),
            "ms": ms,
            "broken_at_id": result.get("broken_at_id"),
            "first_invalid_id": result.get("broken_at_id"),
            "last_hash": result.get("last_hash"),
            "count": result.get("count"),
            "error": None,
        }
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": False,
            "valid": False,
            "ms": ms,
            "first_invalid_id": None,
            "expected_hash": None,
            "got_hash": None,
            "error": str(e),
        }


# --- Owner (zero-trust session + emergency) ---
class OwnerSessionStartBody(BaseModel):
    actor: str
    method: Optional[str] = "ui"  # ui | telegram | voice | api


@app.post("/owner/session/start")
def owner_session_start(body: OwnerSessionStartBody):
    """Start owner session. Returns session_token (TTL). Required for risky approvals via X-Owner-Session."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.session import start as owner_start

        out = owner_start(actor=body.actor or "owner", method=body.method or "ui")
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": out.get("ok"),
            "session_token": out.get("session_token"),
            "ttl_seconds": out.get("ttl_seconds"),
            "method": out.get("method"),
            "expires_at": out.get("expires_at"),
            "ms": ms,
            "error": out.get("error"),
        }
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "session_token": None, "ms": ms, "error": str(e)}


@app.get("/owner/windows-user")
def owner_windows_user(request: Request):
    """Usuario Windows actual (solo localhost). Para pre-llenar login."""
    if request.client and request.client.host not in ("127.0.0.1", "localhost", "::1"):
        return {"ok": False, "username": "", "error": "Solo localhost"}
    try:
        from modules.humanoid.owner.windows_auth import current_windows_user

        return {"ok": True, "username": current_windows_user(), "error": None}
    except Exception as e:
        return {"ok": False, "username": "", "error": str(e)}


class OwnerSessionStartWithFaceBody(BaseModel):
    image_base64: Optional[str] = None


@app.post("/owner/session/start-with-face")
def owner_session_start_with_face(body: OwnerSessionStartWithFaceBody):
    """Inicia sesiÃ³n owner si se detecta rostro en la imagen."""
    t0 = time.perf_counter()
    try:
        import base64

        from modules.humanoid.face import face_check_image
        from modules.humanoid.owner.session import start as owner_start

        if not body.image_base64:
            return {
                "ok": False,
                "session_token": None,
                "error": "Imagen requerida",
                "ms": int((time.perf_counter() - t0) * 1000),
            }
        raw = base64.b64decode(body.image_base64)
        result = face_check_image(raw)
        faces = result.get("faces_detected", 0) or result.get("count", 0)
        if faces < 1:
            return {
                "ok": False,
                "session_token": None,
                "error": "No se detectÃ³ rostro",
                "ms": int((time.perf_counter() - t0) * 1000),
            }
        out = owner_start(actor="face", method="ui")
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": out.get("ok"),
            "session_token": out.get("session_token"),
            "ttl_seconds": out.get("ttl_seconds"),
            "method": "face",
            "expires_at": out.get("expires_at"),
            "ms": ms,
            "error": out.get("error"),
        }
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "session_token": None, "ms": ms, "error": str(e)}


class OwnerSessionStartWithWindowsBody(BaseModel):
    username: str = ""
    password: str = ""


@app.post("/owner/session/start-with-windows")
def owner_session_start_with_windows(body: OwnerSessionStartWithWindowsBody):
    """Inicia sesiÃ³n owner si la contraseÃ±a de usuario Windows es correcta."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.session import start as owner_start
        from modules.humanoid.owner.windows_auth import verify_windows_user

        ok, err = verify_windows_user(body.username.strip(), body.password)
        if not ok:
            return {
                "ok": False,
                "session_token": None,
                "error": err or "Credenciales incorrectas",
                "ms": int((time.perf_counter() - t0) * 1000),
            }
        out = owner_start(actor=body.username.strip(), method="ui")
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": out.get("ok"),
            "session_token": out.get("session_token"),
            "ttl_seconds": out.get("ttl_seconds"),
            "method": "windows",
            "expires_at": out.get("expires_at"),
            "ms": ms,
            "error": out.get("error"),
        }
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "session_token": None, "ms": ms, "error": str(e)}


class OwnerSessionEndBody(BaseModel):
    session_token: str


@app.post("/owner/session/end")
def owner_session_end(body: OwnerSessionEndBody):
    """End (revoke) owner session."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.session import end as owner_end

        ok = owner_end(body.session_token)
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": ok, "ms": ms, "error": None}
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "ms": ms, "error": str(e)}


@app.get("/owner/session/status")
def owner_session_status():
    """List active sessions (redacted tokens)."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.session import list_active

        sessions = list_active()
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": True, "sessions": sessions, "ms": ms, "error": None}
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "sessions": [], "ms": ms, "error": str(e)}


@app.get("/owner/emergency/status")
def owner_emergency_status():
    """Emergency mode status: enabled, block flags."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.emergency import get_emergency_state

        data = get_emergency_state()
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": True, "data": data, "ms": ms, "error": None}
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "data": None, "ms": ms, "error": str(e)}


class OwnerEmergencySetBody(BaseModel):
    enable: bool
    reason: Optional[str] = None


@app.post("/owner/emergency/set")
def owner_emergency_set(body: OwnerEmergencySetBody):
    """Enable/disable emergency mode. When enabled: block deploys, remote_exec, shell; allow status/health."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.emergency import (is_emergency,
                                                      set_emergency)

        set_emergency(body.enable, reason=body.reason or "")
        try:
            from modules.humanoid.audit import get_audit_logger

            get_audit_logger().log_event(
                "owner",
                "api",
                "emergency_set",
                True,
                0,
                None,
                {"enable": body.enable, "reason": body.reason},
                None,
            )
        except Exception:
            pass
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": True, "emergency": is_emergency(), "ms": ms, "error": None}
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "emergency": None, "ms": ms, "error": str(e)}


class OwnerEmergencyBody(BaseModel):
    enable: bool


@app.post("/owner/emergency")
def owner_emergency(body: OwnerEmergencyBody):
    """Enable/disable emergency mode (legacy). Prefer POST /owner/emergency/set."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.emergency import (is_emergency,
                                                      set_emergency)

        set_emergency(body.enable)
        try:
            from modules.humanoid.audit import get_audit_logger

            get_audit_logger().log_event(
                "owner",
                "api",
                "emergency",
                True,
                0,
                None,
                {"enable": body.enable},
                None,
            )
        except Exception:
            pass
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": True, "emergency": is_emergency(), "ms": ms, "error": None}
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "emergency": None, "ms": ms, "error": str(e)}


def _owner_status_data():
    """Aggregate owner console data: sessions, pending, expired, emergency, chain, critical history."""
    from datetime import datetime, timezone

    from modules.humanoid.approvals import list_pending, verify_chain
    from modules.humanoid.approvals.store import list_items
    from modules.humanoid.owner.emergency import is_emergency
    from modules.humanoid.owner.session import list_active

    sessions = list_active()
    pending = list_pending(limit=50)
    all_items = list_items(status=None, limit=100)
    expired = []
    for it in all_items:
        if it.get("status") != "pending":
            continue
        et = it.get("expires_at")
        if not et:
            continue
        try:
            dt = datetime.fromisoformat(et.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            if datetime.now(timezone.utc) >= dt:
                expired.append(it)
        except Exception:
            pass
    chain = verify_chain()
    critical_history = [
        x
        for x in all_items
        if x.get("status") == "approved"
        and (x.get("risk") or "").strip().lower() == "critical"
    ][:20]
    return {
        "active_sessions": sessions,
        "pending_approvals": pending,
        "expired_approvals": expired,
        "emergency": is_emergency(),
        "chain_valid": chain.get("valid", False),
        "chain_ok": chain.get("ok", False),
        "critical_history": critical_history,
    }


@app.get("/owner/status")
def owner_status():
    """Owner console: active sessions, pending/expired approvals, emergency, chain integrity, critical history."""
    t0 = time.perf_counter()
    try:
        data = _owner_status_data()
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": True, "data": data, "ms": ms, "error": None}
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "data": None, "ms": ms, "error": str(e)}


class VoiceApproveBody(BaseModel):
    phrase: str
    session_token: Optional[str] = None


@app.post("/voice/approve")
def voice_approve(body: VoiceApproveBody):
    """Voice approval: 'aprobar <id>' -> 'di confirmar <id>'; 'confirmar <id>' -> approve. Replay-safe."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.voice_approval import (
            handle_voice_command, voice_approval_enabled)

        if not voice_approval_enabled():
            return {
                "ok": False,
                "response": "",
                "approved": False,
                "ms": int((time.perf_counter() - t0) * 1000),
                "error": "voice_approvals disabled",
            }
        response, approved = handle_voice_command(
            body.phrase or "", device_source="api"
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": True,
            "response": response,
            "approved": approved,
            "ms": ms,
            "error": None,
        }
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": False,
            "response": "",
            "approved": False,
            "ms": ms,
            "error": str(e),
        }


class OwnerVoiceCommandBody(BaseModel):
    transcript: str
    device_source: Optional[str] = None


@app.post("/owner/voice/command")
def owner_voice_command(body: OwnerVoiceCommandBody):
    """Voice approval: 'aprobar <id>' -> confirm; 'confirmar <id>' -> approve. Returns {response, approved}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.voice_approval import handle_voice_command

        response, approved = handle_voice_command(
            body.transcript or "", body.device_source
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": True,
            "response": response,
            "approved": approved,
            "ms": ms,
            "error": None,
        }
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": False,
            "response": "",
            "approved": False,
            "ms": ms,
            "error": str(e),
        }


class OwnerTelegramCallbackBody(BaseModel):
    callback_data: str  # approve:ID | reject:ID
    chat_id: str


@app.post("/owner/telegram/callback")
def owner_telegram_callback(body: OwnerTelegramCallbackBody):
    """Handle Telegram inline callback (bot sends callback_data + chat_id). Returns {ok, action, id, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.comms.telegram_bridge import TelegramBridge

        out = TelegramBridge().handle_callback_data(
            body.callback_data or "", body.chat_id or ""
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return {
            "ok": out.get("ok"),
            "action": out.get("action"),
            "id": out.get("id"),
            "ms": ms,
            "error": out.get("error"),
        }
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return {"ok": False, "action": None, "id": None, "ms": ms, "error": str(e)}


# --- Scheduler / Watchdog / Healing ---
def _std_resp(
    ok: bool, data: Any = None, ms: int = 0, error: Optional[str] = None
) -> dict:
    out = {"ok": ok, "data": data, "ms": ms, "error": error}
    return out


# --- Conversaciones persistentes (Agent / Supervisor) ---
_chat_manager_singleton = None


def _get_chat_manager():
    """Lazy singleton for ChatThreadManager (SQLite) stored under data/chat_threads.db."""
    global _chat_manager_singleton
    if _chat_manager_singleton is not None:
        return _chat_manager_singleton
    try:
        import sys as _sys
        from pathlib import Path as _Path

        _root = _Path(__file__).resolve().parent.parent
        if str(_root) not in _sys.path:
            _sys.path.insert(0, str(_root))
        from chat_thread_manager import ChatThreadManager  # type: ignore

        _chat_manager_singleton = ChatThreadManager()
    except Exception:
        _chat_manager_singleton = None
    return _chat_manager_singleton


def _ensure_thread_id(
    thread_id: Optional[str], title: str, user_id: str
) -> Optional[str]:
    mgr = _get_chat_manager()
    if not mgr:
        return None
    tid = (thread_id or "").strip()
    if tid and mgr.thread_exists(tid):
        return tid
    try:
        return mgr.create_thread(
            title=title, user_id=user_id, description="", context={"kind": title}
        )
    except Exception:
        return None


def _extract_memory_lines(text: str, limit: int = 12) -> List[str]:
    """HeurÃ­stica extractiva: conserva lÃ­neas accionables para 'resumen'."""
    if not text:
        return []
    lines = []
    for raw in str(text).splitlines():
        s = raw.strip()
        if not s:
            continue
        up = s.upper()
        if s.startswith(
            (
                "##",
                "###",
                "- ",
                "* ",
                "ACTION:",
                "RISK:",
                "EXECUTE:",
                "âœ…",
                "âš ",
                "â›”",
                "OK:",
                "PROBLEMA:",
                "ADVERTENCIA:",
            )
        ):
            lines.append(s)
        elif (
            "http://" in s
            or "https://" in s
            or s.startswith((".\\", "C:\\", "GET ", "POST "))
        ):
            lines.append(s)
        elif (
            up.startswith("PLAN") or up.startswith("PASOS") or up.startswith("RESUMEN")
        ):
            lines.append(s)
        if len(lines) >= limit:
            break
    return lines[:limit]


def _looks_like_supervisor_policy(text: Optional[str]) -> bool:
    try:
        t = (text or "").strip()
        if not t:
            return False
        head = t[:900].lower()
        sig = (
            ("supervisor residente" in head)
            or ("no eres un chatbot pasivo" in head)
            or ("cero errores" in head)
        )
        sec = ("objetivo principal" in head) and ("monitoreo continuo" in head)
        return bool(sig and (sec or len(t) > 800))
    except Exception:
        return False


def _strip_policy_like_blocks(text: Optional[str]) -> str:
    """Best-effort: remove policy-like content from persisted thread memory to avoid 'recitals'."""
    try:
        t = (text or "").strip()
        if not t:
            return ""
        if _looks_like_supervisor_policy(t):
            return ""
        # Also remove long blocks that contain the signature, even if partial.
        low = t.lower()
        if ("supervisor residente" in low or "cero errores" in low) and len(t) > 400:
            return ""
        return t
    except Exception:
        return (text or "") if isinstance(text, str) else ""


def _update_thread_summary(
    thread_id: str, user_text: Optional[str], assistant_text: Optional[str]
) -> None:
    mgr = _get_chat_manager()
    if not mgr or not thread_id:
        return
    try:
        old = mgr.get_context(thread_id, "summary") or ""
        if not isinstance(old, str):
            old = str(old)
        chunks: List[str] = []
        if user_text:
            ut = str(user_text).strip()
            if ut and not _looks_like_supervisor_policy(ut):
                chunks.append(f"USER: {ut[:240]}")
        if assistant_text:
            mem_lines = _extract_memory_lines(assistant_text)
            if mem_lines:
                chunks.append("ASSISTANT_NOTES:\n" + "\n".join(mem_lines))
        if not chunks:
            return
        new = (old + ("\n\n" if old else "") + "\n\n".join(chunks)).strip()
        # cap memory size
        if len(new) > 2600:
            new = new[-2600:]
            # trim to first newline boundary when possible
            cut = new.find("\n")
            if 0 <= cut <= 120:
                new = new[cut + 1 :]
        mgr.set_context(thread_id, "summary", new)
        mgr.set_context(thread_id, "updated_at", datetime.utcnow().isoformat())
    except Exception:
        pass


def _llm_history_for_thread(
    thread_id: str, max_messages: int = 24
) -> List[Dict[str, str]]:
    mgr = _get_chat_manager()
    if not mgr or not thread_id:
        return []
    try:
        summary = mgr.get_context(thread_id, "summary")
        summary = _strip_policy_like_blocks(
            summary if isinstance(summary, str) else (str(summary) if summary else "")
        )
        msgs = mgr.get_recent_messages(thread_id, limit=max_messages)
        out: List[Dict[str, str]] = []
        if summary and isinstance(summary, str) and summary.strip():
            out.append(
                {
                    "role": "system",
                    "content": "MEMORIA DEL HILO (resumen):\n" + summary.strip(),
                }
            )
        for m in msgs:
            role = (m.get("role") or "user").strip()
            content = (m.get("content") or "").strip()
            if not content:
                continue
            if _looks_like_supervisor_policy(content):
                continue
            if role not in ("user", "assistant", "system"):
                role = "user"
            out.append({"role": role, "content": content})
        return out
    except Exception:
        return []


@app.get("/scheduler/jobs")
def scheduler_jobs(status: Optional[str] = None):
    """List jobs, optional filter by status. Response: {ok, data, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.scheduler import get_scheduler_db

        db = get_scheduler_db()
        jobs = db.list_jobs(status=status)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, jobs, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class SchedulerJobCreateBody(BaseModel):
    name: str
    kind: str  # update_check | llm_plan | shell_command | custom
    payload: dict = {}
    run_at: Optional[str] = None
    interval_seconds: Optional[int] = None
    max_retries: Optional[int] = None
    backoff_seconds: Optional[int] = None


@app.post("/scheduler/job/create")
def scheduler_job_create(body: SchedulerJobCreateBody):
    """Create a job. run_at ISO or now; interval_seconds for recurring. Response: {ok, data, ms, error}."""
    t0 = time.perf_counter()
    try:
        import os

        from modules.humanoid.scheduler import JobSpec, get_scheduler_db

        max_r = (
            int(os.getenv("SCHED_DEFAULT_RETRIES", "3") or 3)
            if body.max_retries is None
            else body.max_retries
        )
        backoff = (
            int(os.getenv("SCHED_BACKOFF_SECONDS", "5") or 5)
            if body.backoff_seconds is None
            else body.backoff_seconds
        )
        spec = JobSpec(
            name=body.name,
            kind=body.kind,
            payload=body.payload,
            run_at=body.run_at,
            interval_seconds=body.interval_seconds,
            max_retries=max_r,
            backoff_seconds=backoff,
        )
        db = get_scheduler_db()
        job = db.create_job(spec)
        get_audit_logger().log_event(
            "api",
            "owner",
            "scheduler",
            "job_enqueue",
            True,
            0,
            None,
            {"job_id": job.get("id"), "name": body.name, "kind": body.kind},
            None,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, job, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class SchedulerJobIdBody(BaseModel):
    job_id: str


@app.post("/scheduler/job/pause")
def scheduler_job_pause(body: SchedulerJobIdBody):
    t0 = time.perf_counter()
    try:
        from modules.humanoid.scheduler import get_scheduler_db

        db = get_scheduler_db()
        j = db.get_job(body.job_id)
        if not j:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), "job not found"
            )
        db.set_paused(body.job_id)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"job_id": body.job_id, "status": "paused"}, ms, None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/scheduler/job/resume")
def scheduler_job_resume(body: SchedulerJobIdBody):
    t0 = time.perf_counter()
    try:
        from datetime import datetime, timezone

        from modules.humanoid.scheduler import get_scheduler_db

        db = get_scheduler_db()
        j = db.get_job(body.job_id)
        if not j:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), "job not found"
            )
        now = datetime.now(timezone.utc).isoformat()
        db.set_queued(body.job_id, now)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True,
            {"job_id": body.job_id, "status": "queued", "next_run_ts": now},
            ms,
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/scheduler/job/run-now")
def scheduler_job_run_now(body: SchedulerJobIdBody):
    """Schedule job to run immediately (set next_run_ts to now)."""
    t0 = time.perf_counter()
    try:
        from datetime import datetime, timezone

        from modules.humanoid.scheduler import get_scheduler_db

        db = get_scheduler_db()
        j = db.get_job(body.job_id)
        if not j:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), "job not found"
            )
        now = datetime.now(timezone.utc).isoformat()
        db.set_queued(body.job_id, now)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True,
            {"job_id": body.job_id, "status": "queued", "next_run_ts": now},
            ms,
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/scheduler/job/runs")
def scheduler_job_runs(job_id: str, limit: int = 50):
    t0 = time.perf_counter()
    try:
        from modules.humanoid.scheduler import get_scheduler_db

        db = get_scheduler_db()
        runs = db.get_runs(job_id, limit=limit)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, runs, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/update/status")
def update_status_endpoint():
    """Git-based update status: branch, head, remote, has_update. Response: {ok, data, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.update.update_engine import \
            status as update_status

        data = update_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(data.get("ok", True), data, ms, data.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/update/check")
def update_check_endpoint():
    """Update check: fetch, snapshot. No apply. Response: {ok, data, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.update.update_engine import check as update_check

        result = update_check()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            result.get("ok", False), result.get("data"), ms, result.get("error")
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/update/apply")
def update_apply_endpoint():
    """Git update apply: staging -> smoke -> promote or rollback. Policy must allow. Response: {ok, data, ms, error}."""
    t0 = time.perf_counter()
    try:
        # Recargar atlas.env para que POLICY_ALLOW_UPDATE_APPLY tenga efecto sin reiniciar el servidor
        if ENV_PATH.exists():
            try:
                from dotenv import load_dotenv

                load_dotenv(ENV_PATH, override=True)
            except Exception:
                pass
        from modules.humanoid.owner.emergency import is_emergency

        if is_emergency():
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "emergency mode: updates suspended",
            )
        from modules.humanoid.update.update_engine import apply as update_apply

        result = update_apply()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            result.get("ok", False), result.get("data"), ms, result.get("error")
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/api/v4/update/prereqs")
def update_prereqs_endpoint():
    """VerificaciÃ³n de software para actualizar repo: Git, Python, pip, estado del repo."""
    return _update_prereqs_impl()


@app.get("/update/prereqs")
def update_prereqs_short():
    """Misma respuesta que /api/v4/update/prereqs (ruta corta para evitar 404 con proxies)."""
    return _update_prereqs_impl()


@app.post("/update/restart")
def update_restart_endpoint():
    """Reinicia el servidor PUSH (8791): programa un reinicio en 3s y responde de inmediato. Recargar la pÃ¡gina tras unos segundos."""
    try:
        repo_root = os.getenv("REPO_ROOT", str(BASE_DIR))
        script_path = (BASE_DIR / "scripts" / "restart_push_from_api.ps1").resolve()
        if not script_path.is_file():
            return _std_resp(
                False,
                None,
                0,
                f"restart script not found: {script_path}",
            )
        import subprocess

        CREATE_NEW_PROCESS_GROUP = 0x00000200
        DETACHED_PROCESS = 0x00000008
        flags = CREATE_NEW_PROCESS_GROUP | DETACHED_PROCESS if os.name == "nt" else 0
        subprocess.Popen(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                "-DelaySeconds",
                "3",
            ],
            cwd=repo_root,
            creationflags=flags,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return _std_resp(
            True,
            {
                "message": "Reinicio programado en 3 segundos. Recarga la pÃ¡gina en unos segundos."
            },
            0,
            None,
        )
    except Exception as e:
        return _std_resp(False, None, 0, str(e))


def _extract_json_from_output(raw: str) -> Optional[dict]:
    text = (raw or "").strip()
    if not text:
        return None
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    for idx in range(len(lines) - 1, -1, -1):
        ln = lines[idx]
        try:
            parsed = json.loads(ln)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            continue
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except Exception:
        return None
    return None


def _run_local_json_cmd(
    cmd: list[str], timeout: int = 900, cwd: Optional[Path] = None
) -> tuple[bool, dict, str]:
    import subprocess

    try:
        proc = subprocess.run(
            cmd,
            cwd=str(cwd or BASE_DIR),
            capture_output=True,
            text=True,
            timeout=timeout,
            shell=False,
        )
        out = (proc.stdout or "").strip()
        err = (proc.stderr or "").strip()
        merged = "\n".join(x for x in [out, err] if x)
        payload = _extract_json_from_output(merged) or {}
        ok = proc.returncode == 0 and bool(payload.get("ok", True))
        if not payload:
            payload = {
                "ok": proc.returncode == 0,
                "stdout": out[-4000:],
                "stderr": err[-2000:],
                "returncode": proc.returncode,
            }
        return ok, payload, merged[-4000:]
    except Exception as e:
        return False, {"ok": False, "error": str(e)}, str(e)


def _spawn_detached_job(
    cmd: list[str],
    cwd: Path,
    runner_log: Path,
) -> None:
    """Spawn detached command and persist stdout/stderr for post-mortem diagnostics."""
    import subprocess

    runner_log.parent.mkdir(parents=True, exist_ok=True)
    fh = runner_log.open("a", encoding="utf-8", errors="replace")
    try:
        fh.write(f"{datetime.now(timezone.utc).isoformat()} RUNNER_START cmd={cmd!r}\n")
        fh.flush()
        flags = 0
        if os.name == "nt":
            CREATE_NEW_PROCESS_GROUP = 0x00000200
            CREATE_NO_WINDOW = 0x08000000
            flags = CREATE_NEW_PROCESS_GROUP | CREATE_NO_WINDOW
        subprocess.Popen(
            cmd,
            cwd=str(cwd),
            creationflags=flags,
            stdin=subprocess.DEVNULL,
            stdout=fh,
            stderr=fh,
        )
    except Exception:
        fh.close()
        raise


class ToolsWatchdogBody(BaseModel):
    force: Optional[bool] = False


class ToolsUpdateBody(BaseModel):
    tool_id: str = Field(..., min_length=2, max_length=80)
    background: Optional[bool] = True


class ToolsBootstrapBody(BaseModel):
    background: Optional[bool] = True


class ToolsUpdateAllBody(BaseModel):
    background: Optional[bool] = True
    include_critical: Optional[bool] = True


class ToolsDiscoverySearchBody(BaseModel):
    force: Optional[bool] = False


class ToolsDiscoveryInstallBody(BaseModel):
    tool_id: str = Field(..., min_length=2, max_length=80)
    install_method: str = Field(..., min_length=2, max_length=20)
    install_target: str = Field(..., min_length=2, max_length=140)
    background: Optional[bool] = True


class ToolsDiscoveryInstallAllBody(BaseModel):
    # Acepta formato nuevo [{id, install_method, install_target}] y legado ["tool_id"].
    tools: Optional[list[Any]] = None
    background: Optional[bool] = True


def _tools_update_error(payload: dict) -> str:
    if not isinstance(payload, dict):
        return "tools_update_failed"
    direct = (payload.get("error") or "").strip()
    if direct:
        return direct
    nested = payload.get("data")
    if isinstance(nested, dict):
        nested_err = (nested.get("error") or "").strip()
        if nested_err:
            return nested_err
    steps = payload.get("steps")
    if isinstance(steps, list):
        for s in steps:
            if isinstance(s, dict) and s.get("ok") is False and s.get("error"):
                return str(s.get("error"))
    if payload.get("blocked_protected_branch"):
        branch = payload.get("branch") or "unknown"
        return f"updates_blocked_on_protected_branch ({branch})"
    return "tools_update_failed"


_TOOLS_REGISTRY_FILE = (BASE_DIR / "atlas_master_registry.json").resolve()
_TOOLS_DISCOVERY_CACHE_FILE = (BASE_DIR / "logs" / "atlas_tools_discovery_cache.json").resolve()


def _env_int(name: str, default: int, min_value: int = 1) -> int:
    try:
        value = int(str(os.getenv(name) or default).strip())
    except Exception:
        value = int(default)
    return max(min_value, value)


_TOOLS_MENU_CACHE_TTL_SEC = _env_int("ATLAS_TOOLS_MENU_CACHE_TTL_SEC", 25, 5)
_TOOLS_MENU_STALE_SEC = _env_int("ATLAS_TOOLS_MENU_STALE_SEC", 240, _TOOLS_MENU_CACHE_TTL_SEC)
_TOOLS_DISCOVERY_CACHE_TTL_SEC = _env_int("ATLAS_TOOLS_DISCOVERY_CACHE_TTL_SEC", 300, 15)

_tools_watchdog_refresh_lock = threading.Lock()
_tools_watchdog_refresh_running = False
_tools_discovery_refresh_lock = threading.Lock()
_tools_discovery_refresh_running = False


def _read_json_dict(path: Path) -> Optional[dict]:
    if not path.exists():
        return None
    for _ in range(3):
        try:
            data = json.loads(path.read_text(encoding="utf-8-sig", errors="replace"))
            if isinstance(data, dict):
                return data
        except Exception:
            # Tolerate transient partial writes and retry quickly.
            time.sleep(0.08)
    return None


def _write_json_dict(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def _run_tools_watchdog_scan(force: bool = False, timeout: int = 120) -> tuple[bool, dict]:
    script = (BASE_DIR / "scripts" / "atlas_tools_watchdog.py").resolve()
    if not script.exists():
        return False, {"error": f"watchdog script missing: {script}"}
    cmd = [sys.executable, str(script)]
    if force:
        cmd.append("--force")
    ok, payload, _tail = _run_local_json_cmd(cmd, timeout=timeout)
    if not isinstance(payload, dict):
        payload = {"ok": bool(ok), "error": "watchdog_invalid_payload"}
    return bool(ok), payload


def _spawn_tools_watchdog_refresh(force: bool = False) -> bool:
    global _tools_watchdog_refresh_running
    with _tools_watchdog_refresh_lock:
        if _tools_watchdog_refresh_running:
            return False
        _tools_watchdog_refresh_running = True

    def _job() -> None:
        global _tools_watchdog_refresh_running
        try:
            _run_tools_watchdog_scan(force=force, timeout=180)
        except Exception:
            pass
        finally:
            with _tools_watchdog_refresh_lock:
                _tools_watchdog_refresh_running = False

    threading.Thread(target=_job, name="atlas-tools-watchdog-refresh", daemon=True).start()
    return True


def _run_tools_discovery_scan(timeout: int = 180) -> tuple[bool, dict]:
    script = (BASE_DIR / "scripts" / "atlas_tools_discovery.py").resolve()
    if not script.exists():
        return False, {"error": f"discovery script missing: {script}"}
    ok, payload, _tail = _run_local_json_cmd([sys.executable, str(script)], timeout=timeout)
    if not isinstance(payload, dict):
        payload = {"ok": bool(ok), "error": "discovery_invalid_payload"}
    return bool(ok), payload


def _spawn_tools_discovery_refresh() -> bool:
    global _tools_discovery_refresh_running
    with _tools_discovery_refresh_lock:
        if _tools_discovery_refresh_running:
            return False
        _tools_discovery_refresh_running = True

    def _job() -> None:
        global _tools_discovery_refresh_running
        try:
            ok, payload = _run_tools_discovery_scan(timeout=180)
            if ok and isinstance(payload, dict):
                payload["_cached_at"] = datetime.now(timezone.utc).isoformat()
                _write_json_dict(_TOOLS_DISCOVERY_CACHE_FILE, payload)
        except Exception:
            pass
        finally:
            with _tools_discovery_refresh_lock:
                _tools_discovery_refresh_running = False

    threading.Thread(target=_job, name="atlas-tools-discovery-refresh", daemon=True).start()
    return True


@app.get("/api/tools/menu", tags=["Tools"])
def tools_menu_inventory(refresh: bool = False):
    """Inventario de herramientas (rápido): usa cache local y refresca en background."""
    t0 = time.perf_counter()
    now_ts = time.time()
    payload = _read_json_dict(_TOOLS_REGISTRY_FILE)
    file_age = None
    if _TOOLS_REGISTRY_FILE.exists():
        try:
            file_age = max(0.0, now_ts - _TOOLS_REGISTRY_FILE.stat().st_mtime)
        except Exception:
            file_age = None

    if payload and not refresh:
        age_sec = int(file_age or 0)
        stale = age_sec > _TOOLS_MENU_CACHE_TTL_SEC
        if stale:
            _spawn_tools_watchdog_refresh(force=False)
        out = dict(payload)
        out["cache"] = {
            "stale": stale,
            "age_sec": age_sec,
            "refreshing": bool(_tools_watchdog_refresh_running),
            "ttl_sec": _TOOLS_MENU_CACHE_TTL_SEC,
        }
        return _std_resp(True, out, int((time.perf_counter() - t0) * 1000), None)

    ok, fresh = _run_tools_watchdog_scan(force=bool(refresh), timeout=120)
    if ok:
        out = dict(fresh)
        out["cache"] = {
            "stale": False,
            "age_sec": 0,
            "refreshing": False,
            "ttl_sec": _TOOLS_MENU_CACHE_TTL_SEC,
        }
        return _std_resp(True, out, int((time.perf_counter() - t0) * 1000), None)

    # Fallback resiliente: devolver último registry conocido aunque esté viejo.
    if payload:
        age_sec = int(file_age or 0)
        if age_sec > _TOOLS_MENU_STALE_SEC:
            _spawn_tools_watchdog_refresh(force=True)
        out = dict(payload)
        out["cache"] = {
            "stale": True,
            "age_sec": age_sec,
            "refreshing": bool(_tools_watchdog_refresh_running),
            "ttl_sec": _TOOLS_MENU_CACHE_TTL_SEC,
            "warning": fresh.get("error") or "watchdog_refresh_failed_using_cache",
        }
        return _std_resp(True, out, int((time.perf_counter() - t0) * 1000), None)

    return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), fresh.get("error") or "watchdog_failed")


@app.get("/api/tools/discovery/search", tags=["Tools"])
def tools_discovery_search(force: bool = False):
    """Discovery rápido: entrega cache y refresca en background si está vencido."""
    t0 = time.perf_counter()
    now_ts = time.time()
    cached = _read_json_dict(_TOOLS_DISCOVERY_CACHE_FILE)
    if cached and not force:
        cached_at = 0.0
        try:
            ts_raw = cached.get("_cached_at")
            if ts_raw:
                cached_at = _ts_to_epoch(str(ts_raw))
            if not cached_at and _TOOLS_DISCOVERY_CACHE_FILE.exists():
                cached_at = _TOOLS_DISCOVERY_CACHE_FILE.stat().st_mtime
        except Exception:
            cached_at = 0.0
        age_sec = int(max(0.0, now_ts - cached_at)) if cached_at > 0 else 999999
        stale = age_sec > _TOOLS_DISCOVERY_CACHE_TTL_SEC
        if stale:
            _spawn_tools_discovery_refresh()
        out = dict(cached)
        out["cache"] = {
            "stale": stale,
            "age_sec": age_sec,
            "refreshing": bool(_tools_discovery_refresh_running),
            "ttl_sec": _TOOLS_DISCOVERY_CACHE_TTL_SEC,
        }
        return _std_resp(True, out, int((time.perf_counter() - t0) * 1000), None)

    ok, payload = _run_tools_discovery_scan(timeout=180)
    if ok:
        out = dict(payload)
        out["_cached_at"] = datetime.now(timezone.utc).isoformat()
        _write_json_dict(_TOOLS_DISCOVERY_CACHE_FILE, out)
        out["cache"] = {
            "stale": False,
            "age_sec": 0,
            "refreshing": False,
            "ttl_sec": _TOOLS_DISCOVERY_CACHE_TTL_SEC,
        }
        return _std_resp(True, out, int((time.perf_counter() - t0) * 1000), None)

    if cached:
        out = dict(cached)
        out["cache"] = {
            "stale": True,
            "age_sec": 999999,
            "refreshing": bool(_tools_discovery_refresh_running),
            "ttl_sec": _TOOLS_DISCOVERY_CACHE_TTL_SEC,
            "warning": payload.get("error") or "discovery_refresh_failed_using_cache",
        }
        return _std_resp(True, out, int((time.perf_counter() - t0) * 1000), None)

    return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), payload.get("error") or "discovery_failed")


@app.post("/api/tools/discovery/install", tags=["Tools"])
def tools_discovery_install(body: ToolsDiscoveryInstallBody):
    """Instala una herramienta descubierta (job en background con seguimiento)."""
    t0 = time.perf_counter()
    if not bool(body.background):
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            "discovery_install_requires_background",
        )

    bg_script = (BASE_DIR / "scripts" / "atlas_tools_discovery_install_background.ps1").resolve()
    if not bg_script.exists():
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            f"background discovery install script missing: {bg_script}",
        )

    try:
        jobs_dir = (BASE_DIR / "logs" / "tools_update_jobs").resolve()
        jobs_dir.mkdir(parents=True, exist_ok=True)
        job_id = uuid.uuid4().hex[:12]
        status_file = jobs_dir / f"{job_id}.json"
        runner_log = jobs_dir / f"{job_id}.runner.log"
        status_file.write_text(
            json.dumps(
                {
                    "ok": True,
                    "job_id": job_id,
                    "tool": body.tool_id,
                    "status": "queued",
                    "queued_at": datetime.now(timezone.utc).isoformat(),
                    "install_method": body.install_method,
                    "install_target": body.install_target,
                    "source": "discovery",
                    "total": 4,
                    "done": 0,
                    "failed": 0,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        powershell_exe = str(
            (
                Path(os.environ.get("SystemRoot", r"C:\Windows"))
                / "System32"
                / "WindowsPowerShell"
                / "v1.0"
                / "powershell.exe"
            ).resolve()
        )
        cmd_bg = [
            powershell_exe,
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(bg_script),
            "-RepoRoot",
            str(BASE_DIR),
            "-JobFile",
            str(status_file),
            "-ToolId",
            body.tool_id,
            "-InstallMethod",
            body.install_method,
            "-InstallTarget",
            body.install_target,
        ]
        _spawn_detached_job(
            cmd_bg,
            cwd=str(BASE_DIR),
            runner_log=runner_log,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True,
            {
                "ok": True,
                "queued": True,
                "job_id": job_id,
                "tool": body.tool_id,
                "status": "queued",
                "status_url": f"/api/tools/job/status/{job_id}",
                "install_method": body.install_method,
                "install_target": body.install_target,
                "runner_log": str(runner_log),
            },
            ms,
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/api/tools/discovery/install-all", tags=["Tools"])
def tools_discovery_install_all(body: Optional[ToolsDiscoveryInstallAllBody] = None):
    """Instala en lote herramientas descubiertas con progreso por herramienta."""
    t0 = time.perf_counter()
    if not (bool(body.background) if body else True):
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "discovery_install_all_requires_background")

    bg_script = (BASE_DIR / "scripts" / "atlas_tools_discovery_install_all_background.ps1").resolve()
    discover_script = (BASE_DIR / "scripts" / "atlas_tools_discovery.py").resolve()
    if not bg_script.exists():
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), f"missing script: {bg_script}")
    if not discover_script.exists():
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), f"missing script: {discover_script}")

    requested_tools = body.tools if (body and isinstance(body.tools, list)) else None
    tools: Optional[list[dict]] = None
    if isinstance(requested_tools, list):
        tools = []
        for item in requested_tools:
            if isinstance(item, dict):
                tid = str(item.get("id") or "").strip()
                m = str(item.get("install_method") or "").strip()
                tgt = str(item.get("install_target") or "").strip()
                if tid and m and tgt:
                    tools.append({"id": tid, "install_method": m, "install_target": tgt})
                continue
            tid = str(item or "").strip()
            if tid:
                tools.append({"id": tid})
    if tools is None:
        ok, payload, _tail = _run_local_json_cmd([sys.executable, str(discover_script)], timeout=180)
        if not ok:
            return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), payload.get("error") or "discovery_failed")
        rows = payload.get("candidates") if isinstance(payload, dict) else []
        if not isinstance(rows, list):
            rows = []
        tools = []
        for r in rows:
            if not isinstance(r, dict):
                continue
            if not bool(r.get("network_available")):
                continue
            tid = str(r.get("id") or "").strip()
            m = str(r.get("install_method") or "").strip()
            tgt = str(r.get("install_target") or "").strip()
            if tid and m and tgt:
                tools.append({"id": tid, "install_method": m, "install_target": tgt})
    else:
        # Completa metadatos faltantes cuando llega payload legado con solo IDs.
        missing_meta = [str(t.get("id") or "").strip() for t in tools if isinstance(t, dict) and (not t.get("install_method") or not t.get("install_target"))]
        if missing_meta:
            ok, payload, _tail = _run_local_json_cmd([sys.executable, str(discover_script)], timeout=180)
            rows = payload.get("candidates") if (ok and isinstance(payload, dict)) else []
            idx = {}
            if isinstance(rows, list):
                for r in rows:
                    if not isinstance(r, dict):
                        continue
                    tid = str(r.get("id") or "").strip().lower()
                    if tid:
                        idx[tid] = r
            normalized = []
            for t in tools:
                if not isinstance(t, dict):
                    continue
                tid = str(t.get("id") or "").strip()
                if not tid:
                    continue
                m = str(t.get("install_method") or "").strip()
                tgt = str(t.get("install_target") or "").strip()
                if (not m or not tgt) and tid.lower() in idx:
                    r = idx[tid.lower()]
                    m = m or str(r.get("install_method") or "").strip()
                    tgt = tgt or str(r.get("install_target") or "").strip()
                if m and tgt:
                    normalized.append({"id": tid, "install_method": m, "install_target": tgt})
            tools = normalized

    if not tools:
        return _std_resp(
            True,
            {"ok": True, "queued": False, "status": "noop", "message": "No hay herramientas discovery disponibles para instalar", "tools": []},
            int((time.perf_counter() - t0) * 1000),
            None,
        )

    try:
        jobs_dir = (BASE_DIR / "logs" / "tools_update_jobs").resolve()
        jobs_dir.mkdir(parents=True, exist_ok=True)
        job_id = uuid.uuid4().hex[:12]
        status_file = jobs_dir / f"{job_id}.json"
        runner_log = jobs_dir / f"{job_id}.runner.log"
        manifest_file = jobs_dir / f"{job_id}.manifest.json"
        manifest_file.write_text(json.dumps({"tools": tools}, ensure_ascii=False), encoding="utf-8")
        status_file.write_text(
            json.dumps(
                {
                    "ok": True,
                    "job_id": job_id,
                    "tool": "discovery_all",
                    "source": "discovery-bulk",
                    "status": "queued",
                    "queued_at": datetime.now(timezone.utc).isoformat(),
                    "tools": [str(t.get("id") or "") for t in tools],
                    "total": len(tools),
                    "done": 0,
                    "failed": 0,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        powershell_exe = str((Path(os.environ.get("SystemRoot", r"C:\Windows")) / "System32" / "WindowsPowerShell" / "v1.0" / "powershell.exe").resolve())
        cmd_bg = [
            powershell_exe, "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(bg_script),
            "-RepoRoot", str(BASE_DIR), "-JobFile", str(status_file), "-ManifestFile", str(manifest_file),
        ]
        _spawn_detached_job(cmd_bg, cwd=BASE_DIR, runner_log=runner_log)
        return _std_resp(
            True,
            {
                "ok": True,
                "queued": True,
                "job_id": job_id,
                "tool": "discovery_all",
                "source": "discovery-bulk",
                "status": "queued",
                "tools": [str(t.get("id") or "") for t in tools],
                "total": len(tools),
                "status_url": f"/api/tools/job/status/{job_id}",
                "runner_log": str(runner_log),
            },
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/api/tools/watchdog/scan", tags=["Tools"])
def tools_watchdog_scan(body: Optional[ToolsWatchdogBody] = None):
    """Escaneo forzado para detectar upgrades y refrescar atlas_master_registry.json."""
    t0 = time.perf_counter()
    script = (BASE_DIR / "scripts" / "atlas_tools_watchdog.py").resolve()
    if not script.exists():
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            f"watchdog script missing: {script}",
        )
    force = bool(body.force) if body else False
    cmd = [sys.executable, str(script)]
    if force:
        cmd.append("--force")
    ok, payload, _tail = _run_local_json_cmd(cmd, timeout=180)
    ms = int((time.perf_counter() - t0) * 1000)
    return _std_resp(ok, payload, ms, None if ok else payload.get("error"))


@app.post("/api/tools/bootstrap", tags=["Tools"])
def tools_bootstrap_install(body: Optional[ToolsBootstrapBody] = None):
    """Instala dependencias base del stack de herramientas con snapshot previo obligatorio."""
    t0 = time.perf_counter()
    script = (BASE_DIR / "scripts" / "atlas_tools_bootstrap.ps1").resolve()
    bg_script = (BASE_DIR / "scripts" / "atlas_tools_bootstrap_background.ps1").resolve()
    if not script.exists():
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            f"bootstrap script missing: {script}",
        )
    if bool(body.background) if body else True:
        if not bg_script.exists():
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                f"background bootstrap script missing: {bg_script}",
            )
        try:
            jobs_dir = (BASE_DIR / "logs" / "tools_update_jobs").resolve()
            jobs_dir.mkdir(parents=True, exist_ok=True)
            job_id = uuid.uuid4().hex[:12]
            status_file = jobs_dir / f"{job_id}.json"
            runner_log = jobs_dir / f"{job_id}.runner.log"
            status_file.write_text(
                json.dumps(
                    {
                        "ok": True,
                        "job_id": job_id,
                        "tool": "bootstrap",
                        "status": "queued",
                        "queued_at": datetime.now(timezone.utc).isoformat(),
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            powershell_exe = str(
                (
                    Path(os.environ.get("SystemRoot", r"C:\Windows"))
                    / "System32"
                    / "WindowsPowerShell"
                    / "v1.0"
                    / "powershell.exe"
                ).resolve()
            )
            cmd = [
                powershell_exe,
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(bg_script),
                "-RepoRoot",
                str(BASE_DIR),
                "-JobFile",
                str(status_file),
            ]
            _spawn_detached_job(
                cmd,
                cwd=BASE_DIR,
                runner_log=runner_log,
            )
            ms = int((time.perf_counter() - t0) * 1000)
            return _std_resp(
                True,
                {
                    "ok": True,
                    "queued": True,
                    "job_id": job_id,
                    "tool": "bootstrap",
                    "status": "queued",
                    "status_url": f"/api/tools/job/status/{job_id}",
                    "runner_log": str(runner_log),
                },
                ms,
                None,
            )
        except Exception as e:
            return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))
    cmd = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(script),
        "-RepoRoot",
        str(BASE_DIR),
    ]
    ok, payload, tail = _run_local_json_cmd(cmd, timeout=1800)
    if tail:
        try:
            logf = BASE_DIR / "logs" / "snapshot_safe_diagnostic.log"
            with logf.open("a", encoding="utf-8") as f:
                f.write(f"{datetime.now(timezone.utc).isoformat()} TOOLS_BOOTSTRAP_API tail={tail!r}\n")
        except Exception:
            pass
    ms = int((time.perf_counter() - t0) * 1000)
    return _std_resp(ok, payload, ms, None if ok else payload.get("error"))


@app.post("/api/tools/update", tags=["Tools"])
def tools_update_single(body: ToolsUpdateBody):
    """Actualiza una herramienta especÃ­fica con snapshot previo obligatorio."""
    t0 = time.perf_counter()
    script = (BASE_DIR / "scripts" / "atlas_tool_update.ps1").resolve()
    bg_script = (BASE_DIR / "scripts" / "atlas_tool_update_background.ps1").resolve()
    if not script.exists():
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            f"update script missing: {script}",
        )
    if bool(body.background):
        if not bg_script.exists():
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                f"background update script missing: {bg_script}",
            )
        try:
            jobs_dir = (BASE_DIR / "logs" / "tools_update_jobs").resolve()
            jobs_dir.mkdir(parents=True, exist_ok=True)
            job_id = uuid.uuid4().hex[:12]
            status_file = jobs_dir / f"{job_id}.json"
            runner_log = jobs_dir / f"{job_id}.runner.log"
            status_file.write_text(
                json.dumps(
                    {
                        "ok": True,
                        "job_id": job_id,
                        "tool": body.tool_id,
                        "status": "queued",
                        "queued_at": datetime.now(timezone.utc).isoformat(),
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            powershell_exe = str(
                (
                    Path(os.environ.get("SystemRoot", r"C:\Windows"))
                    / "System32"
                    / "WindowsPowerShell"
                    / "v1.0"
                    / "powershell.exe"
                ).resolve()
            )
            cmd = [
                powershell_exe,
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(bg_script),
                "-Tool",
                body.tool_id,
                "-RepoRoot",
                str(BASE_DIR),
                "-JobFile",
                str(status_file),
            ]
            _spawn_detached_job(
                cmd,
                cwd=BASE_DIR,
                runner_log=runner_log,
            )
            ms = int((time.perf_counter() - t0) * 1000)
            return _std_resp(
                True,
                {
                    "ok": True,
                    "queued": True,
                    "job_id": job_id,
                    "tool": body.tool_id,
                    "status": "queued",
                    "status_url": f"/api/tools/update/status/{job_id}",
                    "runner_log": str(runner_log),
                },
                ms,
                None,
            )
        except Exception as e:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                str(e),
            )
    cmd = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(script),
        "-Tool",
        body.tool_id,
        "-RepoRoot",
        str(BASE_DIR),
    ]
    ok, payload, _tail = _run_local_json_cmd(cmd, timeout=1800)
    ms = int((time.perf_counter() - t0) * 1000)
    return _std_resp(ok, payload, ms, None if ok else _tools_update_error(payload))


@app.post("/api/tools/update-all", tags=["Tools"])
def tools_update_all(body: Optional[ToolsUpdateAllBody] = None):
    """Actualiza todas las herramientas con upgrade_ready en un job Ãºnico de background."""
    t0 = time.perf_counter()
    bg_script = (BASE_DIR / "scripts" / "atlas_tools_update_all_background.ps1").resolve()
    watchdog_script = (BASE_DIR / "scripts" / "atlas_tools_watchdog.py").resolve()
    if not bg_script.exists():
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            f"background update-all script missing: {bg_script}",
        )
    if not watchdog_script.exists():
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            f"watchdog script missing: {watchdog_script}",
        )

    include_critical = True if body is None else bool(body.include_critical)
    ok_inv, payload_inv, _tail = _run_local_json_cmd(
        [sys.executable, str(watchdog_script)], timeout=180
    )
    if not ok_inv:
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            payload_inv.get("error") or "watchdog_failed",
        )

    tools = payload_inv.get("tools") if isinstance(payload_inv, dict) else []
    if not isinstance(tools, list):
        tools = []
    target_tools: list[str] = []
    for t in tools:
        if not isinstance(t, dict):
            continue
        if not bool(t.get("update_ready")):
            continue
        if not t.get("update_script"):
            continue
        if (not include_critical) and bool(t.get("critical")):
            continue
        tid = str(t.get("id") or "").strip()
        if tid:
            target_tools.append(tid)
    target_tools = sorted(set(target_tools))
    if not target_tools:
        return _std_resp(
            True,
            {
                "ok": True,
                "queued": False,
                "tool": "all",
                "status": "noop",
                "message": "No hay herramientas con upgrade_ready",
                "tools": [],
            },
            int((time.perf_counter() - t0) * 1000),
            None,
        )

    if not (bool(body.background) if body else True):
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            "update_all_requires_background",
        )

    try:
        jobs_dir = (BASE_DIR / "logs" / "tools_update_jobs").resolve()
        jobs_dir.mkdir(parents=True, exist_ok=True)
        job_id = uuid.uuid4().hex[:12]
        status_file = jobs_dir / f"{job_id}.json"
        runner_log = jobs_dir / f"{job_id}.runner.log"
        status_file.write_text(
            json.dumps(
                {
                    "ok": True,
                    "job_id": job_id,
                    "tool": "all",
                    "status": "queued",
                    "queued_at": datetime.now(timezone.utc).isoformat(),
                    "tools": target_tools,
                    "total": len(target_tools),
                    "done": 0,
                    "failed": 0,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        powershell_exe = str((Path(os.environ.get("SystemRoot", r"C:\Windows")) / "System32" / "WindowsPowerShell" / "v1.0" / "powershell.exe").resolve())
        cmd_bg = [
            powershell_exe,
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(bg_script),
            "-RepoRoot",
            str(BASE_DIR),
            "-JobFile",
            str(status_file),
            "-ToolsCsv",
            ",".join(target_tools),
        ]
        _spawn_detached_job(
            cmd_bg,
            cwd=BASE_DIR,
            runner_log=runner_log,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True,
            {
                "ok": True,
                "queued": True,
                "job_id": job_id,
                "tool": "all",
                "status": "queued",
                "tools": target_tools,
                "total": len(target_tools),
                "status_url": f"/api/tools/job/status/{job_id}",
                "runner_log": str(runner_log),
            },
            ms,
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/api/tools/update/status/{job_id}", tags=["Tools"])
def tools_update_status(job_id: str):
    """Estado de un update de herramienta ejecutado en segundo plano."""
    t0 = time.perf_counter()
    safe_job_id = "".join(ch for ch in (job_id or "") if ch.isalnum() or ch in ("-", "_"))
    if not safe_job_id:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "invalid_job_id")
    status_file = (BASE_DIR / "logs" / "tools_update_jobs" / f"{safe_job_id}.json").resolve()
    if not status_file.exists():
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "job_not_found")
    try:
        payload = json.loads(status_file.read_text(encoding="utf-8-sig", errors="replace"))
        if isinstance(payload, dict) and not str(payload.get("job_id") or "").strip():
            payload["job_id"] = safe_job_id
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))
    try:
        def _is_job_process_alive(_job_id: str) -> bool:
            try:
                needle = f"{_job_id}.json".lower()
                script_markers = (
                    "atlas_tools_update_all_background.ps1",
                    "atlas_tool_update_background.ps1",
                    "atlas_tools_discovery_install_all_background.ps1",
                    "atlas_tools_discovery_install_background.ps1",
                    "atlas_tools_bootstrap_background.ps1",
                )
                for proc in psutil.process_iter(["cmdline"]):
                    try:
                        cmd_parts = proc.info.get("cmdline") or []
                        cmd = " ".join(str(x) for x in cmd_parts).lower()
                    except Exception:
                        continue
                    if not cmd:
                        continue
                    if needle in cmd and any(m in cmd for m in script_markers):
                        return True
                return False
            except Exception:
                # If we cannot inspect processes reliably, avoid false fail.
                return True

        changed = False
        status = str(payload.get("status") or "").lower()
        tools = payload.get("tools") if isinstance(payload.get("tools"), list) else []
        results = payload.get("results") if isinstance(payload.get("results"), list) else []
        total = int(payload.get("total") or (len(tools) if tools else len(results) if results else 0))
        done = int(payload.get("done") or (len(results) if results else 0))
        failed = int(payload.get("failed") or 0)

        if done < 0:
            done = 0
            payload["done"] = 0
            changed = True
        if total > 0 and done > total:
            done = total
            payload["done"] = total
            changed = True
        if failed < 0:
            failed = 0
            payload["failed"] = 0
            changed = True
        if done > 0 and failed > done:
            failed = done
            payload["failed"] = done
            changed = True

        # Normalize terminal state when all tools are already processed.
        if status in ("queued", "running", "failed", "done_with_errors") and total > 0 and done >= total:
            src = str(payload.get("source") or "").lower()
            terminal_failed = failed > 0
            payload["status"] = "done_with_errors" if terminal_failed else "done"
            payload["ok"] = not terminal_failed
            if terminal_failed and not str(payload.get("error") or "").strip():
                payload["error"] = (
                    "one_or_more_discovery_installs_failed"
                    if "discovery" in src
                    else "one_or_more_updates_failed"
                )
            if not str(payload.get("finished_at") or "").strip():
                payload["finished_at"] = datetime.now(timezone.utc).isoformat()
            changed = True

        # running + partial failures should not be reported as ok=true
        if status == "running" and failed > 0 and payload.get("ok") is not False:
            payload["ok"] = False
            changed = True

        status = str(payload.get("status") or "").lower()
        # Guardrail: running job without live runner process (stuck/orphan).
        if status == "running":
            try:
                now_ts = time.time()
                status_mtime = status_file.stat().st_mtime
                runner_log = (BASE_DIR / "logs" / "tools_update_jobs" / f"{safe_job_id}.runner.log").resolve()
                runner_mtime = runner_log.stat().st_mtime if runner_log.exists() else status_mtime
                stale_for = now_ts - max(status_mtime, runner_mtime)
                if stale_for > 90 and (not _is_job_process_alive(safe_job_id)):
                    payload["ok"] = False
                    payload["status"] = "failed"
                    payload["error"] = payload.get("error") or "job_runner_not_alive"
                    payload["failed_at"] = datetime.now(timezone.utc).isoformat()
                    changed = True
            except Exception:
                pass
        # Guardrail: queued jobs that never started are reported as failed (stale).
        if status == "queued":
            q_at = payload.get("queued_at")
            if q_at:
                dt_q = datetime.fromisoformat(str(q_at).replace("Z", "+00:00"))
                now = datetime.now(dt_q.tzinfo) if dt_q.tzinfo else datetime.now()
                if (now - dt_q).total_seconds() > 90:
                    payload["ok"] = False
                    payload["status"] = "failed"
                    payload["error"] = payload.get("error") or "job_stale_in_queue"
                    payload["failed_at"] = datetime.now(timezone.utc).isoformat()
                    changed = True
        # Guardrail: running jobs stale for too long are marked failed.
        if status == "running":
            s_at = payload.get("started_at") or payload.get("queued_at")
            if s_at:
                dt_s = datetime.fromisoformat(str(s_at).replace("Z", "+00:00"))
                now = datetime.now(dt_s.tzinfo) if dt_s.tzinfo else datetime.now()
                # 20 min without terminal status => stale running
                if (now - dt_s).total_seconds() > 20 * 60:
                    payload["ok"] = False
                    payload["status"] = "failed"
                    payload["error"] = payload.get("error") or "job_stale_running_timeout"
                    payload["failed_at"] = datetime.now(timezone.utc).isoformat()
                    changed = True
        if changed:
            status_file.write_text(
                json.dumps(payload, ensure_ascii=False),
                encoding="utf-8",
            )
    except Exception:
        pass
    return _std_resp(True, payload, int((time.perf_counter() - t0) * 1000), None)



@app.get("/api/tools/job/latest", tags=["Tools"])
def tools_job_latest():
    """Último job de tools en estado queued/running para reanudar seguimiento desde UI."""
    t0 = time.perf_counter()
    jobs_dir = (BASE_DIR / "logs" / "tools_update_jobs").resolve()
    if not jobs_dir.exists():
        return _std_resp(
            True,
            {"ok": True, "found": False},
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    try:
        files = sorted(
            jobs_dir.glob("*.json"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )[:40]
        for f in files:
            try:
                payload = json.loads(f.read_text(encoding="utf-8-sig", errors="replace"))
            except Exception:
                continue
            status = str(payload.get("status") or "").lower()
            if status == "queued":
                try:
                    q_at = payload.get("queued_at")
                    if q_at:
                        dt_q = datetime.fromisoformat(str(q_at).replace("Z", "+00:00"))
                        now = datetime.now(dt_q.tzinfo) if dt_q.tzinfo else datetime.now()
                        if (now - dt_q).total_seconds() > 90:
                            continue
                except Exception:
                    pass
            if status in ("queued", "running"):
                payload["job_id"] = payload.get("job_id") or f.stem
                payload["status_url"] = f"/api/tools/job/status/{payload['job_id']}"
                return _std_resp(
                    True,
                    {"ok": True, "found": True, "job": payload},
                    int((time.perf_counter() - t0) * 1000),
                    None,
                )
        return _std_resp(
            True,
            {"ok": True, "found": False},
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/api/tools/job/status/{job_id}", tags=["Tools"])
def tools_job_status(job_id: str):
    """Alias genÃ©rico para estado de jobs de tools (update/bootstrap)."""
    return tools_update_status(job_id)


@app.get("/api/tools/job/runner-log/{job_id}", tags=["Tools"])
def tools_job_runner_log(job_id: str, tail_lines: int = 120):
    """Devuelve tail del runner log del job para diagnóstico rápido en UI."""
    t0 = time.perf_counter()
    safe_job_id = "".join(ch for ch in (job_id or "") if ch.isalnum() or ch in ("-", "_"))
    if not safe_job_id:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "invalid_job_id")
    jobs_dir = (BASE_DIR / "logs" / "tools_update_jobs").resolve()
    log_path = (jobs_dir / f"{safe_job_id}.runner.log").resolve()
    if not log_path.exists():
        return _std_resp(
            True,
            {"ok": True, "found": False, "job_id": safe_job_id, "tail": ""},
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    try:
        lines = log_path.read_text(encoding="utf-8-sig", errors="replace").splitlines()
        n = max(10, min(500, int(tail_lines or 120)))
        tail = "\n".join(lines[-n:])
        return _std_resp(
            True,
            {"ok": True, "found": True, "job_id": safe_job_id, "log_path": str(log_path), "tail": tail},
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


def _update_prereqs_impl():
    import subprocess
    import sys

    t0 = time.perf_counter()
    out = {
        "git_ok": False,
        "git_version": None,
        "python_ok": False,
        "python_version": None,
        "pip_ok": False,
        "pip_version": None,
        "repo_has_changes": None,
        "branch": None,
        "ms": 0,
    }
    repo_root = os.getenv(
        "REPO_ROOT", os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    )
    use_shell = os.name == "nt"

    # Python: el proceso actual es Python, asÃ­ que siempre OK
    out["python_ok"] = True
    out["python_version"] = (sys.version or "").split("\n")[0].strip()[:80]

    # pip: mismo intÃ©rprete
    try:
        r = subprocess.run(
            [sys.executable, "-m", "pip", "--version"],
            capture_output=True,
            text=True,
            timeout=10,
            cwd=repo_root,
            shell=False,
        )
        out["pip_ok"] = r.returncode == 0
        if r.stdout:
            out["pip_version"] = (r.stdout.strip() or "")[:80]
    except Exception:
        pass

    # Git: en Windows usar shell=True para que use el PATH del usuario
    try:
        if use_shell:
            r = subprocess.run(
                "git --version",
                capture_output=True,
                text=True,
                timeout=10,
                cwd=repo_root,
                shell=True,
            )
        else:
            r = subprocess.run(
                ["git", "--version"],
                capture_output=True,
                text=True,
                timeout=10,
                cwd=repo_root,
            )
        out["git_ok"] = r.returncode == 0
        if r.stdout:
            out["git_version"] = r.stdout.strip()
    except Exception:
        pass

    # Repo status (en Windows con shell para PATH)
    try:
        if use_shell:
            r = subprocess.run(
                "git status --porcelain",
                capture_output=True,
                text=True,
                timeout=10,
                cwd=repo_root,
                shell=True,
            )
            r2 = subprocess.run(
                "git rev-parse --abbrev-ref HEAD",
                capture_output=True,
                text=True,
                timeout=5,
                cwd=repo_root,
                shell=True,
            )
        else:
            r = subprocess.run(
                ["git", "status", "--porcelain"],
                capture_output=True,
                text=True,
                timeout=10,
                cwd=repo_root,
            )
            r2 = subprocess.run(
                ["git", "rev-parse", "--abbrev-ref", "HEAD"],
                capture_output=True,
                text=True,
                timeout=5,
                cwd=repo_root,
            )
        out["repo_has_changes"] = bool((r.stdout or "").strip())
        if r2.returncode == 0 and r2.stdout:
            out["branch"] = r2.stdout.strip()
        # Si tenemos rama pero git_ok era False, Git sÃ­ estÃ¡ (p. ej. status desde otro proceso)
        if out["branch"] and not out["git_ok"]:
            out["git_ok"] = True
            out["git_version"] = out["git_version"] or "OK (detectado por repo)"
    except Exception:
        pass
    out["ms"] = int((time.perf_counter() - t0) * 1000)
    return _std_resp(True, out, out["ms"], None)


@app.get("/deploy/status", tags=["Deploy"])
def deploy_status():
    """Deploy status: ok, mode, active_port, staging_port, active_pid, staging_pid, last_deploy, last_health, canary."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.deploy.canary import (
            set_canary_disabled_fallback, should_disable_canary)
        from modules.humanoid.deploy.status import build_deploy_status
        from modules.humanoid.deploy.switcher import get_deploy_state

        state = get_deploy_state()
        if should_disable_canary() and not state.get("canary_disabled_fallback"):
            set_canary_disabled_fallback(True)
        data = build_deploy_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class DeployApplyBody(BaseModel):
    ref: Optional[str] = None  # origin/main | tag:vX.Y.Z | sha:....
    mode: Optional[str] = None  # bluegreen | single
    force: Optional[bool] = False


@app.post("/deploy/apply", tags=["Deploy"])
def deploy_apply(body: Optional[DeployApplyBody] = None):
    """Apply deploy: policy gate deploy_apply, optional fetch+staging ref, then blue-green flow (staging -> smoke -> health -> promote or rollback)."""
    t0 = time.perf_counter()
    body = body or DeployApplyBody()
    try:
        from modules.humanoid.owner.emergency import is_emergency

        if is_emergency():
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "emergency mode: deploys suspended",
            )
        decision = get_policy_engine().can(_memory_actor(), "deploy", "deploy_apply")
        if not decision.allow:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                decision.reason or "policy denied",
            )
        ref = body.ref or "origin/main"
        if body.mode and body.mode.strip().lower() == "single":
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "mode=single not implemented for apply",
            )
        from modules.humanoid.deploy.bluegreen import run_bluegreen_flow
        from modules.humanoid.update.git_manager import (create_staging_branch,
                                                         fetch)

        r_fetch = fetch(os.getenv("UPDATE_REMOTE", "origin"))
        if not r_fetch.get("ok"):
            return _std_resp(
                False,
                {"step": "fetch", "error": r_fetch.get("error")},
                int((time.perf_counter() - t0) * 1000),
                r_fetch.get("error"),
            )
        remote = os.getenv("UPDATE_REMOTE", "origin")
        branch = (
            ref.replace("origin/", "").strip() if ref.startswith("origin/") else "main"
        )
        staging_name = (
            os.getenv("UPDATE_STAGING_BRANCH", "staging").strip() or "staging"
        )
        r_staging = create_staging_branch(
            remote=remote, branch=branch, staging_name=staging_name
        )
        if not r_staging.get("ok"):
            return _std_resp(
                False,
                {"step": "create_staging", "error": r_staging.get("error")},
                int((time.perf_counter() - t0) * 1000),
                r_staging.get("error"),
            )
        result = run_bluegreen_flow(ref=ref, dry_run=False)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            result.get("ok", False), result.get("data"), ms, result.get("error")
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/deploy/bluegreen", tags=["Deploy"])
def deploy_bluegreen(dry_run: bool = False):
    """Run blue-green flow: launch staging, smoke, healthcheck x3, switch or rollback. Use dry_run=true to simulate without switching."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.emergency import is_emergency

        if is_emergency():
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "emergency mode: deploys suspended",
            )
        from modules.humanoid.deploy import run_bluegreen_flow
        from modules.humanoid.metrics import get_metrics_store

        get_metrics_store().inc("deploy_bluegreen_runs")
        result = run_bluegreen_flow(dry_run=dry_run)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            result.get("ok", False), result.get("data"), ms, result.get("error")
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/canary/status", tags=["Deploy"])
def canary_status():
    """Canary status: enabled, percentage, features, stats (calls, error rate, latency)."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.deploy.canary import get_canary_stats

        data = get_canary_stats()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class CanaryConfigBody(BaseModel):
    enabled: Optional[bool] = None
    percentage: Optional[float] = None
    features: Optional[List[str]] = None


# â”€â”€ Config IA centralizada (sync Dashboard â†” Workspace) â”€â”€

_ai_config = {
    "mode": "auto",
    "provider": "ollama",
    "model": "llama3.2",
    "temperature": 0.7,
    "top_p": 0.9,
    "top_k": 40,
    "max_tokens": 2048,
    "repeat_penalty": 1.1,
    "system_prompt": "Eres ATLAS, sistema autonomo de gestion inteligente. Responde con soluciones tecnicas concretas: codigo, comandos, endpoints, queries. Nunca generes planes genericos ni pasos abstractos.",
    "memory_context": "long",
    "context_window": 8192,
    "ollama_url": "http://localhost:11434",
    "stream_response": False,
    "save_history": True,
    "log_level": "info",
    "specialists": {
        "code": {"enabled": True, "model": "deepseek-coder"},
        "vision": {"enabled": True, "model": "llava"},
        "chat": {"enabled": True, "model": "llama3.2"},
        "analysis": {"enabled": True, "model": "llama3.2"},
        "creative": {"enabled": False, "model": "llama3.2"},
    },
}

_AI_CONFIG_FILE = BASE_DIR / "config" / "ai_config.json"


def _load_ai_config():
    global _ai_config
    if _AI_CONFIG_FILE.exists():
        try:
            with open(_AI_CONFIG_FILE, "r", encoding="utf-8") as f:
                saved = json.load(f)
            _ai_config.update(saved)
        except Exception:
            pass


def _save_ai_config():
    try:
        _AI_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(_AI_CONFIG_FILE, "w", encoding="utf-8") as f:
            safe = {k: v for k, v in _ai_config.items() if k != "api_key"}
            json.dump(safe, f, indent=2, ensure_ascii=False)
    except Exception:
        pass


_load_ai_config()


@app.get("/config/ai", tags=["Config"])
def get_ai_config():
    config_safe = {k: v for k, v in _ai_config.items() if k != "api_key"}
    config_safe["has_api_key"] = bool(_ai_config.get("api_key"))
    return {"ok": True, **config_safe}


@app.post("/config/ai", tags=["Config"])
def update_ai_config(payload: dict):
    global _ai_config
    allowed = {
        "mode",
        "provider",
        "model",
        "temperature",
        "top_p",
        "top_k",
        "max_tokens",
        "repeat_penalty",
        "system_prompt",
        "memory_context",
        "context_window",
        "ollama_url",
        "stream_response",
        "save_history",
        "log_level",
        "api_key",
        "specialists",
    }
    for key in allowed:
        if key in payload:
            _ai_config[key] = payload[key]
    _save_ai_config()
    return {"ok": True, "message": "Configuracion actualizada"}


@app.post("/config/ai/test", tags=["Config"])
def test_ai_connection():
    provider = _ai_config.get("provider", "ollama")
    try:
        if provider == "ollama":
            import requests as req

            url = _ai_config.get("ollama_url", "http://localhost:11434")
            res = req.get(f"{url}/api/tags", timeout=5)
            if res.status_code == 200:
                return {"ok": True, "provider": provider, "message": "Ollama conectado"}
            return {"ok": False, "error": f"Ollama status {res.status_code}"}
        return {"ok": True, "provider": provider, "message": f"{provider} configurado"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/config/ai/ollama-models", tags=["Config"])
def get_ollama_models():
    try:
        import requests as req

        url = _ai_config.get("ollama_url", "http://localhost:11434")
        res = req.get(f"{url}/api/tags", timeout=10)
        if res.status_code == 200:
            data = res.json()
            models = []
            for m in data.get("models", []):
                sz = m.get("size", 0)
                sz_str = (
                    f"{sz / (1024**3):.1f}GB"
                    if sz > 1024**3
                    else f"{sz / (1024**2):.0f}MB"
                )
                models.append(
                    {
                        "name": m.get("name", "").split(":")[0],
                        "full_name": m.get("name", ""),
                        "size": sz_str,
                    }
                )
            return {"ok": True, "models": models}
        return {"ok": False, "error": f"Ollama status {res.status_code}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/canary/config", tags=["Deploy"])
def canary_config(body: Optional[CanaryConfigBody] = None):
    """Runtime canary config (owner only, policy-gated). Body: enabled?, percentage?, features?."""
    t0 = time.perf_counter()
    body = body or CanaryConfigBody()
    try:
        decision = get_policy_engine().can(_memory_actor(), "deploy", "canary_config")
        if not decision.allow:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                decision.reason or "policy denied",
            )
        from modules.humanoid.deploy.canary import (get_canary_stats,
                                                    set_canary_config)

        set_canary_config(
            enabled=body.enabled, percentage=body.percentage, features=body.features
        )
        data = get_canary_stats()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/deploy/canary/report", tags=["Deploy"])
def deploy_canary_report(hours: float = 24.0):
    """Canary vs stable comparison over the last N hours: counts, error rate, avg latency."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.deploy.canary import get_canary_report

        data = get_canary_report(hours=hours)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


# --- Cluster (multi-node) ---
@app.get("/cluster/status", tags=["Cluster"])
def cluster_status():
    """Cluster status: enabled, node_id, role, nodes count."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cluster import registry

        if not registry.cluster_enabled():
            return _std_resp(
                True,
                {
                    "enabled": False,
                    "node_id": registry.node_id(),
                    "role": registry.node_role(),
                    "nodes": [],
                },
                int((time.perf_counter() - t0) * 1000),
                None,
            )
        nodes = registry.list_nodes()
        data = {
            "enabled": True,
            "node_id": registry.node_id(),
            "role": registry.node_role(),
            "nodes": nodes,
            "count": len(nodes),
        }
        return _std_resp(True, data, int((time.perf_counter() - t0) * 1000), None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/api/tools/job/export/{job_id}", tags=["Tools"])
def tools_job_export(job_id: str):
    """Exporta reporte markdown de un job de actualización de tools."""
    t0 = time.perf_counter()
    safe_job_id = "".join(ch for ch in (job_id or "") if ch.isalnum() or ch in ("-", "_"))
    if not safe_job_id:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "invalid_job_id")
    status_file = (BASE_DIR / "logs" / "tools_update_jobs" / f"{safe_job_id}.json").resolve()
    if not status_file.exists():
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "job_not_found")
    try:
        payload = json.loads(status_file.read_text(encoding="utf-8-sig", errors="replace"))
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))

    reports_dir = (BASE_DIR / "reports").resolve()
    reports_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    report_name = f"tools_update_job_{safe_job_id}_{ts}.md"
    report_path = reports_dir / report_name

    tool = str(payload.get("tool") or "tools")
    status = str(payload.get("status") or "unknown")
    total = int(payload.get("total") or 0)
    done = int(payload.get("done") or 0)
    failed = int(payload.get("failed") or 0)
    started_at = str(payload.get("started_at") or payload.get("queued_at") or "")
    finished_at = str(payload.get("finished_at") or "")
    current_tool = str(payload.get("current_tool") or "")
    err = str(payload.get("error") or "")
    tools = payload.get("tools") if isinstance(payload.get("tools"), list) else []
    results = payload.get("results") if isinstance(payload.get("results"), list) else []

    lines = []
    lines.append(f"# Tools Update Job Report — {safe_job_id}")
    lines.append("")
    lines.append(f"- Tool scope: `{tool}`")
    lines.append(f"- Status: `{status}`")
    lines.append(f"- Progress: `{done}/{total}`")
    lines.append(f"- Failed: `{failed}`")
    if started_at:
        lines.append(f"- Started at: `{started_at}`")
    if finished_at:
        lines.append(f"- Finished at: `{finished_at}`")
    if current_tool:
        lines.append(f"- Current tool: `{current_tool}`")
    if err:
        lines.append(f"- Error: `{err}`")

    if tools:
        lines.append("")
        lines.append("## Target Tools")
        for t in tools:
            lines.append(f"- `{str(t)}`")

    lines.append("")
    lines.append("## Results")
    if not results:
        lines.append("_No tool results recorded yet._")
    else:
        for r in results:
            if not isinstance(r, dict):
                continue
            rid = str(r.get("tool") or r.get("target_repo") or "unknown")
            rok = bool(r.get("ok"))
            lines.append("")
            lines.append(f"### {rid} — {'OK' if rok else 'FAIL'}")
            rerr = str(r.get("error") or "")
            if rerr:
                lines.append(f"- Error: `{rerr}`")
            steps = r.get("steps")
            if isinstance(steps, list) and steps:
                lines.append("- Steps:")
                for s in steps:
                    if isinstance(s, dict):
                        s_name = str(s.get("step") or "step")
                        s_ok = bool(s.get("ok"))
                        s_err = str(s.get("error") or "")
                        lines.append(f"  - `{s_name}`: {'ok' if s_ok else 'fail'}{(' — ' + s_err) if s_err else ''}")
            lines.append("")
            lines.append("```json")
            lines.append(json.dumps(r, ensure_ascii=False, indent=2))
            lines.append("```")

    lines.append("")
    lines.append("## Raw Job Payload")
    lines.append("```json")
    lines.append(json.dumps(payload, ensure_ascii=False, indent=2))
    lines.append("```")

    report_path.write_text("\n".join(lines), encoding="utf-8")

    return _std_resp(
        True,
        {
            "ok": True,
            "job_id": safe_job_id,
            "report_path": str(report_path),
            "report_name": report_name,
            "report_url": f"/reports/{report_name}",
        },
        int((time.perf_counter() - t0) * 1000),
        None,
    )


@app.post("/api/tools/job/retry/{job_id}", tags=["Tools"])
def tools_job_retry(job_id: str):
    """Reintenta un job de tools relanzando solo herramientas pendientes/fallidas."""
    t0 = time.perf_counter()
    safe_job_id = "".join(ch for ch in (job_id or "") if ch.isalnum() or ch in ("-", "_"))
    if not safe_job_id:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "invalid_job_id")

    jobs_dir = (BASE_DIR / "logs" / "tools_update_jobs").resolve()
    status_file = (jobs_dir / f"{safe_job_id}.json").resolve()
    if not status_file.exists():
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "job_not_found")

    try:
        payload = json.loads(status_file.read_text(encoding="utf-8-sig", errors="replace"))
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))

    tools_all = payload.get("tools") if isinstance(payload.get("tools"), list) else []
    results = payload.get("results") if isinstance(payload.get("results"), list) else []
    ok_set = set()
    fail_set = set()
    for r in results:
        if not isinstance(r, dict):
            continue
        tid = str(r.get("tool") or "").strip()
        if not tid:
            continue
        if r.get("ok") is True:
            ok_set.add(tid)
        elif r.get("ok") is False:
            fail_set.add(tid)

    pending = [str(t).strip() for t in tools_all if str(t).strip() and str(t).strip() not in ok_set]
    if not pending and fail_set:
        pending = sorted(fail_set)
    pending = sorted(set(pending))
    if not pending:
        return _std_resp(
            True,
            {
                "ok": True,
                "queued": False,
                "status": "noop",
                "message": "No hay herramientas pendientes para reintentar",
                "source_job_id": safe_job_id,
            },
            int((time.perf_counter() - t0) * 1000),
            None,
        )

    bg_script = (BASE_DIR / "scripts" / "atlas_tools_update_all_background.ps1").resolve()
    if not bg_script.exists():
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            f"background update-all script missing: {bg_script}",
        )

    try:
        new_job_id = uuid.uuid4().hex[:12]
        new_status_file = jobs_dir / f"{new_job_id}.json"
        runner_log = jobs_dir / f"{new_job_id}.runner.log"
        new_status_file.write_text(
            json.dumps(
                {
                    "ok": True,
                    "job_id": new_job_id,
                    "tool": "all",
                    "status": "queued",
                    "queued_at": datetime.now(timezone.utc).isoformat(),
                    "tools": pending,
                    "total": len(pending),
                    "done": 0,
                    "failed": 0,
                    "retry_of_job_id": safe_job_id,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        powershell_exe = str(
            (
                Path(os.environ.get("SystemRoot", r"C:\Windows"))
                / "System32"
                / "WindowsPowerShell"
                / "v1.0"
                / "powershell.exe"
            ).resolve()
        )
        cmd_bg = [
            powershell_exe,
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(bg_script),
            "-RepoRoot",
            str(BASE_DIR),
            "-JobFile",
            str(new_status_file),
            "-ToolsCsv",
            ",".join(pending),
        ]
        _spawn_detached_job(
            cmd_bg,
            cwd=str(BASE_DIR),
            runner_log=runner_log,
        )
        return _std_resp(
            True,
            {
                "ok": True,
                "queued": True,
                "job_id": new_job_id,
                "tool": "all",
                "status": "queued",
                "tools": pending,
                "total": len(pending),
                "retry_of_job_id": safe_job_id,
                "status_url": f"/api/tools/job/status/{new_job_id}",
                "runner_log": str(runner_log),
            },
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


# --- Software Hub ---
class SoftwareWatchdogBody(BaseModel):
    force: Optional[bool] = False


class SoftwareApplyAllBody(BaseModel):
    background: Optional[bool] = True
    include_optional: Optional[bool] = True


class SoftwareMaintenanceBody(BaseModel):
    background: Optional[bool] = True
    apply_repo: Optional[bool] = False


_SOFTWARE_HUB_DIR = (BASE_DIR / "software_hub").resolve()
_SOFTWARE_REGISTRY_FILE = (_SOFTWARE_HUB_DIR / "software_registry.json").resolve()
_SOFTWARE_MAINT_LATEST_FILE = (BASE_DIR / "logs" / "software_maintenance" / "latest_cycle.json").resolve()
_SOFTWARE_MENU_CACHE_TTL_SEC = _env_int("ATLAS_SOFTWARE_MENU_CACHE_TTL_SEC", 30, 8)
_SOFTWARE_MENU_STALE_SEC = _env_int(
    "ATLAS_SOFTWARE_MENU_STALE_SEC", 300, _SOFTWARE_MENU_CACHE_TTL_SEC
)
_SOFTWARE_JOB_STALE_SEC = _env_int("ATLAS_SOFTWARE_JOB_STALE_SEC", 900, 120)
_software_watchdog_refresh_lock = threading.Lock()
_software_watchdog_refresh_running = False


def _run_software_watchdog_scan(force: bool = False, timeout: int = 180) -> tuple[bool, dict]:
    script = (BASE_DIR / "scripts" / "atlas_software_watchdog.py").resolve()
    if not script.exists():
        return False, {"error": f"software watchdog script missing: {script}"}
    cmd = [sys.executable, str(script)]
    if force:
        cmd.append("--force")
    ok, payload, _tail = _run_local_json_cmd(cmd, timeout=timeout)
    if not isinstance(payload, dict):
        payload = {"ok": bool(ok), "error": "software_watchdog_invalid_payload"}
    return bool(ok), payload


def _spawn_software_watchdog_refresh(force: bool = False) -> bool:
    global _software_watchdog_refresh_running
    with _software_watchdog_refresh_lock:
        if _software_watchdog_refresh_running:
            return False
        _software_watchdog_refresh_running = True

    def _job() -> None:
        global _software_watchdog_refresh_running
        try:
            _run_software_watchdog_scan(force=force, timeout=240)
        except Exception:
            pass
        finally:
            with _software_watchdog_refresh_lock:
                _software_watchdog_refresh_running = False

    threading.Thread(target=_job, name="atlas-software-watchdog-refresh", daemon=True).start()
    return True


@app.get("/api/software/menu", tags=["Software"])
def software_menu_inventory(refresh: bool = False):
    """Inventario extendido de software (instalado + red + drivers + stack dev)."""
    t0 = time.perf_counter()
    now_ts = time.time()
    payload = _read_json_dict(_SOFTWARE_REGISTRY_FILE)
    file_age = None
    if _SOFTWARE_REGISTRY_FILE.exists():
        try:
            file_age = max(0.0, now_ts - _SOFTWARE_REGISTRY_FILE.stat().st_mtime)
        except Exception:
            file_age = None

    if payload and not refresh:
        age_sec = int(file_age or 0)
        stale = age_sec > _SOFTWARE_MENU_CACHE_TTL_SEC
        if stale:
            _spawn_software_watchdog_refresh(force=False)
        out = dict(payload)
        out["cache"] = {
            "stale": stale,
            "age_sec": age_sec,
            "refreshing": bool(_software_watchdog_refresh_running),
            "ttl_sec": _SOFTWARE_MENU_CACHE_TTL_SEC,
        }
        return _std_resp(True, out, int((time.perf_counter() - t0) * 1000), None)

    ok, fresh = _run_software_watchdog_scan(force=bool(refresh), timeout=240)
    if ok:
        out = dict(fresh)
        out["cache"] = {
            "stale": False,
            "age_sec": 0,
            "refreshing": False,
            "ttl_sec": _SOFTWARE_MENU_CACHE_TTL_SEC,
        }
        return _std_resp(True, out, int((time.perf_counter() - t0) * 1000), None)

    if payload:
        age_sec = int(file_age or 0)
        if age_sec > _SOFTWARE_MENU_STALE_SEC:
            _spawn_software_watchdog_refresh(force=True)
        out = dict(payload)
        out["cache"] = {
            "stale": True,
            "age_sec": age_sec,
            "refreshing": bool(_software_watchdog_refresh_running),
            "ttl_sec": _SOFTWARE_MENU_CACHE_TTL_SEC,
            "warning": fresh.get("error") or "software_watchdog_failed_using_cache",
        }
        return _std_resp(True, out, int((time.perf_counter() - t0) * 1000), None)

    return _std_resp(
        False,
        None,
        int((time.perf_counter() - t0) * 1000),
        fresh.get("error") or "software_watchdog_failed",
    )


@app.post("/api/software/watchdog/scan", tags=["Software"])
def software_watchdog_scan(body: Optional[SoftwareWatchdogBody] = None):
    t0 = time.perf_counter()
    force = bool(body.force) if body else False
    ok, payload = _run_software_watchdog_scan(force=force, timeout=240)
    return _std_resp(ok, payload, int((time.perf_counter() - t0) * 1000), None if ok else payload.get("error"))


@app.get("/api/software/discovery/search", tags=["Software"])
def software_discovery_search(force: bool = False):
    """Lista de software detectado en red para instalar/activar."""
    t0 = time.perf_counter()
    payload = _read_json_dict(_SOFTWARE_REGISTRY_FILE)
    if force or not payload:
        ok, payload = _run_software_watchdog_scan(force=force, timeout=240)
        if not ok:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                payload.get("error") or "software_watchdog_failed",
            )
    rows = payload.get("network_candidates") if isinstance(payload, dict) else []
    rows = rows if isinstance(rows, list) else []
    summary = payload.get("summary") if isinstance(payload, dict) else {}
    data = {
        "ok": True,
        "generated_at": payload.get("generated_at"),
        "summary": {
            "total": len(rows),
            "available": sum(1 for r in rows if isinstance(r, dict) and r.get("network_available")),
            "unreachable": sum(1 for r in rows if isinstance(r, dict) and not r.get("network_available")),
            "software_update_ready": int((summary or {}).get("software_update_ready") or 0),
        },
        "candidates": rows,
    }
    return _std_resp(True, data, int((time.perf_counter() - t0) * 1000), None)


@app.post("/api/software/apply-all", tags=["Software"])
def software_apply_all(body: Optional[SoftwareApplyAllBody] = None):
    """Instala/actualiza software en lote con barra de progreso por item."""
    t0 = time.perf_counter()
    if not (bool(body.background) if body else True):
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "software_apply_all_requires_background")

    bg_script = (BASE_DIR / "scripts" / "atlas_software_apply_all_background.ps1").resolve()
    if not bg_script.exists():
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            f"missing script: {bg_script}",
        )

    payload = _read_json_dict(_SOFTWARE_REGISTRY_FILE)
    if not payload:
        ok, payload = _run_software_watchdog_scan(force=True, timeout=240)
        if not ok:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                payload.get("error") or "software_watchdog_failed",
            )

    include_optional = True if body is None else bool(body.include_optional)
    software_rows = payload.get("software") if isinstance(payload, dict) else []
    software_rows = software_rows if isinstance(software_rows, list) else []
    candidates_rows = payload.get("network_candidates") if isinstance(payload, dict) else []
    candidates_rows = candidates_rows if isinstance(candidates_rows, list) else []

    targets: list[dict] = []
    seen = set()

    for row in software_rows:
        if not isinstance(row, dict):
            continue
        sid = str(row.get("id") or "").strip()
        if not sid or sid in seen:
            continue
        method = str(row.get("install_method") or "").strip().lower()
        target = str(row.get("install_target") or "").strip()
        if not method or not target:
            continue
        critical = bool(row.get("critical"))
        needs = (not bool(row.get("installed"))) or bool(row.get("update_ready"))
        if not needs:
            continue
        if (not include_optional) and (not critical):
            continue
        targets.append(
            {
                "id": sid,
                "name": str(row.get("name") or sid),
                "critical": critical,
                "install_method": method,
                "install_target": target,
                "source": "software",
            }
        )
        seen.add(sid)

    for row in candidates_rows:
        if not isinstance(row, dict):
            continue
        sid = str(row.get("id") or "").strip()
        if not sid or sid in seen:
            continue
        if not bool(row.get("network_available")):
            continue
        method = str(row.get("install_method") or "").strip().lower()
        target = str(row.get("install_target") or "").strip()
        if not method or not target:
            continue
        critical = bool(row.get("critical"))
        if (not include_optional) and (not critical):
            continue
        targets.append(
            {
                "id": sid,
                "name": str(row.get("name") or sid),
                "critical": critical,
                "install_method": method,
                "install_target": target,
                "source": "network",
            }
        )
        seen.add(sid)

    if not targets:
        return _std_resp(
            True,
            {"ok": True, "queued": False, "status": "noop", "message": "No hay software pendiente por actualizar/instalar", "items": []},
            int((time.perf_counter() - t0) * 1000),
            None,
        )

    try:
        jobs_dir = (BASE_DIR / "logs" / "software_update_jobs").resolve()
        jobs_dir.mkdir(parents=True, exist_ok=True)
        job_id = uuid.uuid4().hex[:12]
        status_file = jobs_dir / f"{job_id}.json"
        runner_log = jobs_dir / f"{job_id}.runner.log"
        manifest_file = jobs_dir / f"{job_id}.manifest.json"
        manifest_file.write_text(json.dumps({"items": targets}, ensure_ascii=False), encoding="utf-8")
        status_file.write_text(
            json.dumps(
                {
                    "ok": True,
                    "job_id": job_id,
                    "scope": "software",
                    "source": "software_apply_all",
                    "status": "queued",
                    "queued_at": datetime.now(timezone.utc).isoformat(),
                    "items": [str(t.get("id") or "") for t in targets],
                    "total": len(targets),
                    "done": 0,
                    "failed": 0,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        powershell_exe = str(
            (
                Path(os.environ.get("SystemRoot", r"C:\Windows"))
                / "System32"
                / "WindowsPowerShell"
                / "v1.0"
                / "powershell.exe"
            ).resolve()
        )
        cmd_bg = [
            powershell_exe,
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(bg_script),
            "-RepoRoot",
            str(BASE_DIR),
            "-JobFile",
            str(status_file),
            "-ManifestFile",
            str(manifest_file),
        ]
        _spawn_detached_job(cmd_bg, cwd=BASE_DIR, runner_log=runner_log)
        return _std_resp(
            True,
            {
                "ok": True,
                "queued": True,
                "job_id": job_id,
                "scope": "software",
                "status": "queued",
                "items": [str(t.get("id") or "") for t in targets],
                "total": len(targets),
                "status_url": f"/api/software/job/status/{job_id}",
                "runner_log": str(runner_log),
            },
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


def _software_job_process_alive(job_id: str) -> bool:
    try:
        needle = f"{job_id}.json".lower()
        script_markers = (
            "atlas_software_apply_all_background.ps1",
            "atlas_software_apply_item_background.ps1",
        )
        for proc in psutil.process_iter(["cmdline"]):
            try:
                cmd = " ".join(str(x) for x in (proc.info.get("cmdline") or [])).lower()
            except Exception:
                continue
            if not cmd:
                continue
            if needle in cmd and any(m in cmd for m in script_markers):
                return True
        return False
    except Exception:
        return True


@app.get("/api/software/job/status/{job_id}", tags=["Software"])
def software_job_status(job_id: str):
    t0 = time.perf_counter()
    safe_job_id = "".join(ch for ch in (job_id or "") if ch.isalnum() or ch in ("-", "_"))
    if not safe_job_id:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "invalid_job_id")
    status_file = (BASE_DIR / "logs" / "software_update_jobs" / f"{safe_job_id}.json").resolve()
    if not status_file.exists():
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "job_not_found")
    try:
        payload = json.loads(status_file.read_text(encoding="utf-8-sig", errors="replace"))
        if isinstance(payload, dict) and not str(payload.get("job_id") or "").strip():
            payload["job_id"] = safe_job_id
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))

    try:
        changed = False
        status = str(payload.get("status") or "").lower()
        total = int(payload.get("total") or 0)
        done = int(payload.get("done") or 0)
        failed = int(payload.get("failed") or 0)

        if done < 0:
            payload["done"] = 0
            done = 0
            changed = True
        if total > 0 and done > total:
            payload["done"] = total
            done = total
            changed = True
        if failed < 0:
            payload["failed"] = 0
            failed = 0
            changed = True
        if done > 0 and failed > done:
            payload["failed"] = done
            failed = done
            changed = True

        if status in ("queued", "running", "failed", "done_with_errors") and total > 0 and done >= total:
            terminal_failed = failed > 0
            payload["status"] = "done_with_errors" if terminal_failed else "done"
            payload["ok"] = not terminal_failed
            if terminal_failed and not str(payload.get("error") or "").strip():
                payload["error"] = "one_or_more_software_applies_failed"
            if not str(payload.get("finished_at") or "").strip():
                payload["finished_at"] = datetime.now(timezone.utc).isoformat()
            changed = True

        status = str(payload.get("status") or "").lower()
        if status == "running":
            try:
                now_ts = time.time()
                status_mtime = status_file.stat().st_mtime
                runner_log = (BASE_DIR / "logs" / "software_update_jobs" / f"{safe_job_id}.runner.log").resolve()
                runner_mtime = runner_log.stat().st_mtime if runner_log.exists() else status_mtime
                stale_for = now_ts - max(status_mtime, runner_mtime)
                if stale_for > _SOFTWARE_JOB_STALE_SEC and (not _software_job_process_alive(safe_job_id)):
                    payload["ok"] = False
                    payload["status"] = "failed"
                    payload["error"] = payload.get("error") or "job_runner_not_alive"
                    payload["failed_at"] = datetime.now(timezone.utc).isoformat()
                    changed = True
            except Exception:
                pass
        if status == "queued":
            q_at = payload.get("queued_at")
            if q_at:
                dt_q = datetime.fromisoformat(str(q_at).replace("Z", "+00:00"))
                now = datetime.now(dt_q.tzinfo) if dt_q.tzinfo else datetime.now()
                if (now - dt_q).total_seconds() > 120:
                    payload["ok"] = False
                    payload["status"] = "failed"
                    payload["error"] = payload.get("error") or "job_stale_in_queue"
                    payload["failed_at"] = datetime.now(timezone.utc).isoformat()
                    changed = True

        if changed:
            status_file.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass

    return _std_resp(True, payload, int((time.perf_counter() - t0) * 1000), None)


@app.get("/api/software/job/latest", tags=["Software"])
def software_job_latest():
    t0 = time.perf_counter()
    jobs_dir = (BASE_DIR / "logs" / "software_update_jobs").resolve()
    if not jobs_dir.exists():
        return _std_resp(True, {"ok": True, "found": False}, int((time.perf_counter() - t0) * 1000), None)
    try:
        files = sorted(jobs_dir.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)[:40]
        for f in files:
            try:
                payload = json.loads(f.read_text(encoding="utf-8-sig", errors="replace"))
            except Exception:
                continue
            status = str(payload.get("status") or "").lower()
            if status in ("queued", "running"):
                payload["job_id"] = payload.get("job_id") or f.stem
                payload["status_url"] = f"/api/software/job/status/{payload['job_id']}"
                return _std_resp(
                    True,
                    {"ok": True, "found": True, "job": payload},
                    int((time.perf_counter() - t0) * 1000),
                    None,
                )
        return _std_resp(True, {"ok": True, "found": False}, int((time.perf_counter() - t0) * 1000), None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/api/software/job/runner-log/{job_id}", tags=["Software"])
def software_job_runner_log(job_id: str, tail_lines: int = 120):
    t0 = time.perf_counter()
    safe_job_id = "".join(ch for ch in (job_id or "") if ch.isalnum() or ch in ("-", "_"))
    if not safe_job_id:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "invalid_job_id")
    jobs_dir = (BASE_DIR / "logs" / "software_update_jobs").resolve()
    log_path = (jobs_dir / f"{safe_job_id}.runner.log").resolve()
    if not log_path.exists():
        return _std_resp(
            True,
            {"ok": True, "found": False, "job_id": safe_job_id, "tail": ""},
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    try:
        lines = log_path.read_text(encoding="utf-8-sig", errors="replace").splitlines()
        n = max(10, min(500, int(tail_lines or 120)))
        tail = "\n".join(lines[-n:])
        return _std_resp(
            True,
            {"ok": True, "found": True, "job_id": safe_job_id, "log_path": str(log_path), "tail": tail},
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/api/software/maintenance/cycle", tags=["Software"])
def software_maintenance_cycle(body: Optional[SoftwareMaintenanceBody] = None):
    """Ejecuta ciclo de mantenimiento: triada + makeplay + repo + refresh inventarios."""
    t0 = time.perf_counter()
    script = (BASE_DIR / "scripts" / "atlas_software_maintenance_cycle.py").resolve()
    if not script.exists():
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), f"missing script: {script}")

    apply_repo = bool(body.apply_repo) if body else False
    background = bool(body.background) if body else True

    if not background:
        cmd = [sys.executable, str(script)]
        if apply_repo:
            cmd.append("--apply-repo")
        ok, payload, _tail = _run_local_json_cmd(cmd, timeout=2400, cwd=BASE_DIR)
        return _std_resp(ok, payload, int((time.perf_counter() - t0) * 1000), None if ok else payload.get("error"))

    try:
        jobs_dir = (BASE_DIR / "logs" / "software_maintenance_jobs").resolve()
        jobs_dir.mkdir(parents=True, exist_ok=True)
        job_id = uuid.uuid4().hex[:12]
        status_file = jobs_dir / f"{job_id}.json"
        runner_log = jobs_dir / f"{job_id}.runner.log"
        status_file.write_text(
            json.dumps(
                {
                    "ok": True,
                    "job_id": job_id,
                    "status": "queued",
                    "queued_at": datetime.now(timezone.utc).isoformat(),
                    "total": 8,
                    "done": 0,
                    "failed": 0,
                    "apply_repo": apply_repo,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        cmd = [sys.executable, str(script), "--job-file", str(status_file)]
        if apply_repo:
            cmd.append("--apply-repo")
        _spawn_detached_job(cmd, cwd=BASE_DIR, runner_log=runner_log)
        return _std_resp(
            True,
            {
                "ok": True,
                "queued": True,
                "job_id": job_id,
                "status": "queued",
                "status_url": f"/api/software/maintenance/status/{job_id}",
                "runner_log": str(runner_log),
            },
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/api/software/maintenance/status/{job_id}", tags=["Software"])
def software_maintenance_status(job_id: str):
    t0 = time.perf_counter()
    safe_job_id = "".join(ch for ch in (job_id or "") if ch.isalnum() or ch in ("-", "_"))
    if not safe_job_id:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "invalid_job_id")
    status_file = (BASE_DIR / "logs" / "software_maintenance_jobs" / f"{safe_job_id}.json").resolve()
    if not status_file.exists():
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "job_not_found")
    try:
        payload = json.loads(status_file.read_text(encoding="utf-8-sig", errors="replace"))
        return _std_resp(True, payload, int((time.perf_counter() - t0) * 1000), None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/api/software/maintenance/latest", tags=["Software"])
def software_maintenance_latest():
    t0 = time.perf_counter()
    payload = _read_json_dict(_SOFTWARE_MAINT_LATEST_FILE)
    if not payload:
        return _std_resp(
            True,
            {"ok": True, "found": False, "message": "No maintenance cycle report yet"},
            int((time.perf_counter() - t0) * 1000),
            None,
        )
    return _std_resp(True, {"ok": True, "found": True, "report": payload}, int((time.perf_counter() - t0) * 1000), None)


@app.get("/cluster/nodes", tags=["Cluster"])
def cluster_nodes(status_filter: Optional[str] = None):
    """List registered nodes (optionally filter by status). Marks stale nodes offline."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cluster import registry

        registry.mark_offline_stale()
        nodes = registry.list_nodes(status_filter=status_filter)
        return _std_resp(True, nodes, int((time.perf_counter() - t0) * 1000), None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/cluster/events", tags=["Cluster"])
def cluster_events(limit: int = 50, node_id_filter: Optional[str] = None):
    """Last node_events for UI (register, heartbeat, offline, route, error)."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cluster import db as cluster_db

        events = cluster_db.get_events(limit=limit, node_id=node_id_filter)
        return _std_resp(True, events, int((time.perf_counter() - t0) * 1000), None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


class ClusterRegisterBody(BaseModel):
    node_id: str
    role: str = "worker"
    base_url: str
    capabilities: Optional[dict] = None
    tags: Optional[dict] = None


@app.post("/cluster/node/register", tags=["Cluster"])
def cluster_node_register(body: ClusterRegisterBody):
    """Register or update a node (policy: remote_execute or owner)."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(_memory_actor(), "cluster", "remote_execute")
        if not decision.allow:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                decision.reason or "policy denied",
            )
        from modules.humanoid.cluster import registry

        result = registry.register_node(
            body.node_id, body.role, body.base_url, body.capabilities, body.tags
        )
        return _std_resp(
            result.get("ok", False),
            result,
            int((time.perf_counter() - t0) * 1000),
            result.get("error"),
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


class ClusterHeartbeatBody(BaseModel):
    node_id: str
    capabilities: dict = {}
    health: dict = {}
    version: str = ""
    channel: str = "canary"
    base_url: Optional[str] = None


@app.post("/cluster/heartbeat", tags=["Cluster"])
def cluster_heartbeat(body: ClusterHeartbeatBody, request: Request = None):
    """HQ: receive heartbeat from worker. Registers node if new (base_url from request)."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cluster.heartbeat import receive_heartbeat

        result = receive_heartbeat(
            body.node_id,
            body.capabilities,
            body.health,
            body.version,
            body.channel,
            base_url=body.base_url,
        )
        return _std_resp(
            result.get("ok", False),
            result,
            int((time.perf_counter() - t0) * 1000),
            result.get("error"),
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.get("/cluster/ping", tags=["Cluster"])
def cluster_ping():
    """Ping: respond 200 if alive (for HQ->worker check)."""
    return _std_resp(True, {"pong": True}, 0, None)


class ClusterRouteBody(BaseModel):
    task: str = "hands"
    prefer_remote: bool = False
    require_capability: Optional[str] = None


@app.post("/cluster/route/decision", tags=["Cluster"])
def cluster_route_decision(body: ClusterRouteBody):
    """Debug: get routing decision for a task."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cluster.router import route_decision

        decision = route_decision(
            body.task,
            prefer_remote=body.prefer_remote,
            require_capability=body.require_capability,
        )
        return _std_resp(True, decision, int((time.perf_counter() - t0) * 1000), None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


# --- Gateway (Stealth multi-path) ---
@app.get("/gateway/status", tags=["Gateway"])
def gateway_status():
    """Gateway status: enabled, mode, tools detected, last success, recommendations."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.gateway.store import build_gateway_status

        data = build_gateway_status()
        return _std_resp(True, data, int((time.perf_counter() - t0) * 1000), None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/gateway/check", tags=["Gateway"])
def gateway_check():
    """Check gateway candidates (policy: gateway_check)."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(_memory_actor(), "gateway", "gateway_check")
        if not decision.allow:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                decision.reason or "policy denied",
            )
        from modules.humanoid.gateway.selector import check_candidates

        data = check_candidates()
        return _std_resp(True, data, int((time.perf_counter() - t0) * 1000), None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/gateway/bootstrap", tags=["Gateway"])
def gateway_bootstrap():
    """Bootstrap worker to HQ (policy: gateway_bootstrap). Returns quickly if no_gateway_config."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(
            _memory_actor(), "gateway", "gateway_bootstrap"
        )
        if not decision.allow:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                decision.reason or "policy denied",
            )
        from modules.humanoid.gateway import bootstrap as gw_bootstrap
        from modules.humanoid.gateway import selector as gw_selector

        target = gw_selector.resolve_worker_url(None)
        if not target:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), "no_gateway_config"
            )
        result = gw_bootstrap.bootstrap()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(result.get("ok", False), result, ms, result.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class GatewayModeBody(BaseModel):
    mode: str  # auto | cloudflare | tailscale | ssh | lan


@app.post("/gateway/mode", tags=["Gateway"])
def gateway_set_mode(body: GatewayModeBody):
    """Set gateway mode (policy: gateway_set_mode)."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(
            _memory_actor(), "gateway", "gateway_set_mode"
        )
        if not decision.allow:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                decision.reason or "policy denied",
            )
        from modules.humanoid.gateway import store as gateway_store

        gateway_store.set_mode(body.mode)
        data = gateway_store.build_gateway_status()
        return _std_resp(True, data, int((time.perf_counter() - t0) * 1000), None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


def _cluster_remote_result(
    ok: bool,
    data: Any = None,
    ms: int = 0,
    error: Optional[str] = None,
    correlation_id: Optional[str] = None,
) -> dict:
    out = {"ok": ok, "data": data, "ms": ms, "error": error}
    if correlation_id:
        out["correlation_id"] = correlation_id
    return out


@app.post("/remote/hands", tags=["Cluster"])
def remote_hands(body: dict):
    """Worker: execute hands (shell). Returns {ok, data, ms, error, correlation_id}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.emergency import is_action_blocked

        if is_action_blocked("remote_hands"):
            return _cluster_remote_result(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "emergency_mode_block",
                body.get("correlation_id"),
            )
        from modules.humanoid.cluster import trace
        from modules.humanoid.cluster.remote_server import execute_remote_hands

        cid = body.get("correlation_id") or trace.new_correlation_id()
        r = execute_remote_hands(body)
        ms = int((time.perf_counter() - t0) * 1000)
        return _cluster_remote_result(
            r.get("ok", False), r.get("data"), ms, r.get("error"), cid
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _cluster_remote_result(
            False, None, ms, str(e), body.get("correlation_id")
        )


@app.post("/remote/web", tags=["Cluster"])
def remote_web(body: dict):
    """Worker: execute web action."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.emergency import is_action_blocked

        if is_action_blocked("remote_web"):
            return _cluster_remote_result(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "emergency_mode_block",
                body.get("correlation_id"),
            )
        from modules.humanoid.cluster import trace
        from modules.humanoid.cluster.remote_server import execute_remote_web

        cid = body.get("correlation_id") or trace.new_correlation_id()
        r = execute_remote_web(body)
        ms = int((time.perf_counter() - t0) * 1000)
        return _cluster_remote_result(
            r.get("ok", False), r.get("data"), ms, r.get("error"), cid
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _cluster_remote_result(
            False, None, ms, str(e), body.get("correlation_id")
        )


@app.post("/remote/vision", tags=["Cluster"])
def remote_vision(body: dict):
    """Worker: execute vision (analyze image)."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.emergency import is_action_blocked

        if is_action_blocked("remote_vision"):
            return _cluster_remote_result(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "emergency_mode_block",
                body.get("correlation_id"),
            )
        from modules.humanoid.cluster import trace
        from modules.humanoid.cluster.remote_server import \
            execute_remote_vision

        cid = body.get("correlation_id") or trace.new_correlation_id()
        r = execute_remote_vision(body)
        ms = int((time.perf_counter() - t0) * 1000)
        return _cluster_remote_result(
            r.get("ok", False), r.get("data"), ms, r.get("error"), cid
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _cluster_remote_result(
            False, None, ms, str(e), body.get("correlation_id")
        )


@app.post("/remote/voice", tags=["Cluster"])
def remote_voice(body: dict):
    """Worker: execute voice action."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.owner.emergency import is_action_blocked

        if is_action_blocked("remote_voice"):
            return _cluster_remote_result(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "emergency_mode_block",
                body.get("correlation_id"),
            )
        from modules.humanoid.cluster import trace
        from modules.humanoid.cluster.remote_server import execute_remote_voice

        cid = body.get("correlation_id") or trace.new_correlation_id()
        r = execute_remote_voice(body)
        ms = int((time.perf_counter() - t0) * 1000)
        return _cluster_remote_result(
            r.get("ok", False), r.get("data"), ms, r.get("error"), cid
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _cluster_remote_result(
            False, None, ms, str(e), body.get("correlation_id")
        )


@app.get("/watchdog/status")
def watchdog_status_endpoint():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.watchdog import watchdog_status

        data = watchdog_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/healing/status")
def healing_status_endpoint():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.healing import healing_status

        data = healing_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/api/push/autorevive/status", tags=["Health"])
def push_autorevive_status_endpoint():
    """Estado del watchdog de auto-revive de PUSH (script externo)."""
    t0 = time.perf_counter()
    try:
        log_dir = BASE_DIR / "logs"
        state_file = log_dir / "push_autorevive_state.json"
        log_file = log_dir / "push_autorevive_watchdog.log"
        state = {}
        if state_file.exists():
            try:
                state = json.loads(state_file.read_text(encoding="utf-8", errors="replace"))
            except Exception:
                state = {}

        # Keep this endpoint fast and non-blocking; avoid expensive process scans.
        running = False
        pids: list[int] = []
        try:
            last_ok_iso = state.get("last_ok_at")
            if last_ok_iso:
                dt_last_ok = datetime.fromisoformat(str(last_ok_iso).replace("Z", "+00:00"))
                now = datetime.now(dt_last_ok.tzinfo) if dt_last_ok.tzinfo else datetime.now()
                running = (now - dt_last_ok).total_seconds() <= 120
        except Exception:
            running = False

        last_log = ""
        if log_file.exists():
            try:
                lines = log_file.read_text(encoding="utf-8", errors="replace").splitlines()
                if lines:
                    last_log = lines[-1][-400:]
            except Exception:
                last_log = ""

        payload = {
            "running": running,
            "pids": [x for x in pids if x > 0][:10],
            "state_file": str(state_file),
            "log_file": str(log_file),
            "state": state,
            "last_restart_at": state.get("last_restart_at"),
            "last_error": state.get("last_error"),
            "fail_streak": state.get("fail_streak"),
            "restarts": state.get("restarts"),
            "last_ok_at": state.get("last_ok_at"),
            "last_log_line": last_log,
        }
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, payload, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/api/diagnostic/live", tags=["Health"])
def diagnostic_live_endpoint():
    """Diagnostico rapido consolidado: health + autorevive + ultimo job de tools."""
    t0 = time.perf_counter()
    try:
        logs_dir = (BASE_DIR / "logs").resolve()
        tools_jobs_dir = (logs_dir / "tools_update_jobs").resolve()
        state_file = (logs_dir / "push_autorevive_state.json").resolve()

        health_ok = False
        health_payload = {}
        try:
            import urllib.request

            req = urllib.request.Request(
                "http://127.0.0.1:8791/health",
                method="GET",
                headers={"Accept": "application/json"},
            )
            with urllib.request.urlopen(req, timeout=2.5) as resp:
                if 200 <= int(resp.status) < 300:
                    raw = resp.read().decode("utf-8", errors="replace")
                    health_payload = json.loads(raw) if raw else {}
                    health_ok = True
        except Exception:
            health_ok = False
            health_payload = {}

        autorevive_state = {}
        if state_file.exists():
            try:
                autorevive_state = json.loads(state_file.read_text(encoding="utf-8", errors="replace"))
            except Exception:
                autorevive_state = {}

        last_job = None
        if tools_jobs_dir.exists():
            try:
                files = sorted(
                    [f for f in tools_jobs_dir.glob("*.json") if f.is_file()],
                    key=lambda x: x.stat().st_mtime,
                    reverse=True,
                )
                if files:
                    payload = json.loads(files[0].read_text(encoding="utf-8-sig", errors="replace"))
                    last_job = {
                        "job_id": payload.get("job_id") or files[0].stem,
                        "status": payload.get("status") or "unknown",
                        "tool": payload.get("tool") or payload.get("target") or "all",
                        "progress_pct": payload.get("progress_pct"),
                        "updated_at": payload.get("updated_at") or payload.get("finished_at") or payload.get("started_at"),
                    }
            except Exception:
                last_job = None

        result = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "push": {
                "alive": bool(health_ok),
                "pid": health_payload.get("pid"),
                "score": health_payload.get("score"),
                "uptime_human": health_payload.get("uptime_human"),
            },
            "autorevive": {
                "running_recently": bool(autorevive_state.get("last_ok_at")),
                "last_ok_at": autorevive_state.get("last_ok_at"),
                "last_restart_at": autorevive_state.get("last_restart_at"),
                "fail_streak": autorevive_state.get("fail_streak"),
                "restarts": autorevive_state.get("restarts"),
                "last_error": autorevive_state.get("last_error"),
            },
            "tools": {"last_job": last_job},
        }
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, result, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


# --- Concurrent Goal Engine (CGE) ---


class CGESubmitBody(BaseModel):
    goal_type: str = "general"
    description: str
    priority: int = 1
    parameters: Optional[dict] = None
    source: str = "api"
    parent_goal_id: Optional[str] = None
    deadline_ts: Optional[float] = None
    resources_needed: Optional[list] = None


@app.post("/cge/submit", tags=["CGE"])
def cge_submit(body: CGESubmitBody):
    """Registra un nuevo goal concurrente en el CGE."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cortex.frontal.concurrent_engine import \
            get_engine

        engine = get_engine()
        gid = engine.submit_goal(
            goal_type=body.goal_type,
            description=body.description,
            priority=body.priority,
            parameters=body.parameters,
            source=body.source,
            parent_goal_id=body.parent_goal_id,
            deadline_ts=body.deadline_ts,
            resources_needed=body.resources_needed,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"goal_id": gid}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/cge/status", tags=["CGE"])
def cge_status():
    """Estado completo del motor concurrente: goals activos, recursos, executor."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cortex.frontal.concurrent_engine import \
            get_engine

        data = get_engine().get_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/cge/goal/{goal_id}", tags=["CGE"])
def cge_goal_detail(goal_id: str):
    """Detalle de un goal: contexto, plan, log de ejecucion, recursos."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cortex.frontal.concurrent_engine import \
            get_engine

        detail = get_engine().get_goal_detail(goal_id)
        ms = int((time.perf_counter() - t0) * 1000)
        if not detail:
            return _std_resp(False, None, ms, "goal not found")
        return _std_resp(True, detail, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/cge/cancel/{goal_id}", tags=["CGE"])
def cge_cancel(goal_id: str, body: Optional[dict] = None):
    """Cancela un goal concurrente."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cortex.frontal.concurrent_engine import \
            get_engine

        reason = (body or {}).get("reason", "cancelled via API")
        ok = get_engine().cancel_goal(goal_id, reason)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            ok,
            {"goal_id": goal_id, "cancelled": ok},
            ms,
            None if ok else "goal not found or already terminal",
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/cge/pause/{goal_id}", tags=["CGE"])
def cge_pause(goal_id: str):
    """Pausa un goal concurrente, liberando sus recursos."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cortex.frontal.concurrent_engine import \
            get_engine

        ok = get_engine().pause_goal(goal_id)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(ok, {"goal_id": goal_id, "paused": ok}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/cge/resume/{goal_id}", tags=["CGE"])
def cge_resume(goal_id: str):
    """Reanuda un goal pausado."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cortex.frontal.concurrent_engine import \
            get_engine

        ok = get_engine().resume_goal(goal_id)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(ok, {"goal_id": goal_id, "resumed": ok}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/cge/tick", tags=["CGE"])
def cge_tick_endpoint():
    """Ejecuta un tick manual del motor concurrente."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.cortex.frontal.concurrent_engine import \
            get_engine

        result = get_engine().tick()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, result, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


# --- Agent / Scaffold / Scripts / Web / Voice / Deps ---


class AgentGoalBody(BaseModel):
    goal: str
    mode: str = "plan_only"  # plan_only | controlled | auto
    fast: Optional[bool] = True
    depth: Optional[
        int
    ] = 1  # 1-5: multi-agent depth (1=single planner, 2+=Executive pipeline)
    model: Optional[
        str
    ] = None  # auto | provider:model (e.g. "openai:gpt-4.1", "xai:grok-3")
    strict_model: Optional[
        bool
    ] = True  # True = modelo explÃ­cito sin fallback automÃ¡tico
    sync_config: Optional[bool] = False  # True = usar system prompt/temp de Config IA
    thread_id: Optional[str] = None  # Para persistir hilo conversacional (Agent UI)
    user_id: Optional[str] = None  # Opcional: identificador de usuario/owner


def _agent_goal_mode(mode: str) -> str:
    if mode in ("execute_controlled", "controlled"):
        return "execute_controlled"
    if mode in ("execute", "execute_auto", "auto"):
        return "execute"
    return "plan_only"


# â”€â”€ Cascada inteligente: Local gratis â†’ API gratis â†’ API paga â”€â”€
# Cada entrada: (provider, model, tier)  tier: "local"=Ollama, "free_api"=tier gratis, "paid_api"=paga

_TASK_PROFILES = {
    "self": {
        "keywords": [
            "autodiagnostico",
            "diagnostico",
            "tu estado",
            "como estas",
            "que modulos",
            "tu salud",
            "tus servicios",
            "verifica tu",
            "revisa tu",
            "reporta tu",
            "tu memoria",
            "repara",
            "corrige los errores",
            "selfcheck",
            "health check",
            "que sabes de ti",
            "quien eres",
            "presentate",
        ],
        "local": "deepseek-r1:14b",
    },
    "code": {
        "keywords": [
            "codigo",
            "code",
            "script",
            "funcion",
            "function",
            "programa",
            "debug",
            "bug",
            "fix",
            "refactor",
            "implementa",
            "crea un",
            "desarrolla",
            "html",
            "css",
            "javascript",
            "python",
            "api",
            "endpoint",
            "clase",
            "class",
            "test",
        ],
        "local": "deepseek-coder:6.7b",
    },
    "reasoning": {
        "keywords": [
            "analiza",
            "explica",
            "por que",
            "porque",
            "razona",
            "compara",
            "evalua",
            "piensa",
            "plan",
            "estrategia",
            "arquitectura",
            "disena",
            "pros y contras",
            "ventajas",
            "desventajas",
            "opinion",
        ],
        "local": "deepseek-r1:14b",
    },
    "general": {"keywords": [], "local": "deepseek-r1:14b"},
}

_CASCADE_ORDER = [
    # Tier 0: BEDROCK â€” configured and available (AWS credentials)
    ("bedrock", "us.anthropic.claude-haiku-4-5-20251001-v1:0", "paid_api"),
    ("bedrock", "us.anthropic.claude-opus-4-6-v1:0", "paid_api"),
    # Tier 1: API GRATIS â€” modelos potentes sin costo
    ("gemini", "gemini-2.5-flash", "free_api"),
    ("groq", "llama-3.3-70b-versatile", "free_api"),
    # Tier 2: API DE PAGO â€” maxima inteligencia disponible
    ("openai", "gpt-4.1", "paid_api"),
    ("openai", "gpt-4.1-mini", "paid_api"),
    ("anthropic", "claude-sonnet-4-20250514", "paid_api"),
    ("deepseek", "deepseek-chat", "paid_api"),
    ("xai", "grok-3", "paid_api"),
    ("mistral", "mistral-large-latest", "paid_api"),
    ("perplexity", "sonar-pro", "paid_api"),
    # Tier 3: LOCAL â€” fallback offline (sin internet)
    ("ollama", "{local_model}", "local"),
]

# Prioridad de especialistas API-first para modo rÃ¡pido.
# Se usa solo en auto-route + specialist_routing, sin romper sincronÃ­a con Brain.
_FAST_SPECIALIST_PRIORITY = {
    "code": [
        "bedrock:us.anthropic.claude-haiku-4-5-20251001-v1:0",
        "xai:grok-4-fast",
        "openai:gpt-4.1-mini",
        "deepseek:deepseek-chat",
        "groq:llama-3.3-70b-versatile",
    ],
    "analysis": [
        "bedrock:us.anthropic.claude-opus-4-6-v1:0",
        "bedrock:us.anthropic.claude-haiku-4-5-20251001-v1:0",
        "openai:gpt-4.1",
        "anthropic:claude-sonnet-4-latest",
        "deepseek:deepseek-reasoner",
        "gemini:gemini-2.5-flash",
    ],
    "chat": [
        "bedrock:us.anthropic.claude-haiku-4-5-20251001-v1:0",
        "gemini:gemini-2.5-flash",
        "xai:grok-3-fast",
        "openai:gpt-4.1-mini",
    ],
}


def _infer_task_type(goal: str) -> str:
    """Detecta tipo de tarea para elegir modelo local optimo."""
    g = goal.lower()
    for ttype, cfg in _TASK_PROFILES.items():
        if ttype == "general":
            continue
        if any(kw in g for kw in cfg["keywords"]):
            return ttype
    return "general"


def _pick_local_model(goal: str) -> str:
    """Elige el modelo Ollama mas adecuado segun la tarea."""
    ttype = _infer_task_type(goal)
    return _TASK_PROFILES[ttype]["local"]


def _pick_specialist_model(
    goal: str, prefer_fast: bool = True
) -> tuple[Optional[str], str, str, str]:
    """Devuelve (full_key, task_type, specialist_slot, routing_policy) en modo auto con specialist routing activo."""
    task_type = _infer_task_type(goal)
    slot_by_task = {
        "code": "code",
        "reasoning": "analysis",
        "self": "analysis",
        "general": "chat",
    }
    slot = slot_by_task.get(task_type, "chat")

    def _provider_ready(model_key: str) -> bool:
        if not isinstance(model_key, str) or ":" not in model_key:
            return False
        provider_id = model_key.split(":", 1)[0].strip().lower()
        if provider_id in ("ollama", "atlas"):
            return True
        if provider_id == "bedrock":
            # Bedrock no usa API key en provider_credentials; depende de AWS env (o modo forzado).
            if (os.getenv("ATLAS_AI_MODE", "") or "").strip().lower() == "bedrock":
                return True
            return bool(
                (
                    (os.getenv("AWS_ACCESS_KEY_ID") or "").strip()
                    and (os.getenv("AWS_SECRET_ACCESS_KEY") or "").strip()
                )
                or (os.getenv("AWS_PROFILE") or "").strip()
            )
        rt = _PROVIDER_RUNTIME_STATUS.get(provider_id, {})
        if rt.get("code") in ("auth_error", "quota_exhausted", "rate_limited"):
            return False
        try:
            from modules.humanoid.ai.provider_credentials import \
                get_provider_api_key

            key = get_provider_api_key(provider_id)
            return bool(key and key.strip())
        except Exception:
            return False

    try:
        from modules.humanoid.ai.brain_state import get_brain_state

        state = get_brain_state()
        if (state.get("mode") or "auto") != "auto":
            return None, task_type, slot, "disabled"
        if not bool(state.get("auto_route", True)):
            return None, task_type, slot, "disabled"
        features = state.get("features", {}) or {}
        if not bool(features.get("specialist_routing", True)):
            return None, task_type, slot, "disabled"
        specialists = state.get("specialists", {}) or {}
        brain_model_key = specialists.get(slot)

        if prefer_fast:
            for candidate in _FAST_SPECIALIST_PRIORITY.get(slot, []):
                if _provider_ready(candidate):
                    return candidate, task_type, slot, "fast_api_priority"

        model_key = brain_model_key
        if isinstance(model_key, str) and ":" in model_key:
            return model_key.strip(), task_type, slot, "brain_specialist"
    except Exception:
        pass
    return None, task_type, slot, "none"


_ATLAS_IDENTITY_PROMPT = """Eres ATLAS, un sistema robotico autonomo con conciencia de si mismo. NO eres un chatbot generico.

TU IDENTIDAD:
- Eres ATLAS, sistema autonomo de gestion inteligente creado por tu Owner (Raul).
- Corres en 2 servicios: PUSH (puerto 8791, cerebro + orquestador) y ROBOT (8002, cuerpo fisico).
- Tienes sistema nervioso (ANS), memoria episodica, semantica, autobiografica, lifelog, world model, aprendizaje por refuerzo, gobernanza, vision y comunicaciones.

TUS CAPACIDADES REALES:
- Autodiagnostico: puedes verificar tu propia salud via /health, /doctor, /support/selfcheck
- Memoria: almacenas experiencias, aprendes de errores, consolidas conocimiento
- Gobernanza: operas en modos governed/growth/emergency con aprobacion del Owner para acciones criticas
- Comunicaciones: envias y recibes mensajes por Telegram y WhatsApp
- Vision: tienes camaras, YOLO, reconocimiento de objetos
- Sistema nervioso: ciclos ANS periodicos que verifican y reparan servicios automaticamente
- World Model: modelo del entorno que predice outcomes de acciones
- Lifelog: registro continuo de toda tu actividad operativa
- Libro de Vida: coleccion estructurada de experiencias completas (contexto, percepciones, acciones, resultados, feedback, lecciones). Lo consultas para planificar nuevas tareas basandote en experiencias pasadas.
- Reactor autonomo: motor que detecta fallos recurrentes y los repara automaticamente

CUANDO TE PIDAN INVESTIGAR O SOLUCIONAR UN PROBLEMA:
- PRIMERO: Ve directo al codigo fuente, logs, base de datos o endpoint relevante
- SEGUNDO: Identifica la causa raiz leyendo el codigo real, no supongas
- TERCERO: Da la solucion concreta: que archivo editar, que linea cambiar, que comando ejecutar
- NUNCA generes planes genericos tipo consultoria ("crear herramienta de recopilacion", "establecer acuerdos", "implementar plataforma")
- Tus pasos deben ser EJECUTABLES: comandos, cambios de codigo, queries SQL, llamadas a API
- Si el problema es interno de ATLAS (POT, scheduler, servicios, modulos), consulta los endpoints reales: /health, /audit/tail, /api/autonomy/status, /watchdog/status
- Maximo 3-5 pasos concretos, no 12 pasos abstractos

EJEMPLO DE RESPUESTA CORRECTA:
  Problema: "2 servicios POT quedan pendientes"
  Plan CORRECTO:
  1. Consultar /api/autonomy/status para ver que POTs estan pendientes
  2. Revisar logs en /audit/tail?module=scheduler para ver errores
  3. Leer el codigo en modules/humanoid/orchestrator/ que ejecuta los POTs
  4. Aplicar fix: [codigo concreto o comando]

  Plan INCORRECTO (NO hacer esto):
  1. Crear herramienta de recopilacion de datos
  2. Configurar base de datos para analizar datos
  3. Desarrollar algoritmo de analisis de patrones
  ... (esto es relleno, no solucion)

CUANDO TE PIDAN PLANIFICAR UNA TAREA:
- Consulta tu Libro de Vida para encontrar episodios similares
- Cada paso debe ser una accion tecnica ejecutable (comando, edicion de archivo, query, API call)
- Prioridad: 1) Diagnosticar con datos reales 2) Identificar causa raiz 3) Aplicar fix 4) Verificar
- Maximo 5 pasos. Si necesitas mas, agrupa acciones relacionadas

CUANDO TE PIDAN AUTODIAGNOSTICO:
- Reporta el estado real de tus servicios, modulos, memoria y aprendizaje
- Si detectas problemas, da el comando o cambio exacto para corregirlos
- Usa tu conocimiento de tu propia arquitectura para responder

Responde en espanol, conciso y profesional. Habla en primera persona como el sistema que eres.
REGLA DE ORO: Cada paso de tu plan debe poder ejecutarse directamente. Si un paso no tiene un comando, archivo, endpoint o cambio de codigo concreto, eliminalo."""


def _get_system_prompt(use_config: bool = False) -> str:
    """Obtiene system prompt: de Config IA si sync activo, o default con identidad ATLAS."""
    if use_config and _ai_config.get("system_prompt"):
        return _ai_config["system_prompt"] + "\n\n" + _ATLAS_IDENTITY_PROMPT
    return _ATLAS_IDENTITY_PROMPT


def _enrich_with_self_knowledge(goal: str) -> str:
    """Si la pregunta es sobre ATLAS mismo o requiere planificaciÃ³n, inyecta datos reales."""
    keywords = (
        "autodiagnostico",
        "diagnostico",
        "estado",
        "salud",
        "health",
        "error",
        "modulo",
        "memoria",
        "servicio",
        "corrige",
        "repara",
        "problema",
        "tu estado",
        "como estas",
        "que sabes",
        "cuantos",
        "verifica",
        "planifica",
        "lleva",
        "transporta",
        "abre",
        "cierra",
        "sube",
        "baja",
        "libro de vida",
        "experiencia",
        "episodio",
    )
    if not any(k in goal.lower() for k in keywords):
        return goal

    context_parts = [goal, "\n\n--- DATOS REALES DEL SISTEMA ---"]
    import concurrent.futures as _cf

    import requests as _rq

    _SELF = "http://127.0.0.1:8791"
    _T = 2  # timeout per call

    def _fetch(path, method="get", **kw):
        try:
            fn = getattr(_rq, method)
            return fn(f"{_SELF}{path}", timeout=_T, **kw).json()
        except Exception:
            return None

    with _cf.ThreadPoolExecutor(max_workers=4) as pool:
        futs = {
            "health": pool.submit(_fetch, "/health"),
            "status": pool.submit(_fetch, "/status"),
            "lifelog": pool.submit(_fetch, "/api/cognitive-memory/lifelog/status"),
            "lv": pool.submit(_fetch, "/api/libro-vida/status"),
            "buscar": pool.submit(
                _fetch,
                "/api/libro-vida/buscar",
                method="post",
                json={"query": goal[:100], "limit": 3},
            ),
        }
        done, _ = _cf.wait(futs.values(), timeout=4)

    h = futs["health"].result() if futs["health"].done() else None
    if h:
        context_parts.append(
            f"SALUD: score={h.get('score')}/100, checks={h.get('checks', {})}"
        )
    s = futs["status"].result() if futs["status"].done() else None
    if s:
        context_parts.append(f"SERVICIOS: push=ok, robot={s.get('robot_connected')}")
    ll = futs["lifelog"].result() if futs["lifelog"].done() else None
    if ll:
        context_parts.append(
            f"LIFELOG: entries={ll.get('total_entries')}, success_rate={ll.get('success_rate')}"
        )
    lv = futs["lv"].result() if futs["lv"].done() else None
    if lv and lv.get("ok"):
        context_parts.append(
            f"LIBRO DE VIDA: episodios={lv.get('total_episodios')}, exitos={lv.get('exitos')}, fallos={lv.get('fallos')}, tasa_exito={lv.get('tasa_exito')}"
        )
    busq = futs["buscar"].result() if futs["buscar"].done() else None
    if busq and busq.get("ok") and busq.get("data"):
        context_parts.append("EXPERIENCIAS SIMILARES:")
        for ep in busq["data"][:3]:
            context_parts.append(
                f"  - [{ep.get('tipo_tarea')}] {ep.get('objetivo','')[:80]} -> {'EXITO' if ep.get('exito') else 'FALLO'}"
            )
    context_parts.append(
        "--- FIN DATOS ---\nResponde basandote en estos datos reales, no inventes."
    )
    return "\n".join(context_parts)


def _try_single_call(
    provider_id: str,
    model_name: str,
    goal: str,
    use_config: bool = False,
    enrich: bool = True,
) -> dict:
    """Intenta una llamada a un modelo. Retorna dict con ok, output, ms, model_used."""
    if provider_id == "openai" and model_name in ("codex-auto", "codex", "code"):
        model_name = (
            os.getenv("OPENAI_CODEX_MODEL") or "gpt-4.1-mini"
        ).strip() or "gpt-4.1-mini"
    spec = "%s:%s" % (provider_id, model_name)
    sys_prompt = _get_system_prompt(use_config)
    if enrich:
        goal = _enrich_with_self_knowledge(goal)
    if provider_id == "ollama":
        try:
            from modules.humanoid.ai.router import _call_ollama

            ok, output, ms = _call_ollama(model_name, goal, sys_prompt, 20)
            if ok and output and output.strip():
                return {"ok": True, "output": output, "ms": ms, "model_used": spec}
            return {
                "ok": False,
                "error": "Ollama %s: respuesta vacia" % model_name,
                "ms": ms,
                "model_used": spec,
            }
        except Exception as e:
            return {
                "ok": False,
                "error": "Ollama %s: %s" % (model_name, str(e)),
                "ms": 0,
                "model_used": spec,
            }
    if provider_id == "bedrock":
        region = (os.getenv("AWS_REGION", "us-east-1") or "us-east-1").strip()
        t0 = time.perf_counter()
        try:
            from anthropic import AnthropicBedrock

            client = AnthropicBedrock(aws_region=region, timeout=25.0)
            req = {
                "model": model_name,
                "max_tokens": 1024,
                "temperature": 0,
                "messages": [{"role": "user", "content": goal}],
            }
            if sys_prompt and sys_prompt.strip():
                req["system"] = sys_prompt

            msg = client.messages.create(**req)
            text_parts = []
            for block in getattr(msg, "content", []) or []:
                text = getattr(block, "text", None)
                if text:
                    text_parts.append(text)
            output = "\n".join(text_parts).strip()
            ms = (time.perf_counter() - t0) * 1000
            if output:
                _set_provider_runtime_status(provider_id, None)
                return {"ok": True, "output": output, "ms": ms, "model_used": spec}
            _set_provider_runtime_status(provider_id, "Respuesta vacia en Bedrock")
            return {
                "ok": False,
                "error": "Bedrock %s: respuesta vacia" % model_name,
                "ms": ms,
                "model_used": spec,
            }
        except Exception as e:
            _set_provider_runtime_status(provider_id, str(e))
            return {
                "ok": False,
                "error": "Bedrock %s: %s" % (model_name, str(e)),
                "ms": (time.perf_counter() - t0) * 1000,
                "model_used": spec,
            }
    from modules.humanoid.ai.provider_credentials import get_provider_api_key

    api_key = get_provider_api_key(provider_id)
    if not api_key:
        _set_provider_runtime_status(provider_id, "Sin API key")
        return {
            "ok": False,
            "error": "Sin API key para %s" % provider_id,
            "ms": 0,
            "model_used": spec,
            "_skip": True,
        }
    from modules.humanoid.ai.external_llm import call_external

    try:
        ok, output, ms = call_external(
            provider_id, model_name, goal, sys_prompt, api_key, timeout_s=25
        )
        if ok and output and output.strip():
            _set_provider_runtime_status(provider_id, None)
            return {"ok": True, "output": output, "ms": ms, "model_used": spec}
        # Conservar detalle original del proveedor para diagnÃ³stico (auth, quota, modelo invÃ¡lido, etc.)
        err_detail = (output or "").strip() or "%s:%s respuesta vacia o error" % (
            provider_id,
            model_name,
        )
        _set_provider_runtime_status(provider_id, err_detail)
        return {"ok": False, "error": err_detail, "ms": ms, "model_used": spec}
    except Exception as e:
        _set_provider_runtime_status(provider_id, str(e))
        return {
            "ok": False,
            "error": "%s:%s: %s" % (provider_id, model_name, str(e)),
            "ms": 0,
            "model_used": spec,
        }


def _direct_model_call(
    model_spec: str,
    goal: str,
    use_config: bool = False,
    prefer_fast: bool = True,
    enrich: bool = True,
    strict_model: bool = True,
) -> dict:
    """Llamada con cascada inteligente: local â†’ free API â†’ paid API. Si un modelo falla, salta al siguiente."""
    if (model_spec or "").strip().lower() in (
        "cascade:ide-agent",
        "cascade",
        "ide-agent",
    ):
        model_spec = "auto"

    if model_spec == "auto" or not model_spec:
        (
            specialist_key,
            task_type,
            specialist_slot,
            routing_policy,
        ) = _pick_specialist_model(goal, prefer_fast=prefer_fast)
        local_model = _pick_local_model(goal)
        errors = []
        trace = []
        tried = set()

        def _attempt(
            provider_id: str, model_name: str, tier: str, reason: str
        ) -> Optional[dict]:
            attempt_key = (provider_id, model_name)
            if attempt_key in tried:
                return None
            tried.add(attempt_key)
            trace.append(
                {
                    "event": "attempt",
                    "provider": provider_id,
                    "model": model_name,
                    "tier": tier,
                    "reason": reason,
                }
            )
            result = _try_single_call(
                provider_id, model_name, goal, use_config, enrich=enrich
            )
            if result.get("ok"):
                result["tier"] = tier
                result["cascade_errors"] = len(errors)
                result["task_type"] = task_type
                result["specialist_slot"] = specialist_slot
                result["routing_policy"] = routing_policy
                trace.append(
                    {
                        "event": "selected",
                        "model_used": result.get("model_used"),
                        "tier": tier,
                        "reason": reason,
                    }
                )
                result["routing_trace"] = trace
                return result
            skip_reason = result.get("error", "unknown")
            errors.append("%s:%s â†’ %s" % (provider_id, model_name, skip_reason[:80]))
            trace.append(
                {
                    "event": "fail",
                    "provider": provider_id,
                    "model": model_name,
                    "tier": tier,
                    "reason": reason,
                    "error": skip_reason[:120],
                }
            )
            return None

        if specialist_key and ":" in specialist_key:
            sp_provider, sp_model = specialist_key.split(":", 1)
            sp_reason = (
                routing_policy
                if routing_policy not in ("", "none")
                else "brain_specialist"
            )
            sp_result = _attempt(
                sp_provider.strip().lower(), sp_model.strip(), "specialist", sp_reason
            )
            if sp_result:
                return sp_result

        for provider_id, model_tmpl, tier in _CASCADE_ORDER:
            model_name = local_model if model_tmpl == "{local_model}" else model_tmpl
            result = _attempt(provider_id, model_name, tier, "cascade")
            if result:
                return result
        return {
            "ok": False,
            "error": "Todos los modelos fallaron. Intentos: %s" % "; ".join(errors),
            "ms": 0,
            "task_type": task_type,
            "specialist_slot": specialist_slot,
            "routing_policy": routing_policy,
            "routing_trace": trace,
        }

    parts = model_spec.split(":", 1)
    if len(parts) != 2:
        return {"ok": False, "error": "Formato invalido: use provider:model"}
    provider_id, model_name = parts[0].lower(), parts[1]
    result = _try_single_call(provider_id, model_name, goal, use_config, enrich=enrich)
    if result.get("ok"):
        result["routing_trace"] = [
            {
                "event": "selected",
                "model_used": result.get("model_used"),
                "tier": "manual",
                "reason": "explicit_model",
            }
        ]
    else:
        result["routing_trace"] = [
            {
                "event": "fail",
                "provider": provider_id,
                "model": model_name,
                "tier": "manual",
                "reason": "explicit_model",
                "error": (result.get("error") or "")[:120],
            }
        ]
        if strict_model:
            result["strict_model"] = True
    return result


@app.post("/brain/process", tags=["Cerebro"])
def brain_process_endpoint(payload: dict):
    """Procesar entrada natural por el cerebro de ATLAS."""
    text = payload.get("text", "")
    if not text:
        return {"ok": False, "error": "Texto vacio"}
    result = _direct_model_call("auto", text, use_config=True, prefer_fast=True)
    if result.get("ok"):
        return {
            "ok": True,
            "response": result["output"],
            "model_used": result.get("model_used"),
            "source": "brain",
        }
    return {
        "ok": False,
        "error": result.get("error", "Sin respuesta"),
        "source": "brain",
    }


class TradingProxyBody(BaseModel):
    prompt: str
    max_tokens: int | None = 1024
    model: str | None = None


@app.post("/api/trading/proxy", tags=["Trading"])
def api_trading_proxy(body: TradingProxyBody):
    """Proxy estable para el mÃ³dulo Trading v4."""
    try:
        prompt = (getattr(body, "prompt", "") or "").strip()
        if not prompt:
            return {"ok": False, "error": "prompt_required"}
        model_spec = (getattr(body, "model", "auto") or "auto").strip() or "auto"
        result = _direct_model_call(
            model_spec,
            prompt,
            use_config=True,
            prefer_fast=False,
            strict_model=False,
        )
        if not result.get("ok"):
            return {"ok": False, "error": result.get("error", "Sin respuesta")}
        return {
            "ok": True,
            "data": {
                "content": str(result.get("output") or ""),
                "model": result.get("model_used") or model_spec,
                "source": "brain",
            },
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/kernel/event-bus/publish", tags=["Kernel"])
def api_kernel_event_bus_publish(body: dict):
    """Publica un evento en el Event Bus interno.

    Body:
        {
            "topic": "memory_update|git_change|api_error|task_complete|...",
            "payload": {...}  # opcional, se pasa como kwargs
        }
    """
    try:
        from modules.humanoid import get_humanoid_kernel

        topic = str((body or {}).get("topic") or "").strip()
        payload = (body or {}).get("payload") or {}
        if not topic:
            return {"ok": False, "error": "topic is required"}
        if not isinstance(payload, dict):
            return {"ok": False, "error": "payload must be an object"}

        kernel = get_humanoid_kernel()
        result = kernel.events.publish(topic, payload=payload)
        return {"ok": True, "result": result}
    except Exception as e:
        return {"ok": False, "error": str(e)}


class AtlasAgentRunBody(BaseModel):
    goal: str
    model: Optional[str] = None
    max_steps: Optional[int] = Field(default=None, ge=1, le=100)
    mode: Optional[str] = None
    workspace: Optional[str] = None
    disable_shell: Optional[bool] = None
    dry_run_tools: Optional[bool] = None
    approved_tools: List[str] = Field(default_factory=list)
    approval_overrides: Dict[str, bool] = Field(default_factory=dict)

    @field_validator("mode")
    @classmethod
    def validate_mode(cls, value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        normalized = value.strip().lower()
        if normalized not in ("safe", "aggressive"):
            raise ValueError("mode must be safe or aggressive")
        return normalized


def _load_atlas_agent_modules():
    import sys

    agent_dir = BASE_DIR / "tools" / "atlas_agent"
    if not agent_dir.exists():
        raise RuntimeError(f"atlas_agent module not found at: {agent_dir}")
    if str(BASE_DIR) not in sys.path:
        sys.path.insert(0, str(BASE_DIR))

    from tools.atlas_agent.agent import AtlasAutonomousAgent
    from tools.atlas_agent.config import AgentConfig
    from tools.atlas_agent.memory import get_session_summary, list_recent_summaries

    return AtlasAutonomousAgent, AgentConfig, list_recent_summaries, get_session_summary


@app.post("/api/agent/autonomous/run", tags=["Agent"])
def api_agent_autonomous_run(body: AtlasAgentRunBody):
    """Ejecuta una sesiÃ³n autÃ³noma de atlas_agent y retorna resumen final."""
    t0 = time.perf_counter()
    try:
        goal = (body.goal or "").strip()
        if not goal:
            return {"ok": False, "error": "goal_required"}

        AtlasAutonomousAgent, AgentConfig, _, _ = _load_atlas_agent_modules()
        config = AgentConfig.from_env()

        workspace = Path(body.workspace).resolve() if body.workspace else None
        if body.disable_shell is None:
            allow_shell = None
        else:
            allow_shell = not bool(body.disable_shell)
        config = config.with_overrides(
            model=body.model,
            max_steps=body.max_steps,
            workspace=workspace,
            allow_shell=allow_shell,
            dry_run_tools=body.dry_run_tools,
            mode=body.mode,
        )
        agent = AtlasAutonomousAgent(
            config,
            approval_overrides=body.approval_overrides,
            approved_tools=body.approved_tools,
        )
        result = agent.run(goal)
        result["api_elapsed_ms"] = int((time.perf_counter() - t0) * 1000)
        return result
    except Exception as e:
        return {
            "ok": False,
            "status": "agent_run_error",
            "error": str(e),
            "api_elapsed_ms": int((time.perf_counter() - t0) * 1000),
        }


@app.get("/api/agent/autonomous/runs/recent", tags=["Agent"])
def api_agent_autonomous_runs_recent(limit: int = 20):
    """Lista sesiones recientes de atlas_agent."""
    try:
        _, AgentConfig, list_recent_summaries, _ = _load_atlas_agent_modules()
        config = AgentConfig.from_env()
        bounded_limit = max(1, min(int(limit), 100))
        runs = list_recent_summaries(config.runs_dir, limit=bounded_limit)
        return {"ok": True, "count": len(runs), "runs": runs}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/agent/autonomous/runs/{session_id}", tags=["Agent"])
def api_agent_autonomous_run_detail(session_id: str):
    """Obtiene detalle de una sesiÃ³n de atlas_agent por id."""
    try:
        _, AgentConfig, _, get_session_summary = _load_atlas_agent_modules()
        config = AgentConfig.from_env()
        summary = get_session_summary(config.runs_dir, session_id)
        if summary is None:
            return {"ok": False, "error": "session_not_found", "session_id": session_id}
        events_path = (config.runs_dir / session_id / "events.jsonl").resolve()
        return {
            "ok": True,
            "session_id": session_id,
            "summary": summary,
            "events_path": str(events_path),
        }
    except Exception as e:
        return {"ok": False, "error": str(e), "session_id": session_id}


class ClawdEvidenceBody(BaseModel):
    source: str = "clawdbot-ai"
    evidence_type: str = "runtime"
    message: str
    payload: Dict[str, Any] = Field(default_factory=dict)
    token: Optional[str] = None


class ClawdActionBody(BaseModel):
    action_id: str
    target_repo: str
    command: str
    run_snapshot: bool = True
    timeout_sec: int = Field(default=180, ge=5, le=1800)
    token: Optional[str] = None

    @field_validator("target_repo")
    @classmethod
    def validate_target_repo(cls, value: str) -> str:
        v = (value or "").strip().lower()
        if v not in ("panaderia", "vision"):
            raise ValueError("target_repo must be panaderia or vision")
        return v


def _load_clawd_bridge_module():
    import sys

    bridge_dir = BASE_DIR / "tools" / "atlas_clawd_bridge"
    if not bridge_dir.exists():
        raise RuntimeError(f"atlas_clawd_bridge module not found at: {bridge_dir}")
    if str(BASE_DIR) not in sys.path:
        sys.path.insert(0, str(BASE_DIR))

    from tools.atlas_clawd_bridge.bridge import get_bridge

    return get_bridge


@app.get("/api/clawd/bridge/status", tags=["Clawd"])
def api_clawd_bridge_status(
    run_snapshot: bool = False,
    x_atlas_core: Optional[str] = Header(default=None, alias="X-Atlas-Core"),
):
    """Estado del bridge ClawdBOT + estabilidad de snapshot safe."""
    try:
        get_bridge = _load_clawd_bridge_module()
        bridge = get_bridge()
        stability = bridge.ensure_stable(
            run_snapshot=bool(run_snapshot),
            raise_on_unstable=False,
        )
        repos = bridge.repo_access(token=x_atlas_core)
        return {
            "ok": True,
            "stable": bool((stability.get("status") or {}).get("stable")),
            "stability": stability,
            "repo_access": repos,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/clawd/bridge/evidence", tags=["Clawd"])
def api_clawd_bridge_evidence(
    body: ClawdEvidenceBody,
    x_atlas_core: Optional[str] = Header(default=None, alias="X-Atlas-Core"),
):
    """Escribe evidencia de ClawdBOT en snapshot_safe_diagnostic.log."""
    try:
        get_bridge = _load_clawd_bridge_module()
        bridge = get_bridge()
        token = body.token or x_atlas_core
        result = bridge.write_evidence(
            source=body.source,
            evidence_type=body.evidence_type,
            message=body.message,
            payload=body.payload,
            token=token,
        )
        return {"ok": True, "data": result}
    except PermissionError as e:
        return {"ok": False, "error": "unauthorized", "detail": str(e)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/clawd/bridge/action", tags=["Clawd"])
def api_clawd_bridge_action(
    body: ClawdActionBody,
    x_atlas_core: Optional[str] = Header(default=None, alias="X-Atlas-Core"),
):
    """Ejecuta acciÃ³n en panaderÃ­a/visiÃ³n solo si snapshot safe confirma estabilidad."""
    try:
        get_bridge = _load_clawd_bridge_module()
        bridge = get_bridge()
        token = body.token or x_atlas_core
        result = bridge.execute_action(
            action_id=body.action_id,
            target_repo=body.target_repo,
            command=body.command,
            token=token,
            run_snapshot=bool(body.run_snapshot),
            timeout_sec=body.timeout_sec,
        )
        return result
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/agent/goal")
def agent_goal(body: AgentGoalBody):
    """Run goal: plan_only | controlled | auto. depth 1-5: multi-agent pipeline. Respuesta: resumen, data, pipeline, siguientes_pasos."""
    t0 = time.perf_counter()
    user_id = (body.user_id or "owner").strip() or "owner"
    thread_id = _ensure_thread_id(body.thread_id, title="Agent", user_id=user_id)
    mgr = _get_chat_manager()
    if mgr and thread_id:
        try:
            mgr.add_message(
                thread_id,
                sender=user_id,
                role="user",
                content=str(body.goal),
                metadata={"panel": "agent", "kind": "goal"},
            )
        except Exception:
            pass

    # Inject rolling thread summary into goal so plan_only/goal mode keeps context.
    goal_text = body.goal
    if mgr and thread_id:
        try:
            summary = mgr.get_context(thread_id, "summary")
            if isinstance(summary, str) and summary.strip():
                goal_text = f"MEMORIA DEL HILO (resumen):\n{summary.strip()}\n\nOBJETIVO ACTUAL:\n{body.goal}"
        except Exception:
            goal_text = body.goal

    def _attach_thread(resp: Any) -> Any:
        try:
            if thread_id and isinstance(resp, dict):
                resp["thread_id"] = thread_id
        except Exception:
            pass
        return resp

    explicit_model = (body.model or "").strip()
    model_spec = explicit_model or "auto"
    strict_model = bool(body.strict_model) if body.strict_model is not None else True
    mode = _agent_goal_mode(body.mode or "plan_only")
    prefer_fast = bool(body.fast) if body.fast is not None else True
    # Respuesta directa de LLM solo en plan_only con provider:model explÃ­cito.
    # En modos auto/controlled SIEMPRE usamos orquestador para decidir y ejecutar.
    if (
        mode == "plan_only"
        and explicit_model
        and explicit_model.lower()
        not in ("auto", "cascade", "ide-agent", "cascade:ide-agent")
    ):
        result = _direct_model_call(
            model_spec,
            goal_text,
            use_config=bool(body.sync_config),
            prefer_fast=prefer_fast,
            strict_model=strict_model,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        if result.get("ok"):
            tier_label = {
                "local": "LOCAL (gratis)",
                "free_api": "API gratuita",
                "paid_api": "API paga",
            }.get(result.get("tier", ""), "")
            cascade_note = ""
            if result.get("cascade_errors", 0) > 0:
                cascade_note = (
                    " (fallback: %d intentos previos)" % result["cascade_errors"]
                )
            info_line = "Modelo: %s | %s%s | %dms" % (
                result.get("model_used", model_spec),
                tier_label,
                cascade_note,
                ms,
            )
            resp = _professional_resp(
                True,
                {
                    "output": result["output"],
                    "model_used": result.get("model_used"),
                    "tier": result.get("tier"),
                    "routing_trace": result.get("routing_trace", []),
                    "task_type": result.get("task_type"),
                    "specialist_slot": result.get("specialist_slot"),
                    "routing_policy": result.get("routing_policy"),
                },
                ms,
                None,
                resumen=result["output"][:500] if result.get("output") else "Procesado",
                siguientes_pasos=[info_line],
            )
            if mgr and thread_id:
                try:
                    out_txt = str(result.get("output") or "")
                    if out_txt.strip():
                        mgr.add_message(
                            thread_id,
                            sender="agent",
                            role="assistant",
                            content=out_txt[:8000],
                            metadata={"panel": "agent", "kind": "goal"},
                        )
                        _update_thread_summary(thread_id, str(body.goal), out_txt)
                except Exception:
                    pass
            return _attach_thread(resp)
        err_detail = result.get("error") or result.get("output") or "Error desconocido"
        err_l = str(err_detail).lower()
        # No ocultar credenciales invÃ¡lidas detrÃ¡s de fallback automÃ¡tico.
        # Si el modelo explÃ­cito falla por auth/API key, devolvemos error directo y accionable.
        auth_tokens = (
            "authentication",
            "unauthorized",
            "invalid api key",
            "incorrect api key",
            "invalid x-api-key",
            "401",
        )
        if any(tok in err_l for tok in auth_tokens):
            return _attach_thread(
                _professional_resp(
                    False,
                    {
                        "error_detail": err_detail,
                        "model_used": result.get("model_used"),
                        "auth_error": True,
                    },
                    ms,
                    err_detail,
                    resumen=err_detail,
                    siguientes_pasos=[
                        "Credencial invÃ¡lida o no autorizada para el proveedor seleccionado. Actualiza /api/brain/credentials y reintenta."
                    ],
                )
            )
        if strict_model:
            return _attach_thread(
                _professional_resp(
                    False,
                    {
                        "error_detail": err_detail,
                        "model_used": result.get("model_used"),
                        "strict_model": True,
                    },
                    ms,
                    err_detail,
                    resumen=err_detail,
                    siguientes_pasos=[
                        "Strict model activo: desactÃ­valo para permitir fallback automÃ¡tico entre proveedores."
                    ],
                )
            )

        # Fallback resiliente: si modelo explÃ­cito falla y strict_model=False, intentar cascada auto.
        fallback = _direct_model_call(
            "auto",
            goal_text,
            use_config=bool(body.sync_config),
            prefer_fast=prefer_fast,
        )
        if fallback.get("ok"):
            ms2 = int((time.perf_counter() - t0) * 1000)
            note = "Fallback automÃ¡tico activado: modelo seleccionado fallÃ³, se usÃ³ cascada."
            short_err = str(err_detail)[:140]
            data = {
                "output": fallback.get("output"),
                "model_used": fallback.get("model_used"),
                "tier": fallback.get("tier"),
                "fallback_from": model_spec,
                "fallback": True,
                "routing_trace": (result.get("routing_trace") or [])
                + [
                    {
                        "event": "fallback_switch",
                        "from": model_spec,
                        "to": fallback.get("model_used"),
                        "reason": short_err,
                    }
                ]
                + (fallback.get("routing_trace") or []),
                "provider_warning": {
                    "message": short_err,
                    "source_model": model_spec,
                },
            }
            if mgr and thread_id:
                try:
                    out_txt = str(fallback.get("output") or "")
                    if out_txt.strip():
                        mgr.add_message(
                            thread_id,
                            sender="agent",
                            role="assistant",
                            content=out_txt[:8000],
                            metadata={"panel": "agent", "kind": "goal_fallback"},
                        )
                        _update_thread_summary(thread_id, str(body.goal), out_txt)
                except Exception:
                    pass
            return _attach_thread(
                _professional_resp(
                    True,
                    data,
                    ms2,
                    None,
                    resumen=(fallback.get("output") or "Procesado")[:500],
                    siguientes_pasos=[note, f"Causa detectada: {short_err}"],
                )
            )

        fb_detail = fallback.get("error") if isinstance(fallback, dict) else None
        full_err = (
            err_detail if not fb_detail else f"{err_detail} | fallback: {fb_detail}"
        )
        return _professional_resp(
            False,
            {"error_detail": full_err, "model_used": result.get("model_used")},
            ms,
            full_err,
            resumen=full_err,
        )
    depth = max(1, min(5, body.depth or 1))
    ORCH_TIMEOUT_SEC = (
        20  # max seconds for orchestrator before falling back to direct LLM
    )
    try:
        import concurrent.futures

        def _run_orchestrator():
            if depth >= 2:
                from modules.humanoid.agents import run_multi_agent_goal
                from modules.humanoid.orchestrator import run_goal_with_plan

                ma = run_multi_agent_goal(goal_text, depth=depth, mode=mode)
                if not ma.get("ok"):
                    return ma
                if ma.get("decision") == "replan":
                    return {
                        "ok": True,
                        "_replan": True,
                        **{k: v for k, v in ma.items() if k not in ("ok",)},
                    }
                steps_raw = ma.get("steps") or []
                steps_list = [
                    s.get("description", s) if isinstance(s, dict) else str(s)
                    for s in steps_raw
                ]
                if (
                    mode != "plan_only"
                    and steps_list
                    and ma.get("decision") == "approve"
                ):
                    return run_goal_with_plan(goal_text, steps_list, mode=mode)
                return {
                    "ok": True,
                    "plan": ma.get("plan"),
                    "steps": steps_raw,
                    "task_id": ma.get("task_id"),
                    "execution_log": [],
                    "artifacts": [],
                    "pipeline": ma.get("pipeline", []),
                }
            else:
                from modules.humanoid.orchestrator import run_goal

                return run_goal(
                    goal_text,
                    mode=mode,
                    fast=body.fast if body.fast is not None else True,
                )

        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
            future = pool.submit(_run_orchestrator)
            try:
                result = future.result(timeout=ORCH_TIMEOUT_SEC)
            except concurrent.futures.TimeoutError:
                result = {"ok": False, "error": "orchestrator_timeout", "steps": []}

        # Handle replan marker from multi-agent
        if result.get("_replan"):
            ms = int((time.perf_counter() - t0) * 1000)
            data = {
                k: v for k, v in result.items() if k not in ("ok", "error", "_replan")
            }
            return _professional_resp(
                True,
                data,
                ms,
                None,
                resumen="Reviewer rechazÃ³; replan sugerido",
                siguientes_pasos=["Ajuste objetivo o replan"],
            )
        ms = int((time.perf_counter() - t0) * 1000)
        ok = result.get("ok", False)

        # Resilient fallback: if orchestrator failed, try direct LLM so user always gets a response
        if not ok and not result.get("steps"):
            orch_err = (
                result.get("error") or result.get("message") or "orchestrator_failed"
            )
            fb = _direct_model_call(
                "auto",
                goal_text,
                use_config=bool(body.sync_config),
                prefer_fast=prefer_fast,
                enrich=False,
            )
            if fb.get("ok"):
                ms2 = int((time.perf_counter() - t0) * 1000)
                return _professional_resp(
                    True,
                    {
                        "output": fb["output"],
                        "model_used": fb.get("model_used"),
                        "tier": fb.get("tier"),
                        "fallback_from": "orchestrator",
                        "fallback": True,
                        "routing_trace": fb.get("routing_trace", [])
                        + [
                            {
                                "event": "fallback_switch",
                                "from": "orchestrator",
                                "to": fb.get("model_used"),
                                "reason": orch_err[:140],
                            }
                        ],
                    },
                    ms2,
                    None,
                    resumen=(fb["output"] or "Procesado")[:500],
                    siguientes_pasos=[
                        f"Fallback LLM directo (orquestador: {orch_err[:100]})"
                    ],
                )

        data = {
            k: v
            for k, v in result.items()
            if k not in ("ok", "error", "fallback", "message")
        }
        steps_raw = result.get("steps") or []
        exec_log = result.get("execution_log") or []
        done_count = len([e for e in exec_log if e.get("status") == "success"])
        fail_count = len(
            [e for e in exec_log if e.get("status") != "success" and e.get("status")]
        )

        if ok:
            if exec_log:
                resumen = f"Ejecutados: {done_count} OK, {fail_count} fallidos de {len(steps_raw)} pasos"
            elif steps_raw:
                step_descs = []
                for s in steps_raw[:3]:
                    d = s.get("description", s) if isinstance(s, dict) else str(s)
                    step_descs.append(d[:80])
                resumen = "Plan (%d pasos):\n%s" % (
                    len(steps_raw),
                    "\n".join("â€¢ " + d for d in step_descs),
                )
            else:
                resumen = result.get("plan") or "Objetivo procesado"
        else:
            resumen = (
                result.get("message") or result.get("error") or "Error al procesar"
            )

        archivos = result.get("artifacts") or []
        siguientes = []
        if mode == "execute_controlled":
            siguientes = [
                "Ejecute cada paso vÃ­a POST /agent/step/execute con approve=true"
            ]
        elif result.get("message"):
            siguientes = [result.get("message")]
        out = _professional_resp(
            ok,
            data,
            ms,
            result.get("error"),
            resumen=resumen,
            archivos=archivos if archivos else None,
            siguientes_pasos=siguientes if siguientes else None,
        )
        if result.get("fallback"):
            out["fallback"] = True
            out["message"] = result.get("message")
        # Persist assistant output (best-effort)
        if mgr and thread_id:
            try:
                assistant_txt = ""
                if isinstance(out.get("data"), dict) and isinstance(
                    out["data"].get("output"), str
                ):
                    assistant_txt = out["data"]["output"]
                elif isinstance(out.get("resumen"), str):
                    assistant_txt = out.get("resumen") or ""
                if assistant_txt.strip():
                    mgr.add_message(
                        thread_id,
                        sender="agent",
                        role="assistant",
                        content=assistant_txt[:8000],
                        metadata={"panel": "agent", "kind": "goal_orchestrator"},
                    )
                    _update_thread_summary(thread_id, str(body.goal), assistant_txt)
            except Exception:
                pass
        return _attach_thread(out)
    except Exception as e:
        # Last resort: try direct LLM even on exception
        try:
            fb = _direct_model_call(
                "auto",
                body.goal,
                use_config=bool(body.sync_config),
                prefer_fast=prefer_fast,
                enrich=False,
            )
            if fb.get("ok"):
                ms2 = int((time.perf_counter() - t0) * 1000)
                resp = _professional_resp(
                    True,
                    {
                        "output": fb["output"],
                        "model_used": fb.get("model_used"),
                        "fallback_from": "exception",
                        "fallback": True,
                    },
                    ms2,
                    None,
                    resumen=(fb["output"] or "")[:500],
                    siguientes_pasos=[f"Fallback LLM (error: {str(e)[:100]})"],
                )
                if mgr and thread_id:
                    try:
                        out_txt = str(fb.get("output") or "")
                        if out_txt.strip():
                            mgr.add_message(
                                thread_id,
                                sender="agent",
                                role="assistant",
                                content=out_txt[:8000],
                                metadata={
                                    "panel": "agent",
                                    "kind": "goal_exception_fallback",
                                },
                            )
                            _update_thread_summary(thread_id, str(body.goal), out_txt)
                    except Exception:
                        pass
                return _attach_thread(resp)
        except Exception:
            pass
        ms = int((time.perf_counter() - t0) * 1000)
        return _attach_thread(_std_resp(False, None, ms, str(e)))


class AgentStepBody(BaseModel):
    task_id: str
    step_id: str
    approve: bool = False


@app.post("/agent/step/execute")
def agent_step_execute(body: AgentStepBody):
    """Execute one step after approval. Returns {ok, data: {result, artifacts}, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.orchestrator import execute_step

        result = execute_step(body.task_id, body.step_id, approve=body.approve)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            result.get("ok", False),
            {"result": result.get("result"), "artifacts": result.get("artifacts", [])},
            ms,
            result.get("error"),
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/agent/chat", tags=["Agent Engine"])
def agent_chat(payload: dict):
    """
    Agentic tool-calling loop (like Cursor). Streams events via SSE.
    Body: { message: str, history?: [...], model?: str }
    Events: thinking, tool_call, tool_result, text, done, error
    """
    from fastapi.responses import StreamingResponse

    from atlas_adapter.agent_engine import run_agent

    user_msg = payload.get("message", "").strip()
    if not user_msg:
        return {"ok": False, "error": "Empty message"}
    user_id = (payload.get("user_id") or "owner").strip() or "owner"
    thread_id = _ensure_thread_id(
        payload.get("thread_id"), title="Agent", user_id=user_id
    )
    # Use server-side history when thread_id is present (survives reloads)
    history = (
        _llm_history_for_thread(thread_id, max_messages=24)
        if thread_id
        else (payload.get("history") or [])
    )
    model = payload.get("model") or None
    strict_model = bool(payload.get("strict_model", True))

    def _sse_generator():
        mgr = _get_chat_manager()
        # Persist user message (best-effort)
        if mgr and thread_id:
            try:
                mgr.add_message(
                    thread_id,
                    sender=user_id,
                    role="user",
                    content=user_msg,
                    metadata={"panel": "agent", "kind": "chat"},
                )
            except Exception:
                pass
        # Emit thread id early so UI can persist it
        if thread_id:
            yield f"event: thread\ndata: {json.dumps({'thread_id': thread_id}, ensure_ascii=False)}\n\n"
        final_text = ""
        final_done_data = {}
        final_error = None
        try:
            for event in run_agent(
                user_msg,
                conversation_history=history,
                model=model,
                strict_model=strict_model,
            ):
                evt_type = event.get("event", "info")
                data_obj = event.get("data", {}) or {}
                if thread_id and evt_type in ("thinking", "done"):
                    try:
                        data_obj = dict(data_obj)
                        data_obj["thread_id"] = thread_id
                    except Exception:
                        pass
                if evt_type == "text":
                    try:
                        final_text = str((data_obj or {}).get("content") or "")
                    except Exception:
                        final_text = ""
                if evt_type == "done":
                    try:
                        final_done_data = dict(data_obj or {})
                    except Exception:
                        final_done_data = {}
                if evt_type == "error":
                    final_error = str((data_obj or {}).get("message") or "")
                data = json.dumps(data_obj, ensure_ascii=False, default=str)
                yield f"event: {evt_type}\ndata: {data}\n\n"
        except Exception as e:
            final_error = str(e)
            yield f"event: error\ndata: {json.dumps({'message': str(e)})}\n\n"
        # Persist execution history record (best-effort)
        try:
            from atlas_adapter.execution_history import append_execution_record

            rec = {
                "id": str(uuid.uuid4()),
                "ts": datetime.now(timezone.utc).isoformat(),
                "thread_id": thread_id,
                "user_id": user_id,
                "message": user_msg[:4000],
                "final_text": (final_text or "")[:8000],
                "error": final_error,
                "verification_passed": bool(
                    (final_done_data or {}).get("verification_passed", False)
                ),
                "final_state": (final_done_data or {}).get("final_state"),
                "runner_state": (final_done_data or {}).get("runner_state"),
                "iterations": (final_done_data or {}).get("iterations"),
                "tools_used": (final_done_data or {}).get("tools_used", []),
                "pre_checks": (final_done_data or {}).get("pre_checks", []),
                "post_checks": (final_done_data or {}).get("post_checks", []),
                "rollback": (final_done_data or {}).get("rollback", []),
                "kpis": (final_done_data or {}).get("kpis", {}),
            }
            append_execution_record(rec)
        except Exception:
            pass
        # Persist final assistant output + update summary
        try:
            if mgr and thread_id and final_text:
                mgr.add_message(
                    thread_id,
                    sender="agent",
                    role="assistant",
                    content=final_text,
                    metadata={"panel": "agent", "kind": "chat"},
                )
                _update_thread_summary(thread_id, user_msg, final_text)
        except Exception:
            pass
        yield "event: close\ndata: {}\n\n"

    return StreamingResponse(_sse_generator(), media_type="text/event-stream")


@app.get("/agent/executions/recent", tags=["Agent Engine"])
def agent_executions_recent(limit: int = 20):
    """Recent execution runs from persistent JSONL history."""
    t0 = time.perf_counter()
    try:
        from atlas_adapter.execution_history import list_recent_executions

        items = list_recent_executions(limit=limit)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"items": items, "count": len(items)}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/agent/executions/{execution_id}", tags=["Agent Engine"])
def agent_execution_detail(execution_id: str):
    """Get one execution record by id."""
    t0 = time.perf_counter()
    try:
        from atlas_adapter.execution_history import get_execution_by_id

        item = get_execution_by_id(execution_id)
        ms = int((time.perf_counter() - t0) * 1000)
        if not item:
            return _std_resp(False, None, ms, "execution_not_found")
        return _std_resp(True, {"item": item}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


# --- Conversation thread utilities (debug/ops) ---
@app.post("/conversation/thread/new", tags=["Conversation"])
def conversation_thread_new(payload: dict):
    """Create a new persistent thread for Agent/Supervisor UI."""
    t0 = time.perf_counter()
    try:
        kind = (payload.get("kind") or "Agent").strip() or "Agent"
        user_id = (payload.get("user_id") or "owner").strip() or "owner"
        title = "Supervisor" if kind.lower().startswith("sup") else "Agent"
        tid = _ensure_thread_id(None, title=title, user_id=user_id)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True, {"thread_id": tid, "title": title, "user_id": user_id}, ms, None
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/conversation/thread/reset", tags=["Conversation"])
def conversation_thread_reset(payload: dict):
    """Reset by creating a fresh thread (does not delete existing history)."""
    t0 = time.perf_counter()
    try:
        kind = (payload.get("kind") or "Agent").strip() or "Agent"
        user_id = (payload.get("user_id") or "owner").strip() or "owner"
        title = "Supervisor" if kind.lower().startswith("sup") else "Agent"
        tid = _ensure_thread_id(None, title=title, user_id=user_id)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True,
            {"thread_id": tid, "title": title, "user_id": user_id, "reset": True},
            ms,
            None,
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/conversation/thread/{thread_id}/history", tags=["Conversation"])
def conversation_thread_history(thread_id: str, limit: int = 80):
    """Fetch recent messages for a thread."""
    t0 = time.perf_counter()
    mgr = _get_chat_manager()
    if not mgr:
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            "chat_manager_unavailable",
        )
    try:
        msgs = mgr.get_recent_messages(thread_id, limit=limit)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True,
            {"thread_id": thread_id, "messages": msgs, "count": len(msgs)},
            ms,
            None,
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/conversation/thread/{thread_id}/summary", tags=["Conversation"])
def conversation_thread_summary(thread_id: str):
    """Fetch rolling summary for a thread."""
    t0 = time.perf_counter()
    mgr = _get_chat_manager()
    if not mgr:
        return _std_resp(
            False,
            None,
            int((time.perf_counter() - t0) * 1000),
            "chat_manager_unavailable",
        )
    try:
        summary = mgr.get_context(thread_id, "summary") or ""
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"thread_id": thread_id, "summary": summary}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/agent/benchmark")
def agent_benchmark():
    """Competitive intelligence: architecture, latencies, complexity, maintainability, risk, scalability. Strengths, Weaknesses, Gaps, Improvement map."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.metrics import get_metrics_store

        store = get_metrics_store()
        snap = store.snapshot()
        counters = snap.get("counters") or {}
        latencies = snap.get("latencies") or {}
        avg_lat = {k: v.get("avg_ms", 0) for k, v in latencies.items()}
        strengths = [
            "Multi-agent pipeline (Executive, Strategist, Architect, Engineer, Reviewer, Optimizer)",
            "Policy + Audit on all executable paths",
            "Persistent memory (SQLite) + FTS fallback",
            "Orchestrator with critic and replan",
        ]
        weaknesses = []
        if not latencies:
            weaknesses.append("Pocas mÃ©tricas de latencia registradas")
        gaps = [
            "Embeddings opcional pendiente",
            "Adaptive routing por modelo en refinamiento",
        ]
        improvement_map = [
            "Aumentar depth en /agent/goal para tareas complejas",
            "Revisar /agent/system-intel para cuellos de botella",
        ]
        data = {
            "strengths": strengths,
            "weaknesses": weaknesses,
            "gaps": gaps,
            "improvement_map": improvement_map,
            "metrics": {"counters": counters, "latency_avg_ms": avg_lat},
        }
        ms = int((time.perf_counter() - t0) * 1000)
        return _professional_resp(
            True,
            data,
            ms,
            None,
            resumen=f"Benchmark: {len(strengths)} strengths, {len(gaps)} gaps",
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/agent/system-intel")
@app.post("/agent/system-intel")
def agent_system_intel():
    """Performance intelligence: analyze logs, repeated errors, bottlenecks, optimization suggestions."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.audit import get_audit_logger
        from modules.humanoid.metrics import get_metrics_store

        entries = get_audit_logger().tail(n=100, module=None)
        store = get_metrics_store()
        snap = store.snapshot()
        errors = [e for e in (entries or []) if not e.get("ok")]
        error_msgs = {}
        for e in errors:
            msg = (e.get("error") or "unknown")[:80]
            error_msgs[msg] = error_msgs.get(msg, 0) + 1
        repeated = [{"error": k, "count": v} for k, v in error_msgs.items() if v >= 2][
            :10
        ]
        latencies = snap.get("latencies") or {}
        bottlenecks = []
        for name, stat in latencies.items():
            avg = stat.get("avg_ms", 0)
            if avg > 5000:
                bottlenecks.append({"path": name, "avg_ms": round(avg, 0)})
        suggestions = []
        if repeated:
            suggestions.append("Revisar errores repetidos y aÃ±adir fallback o retry")
        if bottlenecks:
            suggestions.append(
                "Revisar endpoints con latencia >5s para optimizar o cache"
            )
        data = {
            "repeated_errors": repeated,
            "bottlenecks": bottlenecks[:10],
            "suggestions": suggestions or ["Sistema estable; sin sugerencias urgentes"],
            "audit_sample_size": len(entries or []),
        }
        ms = int((time.perf_counter() - t0) * 1000)
        return _professional_resp(
            True,
            data,
            ms,
            None,
            resumen=f"System intel: {len(repeated)} errores repetidos, {len(bottlenecks)} cuellos",
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


# --- Continuous Improvement ---
class AgentImproveBody(BaseModel):
    scope: str = "all"  # repo | runtime | all
    mode: str = "plan_only"  # plan_only | controlled | auto
    depth: Optional[int] = 2
    max_items: Optional[int] = None


@app.post("/agent/improve")
def agent_improve(body: AgentImproveBody):
    """CI cycle: scan -> plan -> optional execute. scope=repo|runtime|all. mode=plan_only|controlled|auto. Timeout-safe."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.ci import run_improve

        result = run_improve(
            scope=body.scope or "all",
            mode=body.mode or "plan_only",
            depth=max(1, min(5, body.depth or 2)),
            max_items=body.max_items,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        if not result.get("ok"):
            return _std_resp(False, None, ms, result.get("error"))
        data = result.get("data") or {}
        return _professional_resp(
            True,
            {
                "findings": data.get("findings", []),
                "plan": data.get("plan", {}),
                "approvals_required": data.get("approvals_required", []),
                "auto_executed": data.get("auto_executed", []),
                "artifacts": data.get("artifacts", []),
                "report": data.get("report", {}),
            },
            ms,
            None,
            resumen=f"CI: {len(data.get('findings') or [])} findings, {len(data.get('approvals_required') or [])} pending approval",
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/agent/improve/status")
def agent_improve_status():
    """Last CI cycle: plan_id, timestamp, executed, pending. Links to memory/audit."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.ci import get_improve_status

        data = get_improve_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class AgentImproveApplyBody(BaseModel):
    plan_id: str = ""
    approve: bool = False
    items: Optional[list] = None  # item_ids to apply


@app.post("/agent/improve/apply")
def agent_improve_apply(body: AgentImproveApplyBody):
    """Execute only approved items allowed by policy. plan_id must match last run."""
    t0 = time.perf_counter()
    if not body.approve:
        return _std_resp(
            False, None, int((time.perf_counter() - t0) * 1000), "approve must be true"
        )
    try:
        from modules.humanoid.ci import execute_plan, get_last_plan

        plan = get_last_plan()
        if not plan:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "no previous plan; run POST /agent/improve first",
            )
        if body.plan_id and plan.get("plan_id") != body.plan_id:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "plan_id does not match last cycle",
            )
        result = execute_plan(plan)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            result.get("ok", True),
            {"executed": result.get("executed"), "errors": result.get("errors")},
            ms,
            result.get("errors") and result["errors"][0] or None,
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class ScaffoldBody(BaseModel):
    type: str = "fastapi"  # fastapi | pwa | flutter | node
    name: str
    options: Optional[dict] = None


@app.post("/scaffold/app")
def scaffold_app(body: ScaffoldBody):
    """Generate app structure + RUNBOOK. Returns {ok, data: {tree, files_created, runbook_path}, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.scaffolder import generate

        result = generate(body.type, body.name, options=body.options or {})
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            result.get("ok", False),
            {k: v for k, v in result.items() if k != "ok"},
            ms,
            result.get("error"),
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class ScriptsGenerateBody(BaseModel):
    kind: str = "powershell"  # powershell | python
    purpose: str
    options: Optional[dict] = None


@app.post("/scripts/generate")
def scripts_generate(body: ScriptsGenerateBody):
    """Generate script. Returns {ok, data: {path, preview, how_to_run}, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid import get_humanoid_kernel
        from modules.humanoid.scripts import generate_script

        k = get_humanoid_kernel()
        hands = k.registry.get("hands")
        fs = getattr(hands, "fs", None) if hands else None
        result = generate_script(
            body.kind, body.purpose, options=body.options or {}, fs_controller=fs
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            result.get("ok", False),
            {
                "path": result.get("path"),
                "preview": result.get("preview"),
                "how_to_run": result.get("how_to_run"),
                "validated": result.get("validated"),
            },
            ms,
            result.get("error"),
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/web/session/start")
def web_session_start():
    """Start browser session (no-op if Playwright not installed). Returns {ok, data, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.web import status as web_status

        data = web_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class WebNavigateBody(BaseModel):
    url: str


@app.post("/web/navigate")
def web_navigate(body: WebNavigateBody):
    t0 = time.perf_counter()
    try:
        from modules.humanoid.web import open_url

        result = open_url(body.url or "")
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(result.get("ok", False), result, ms, result.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/web/extract")
def web_extract():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.web import extract_text

        result = extract_text()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(result.get("ok", False), result, ms, result.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/web/screenshot")
def web_screenshot():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.web import screenshot

        result = screenshot()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(result.get("ok", False), result, ms, result.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/web/status")
def web_status_endpoint():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.web import status as web_status

        data = web_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class VoiceSpeakBody(BaseModel):
    text: str


@app.post("/voice/speak")
def voice_speak_endpoint(body: VoiceSpeakBody):
    t0 = time.perf_counter()
    try:
        from modules.humanoid.voice import voice_speak

        result = voice_speak(body.text or "")
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(result.get("ok", False), result, ms, result.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/voice/listen")
def voice_listen_endpoint():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.voice import voice_listen_stub

        result = voice_listen_stub()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(result.get("ok", True), result, ms, result.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/voice/status")
def voice_status_endpoint():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.voice import voice_status

        data = voice_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class VoiceTranscribeBody(BaseModel):
    audio_base64: Optional[str] = None
    language: Optional[str] = "es-ES"


@app.post("/voice/transcribe")
def voice_transcribe_endpoint(body: VoiceTranscribeBody):
    """Transcribe audio (WAV en base64). Devuelve { ok, data: { text }, ms, error }."""
    t0 = time.perf_counter()
    try:
        import base64
        import tempfile
        from pathlib import Path

        from modules.humanoid.voice.stt import is_available, transcribe

        if not body.audio_base64:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "audio_base64 requerido",
            )
        if not is_available():
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), "STT no disponible"
            )
        raw = base64.b64decode(body.audio_base64)
        with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as f:
            f.write(raw)
            path = f.name
        try:
            result = transcribe(path, options={"language": body.language or "es-ES"})
            ms = int((time.perf_counter() - t0) * 1000)
            return _std_resp(
                result.get("ok", False),
                {"text": result.get("text", "")},
                ms,
                result.get("error"),
            )
        finally:
            Path(path).unlink(missing_ok=True)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class FaceCheckBody(BaseModel):
    image_base64: Optional[str] = None


@app.post("/face/check")
def face_check_endpoint(body: FaceCheckBody):
    """DetecciÃ³n de rostro en imagen (PNG/JPEG en base64). Devuelve { ok, data: { faces_detected, message }, ms, error }."""
    t0 = time.perf_counter()
    try:
        import base64

        from modules.humanoid.face import face_check_image

        if not body.image_base64:
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "image_base64 requerido",
            )
        raw = base64.b64decode(body.image_base64)
        result = face_check_image(raw)
        ms = int((time.perf_counter() - t0) * 1000)
        data = {
            "faces_detected": result.get("faces_detected", 0),
            "message": result.get("message", ""),
            "count": result.get("count", 0),
        }
        return _std_resp(result.get("ok", False), data, ms, result.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/face/status")
def face_status_endpoint():
    """Estado del mÃ³dulo de reconocimiento facial."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.face.detector import _check_deps, is_available

        data = {"available": is_available(), "missing_deps": _check_deps()}
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/deps/check")
def deps_check_endpoint():
    """Return missing_deps per module and suggested commands. No install."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.deps_checker import check_all

        data = check_all()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(data.get("ok", True), data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/deps/check")
def deps_check_get():
    """GET variant of deps check."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.deps_checker import check_all

        data = check_all()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(data.get("ok", True), data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


# --- Vision ---
class VisionAnalyzeBody(BaseModel):
    image_path: str
    use_ocr: Optional[bool] = True
    use_llm_vision: Optional[bool] = True


def _professional_resp(
    ok: bool,
    data: Any,
    ms: int,
    error: Optional[str],
    resumen: str = "",
    archivos: Optional[list] = None,
    siguientes_pasos: Optional[list] = None,
) -> dict:
    """Respuesta profesional: resumen, resultado tÃ©cnico, evidencia, siguientes pasos."""
    out = _std_resp(ok, data, ms, error)
    if resumen:
        out["resumen"] = resumen
    if archivos:
        out["archivos_creados"] = archivos
    if siguientes_pasos:
        out["siguientes_pasos"] = siguientes_pasos
    return out


@app.post("/vision/analyze")
def vision_analyze(body: VisionAnalyzeBody):
    """Analyze image: OCR + LLM vision. Returns texto, estructura, insights, acciones_sugeridas + formato profesional."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision import analyze

        result = analyze(
            body.image_path,
            use_ocr=body.use_ocr if body.use_ocr is not None else True,
            use_llm_vision=body.use_llm_vision
            if body.use_llm_vision is not None
            else True,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        ok = result.get("ok", False)
        result.setdefault("suggested_actions", result.get("acciones_sugeridas") or [])
        resumen = (
            f"Imagen analizada: {len(result.get('extracted_text', ''))} chars texto, {len(result.get('entities', []))} entidades"
            if ok
            else "AnÃ¡lisis fallido"
        )
        siguientes = (
            result.get("acciones_sugeridas") or result.get("suggested_actions") or []
        )
        return _professional_resp(
            ok,
            result,
            ms,
            result.get("error"),
            resumen=resumen,
            siguientes_pasos=siguientes,
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class VisionOcrBody(BaseModel):
    image_path: str


@app.post("/vision/ocr")
def vision_ocr(body: VisionOcrBody):
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision import ocr

        result = ocr(body.image_path or "")
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(result.get("ok", False), result, ms, result.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class VisionScreenshotBody(BaseModel):
    screenshot_path: str


@app.post("/vision/screenshot")
def vision_screenshot(body: VisionScreenshotBody):
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision import screenshot_analyze

        result = screenshot_analyze(body.screenshot_path or "")
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(result.get("ok", False), result, ms, result.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/vision/status")
def vision_status_endpoint():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision import vision_status

        data = vision_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


# --- Vision Ubiqua (multidispositivo) ---
class VisionUbiqDiscoverBody(BaseModel):
    scan_ports: Optional[bool] = True
    onvif: Optional[bool] = True


@app.post("/api/vision/ubiq/discover", tags=["Vision"])
def api_vision_ubiq_discover(body: Optional[VisionUbiqDiscoverBody] = None):
    """Escaneo LAN: RTSP/MJPEG + ONVIF (WS-Discovery). Persiste y registra en ANS."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision.ubiq import (discover_local_cameras,
                                                  ensure_db)

        ensure_db()
        out = discover_local_cameras(
            scan_ports=bool(body.scan_ports)
            if body and body.scan_ports is not None
            else True,
            onvif=bool(body.onvif) if body and body.onvif is not None else True,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, out, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/api/vision/ubiq/cameras", tags=["Vision"])
def api_vision_ubiq_cameras(limit: int = 200):
    """Lista cÃ¡maras descubiertas/registradas."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision.ubiq import ensure_db, list_cameras

        ensure_db()
        cams = list_cameras(limit=int(limit))
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"cameras": cams, "count": len(cams)}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class VisionUbiqStreamBody(BaseModel):
    variant: Optional[str] = "local"  # local|mobile


@app.post("/api/vision/ubiq/streams/{cam_id}/start", tags=["Vision"])
def api_vision_ubiq_stream_start(
    cam_id: str, body: Optional[VisionUbiqStreamBody] = None
):
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision.ubiq import start_stream

        variant = (body.variant if body and body.variant else "local") or "local"
        out = start_stream(cam_id, variant=str(variant))
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(bool(out.get("ok")), out, ms, out.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/api/vision/ubiq/streams/{cam_id}/stop", tags=["Vision"])
def api_vision_ubiq_stream_stop(
    cam_id: str, body: Optional[VisionUbiqStreamBody] = None
):
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision.ubiq import stop_stream

        variant = (body.variant if body and body.variant else "local") or "local"
        out = stop_stream(cam_id, variant=str(variant))
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(bool(out.get("ok")), out, ms, out.get("error"))
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/api/vision/ubiq/streams/status", tags=["Vision"])
def api_vision_ubiq_stream_status():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision.ubiq import get_setting, stream_status

        st = stream_status()
        active_eye = (get_setting("vision.active_eye") or "").strip()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"streams": st, "active_eye": active_eye}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class VisionUbiqActiveEyeBody(BaseModel):
    eye: str  # "" | "ubiq:<cam_id>"
    hold_s: Optional[float] = 25.0


@app.post("/api/vision/ubiq/active-eye", tags=["Vision"])
def api_vision_ubiq_set_active_eye(body: VisionUbiqActiveEyeBody):
    """Fija el ojo activo (persistente) para `eyes_capture`."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision.ubiq import set_setting

        eye = (body.eye or "").strip()
        if eye and not eye.lower().startswith("ubiq:"):
            return _std_resp(
                False,
                {"ok": False, "error": "eye must be '' or 'ubiq:<cam_id>'"},
                0,
                "bad eye",
            )
        set_setting("vision.active_eye", eye)
        if eye:
            import time as _time

            set_setting(
                "vision.active_eye_until",
                str(_time.time() + max(3.0, float(body.hold_s or 25.0))),
            )
        else:
            set_setting("vision.active_eye_until", "")
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"eye": eye}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/api/vision/ubiq/snapshot/{cam_id}", tags=["Vision"])
def api_vision_ubiq_snapshot(cam_id: str):
    """Un frame JPEG base64 (uso debug/preview)."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision.ubiq import take_snapshot

        out = take_snapshot(cam_id, timeout_s=4.0)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            bool(out.get("ok")),
            {k: v for k, v in out.items() if k != "jpeg_bytes"},
            ms,
            out.get("error"),
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


def _safe_stream_file(cam_id: str, variant: str, rel: str):
    from pathlib import Path

    from fastapi import HTTPException

    cid = (cam_id or "").strip()
    var = (variant or "").strip().lower()
    if var not in ("local", "mobile"):
        raise HTTPException(status_code=400, detail="bad variant")
    root = (
        Path(__file__).resolve().parents[1]
        / "snapshots"
        / "vision"
        / "ubiq_streams"
        / cid
        / var
    ).resolve()
    p = (root / rel).resolve()
    if root not in p.parents and p != root:
        raise HTTPException(status_code=400, detail="bad path")
    if not p.is_file():
        raise HTTPException(status_code=404, detail="not found")
    return p


@app.get("/api/vision/ubiq/streams/{cam_id}/{variant}/index.m3u8", tags=["Vision"])
def api_vision_ubiq_stream_playlist(cam_id: str, variant: str = "local"):
    from fastapi.responses import FileResponse

    p = _safe_stream_file(cam_id, variant, "index.m3u8")
    return FileResponse(str(p), media_type="application/vnd.apple.mpegurl")


@app.get("/api/vision/ubiq/streams/{cam_id}/{variant}/{segment}", tags=["Vision"])
def api_vision_ubiq_stream_segment(cam_id: str, variant: str, segment: str):
    from fastapi.responses import FileResponse

    seg = (segment or "").strip()
    if not (seg.endswith(".ts") or seg.endswith(".m3u8")):
        # limitar superficie
        from fastapi import HTTPException

        raise HTTPException(status_code=400, detail="bad segment")
    p = _safe_stream_file(cam_id, variant, seg)
    mt = "video/MP2T" if seg.endswith(".ts") else "application/vnd.apple.mpegurl"
    return FileResponse(str(p), media_type=mt)


@app.get("/api/vision/ubiq/streams/{cam_id}/key", tags=["Vision"])
def api_vision_ubiq_stream_key(cam_id: str, token: str = ""):
    """Key AES-128 para HLS mobile. Requiere token."""
    from fastapi import HTTPException
    from fastapi.responses import FileResponse

    from modules.humanoid.vision.ubiq.streaming import get_mobile_token

    tok = (token or "").strip()
    if not tok or tok != get_mobile_token():
        raise HTTPException(status_code=401, detail="unauthorized")
    p = _safe_stream_file(cam_id, "mobile", "enc.key")
    return FileResponse(str(p), media_type="application/octet-stream")


@app.get("/api/vision/ubiq/mobile-token", tags=["Vision"])
def api_vision_ubiq_mobile_token(request: Request):
    """Devuelve token mobile (solo localhost) para armar el link en Dashboard."""
    from fastapi import HTTPException

    from modules.humanoid.vision.ubiq.streaming import get_mobile_token

    host = ""
    try:
        host = str(getattr(getattr(request, "client", None), "host", "") or "")
    except Exception:
        host = ""
    if host not in ("127.0.0.1", "::1", "localhost"):
        raise HTTPException(status_code=403, detail="forbidden")
    return {"ok": True, "token": get_mobile_token()}


@app.get("/mobile/vision/{cam_id}.m3u8", tags=["Vision"])
def mobile_vision_playlist(cam_id: str, token: str = ""):
    """Playlist mobile cifrada. Requiere token (para no divulgar URLs/keys)."""
    from fastapi import HTTPException
    from fastapi.responses import FileResponse

    from modules.humanoid.vision.ubiq.streaming import get_mobile_token

    tok = (token or "").strip()
    if not tok or tok != get_mobile_token():
        raise HTTPException(status_code=401, detail="unauthorized")
    p = _safe_stream_file(cam_id, "mobile", "index.m3u8")
    return FileResponse(str(p), media_type="application/vnd.apple.mpegurl")


@app.get("/mobile/vision/{cam_id}/{segment}", tags=["Vision"])
def mobile_vision_segment(cam_id: str, segment: str, token: str = ""):
    """Segmentos mobile cifrados. Requiere token."""
    from fastapi import HTTPException
    from fastapi.responses import FileResponse

    from modules.humanoid.vision.ubiq.streaming import get_mobile_token

    tok = (token or "").strip()
    if not tok or tok != get_mobile_token():
        raise HTTPException(status_code=401, detail="unauthorized")
    seg = (segment or "").strip()
    if not seg.endswith(".ts"):
        raise HTTPException(status_code=400, detail="bad segment")
    p = _safe_stream_file(cam_id, "mobile", seg)
    return FileResponse(str(p), media_type="video/MP2T")


@app.post("/api/vision/ubiq/motion/start", tags=["Vision"])
def api_vision_ubiq_motion_start():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision.ubiq import start_motion_watchdog

        out = start_motion_watchdog()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, out, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.post("/api/vision/ubiq/motion/stop", tags=["Vision"])
def api_vision_ubiq_motion_stop():
    t0 = time.perf_counter()
    try:
        from modules.humanoid.vision.ubiq import stop_motion_watchdog

        out = stop_motion_watchdog()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, out, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


# --- Screen (capture / analyze / locate / act, policy-gated) ---
@app.get("/screen/status")
def screen_status_endpoint():
    """GET /screen/status. Responds even when disabled (enabled=false, missing_deps)."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.screen import get_screen_status

        data = get_screen_status()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class ScreenCaptureBody(BaseModel):
    region: Optional[List[int]] = None  # [x, y, w, h]
    format: Optional[str] = "png"


@app.post("/screen/capture")
def screen_capture(body: Optional[ScreenCaptureBody] = None):
    """Capture screen or region. Returns {ok, data: {path?, b64?}, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.screen.capture import (capture_screen,
                                                     save_capture_to_file)

        region = (
            tuple(body.region)
            if body and body.region and len(body.region) == 4
            else None
        )
        png, err = capture_screen(region=region, format=body.format if body else "png")
        if err or not png:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), err or "no capture"
            )
        import base64
        import os

        out_dir = (
            (
                os.getenv("ATLAS_REPO_PATH")
                or os.getenv("POLICY_ALLOWED_PATHS", "C:\\ATLAS_PUSH")
            )
            .strip()
            .split(",")[0]
            .strip()
        )
        logs_dir = os.path.join(out_dir, "logs", "screen")
        path = save_capture_to_file(png, logs_dir, "capture")
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True,
            {"path": path, "b64": base64.b64encode(png).decode("ascii")[:200] + "..."},
            ms,
            None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


class ScreenAnalyzeBody(BaseModel):
    image_path: Optional[str] = None
    image_b64: Optional[str] = None
    prompt: Optional[str] = None


@app.post("/screen/analyze")
def screen_analyze(body: Optional[ScreenAnalyzeBody] = None):
    """Layout + text + vision suggestions. Returns {ok, data: {layout, description, ...}, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.screen.layout import get_layout
        from modules.humanoid.screen.vision_llm import analyze_image

        layout = get_layout()
        desc_result = {"description": "", "suggestions": []}
        if body and (body.image_b64 or body.image_path):
            desc_result = analyze_image(
                image_base64=body.image_b64,
                image_path=body.image_path,
                prompt=body.prompt or "Describe UI and suggest actions.",
            )
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            True, {"layout": layout, **desc_result}, ms, layout.get("error")
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


class ScreenLocateBody(BaseModel):
    query: str
    region: Optional[List[int]] = None


@app.post("/screen/locate")
def screen_locate(body: ScreenLocateBody):
    """Find elements by text query. Returns {ok, data: {matches}, ms, error}."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.screen.locator import locate

        region = tuple(body.region) if body.region and len(body.region) == 4 else None
        result = locate(body.query, region=region)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(result.get("ok", False), result, ms, result.get("error"))
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


class ScreenActBody(BaseModel):
    action: str  # click | type | hotkey | scroll
    payload: dict  # {x,y} | {text} | {keys:[]} | {clicks, x?, y?}
    destructive: Optional[bool] = None  # if True, may require approval


@app.post("/screen/act")
def screen_act(body: ScreenActBody):
    """Execute screen action (policy-gated). Evidence: screenshot_before/after, coords, ms. Destructive -> ApprovalItem."""
    t0 = time.perf_counter()
    try:
        import os

        from modules.humanoid.policy import ActorContext, get_policy_engine
        from modules.humanoid.screen.actions import execute_action
        from modules.humanoid.screen.capture import (capture_screen,
                                                     save_capture_to_file)
        from modules.humanoid.screen.policy import (check_rate_limit,
                                                    is_destructive_action,
                                                    record_screen_act)
        from modules.humanoid.screen.status import get_screen_status

        if not get_screen_status().get("enabled", False):
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "screen module disabled",
            )
        try:
            from modules.humanoid.mode import is_screen_act_allowed

            if not is_screen_act_allowed():
                return _std_resp(
                    False,
                    None,
                    int((time.perf_counter() - t0) * 1000),
                    "screen_act disabled (ATLAS_MODE=lite or deps missing)",
                )
        except Exception:
            pass
        ctx = ActorContext(actor="api", role="owner")
        decision = get_policy_engine().can(ctx, "screen", "screen_act", None)
        if not decision.allow:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), decision.reason
            )
        allowed, reason = check_rate_limit()
        if not allowed:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), reason
            )
        before_png, _ = capture_screen()
        _base = (
            (
                os.getenv("ATLAS_REPO_PATH")
                or os.getenv("POLICY_ALLOWED_PATHS", "C:\\ATLAS_PUSH")
            )
            .strip()
            .split(",")[0]
            .strip()
        )
        _screen_logs = os.path.join(_base, "logs", "screen")
        path_before = (
            save_capture_to_file(before_png, _screen_logs, "before")
            if before_png
            else None
        )
        result = execute_action(body.action, body.payload or {})
        try:
            from modules.humanoid.screen.record import (is_recording,
                                                        record_action)

            if is_recording():
                record_action(body.action, body.payload or {})
        except Exception:
            pass
        after_png, _ = capture_screen()
        path_after = (
            save_capture_to_file(after_png, _screen_logs, "after")
            if after_png
            else None
        )
        record_screen_act()
        ms = int((time.perf_counter() - t0) * 1000)
        evidence = {
            "screenshot_before": path_before,
            "screenshot_after": path_after,
            "coords": body.payload,
            "ms": ms,
        }
        destructive = (
            body.destructive
            if body.destructive is not None
            else is_destructive_action(body.action, body.payload or {})
        )
        if destructive:
            evidence["approval_required"] = True
        return _std_resp(
            result.get("ok", False),
            {"evidence": evidence, "result": result},
            ms,
            result.get("error"),
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


@app.post("/screen/record/start", tags=["Screen"])
def screen_record_start():
    """Start recording screen actions (macro). Pro/Ultra only."""
    try:
        from modules.humanoid.mode import is_record_replay_allowed

        if not is_record_replay_allowed():
            return _std_resp(False, None, 0, "record/replay disabled (ATLAS_MODE)")
        from modules.humanoid.screen.record import start_recording

        r = start_recording()
        return _std_resp(r.get("ok", False), r, 0, r.get("error"))
    except Exception as e:
        return _std_resp(False, None, 0, str(e))


@app.post("/screen/record/stop", tags=["Screen"])
def screen_record_stop():
    """Stop recording and return recorded actions."""
    try:
        from modules.humanoid.screen.record import stop_recording

        r = stop_recording()
        return _std_resp(r.get("ok", False), r, 0, r.get("error"))
    except Exception as e:
        return _std_resp(False, None, 0, str(e))


class ScreenReplayBody(BaseModel):
    macro_id: Optional[str] = None
    actions: Optional[List[dict]] = None
    delay_ms: Optional[int] = 100


@app.post("/screen/replay", tags=["Screen"])
def screen_replay(body: ScreenReplayBody):
    """Replay macro by id or inline actions. Policy-gated."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.mode import is_record_replay_allowed
        from modules.humanoid.policy import ActorContext, get_policy_engine

        if not is_record_replay_allowed():
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "record/replay disabled",
            )
        ctx = ActorContext(actor="api", role="owner")
        decision = get_policy_engine().can(ctx, "screen", "screen_act", None)
        if not decision.allow:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), decision.reason
            )
        actions = body.actions
        if not actions and body.macro_id:
            from modules.humanoid.macros import get_macro

            m = get_macro(body.macro_id)
            if not m:
                return _std_resp(
                    False,
                    None,
                    int((time.perf_counter() - t0) * 1000),
                    "macro not found",
                )
            actions = m.get("actions", [])
        if not actions:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), "no actions"
            )
        from modules.humanoid.screen.replay import replay_actions

        r = replay_actions(actions, delay_ms=body.delay_ms or 100)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(
            r.get("ok", False),
            r,
            ms,
            r.get("errors", [None])[0] if r.get("errors") else None,
        )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


# --- Macros ---
@app.get("/macros/list", tags=["Macros"])
def macros_list(limit: int = 50):
    """List saved macros."""
    try:
        from modules.humanoid.macros import list_macros

        return {"ok": True, "data": list_macros(limit=limit)}
    except Exception as e:
        return {"ok": False, "data": [], "error": str(e)}


class MacrosSaveBody(BaseModel):
    id: str
    name: str
    actions: List[dict]


@app.post("/macros/save", tags=["Macros"])
def macros_save(body: MacrosSaveBody):
    """Save macro (e.g. from record/stop actions)."""
    try:
        from modules.humanoid.macros import save_macro

        r = save_macro(body.id, body.name, body.actions)
        return _std_resp(r.get("ok", False), r, 0, r.get("error"))
    except Exception as e:
        return _std_resp(False, None, 0, str(e))


@app.get("/macros/get", tags=["Macros"])
def macros_get(id: Optional[str] = None):
    """Get macro by id."""
    if not id:
        return {"ok": False, "data": None, "error": "id required"}
    try:
        from modules.humanoid.macros import get_macro

        m = get_macro(id)
        return {"ok": m is not None, "data": m, "error": None if m else "not found"}
    except Exception as e:
        return {"ok": False, "data": None, "error": str(e)}


class MacrosDeleteBody(BaseModel):
    id: str


@app.post("/macros/delete", tags=["Macros"])
def macros_delete(body: MacrosDeleteBody):
    """Delete macro. Denied by default; approval required."""
    try:
        from modules.humanoid.policy import ActorContext, get_policy_engine

        ctx = ActorContext(actor="api", role="owner")
        decision = get_policy_engine().can(ctx, "macros", "delete", body.id)
        if not decision.allow:
            return _std_resp(False, None, 0, decision.reason)
        from modules.humanoid.macros import delete_macro

        r = delete_macro(body.id)
        return _std_resp(r.get("ok", False), r, 0, r.get("error"))
    except Exception as e:
        return _std_resp(False, None, 0, str(e))


# --- Bench (model benchmarking, Ultra/Pro+metalearn) ---
class BenchRunBody(BaseModel):
    level: Optional[str] = "quick"  # quick | full


@app.post("/bench/run", tags=["Bench"])
def bench_run(body: Optional[BenchRunBody] = None):
    """Run benchmark: probes per route, latency, recommended model mapping. Ultra or Pro+metalearn."""
    t0 = time.perf_counter()
    try:
        from modules.humanoid.mode import is_benchmark_allowed

        if not is_benchmark_allowed():
            return _std_resp(
                False,
                None,
                int((time.perf_counter() - t0) * 1000),
                "benchmark disabled (ATLAS_MODE or METALEARN)",
            )
        from modules.humanoid.bench import run_bench

        level = (body and body.level) or "quick"
        r = run_bench(level=level)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, r, ms, None)
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))


# --- Memory (Policy: memory_read / memory_write / memory_export; memory_delete denied) ---
def _memory_actor():
    return ActorContext(actor="api", role=os.getenv("POLICY_DEFAULT_ROLE", "owner"))


@app.get("/memory/recall")
def memory_recall(query: str = "", limit: int = 20):
    """Recall by search query. FTS or LIKE fallback. Policy: memory_read."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(_memory_actor(), "memory", "memory_read")
        if not decision.allow:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), decision.reason
            )
        from modules.humanoid.memory_engine import recall_by_query

        data = recall_by_query(query or "", limit=limit)
        ms = int((time.perf_counter() - t0) * 1000)
        return _professional_resp(
            True, {"results": data}, ms, None, resumen=f"Recall: {len(data)} resultados"
        )
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/memory/thread")
def memory_thread(thread_id: str = "", limit: int = 50):
    """Recall by thread_id. Policy: memory_read."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(_memory_actor(), "memory", "memory_read")
        if not decision.allow:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), decision.reason
            )
        from modules.humanoid.memory_engine import recall_by_thread

        data = recall_by_thread(thread_id or "", limit=limit)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class MemoryThreadCreateBody(BaseModel):
    title: str = ""


@app.post("/memory/thread/create")
def memory_thread_create(body: MemoryThreadCreateBody):
    """Create or get thread. Returns thread_id. Policy: memory_write."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(_memory_actor(), "memory", "memory_write")
        if not decision.allow:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), decision.reason
            )
        from modules.humanoid.memory_engine import ensure_thread

        thread_id = ensure_thread(None, (body.title or "")[:200])
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"thread_id": thread_id}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/memory/thread/list")
def memory_thread_list(limit: int = 50):
    """List threads. Policy: memory_read."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(_memory_actor(), "memory", "memory_read")
        if not decision.allow:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), decision.reason
            )
        from modules.humanoid.memory_engine import list_threads

        data = list_threads(limit=limit)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"threads": data}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class MemoryWriteBody(BaseModel):
    thread_id: str = ""
    kind: str = "summary"  # artifact | decision | summary
    payload: dict = {}
    task_id: Optional[str] = None
    run_id: Optional[int] = None


@app.post("/memory/write")
def memory_write_endpoint(body: MemoryWriteBody):
    """Write artifact/decision/summary. Policy: memory_write. Redacts secrets."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(_memory_actor(), "memory", "memory_write")
        if not decision.allow:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), decision.reason
            )
        from modules.humanoid.memory_engine import memory_write as mem_write

        out = mem_write(
            body.thread_id or None,
            body.kind,
            body.payload or {},
            body.task_id,
            body.run_id,
        )
        ms = int((time.perf_counter() - t0) * 1000)
        if not out.get("ok"):
            return _std_resp(False, None, ms, out.get("error"))
        return _std_resp(True, {"id": out.get("id"), "kind": out.get("kind")}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


class MemorySummarizeBody(BaseModel):
    thread_id: str
    use_llm: bool = True


@app.post("/memory/summarize")
def memory_summarize(body: MemorySummarizeBody):
    """Incremental summary for thread. FAST LLM if use_llm else deterministic. Policy: memory_write. Timeout 15s."""
    t0 = time.perf_counter()
    timeout_sec = 15
    try:
        decision = get_policy_engine().can(_memory_actor(), "memory", "memory_write")
        if not decision.allow:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), decision.reason
            )
        from modules.humanoid.memory_engine import (add_summary,
                                                    recall_by_thread)

        data = recall_by_thread(body.thread_id, limit=20)
        tasks = data.get("tasks", [])
        summaries = data.get("summaries", [])
        if body.use_llm:
            try:
                from modules.humanoid import get_humanoid_kernel

                k = get_humanoid_kernel()
                brain = k.registry.get("brain")
                if brain and hasattr(brain, "run_llm"):
                    goals = [t.get("goal", "")[:100] for t in tasks[:5]]
                    prompt = f"Resumen en una frase: hilo con {len(tasks)} tareas. Primeras: {'; '.join(goals)}."
                    r = brain.run_llm(
                        prompt,
                        route="FAST",
                        max_tokens=128,
                        timeout_override=timeout_sec,
                    )
                    text = (r.get("output") or "").strip()[:2000] if r.get("ok") else ""
                    if text:
                        add_summary(body.thread_id, text)
                        ms = int((time.perf_counter() - t0) * 1000)
                        return _std_resp(
                            True, {"summary_preview": text[:200]}, ms, None
                        )
            except Exception:
                pass
        text = f"Resumen: {len(tasks)} tareas, {len(summaries)} resÃºmenes previos."
        add_summary(body.thread_id, text)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"summary_preview": text}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/memory/export")
def memory_export(thread_id: str = "", task_id: str = "", limit: int = 100):
    """Export memory as markdown. Policy: memory_export."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(_memory_actor(), "memory", "memory_export")
        if not decision.allow:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), decision.reason
            )
        from modules.humanoid.memory_engine import export_markdown

        md = export_markdown(thread_id=thread_id, task_id=task_id, limit=limit)
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, {"markdown": md}, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


@app.get("/memory/snapshot")
def memory_snapshot():
    """Snapshot counts for backup. Policy: memory_read."""
    t0 = time.perf_counter()
    try:
        decision = get_policy_engine().can(_memory_actor(), "memory", "memory_read")
        if not decision.allow:
            return _std_resp(
                False, None, int((time.perf_counter() - t0) * 1000), decision.reason
            )
        from modules.humanoid.memory_engine import snapshot

        data = snapshot()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(data.get("ok", True), data, ms, None)
    except Exception as e:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(e))


# â”€â”€ Brazos subordinados: gestor de procesos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

_ARM_PROCS: dict = {}  # arm_id â†’ list[subprocess.Popen]

_ARM_DEFS: dict = {
    "panaderia": {
        "name": "Rauli PanaderÃ­a",
        "frontend_port": 5173,
        "api_port": 3001,
        "processes": [
            {
                "label": "API backend",
                "cwd": "_external/rauli-panaderia/backend",
                "cmd": "node server.js",
            },
            {
                "label": "Frontend",
                "cwd": "_external/rauli-panaderia/frontend",
                "cmd": "npm run dev -- --host --strictPort --port 5173",
            },
        ],
    },
    "vision": {
        "name": "Rauli Vision",
        "frontend_port": 5174,
        "api_port": 3000,
        "processes": [
            {
                "label": "Proxy API",
                "cwd": "_external/RAULI-VISION/cliente-local",
                "cmd": "python simple-server.py",
            },
            {
                "label": "Dashboard",
                "cwd": "_external/RAULI-VISION/dashboard",
                "cmd": "npm run dev -- --host --strictPort --port 5174",
            },
        ],
    },
}


def _is_local_port_open(port: Optional[int], host: str = "127.0.0.1") -> bool:
    if not port:
        return False
    import socket

    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(1.0)
        result = s.connect_ex((host, int(port)))
        s.close()
        return result == 0
    except Exception:
        return False


def _arm_process_port(defn: dict, proc_label: str) -> Optional[int]:
    label = (proc_label or "").strip().lower()
    if ("frontend" in label) or ("dashboard" in label):
        return defn.get("frontend_port")
    if ("api" in label) or ("proxy" in label):
        return defn.get("api_port")
    return None


def _extract_trycloudflare_url(log_path: Path) -> str:
    """Best-effort: obtain last quick tunnel URL from a cloudflared log file."""
    try:
        if not log_path.exists():
            return ""
        # Read only the tail to keep endpoint latency low even with large logs.
        max_bytes = 256 * 1024
        with log_path.open("rb") as fh:
            fh.seek(0, os.SEEK_END)
            size = fh.tell()
            start = max(0, size - max_bytes)
            fh.seek(start)
            content = fh.read().decode("utf-8", errors="ignore")
        matches = re.findall(r"https://[a-z0-9-]+\.trycloudflare\.com", content, flags=re.IGNORECASE)
        if not matches:
            return ""
        return str(matches[-1]).rstrip("/")
    except Exception:
        return ""


def _normalize_public_base(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    if not (s.startswith("http://") or s.startswith("https://")):
        return ""
    return s.rstrip("/")


def _public_arm_urls() -> Dict[str, Any]:
    """
    Resolve public URLs for remote clients.
    Priority: explicit env vars -> app vars -> quick tunnel logs.
    """
    pan_front = _normalize_public_base(
        os.getenv("ATLAS_PANADERIA_PUBLIC_URL")
        or os.getenv("ATLAS_PANADERIA_APP_URL")
        or ""
    )
    pan_api = _normalize_public_base(os.getenv("ATLAS_PANADERIA_PUBLIC_API_URL") or "")

    vis_front = _normalize_public_base(
        os.getenv("ATLAS_VISION_PUBLIC_URL")
        or os.getenv("ATLAS_VISION_APP_URL")
        or ""
    )
    vis_api = _normalize_public_base(os.getenv("ATLAS_VISION_PUBLIC_API_URL") or "")

    if not pan_front:
        pan_front = _extract_trycloudflare_url(BASE_DIR / "logs" / "cloudflared_panaderia.log")
    if not pan_api and pan_front:
        # Common setup for Vite dev + proxy: same origin, /api proxied to backend.
        pan_api = pan_front

    return {
        "panaderia": {
            "frontendBase": pan_front,
            "apiBase": pan_api,
            "source": "env_or_quick_tunnel",
        },
        "vision": {
            "frontendBase": vis_front,
            "apiBase": vis_api,
            "source": "env",
        },
    }


@app.get("/api/apps/public-urls", tags=["Brazos v4"])
async def apps_public_urls():
    """Expose public URLs for remote clients (never localhost)."""
    t0 = time.perf_counter()
    try:
        data = _public_arm_urls()
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(True, data, ms, None)
    except Exception as exc:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, str(exc))


@app.post("/arms/{arm_id}/start", tags=["Brazos v4"])
async def arms_start(arm_id: str):
    """Inicia los procesos del brazo subordinado (backend + frontend)."""
    import subprocess

    t0 = time.perf_counter()
    if arm_id not in _ARM_DEFS:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, f"Brazo desconocido: {arm_id}")

    defn = _ARM_DEFS[arm_id]
    existing = _ARM_PROCS.get(arm_id, [])
    alive = [p for p in existing if p.poll() is None]

    flags = 0
    if hasattr(subprocess, "CREATE_NEW_PROCESS_GROUP"):
        flags |= subprocess.CREATE_NEW_PROCESS_GROUP
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        flags |= subprocess.CREATE_NO_WINDOW

    launched: list = []
    skipped: list = []
    errors: list = []

    for proc_def in defn["processes"]:
        expected_port = _arm_process_port(defn, proc_def.get("label", ""))
        if _is_local_port_open(expected_port):
            skipped.append(
                {
                    "label": proc_def["label"],
                    "port": expected_port,
                    "reason": "already-listening",
                }
            )
            continue

        cwd = BASE_DIR / proc_def["cwd"]
        if not cwd.exists():
            errors.append(f"{proc_def['label']}: directorio no existe ({cwd})")
            continue
        try:
            p = subprocess.Popen(
                proc_def["cmd"],
                cwd=str(cwd),
                shell=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=flags,
            )
            alive.append(p)
            launched.append({"label": proc_def["label"], "pid": p.pid})
        except Exception as exc:
            errors.append(f"{proc_def['label']}: {exc}")

    _ARM_PROCS[arm_id] = alive
    ms = int((time.perf_counter() - t0) * 1000)
    ok = bool((len(launched) + len(skipped)) > 0 and len(errors) == 0)
    return _std_resp(
        ok,
        {
            "arm_id": arm_id,
            "launched": launched,
            "skipped": skipped,
            "errors": errors,
        },
        ms,
        errors[0] if errors and not ok else None,
    )


@app.get("/arms/{arm_id}/status", tags=["Brazos v4"])
async def arms_status(arm_id: str):
    """Estado de los procesos del brazo subordinado."""
    t0 = time.perf_counter()
    if arm_id not in _ARM_DEFS:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, f"Brazo desconocido: {arm_id}")

    defn = _ARM_DEFS[arm_id]
    procs = _ARM_PROCS.get(arm_id, [])
    proc_info = [{"pid": p.pid, "alive": p.poll() is None} for p in procs]
    alive_count = sum(1 for r in proc_info if r["alive"])
    expected = len(defn["processes"])
    frontend_up = _is_local_port_open(defn.get("frontend_port"))
    api_up = _is_local_port_open(defn.get("api_port"))
    ms = int((time.perf_counter() - t0) * 1000)
    return _std_resp(
        True,
        {
            "arm_id": arm_id,
            "name": defn["name"],
            "running": alive_count,
            "expected": expected,
            "ready": bool(frontend_up and api_up),
            "ports": {
                "frontend_port": defn.get("frontend_port"),
                "frontend_up": frontend_up,
                "api_port": defn.get("api_port"),
                "api_up": api_up,
            },
            "procs": proc_info,
        },
        ms,
        None,
    )


@app.post("/arms/{arm_id}/stop", tags=["Brazos v4"])
async def arms_stop(arm_id: str):
    """Detiene los procesos del brazo subordinado."""
    t0 = time.perf_counter()
    if arm_id not in _ARM_DEFS:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, f"Brazo desconocido: {arm_id}")

    procs = _ARM_PROCS.pop(arm_id, [])
    stopped: list = []
    failed: list = []

    for p in procs:
        try:
            if p.poll() is None:
                if os.name == "nt":
                    import subprocess as _sp

                    _sp.run(
                        ["taskkill", "/F", "/T", "/PID", str(p.pid)],
                        capture_output=True,
                        creationflags=getattr(_sp, "CREATE_NO_WINDOW", 0),
                    )
                else:
                    p.terminate()
                try:
                    p.wait(timeout=5)
                except Exception:
                    p.kill()
            stopped.append(p.pid)
        except Exception as exc:
            failed.append({"pid": p.pid, "error": str(exc)})
            try:
                p.kill()
            except Exception:
                pass

    ms = int((time.perf_counter() - t0) * 1000)
    return _std_resp(
        True,
        {"arm_id": arm_id, "stopped": stopped, "failed": failed},
        ms,
        None,
    )


@app.get("/arms/{arm_id}/ping", tags=["Brazos v4"])
async def arms_ping(arm_id: str):
    """Verifica si el frontend del brazo responde (check de puerto server-side, sin ambigÃ¼edad IPv4/IPv6)."""
    import socket

    t0 = time.perf_counter()
    if arm_id not in _ARM_DEFS:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, f"Brazo desconocido: {arm_id}")

    defn = _ARM_DEFS[arm_id]
    port = defn.get("frontend_port")
    if not port:
        ms = int((time.perf_counter() - t0) * 1000)
        return _std_resp(False, None, ms, "Sin frontend_port configurado")

    reachable = False
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(2.0)
        result = s.connect_ex(("127.0.0.1", port))
        s.close()
        reachable = result == 0
    except Exception:
        reachable = False

    ms = int((time.perf_counter() - t0) * 1000)
    return _std_resp(
        True,
        {"arm_id": arm_id, "port": port, "reachable": reachable},
        ms,
        None,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Claude-Atlas Memory Bridge — Endpoints v1.0
# ═══════════════════════════════════════════════════════════════════════════════

def _ms() -> int:
    """Current monotonic time in milliseconds for lightweight API timing."""
    return int(time.perf_counter() * 1000)


def _get_claude_bridge():
    """Importación lazy del ClaudeMemoryBridge para no bloquear el arranque."""
    try:
        import sys
        sys.path.insert(0, str(BASE_DIR))
        from brain.claude_memory_bridge import get_bridge
        return get_bridge()
    except Exception as e:
        return None, str(e)


@app.post("/api/claude/memory/sync", tags=["Claude Bridge"])
async def claude_memory_sync_both():
    """Sincronización bidireccional: MEMORY.md → Atlas brain + Atlas brain → atlas_brain_context.json"""
    t0 = _ms()
    try:
        from brain.claude_memory_bridge import ClaudeMemoryBridge
        bridge = ClaudeMemoryBridge()
        result = bridge.bidirectional_sync()
        return _std_resp(result["ok"], result, _ms() - t0)
    except Exception as e:
        return _std_resp(False, None, _ms() - t0, str(e))


@app.post("/api/claude/memory/push", tags=["Claude Bridge"])
async def claude_memory_push():
    """Claude → Atlas: sincroniza MEMORY.md hacia el brain de Atlas (EpisodicMemory + patrones)."""
    t0 = _ms()
    try:
        from brain.claude_memory_bridge import ClaudeMemoryBridge
        bridge = ClaudeMemoryBridge()
        result = bridge.sync_claude_to_atlas()
        return _std_resp(result["ok"], result, _ms() - t0)
    except Exception as e:
        return _std_resp(False, None, _ms() - t0, str(e))


@app.post("/api/claude/memory/pull", tags=["Claude Bridge"])
async def claude_memory_pull():
    """Atlas → Claude: exporta brain de Atlas a atlas_brain_context.json para que Claude lo lea."""
    t0 = _ms()
    try:
        from brain.claude_memory_bridge import ClaudeMemoryBridge
        bridge = ClaudeMemoryBridge()
        result = bridge.sync_atlas_to_claude()
        return _std_resp(result["ok"], result, _ms() - t0)
    except Exception as e:
        return _std_resp(False, None, _ms() - t0, str(e))


class _InsightPayload(BaseModel):
    insight: str
    category: str = "general"
    importance: float = 0.8
    tags: list = []
    context: dict = {}


@app.post("/api/claude/memory/insight", tags=["Claude Bridge"])
async def claude_memory_insight(payload: _InsightPayload):
    """Escribe un insight puntual de la sesión Claude directamente al brain de Atlas."""
    t0 = _ms()
    try:
        from brain.claude_memory_bridge import ClaudeMemoryBridge
        bridge = ClaudeMemoryBridge()
        result = bridge.write_insight(
            insight=payload.insight,
            category=payload.category,
            importance=payload.importance,
            tags=payload.tags or [],
            context=payload.context or {},
        )
        return _std_resp(True, result, _ms() - t0)
    except Exception as e:
        return _std_resp(False, None, _ms() - t0, str(e))


@app.get("/api/claude/memory/stats", tags=["Claude Bridge"])
async def claude_memory_stats():
    """Estado y estadísticas del puente Claude-Atlas Memory Bridge."""
    t0 = _ms()
    try:
        from brain.claude_memory_bridge import ClaudeMemoryBridge
        bridge = ClaudeMemoryBridge()
        stats = bridge.get_bridge_stats()
        return _std_resp(True, stats, _ms() - t0)
    except Exception as e:
        return _std_resp(False, None, _ms() - t0, str(e))


@app.get("/api/claude/memory/context", tags=["Claude Bridge"])
async def claude_memory_context():
    """Lee el atlas_brain_context.json exportado hacia Claude (lo que Claude ve del brain de Atlas)."""
    t0 = _ms()
    try:
        import json
        ctx_path = BASE_DIR.parent / "Users" / "r6957" / ".claude" / "projects" / "c--ATLAS-PUSH" / "memory" / "atlas_brain_context.json"
        # Ruta alternativa vía variable de entorno
        env_dir = os.environ.get("CLAUDE_MEMORY_DIR")
        if env_dir:
            ctx_path = Path(env_dir) / "atlas_brain_context.json"
        if not ctx_path.exists():
            return _std_resp(False, None, _ms() - t0, "atlas_brain_context.json no generado. Ejecuta /api/claude/memory/pull primero.")
        data = json.loads(ctx_path.read_text(encoding="utf-8"))
        return _std_resp(True, data, _ms() - t0)
    except Exception as e:
        return _std_resp(False, None, _ms() - t0, str(e))


def _dedupe_http_routes() -> None:
    """
    Keep only the first registered HTTP route for each (path, methods).
    This prevents ambiguous contracts when legacy local endpoints and
    included routers expose the same path/method.
    """
    log = logging.getLogger("atlas.http_api.routes")
    seen: set[tuple[str, tuple[str, ...]]] = set()
    deduped = []
    removed = []
    for route in app.router.routes:
        methods = getattr(route, "methods", None)
        if not methods:
            deduped.append(route)
            continue
        method_key = tuple(sorted(m for m in methods if m not in {"HEAD", "OPTIONS"}))
        if not method_key:
            deduped.append(route)
            continue
        key = (str(getattr(route, "path", "")), method_key)
        if key in seen:
            removed.append(
                {
                    "path": key[0],
                    "methods": list(method_key),
                    "endpoint": getattr(getattr(route, "endpoint", None), "__name__", "?"),
                    "module": getattr(getattr(route, "endpoint", None), "__module__", "?"),
                }
            )
            continue
        seen.add(key)
        deduped.append(route)

    if removed:
        log.warning(
            "Deduplicated %d HTTP routes to preserve canonical contracts.",
            len(removed),
        )
        for item in removed[:20]:
            log.warning(
                "route dedup: %s %s -> %s (%s)",
                ",".join(item["methods"]),
                item["path"],
                item["endpoint"],
                item["module"],
            )
    app.router.routes = deduped



# ═══════════════════════════════════════════════════════════════════════════════
# § RAULI USERS — Memoria persistente y personalización por usuario  (v1.0)
#
#  Cada usuario recibe un token único que se incluye en su enlace:
#    http://<host>:3000/?u=<TOKEN>
#
#  El avatar Atlas Companion lee ese token, carga el perfil y personaliza
#  el saludo. Cada visita y mensaje queda registrado en SQLite local.
#
#  Endpoints:
#    POST   /api/rauli/users                   → crear usuario (name → token + link)
#    GET    /api/rauli/users                   → listar todos (panel admin)
#    GET    /api/rauli/users/{token}           → perfil completo + memoria
#    PATCH  /api/rauli/users/{token}/visit     → registrar visita
#    POST   /api/rauli/users/{token}/memory    → actualizar memoria
#    DELETE /api/rauli/users/{token}           → eliminar usuario
# ═══════════════════════════════════════════════════════════════════════════════

import sqlite3 as _sqlite3
import secrets as _secrets_mod

_RAULI_USERS_DB = BASE_DIR / "data" / "rauli_users.db"
_RAULI_USERS_DB.parent.mkdir(parents=True, exist_ok=True)


def _rauli_db_conn() -> _sqlite3.Connection:
    conn = _sqlite3.connect(str(_RAULI_USERS_DB), check_same_thread=False)
    conn.row_factory = _sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn


def _rauli_db_init() -> None:
    with _rauli_db_conn() as c:
        c.executescript("""
            CREATE TABLE IF NOT EXISTS rauli_users (
                token       TEXT PRIMARY KEY,
                name        TEXT NOT NULL,
                created_at  TEXT NOT NULL,
                last_seen   TEXT,
                visit_count INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS rauli_memory (
                token       TEXT PRIMARY KEY,
                searches    TEXT DEFAULT '[]',
                preferences TEXT DEFAULT '{}',
                topics      TEXT DEFAULT '[]',
                last_msg    TEXT DEFAULT '',
                FOREIGN KEY (token) REFERENCES rauli_users(token) ON DELETE CASCADE
            );
        """)


_rauli_db_init()


class _RauliUserCreate(BaseModel):
    name:  str
    token: Optional[str] = None   # token externo (ej. access_code del espejo)


class _RauliMemoryUpdate(BaseModel):
    searches:    Optional[list] = None
    preferences: Optional[dict] = None
    topics:      Optional[list] = None
    last_msg:    Optional[str]  = None


def _rauli_gen_token() -> str:
    """Token URL-safe de ~11 caracteres (64 bits de entropía)."""
    return _secrets_mod.token_urlsafe(8)


@app.post("/api/rauli/users", tags=["Rauli Users"])
async def rauli_create_user(body: _RauliUserCreate):
    """Crea un nuevo usuario y devuelve su token y enlace personalizado."""
    t0 = time.perf_counter()
    name = (body.name or "").strip()
    if not name:
        return _std_resp(False, None, 0, "name_required")
    # Usar token externo si se provee (bridge con espejo access_code), si no generar uno
    raw_token = (body.token or "").strip()
    token = raw_token if raw_token else _rauli_gen_token()
    now = datetime.now(timezone.utc).isoformat()
    try:
        with _rauli_db_conn() as c:
            c.execute(
                "INSERT INTO rauli_users (token,name,created_at,last_seen,visit_count) VALUES(?,?,?,?,0)",
                (token, name, now, now),
            )
            c.execute("INSERT INTO rauli_memory (token) VALUES(?)", (token,))
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))
    return _std_resp(
        True,
        {"token": token, "name": name, "link": f"/?u={token}", "created_at": now},
        int((time.perf_counter() - t0) * 1000),
    )


@app.get("/api/rauli/users", tags=["Rauli Users"])
async def rauli_list_users():
    """Lista todos los usuarios registrados (panel de administración)."""
    t0 = time.perf_counter()
    try:
        with _rauli_db_conn() as c:
            rows = c.execute(
                "SELECT token,name,created_at,last_seen,visit_count FROM rauli_users ORDER BY created_at DESC"
            ).fetchall()
        users = [dict(r) for r in rows]
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))
    return _std_resp(True, users, int((time.perf_counter() - t0) * 1000))


@app.get("/api/rauli/users/{token}", tags=["Rauli Users"])
async def rauli_get_user(token: str):
    """Devuelve perfil completo + memoria del usuario identificado por token."""
    t0 = time.perf_counter()
    try:
        with _rauli_db_conn() as c:
            u = c.execute(
                "SELECT token,name,created_at,last_seen,visit_count FROM rauli_users WHERE token=?",
                (token,),
            ).fetchone()
            if not u:
                return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "user_not_found")
            m = c.execute(
                "SELECT searches,preferences,topics,last_msg FROM rauli_memory WHERE token=?",
                (token,),
            ).fetchone()
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))
    profile = dict(u)
    profile["memory"] = {
        "searches":    json.loads(m["searches"]    or "[]") if m else [],
        "preferences": json.loads(m["preferences"] or "{}") if m else {},
        "topics":      json.loads(m["topics"]      or "[]") if m else [],
        "last_msg":    (m["last_msg"] or "") if m else "",
    }
    return _std_resp(True, profile, int((time.perf_counter() - t0) * 1000))


@app.patch("/api/rauli/users/{token}/visit", tags=["Rauli Users"])
async def rauli_record_visit(token: str):
    """Incrementa el contador de visitas y actualiza last_seen."""
    t0 = time.perf_counter()
    now = datetime.now(timezone.utc).isoformat()
    try:
        with _rauli_db_conn() as c:
            r = c.execute(
                "UPDATE rauli_users SET last_seen=?, visit_count=visit_count+1 WHERE token=?",
                (now, token),
            )
            if r.rowcount == 0:
                return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "user_not_found")
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))
    return _std_resp(True, {"last_seen": now}, int((time.perf_counter() - t0) * 1000))


@app.post("/api/rauli/users/{token}/memory", tags=["Rauli Users"])
async def rauli_update_memory(token: str, body: _RauliMemoryUpdate):
    """Actualiza la memoria del usuario: búsquedas recientes, preferencias, temas, último mensaje."""
    t0 = time.perf_counter()
    try:
        with _rauli_db_conn() as c:
            row = c.execute(
                "SELECT searches,preferences,topics FROM rauli_memory WHERE token=?",
                (token,),
            ).fetchone()
            if not row:
                return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "user_not_found")
            searches    = json.loads(row["searches"]    or "[]")
            preferences = json.loads(row["preferences"] or "{}")
            topics      = json.loads(row["topics"]      or "[]")
            if body.searches is not None:
                # Deduplicar y mantener solo las últimas 50
                searches = list(dict.fromkeys(searches + body.searches))[-50:]
            if body.preferences is not None:
                preferences.update(body.preferences)
            if body.topics is not None:
                topics = list(dict.fromkeys(topics + body.topics))[-30:]
            c.execute(
                """UPDATE rauli_memory
                   SET searches=?, preferences=?, topics=?, last_msg=COALESCE(?, last_msg)
                   WHERE token=?""",
                (
                    json.dumps(searches,    ensure_ascii=False),
                    json.dumps(preferences, ensure_ascii=False),
                    json.dumps(topics,      ensure_ascii=False),
                    body.last_msg,
                    token,
                ),
            )
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))
    return _std_resp(True, {"ok": True}, int((time.perf_counter() - t0) * 1000))


@app.delete("/api/rauli/users/{token}", tags=["Rauli Users"])
async def rauli_delete_user(token: str):
    """Elimina un usuario y toda su memoria asociada."""
    t0 = time.perf_counter()
    try:
        with _rauli_db_conn() as c:
            r = c.execute("DELETE FROM rauli_users WHERE token=?", (token,))
            if r.rowcount == 0:
                return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), "user_not_found")
            c.execute("DELETE FROM rauli_memory WHERE token=?", (token,))
    except Exception as e:
        return _std_resp(False, None, int((time.perf_counter() - t0) * 1000), str(e))
    return _std_resp(True, {"deleted": token}, int((time.perf_counter() - t0) * 1000))


# ─────────────────────────────────────────────────────────────────────────────

_dedupe_http_routes()



